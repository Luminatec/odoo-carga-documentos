"""
Carga de documentos → Odoo
Luminatec / GPowerByte
v3 — Auth por email/contraseña + API key de servicio central
"""

import streamlit as st
import xmlrpc.client
import base64
import re
from io import BytesIO
from datetime import datetime as _dt_now

import pandas as pd

st.set_page_config(
    page_title="Luminatec · Odoo",
    page_icon="🛒",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
  /* ── Variables de marca Luminatec ─────────────────────────── */
  :root {
    --lumi-red:    #FD0029;
    --lumi-red-dk: #C21F34;
    --lumi-gold:   #F2C800;
    --lumi-dark:   #111111;
    --lumi-gray:   #F7F7F7;
    --lumi-border: #E0E0E0;
  }

  /* ── Sidebar ──────────────────────────────────────────────── */
  [data-testid="stSidebar"] {
    background: var(--lumi-dark) !important;
    border-right: 3px solid var(--lumi-red) !important;
  }
  [data-testid="stSidebar"] label,
  [data-testid="stSidebar"] .stMarkdown,
  [data-testid="stSidebar"] p,
  [data-testid="stSidebar"] small,
  [data-testid="stSidebar"] span,
  [data-testid="stSidebar"] .stCaption { color: #dddddd !important; }
  [data-testid="stSidebar"] h1,
  [data-testid="stSidebar"] h2,
  [data-testid="stSidebar"] h3 { color: #ffffff !important; }
  [data-testid="stSidebar"] .stTextInput input {
    background: #1e1e1e !important; color: #eee !important;
    border-color: #444 !important; border-radius: 6px !important;
  }
  [data-testid="stSidebar"] .stTextInput input:focus {
    border-color: var(--lumi-red) !important;
    box-shadow: 0 0 0 2px rgba(253,0,41,0.25) !important;
  }
  [data-testid="stSidebar"] .stButton button,
  [data-testid="stSidebar"] .stFormSubmitButton button {
    background: var(--lumi-red) !important; color: white !important;
    border: none !important; font-weight: 700 !important;
    border-radius: 6px !important; letter-spacing: 0.3px;
    transition: background 0.15s ease !important;
  }
  [data-testid="stSidebar"] .stButton button:hover,
  [data-testid="stSidebar"] .stFormSubmitButton button:hover {
    background: var(--lumi-red-dk) !important;
  }
  [data-testid="stSidebar"] hr { border-color: #333 !important; }

  /* ── Logo / branding sidebar ──────────────────────────────── */
  .lumi-sidebar-logo {
    display: flex; align-items: center; gap: 10px;
    padding: 8px 0 14px 0;
  }
  .lumi-logo-text {
    font-size: 1.6rem; font-weight: 900;
    color: var(--lumi-red) !important; letter-spacing: -1px;
  }
  .lumi-logo-dot { color: var(--lumi-gold) !important; font-size: 1.8rem; }

  /* ── Título principal ─────────────────────────────────────── */
  .main-title {
    font-size: 1.85rem; font-weight: 900;
    color: var(--lumi-red); letter-spacing: -1px;
    border-bottom: 3px solid var(--lumi-gold);
    padding-bottom: 6px; margin-bottom: 4px;
  }
  .main-title span { color: var(--lumi-gold); }

  /* ── Tabs ─────────────────────────────────────────────────── */
  .stTabs [data-baseweb="tab-list"] {
    gap: 4px;
    border-bottom: 2px solid var(--lumi-border) !important;
  }
  .stTabs [data-baseweb="tab"] {
    font-weight: 600; font-size: 0.88rem;
    border-radius: 6px 6px 0 0 !important;
    padding: 8px 16px !important;
    color: #555 !important;
  }
  .stTabs [aria-selected="true"] {
    background: var(--lumi-red) !important;
    color: white !important;
    border-bottom: 2px solid var(--lumi-red) !important;
  }

  /* ── File uploader ────────────────────────────────────────── */
  [data-testid="stFileUploader"] {
    border: 2px dashed rgba(253,0,41,0.3) !important;
    border-radius: 10px !important;
    padding: 8px !important;
    background: rgba(253,0,41,0.02) !important;
  }
  [data-testid="stFileUploader"]:hover {
    border-color: var(--lumi-red) !important;
  }

  /* ── Chips / badges ───────────────────────────────────────── */
  .admin-badge {
    display: inline-block;
    background: var(--lumi-gold); color: #111;
    font-size: 0.68rem; font-weight: 800;
    padding: 3px 10px; border-radius: 10px;
    text-transform: uppercase; letter-spacing: 0.6px;
  }
  .user-chip {
    display: inline-block;
    background: #1e1e1e; color: #eee;
    font-size: 0.78rem; padding: 4px 12px;
    border-radius: 20px; border: 1px solid #444;
    margin-bottom: 4px;
  }

  /* ── Métricas ─────────────────────────────────────────────── */
  [data-testid="stMetric"] {
    background: var(--lumi-gray);
    border-radius: 8px; padding: 10px 14px;
    border-left: 4px solid var(--lumi-red);
  }
  [data-testid="stMetricValue"] { color: var(--lumi-red) !important; font-weight: 800 !important; }

  /* ── Botones primarios (fuera del sidebar) ────────────────── */
  .stButton > button[kind="primary"],
  .stFormSubmitButton > button {
    background: var(--lumi-red) !important;
    color: white !important; font-weight: 700 !important;
    border: none !important; border-radius: 6px !important;
  }
  .stButton > button[kind="primary"]:hover,
  .stFormSubmitButton > button:hover {
    background: var(--lumi-red-dk) !important;
  }

  /* ── Expanders ────────────────────────────────────────────── */
  [data-testid="stExpander"] summary {
    font-weight: 600; color: var(--lumi-red);
  }
</style>
""", unsafe_allow_html=True)

ODOO_URL = "https://gpowerbyte-luminatec.odoo.com"
ODOO_DB  = "gpowerbyte-luminatec-master-22753148"
TEST_ODOO_URL = "https://gpowerbyte-luminatec-test-31645353.dev.odoo.com"
TEST_ODOO_DB  = "gpowerbyte-luminatec-test-31645353"

# Emails con acceso a Importaciones.
# BASE_ADMIN_EMAILS siempre tienen acceso, independientemente del secret.
# Se puede agregar más emails via st.secrets["ADMIN_EMAILS"] (separados por coma).
BASE_ADMIN_EMAILS = {"ivarela@luminatec.com", "dario@luminatec.com", "dalonso@luminatec.com"}
_raw_admin = st.secrets.get("ADMIN_EMAILS", "")
ADMIN_EMAILS = BASE_ADMIN_EMAILS | {e.strip().lower() for e in _raw_admin.split(",") if e.strip()}

MIMETYPES = {
    "pdf":  "application/pdf",
    "jpg":  "image/jpeg",
    "jpeg": "image/jpeg",
    "png":  "image/png",
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "xls":  "application/vnd.ms-excel",
}

LC_PRODUCTS = {
    20241: "CMV - Agente de carga",
    20242: "CMV - Honorarios despachante",
    20243: "CMV - Otros",
    20244: "CMV - Terminal portuaria",
    20245: "CMV - Tasas y Derechos",
    20276: "Despacho Cta Transitoria 21%",
    20277: "Despacho Cta Transitoria 10,5%",
}

ETAPAS_DEF = [
    ("0",    "Etapa 0",    "OC PETDUR confirmada y bloqueada"),
    ("1",    "Etapa 1",    "Bill PETDUR posted (DR 3283 / CR Prov USD)"),
    ("2",    "Etapa 2",    "DI AFIP posted + OP automática BA#26"),
    ("2a",   "Etapa 2a",   "Bills de nacionalización (TRICE, T4, MUNDO COMEX...)"),
    ("2.5",  "Etapa 2.5",  "Reclasificación Tránsito si TC factura ≠ TC despacho"),
    ("3",    "Etapa 3",    "Picking IN validado → WH/PreIngreso"),
    ("T3.5", "Etapa T3.5", "Reclasificación Tránsito → Cuenta Puente"),
    ("4",    "Etapa 4 Bis","Landed Cost validado (by_current_cost_price)"),
    ("5",    "Etapa 5",    "Internal Transfer → WH/Disponible"),
    ("6",    "Etapa 6",    "Acta CFO firmada — 15 checks Decálogo"),
]

DECALOGO = [
    "Cuenta Puente Recepciones (3284) cohorte = $0",
    "Mercadería en Tránsito (3283) cohorte = $0",
    "BAs #8 y #11 sin línea standard_price (verificado hoy)",
    "Productos LC 20241-20245 con expense_id = 3284",
    "Productos LC 20276-20277 con expense_id = 3284",
    "Todos los LCs usaron split_method = by_current_cost_price",
    "NO se usó by_quantity en ningún LC",
    "NO se crearon asientos correctivos preventivos",
    "Costo USD/u del lote coincide con modelo CFO (delta < $0.10 USD/u)",
    "WAC ponderado verificado en libro mayor estándar Odoo UI",
    "Suppliers AFIP residual = deuda real del DI",
    "BA #23 actualizó x_studio_ppp al validar el Landed Cost",
    "Picking IN en estado Done / WH/PreIngreso confirmado",
    "Internal Transfer ejecutado → stock en WH/Disponible",
    "Referencia LUMI_XXX en todos los asientos de la cohorte",
]


# ───────────────────────────────────────────────────
# PROXY DE MODELOS (stateless, compartido entre usuarios)
# ───────────────────────────────────────────────────
def get_models_proxy():
    """ServerProxy para account.move, etc. Es stateless — uid y password van por llamada."""
    return xmlrpc.client.ServerProxy(f"{ODOO_URL}/xmlrpc/2/object", allow_none=True)

def odoo_authenticate(email: str, password: str):
    """
    Autentica al usuario contra Odoo XML-RPC.
    Devuelve (uid, "") si OK, (None, mensaje_error) si falla.
    """
    try:
        common = xmlrpc.client.ServerProxy(f"{ODOO_URL}/xmlrpc/2/common", allow_none=True)
        uid = common.authenticate(ODOO_DB, email.strip().lower(), password, {})
        if uid:
            return uid, ""
        return None, "Email o contraseña incorrectos."
    except Exception as e:
        return None, f"No se pudo conectar a Odoo: {e}"

def verify_user(email, password):
    """
    Verifica email + contraseña.
    Prioridad:
      1. Odoo JSON-RPC /web/session/authenticate (contraseña web normal, sin API key)
      2. APP_USERS  = "email1:clave1,email2:clave2"  (fallback manual en secrets)
      3. APP_PASSWORD = "clave_compartida"            (fallback global)
    """
    import hashlib, urllib.request, json as _json
    email = email.strip().lower()

    # ── 1. Autenticación via Odoo JSON-RPC (misma clave que el navegador) ────
    try:
        _payload = _json.dumps({
            "jsonrpc": "2.0", "method": "call", "id": 1,
            "params": {"db": ODOO_DB, "login": email, "password": password}
        }).encode()
        _req = urllib.request.Request(
            f"{ODOO_URL}/web/session/authenticate",
            data=_payload,
            headers={"Content-Type": "application/json"},
            method="POST"
        )
        with urllib.request.urlopen(_req, timeout=8) as _resp:
            _data = _json.loads(_resp.read())
        _uid = (_data.get("result") or {}).get("uid")
        if _uid:
            return True, ""
    except Exception:
        pass  # si Odoo no responde, cae al fallback

    # ── 2. APP_USERS en secrets (fallback) ───────────────────────────────────
    pw_hash = hashlib.sha256(password.encode()).hexdigest()
    app_users_raw = st.secrets.get("APP_USERS", "")
    if app_users_raw:
        user_map = {}
        for entry in app_users_raw.split(","):
            parts = entry.strip().split(":", 1)
            if len(parts) == 2:
                user_map[parts[0].strip().lower()] = parts[1].strip()
        if email in user_map:
            expected = user_map[email]
            if password == expected or pw_hash == expected:
                return True, ""
            return False, "Contraseña incorrecta."
        return False, "Email no autorizado o contraseña incorrecta."

    # ── 3. APP_PASSWORD (clave global compartida) ────────────────────────────
    app_password = st.secrets.get("APP_PASSWORD", "")
    if app_password:
        if password == app_password or pw_hash == app_password:
            return True, ""
        return False, "Contraseña incorrecta."

    return False, "Contraseña incorrecta."

def call(models, uid, api_key, model, method, args, kw=None):
    return models.execute_kw(ODOO_DB, uid, api_key, model, method, args, kw or {})

@st.cache_data(ttl=300, show_spinner=False)
def search_partners(models_url, uid, api_key, name, limit=8):
    m = xmlrpc.client.ServerProxy(models_url, allow_none=True)
    rows = m.execute_kw(ODOO_DB, uid, api_key, "res.partner", "search_read",
        [[("name", "ilike", name), ("active", "=", True)]],
        {"fields": ["id", "name"], "limit": limit, "order": "name asc"})
    return [(r["id"], r["name"]) for r in rows]

@st.cache_data(ttl=600, show_spinner=False)
def get_all_accounts(models_url, uid, api_key):
    """Carga todas las cuentas contables activas de Odoo (cacheado 10 min)."""
    try:
        m = xmlrpc.client.ServerProxy(models_url, allow_none=True)
        rows = m.execute_kw(ODOO_DB, uid, api_key, "account.account", "search_read",
            [[("deprecated", "=", False)]],
            {"fields": ["id", "code", "name"], "order": "code asc"})
        return [(r["id"], f"{r['code']}  {r['name']}") for r in rows]
    except Exception:
        return []

@st.cache_data(ttl=120, show_spinner=False)
def get_partner_default_account(models_url, uid, api_key, partner_id):
    """Devuelve (account_id, 'CODE  Name') de la cuenta de gasto por defecto del proveedor.
    Estrategia:
      1. product.supplierinfo donde partner_id -> product_tmpl_id
         -> property_account_expense_id del product.template (pestaña Contabilidad en Odoo)
      2. Fallback: account de la categoria del producto (product.category)
    Solo usa cuentas configuradas en el producto — NO recurre a facturas históricas.
    Retorna None si no se puede determinar."""
    try:
        m = xmlrpc.client.ServerProxy(models_url, allow_none=True)

        # --- Estrategia 1: producto del proveedor → cuenta de gasto del producto ---
        sinfo = m.execute_kw(ODOO_DB, uid, api_key, "product.supplierinfo", "search_read",
            [[("partner_id", "=", partner_id)]],
            {"fields": ["product_tmpl_id"], "limit": 5, "order": "id asc"})

        for si in (sinfo or []):
            tmpl_id = (si.get("product_tmpl_id") or [None])[0]
            if not tmpl_id:
                continue
            # Leer cuenta de gasto Y categoría (para fallback)
            tmpls = m.execute_kw(ODOO_DB, uid, api_key, "product.template", "read",
                [[tmpl_id]],
                {"fields": ["property_account_expense_id", "categ_id"]})
            if not tmpls:
                continue
            tmpl = tmpls[0]

            # property_account_expense_id (pestaña Contabilidad del producto)
            acct = tmpl.get("property_account_expense_id")
            if acct and isinstance(acct, (list, tuple)) and acct[0]:
                accts = m.execute_kw(ODOO_DB, uid, api_key, "account.account", "read",
                    [[acct[0]]], {"fields": ["code", "name", "deprecated"]})
                if accts and not accts[0].get("deprecated"):
                    return (accts[0]["id"], f"{accts[0]['code']}  {accts[0]['name']}")

            # Fallback: cuenta de gasto de la categoria del producto
            categ = tmpl.get("categ_id")
            categ_id = (categ[0] if isinstance(categ, (list, tuple)) else categ) if categ else None
            if categ_id:
                cats = m.execute_kw(ODOO_DB, uid, api_key, "product.category", "read",
                    [[categ_id]],
                    {"fields": ["property_account_expense_categ_id"]})
                if cats:
                    categ_acct = cats[0].get("property_account_expense_categ_id")
                    if categ_acct and isinstance(categ_acct, (list, tuple)) and categ_acct[0]:
                        accts = m.execute_kw(ODOO_DB, uid, api_key, "account.account", "read",
                            [[categ_acct[0]]], {"fields": ["code", "name", "deprecated"]})
                        if accts and not accts[0].get("deprecated"):
                            return (accts[0]["id"], f"{accts[0]['code']}  {accts[0]['name']}")
    except Exception:
        pass
    return None



@st.cache_data(ttl=300, show_spinner=False)
def get_expense_products(models_url, uid, api_key):
    """Carga variantes de productos activos y aptos para compra (product.product → IDs válidos para lineas de factura)."""
    try:
        m = xmlrpc.client.ServerProxy(models_url, allow_none=True)
        rows = m.execute_kw(ODOO_DB, uid, api_key, "product.product", "search_read",
            [[("active", "=", True), ("purchase_ok", "=", True), ("type", "=", "service")]],
            {"fields": ["id", "name", "default_code"], "order": "name asc", "limit": 500})
        result = []
        for r in rows:
            code = r.get("default_code") or ""
            label = f"[{code}] {r['name']}" if code else r["name"]
            result.append((r["id"], label))
        return result
    except Exception:
        return []

@st.cache_data(ttl=120, show_spinner=False)
def get_partner_default_product(models_url, uid, api_key, partner_id):
    """Devuelve (product_product_id, label) del primer producto configurado para el proveedor."""
    try:
        m = xmlrpc.client.ServerProxy(models_url, allow_none=True)
        # Intentar campo product_id (variante específica) primero
        sinfo = m.execute_kw(ODOO_DB, uid, api_key, "product.supplierinfo", "search_read",
            [[("partner_id", "=", partner_id)]],
            {"fields": ["product_tmpl_id", "product_id"], "limit": 1, "order": "id asc"})
        if sinfo:
            row = sinfo[0]
            # product_id en supplierinfo es Many2one a product.product (puede ser False)
            prod_var = row.get("product_id")
            if prod_var and isinstance(prod_var, (list, tuple)) and prod_var[0]:
                return (prod_var[0], prod_var[1])
            # Sino buscar primera variante activa del template
            tmpl = row.get("product_tmpl_id")
            if tmpl and isinstance(tmpl, (list, tuple)) and tmpl[0]:
                variants = m.execute_kw(ODOO_DB, uid, api_key, "product.product", "search_read",
                    [[("product_tmpl_id", "=", tmpl[0]), ("active", "=", True)]],
                    {"fields": ["id", "name"], "limit": 1})
                if variants:
                    return (variants[0]["id"], variants[0]["name"])
    except Exception:
        pass
    return None

@st.cache_data(ttl=300, show_spinner=False)
def get_analytic_accounts(models_url, uid, api_key):
    """Carga cuentas analíticas activas (Centros de Costo) de Odoo."""
    try:
        m = xmlrpc.client.ServerProxy(models_url, allow_none=True)
        rows = m.execute_kw(ODOO_DB, uid, api_key, "account.analytic.account", "search_read",
            [[("active", "=", True)]],
            {"fields": ["id", "name", "code"], "order": "name asc"})
        result = []
        for r in rows:
            label = f"{r['code']}  {r['name']}" if r.get("code") else r["name"]
            result.append((r["id"], label))
        return result
    except Exception:
        return []

@st.cache_data(ttl=3600, show_spinner=False)
def get_currency_id(models_url, uid, api_key, name):
    """Retorna el ID de la moneda por nombre (ej: 'USD', 'ARS')."""
    try:
        m = xmlrpc.client.ServerProxy(models_url, allow_none=True)
        rows = m.execute_kw(ODOO_DB, uid, api_key, "res.currency", "search_read",
            [[("name", "=", name)]],
            {"fields": ["id", "name"], "limit": 1})
        return rows[0]["id"] if rows else None
    except Exception:
        return None

@st.cache_data(ttl=120, show_spinner=False)
def search_purchase_orders(models_url, uid, api_key, query):
    m = xmlrpc.client.ServerProxy(models_url, allow_none=True)
    rows = m.execute_kw(ODOO_DB, uid, api_key, "purchase.order", "search_read",
        [[("name", "ilike", query), ("state", "in", ["purchase", "done"])]],
        {"fields": ["id", "name", "partner_id", "date_order", "amount_total"], "limit": 10})
    return rows

@st.cache_data(ttl=60, show_spinner=False)
def get_pickings_for_po(models_url, uid, api_key, po_id):
    m = xmlrpc.client.ServerProxy(models_url, allow_none=True)
    rows = m.execute_kw(ODOO_DB, uid, api_key, "stock.picking", "search_read",
        [[("purchase_id", "=", po_id), ("state", "!=", "cancel")]],
        {"fields": ["id", "name", "state", "location_dest_id"], "limit": 10})
    return rows

@st.cache_data(ttl=60, show_spinner=False)
def get_bills_for_carpeta(models_url, uid, api_key, carpeta_ref):
    m = xmlrpc.client.ServerProxy(models_url, allow_none=True)
    rows = m.execute_kw(ODOO_DB, uid, api_key, "account.move", "search_read",
        [[("move_type", "=", "in_invoice"), ("ref", "ilike", carpeta_ref), ("state", "!=", "cancel")]],
        {"fields": ["id", "name", "partner_id", "invoice_date", "amount_total", "state", "journal_id"], "limit": 50})
    return rows

def attach_file(models, uid, api_key, res_model, res_id, filename, file_bytes, mimetype):
    call(models, uid, api_key, "ir.attachment", "create", [{
        "name": filename, "res_model": res_model, "res_id": res_id,
        "datas": base64.b64encode(file_bytes).decode(), "mimetype": mimetype,
    }])


def create_purchase_order_petdur(models, uid, api_key, carpeta_id,
                                  currency_id=None, tc_usd=None,
                                  filename=None, file_bytes=None, mimetype=None,
                                  lineas=None):
    """
    Crea OC PETDUR en draft (purchase.order).
    currency_id: pasar SOLO si se tiene el ID real de USD (no un fallback arbitrario).
    tc_usd: cotización ARS/USD del día (x_studio_cotizacion_dolar) — requerido por Odoo.
    lineas: lista de {"descripcion", "cantidad", "precio_unit"} del documento PETDUR.
    Retorna el ID del purchase.order.
    """
    import datetime as _dt_po
    po_vals = {
        "partner_id":  49328,       # PETDUR CORPORATION S.A.
        "partner_ref": carpeta_id,  # clave de búsqueda en load_carpeta_full
    }
    if currency_id:
        po_vals["currency_id"] = currency_id
    if tc_usd and float(tc_usd) > 0:
        po_vals["x_studio_cotizacion_dolar"] = float(tc_usd)
    po_id = call(models, uid, api_key, "purchase.order", "create", [po_vals])

    # Agregar líneas al pedido si vienen del documento PETDUR
    if lineas:
        _today_po = _dt_po.date.today().isoformat()
        for ln in lineas:
            try:
                desc  = str(ln.get("descripcion") or "").strip() or carpeta_id
                qty   = float(ln.get("cantidad")   or 1)
                price = float(ln.get("precio_unit") or 0)
                call(models, uid, api_key, "purchase.order.line", "create", [{
                    "order_id":     po_id,
                    "name":         desc,
                    "product_qty":  qty,
                    "price_unit":   price,
                    "date_planned": _today_po,
                    "product_uom":  1,          # UdM = Unidad (ID 1 en Odoo estándar)
                }])
            except Exception:
                pass  # si falla una línea, seguir con las demás

    if filename and file_bytes:
        try:
            attach_file(models, uid, api_key, "purchase.order", po_id, filename, file_bytes, mimetype)
        except Exception:
            pass
    return po_id


# ── Mapeo partner_id → tipo para facturas de importación ────────────────────
PARTNER_TO_TIPO = {
    49328: {"tipo": "petdur", "etapa": "1",  "label": "PETDUR"},
    9:     {"tipo": "di_afip","etapa": "2",  "label": "AFIP"},
    48825: {"tipo": "nac",    "etapa": "2a", "label": "TRICE"},
    48828: {"tipo": "nac",    "etapa": "2a", "label": "Terminal 4"},
    48826: {"tipo": "nac",    "etapa": "2a", "label": "Mundo Comex"},
    48827: {"tipo": "nac",    "etapa": "2a", "label": "SENASA"},
}

# ── CUIT → partner (sin guiones, 11 dígitos) ────────────────────────────────
CUIT_TO_PARTNER = {
    "30711100314": {"tipo":"nac",    "label":"Bill TRICE Transport (Etapa 2a)", "partner_id":48825,"journal_id":10,"doc_type":None},
    "30717845419": {"tipo":"nac",    "label":"Bill Mundo Comex (Etapa 2a)",     "partner_id":48826,"journal_id":10,"doc_type":None},
    "30678196165": {"tipo":"nac",    "label":"Bill Terminal 4 SA (Etapa 2a)",   "partner_id":48828,"journal_id":10,"doc_type":None},
    "33546700939": {"tipo":"nac",    "label":"Bill SENASA (Etapa 2a)",          "partner_id":48827,"journal_id":10,"doc_type":None},
}

def _parse_odoo_rate(r):
    """
    Extrae el TC ARS/USD de un registro res.currency.rate.
    Odoo almacena la tasa de distintas formas según versión y localización:
      - inverse_company_rate: ARS por 1 USD (ej: 1417.0)  ← Odoo 16+
      - company_rate:         ARS por 1 USD (ej: 1417.0)  ← alternativo
      - rate:                 USD por 1 ARS (ej: 0.000706) ← inverso
    Siempre validamos que el resultado sea > 100 para descartar defaults de 1.0.
    """
    for field in ("inverse_company_rate", "company_rate"):
        v = r.get(field)
        if v and v is not False:
            fv = float(v)
            if fv > 100:
                return fv
            if 0 < fv < 0.01:          # almacenado como su propio inverso
                return 1.0 / fv
    # rate suele ser 1/TC (USD por ARS)
    rate = r.get("rate")
    if rate and rate is not False:
        fv = float(rate)
        if fv > 100:                    # ya está en ARS/USD directamente
            return fv
        if 0 < fv < 0.01:              # correcto: 1/TC ≈ 0.000706
            return 1.0 / fv
    return None

@st.cache_data(ttl=300, show_spinner=False)
def get_usd_rate_odoo(models_url, uid, api_key, date_str):
    """TC ARS/USD del día o el más reciente anterior en Odoo."""
    try:
        m = xmlrpc.client.ServerProxy(models_url, allow_none=True)
        rows = m.execute_kw(ODOO_DB, uid, api_key, "res.currency.rate", "search_read",
            [[("currency_id.name", "=", "USD"), ("name", "<=", date_str)]],
            {"fields": ["name", "rate", "inverse_company_rate", "company_rate"],
             "limit": 1, "order": "name desc"})
        if rows:
            tc = _parse_odoo_rate(rows[0])
            if tc:
                return tc, rows[0]["name"]
    except Exception:
        pass
    return None, None

@st.cache_data(ttl=120, show_spinner=False)
def get_po_lines(models_url, uid, api_key, po_id):
    """Líneas de una OC con productos, cantidades y precios."""
    try:
        m = xmlrpc.client.ServerProxy(models_url, allow_none=True)
        return m.execute_kw(ODOO_DB, uid, api_key, "purchase.order.line", "search_read",
            [[("order_id", "=", po_id), ("state", "!=", "cancel")]],
            {"fields": ["product_id", "product_qty", "price_unit", "price_subtotal", "name"]})
    except Exception:
        return []

@st.cache_data(ttl=120, show_spinner=False)
def load_carpeta_full(models_url, uid, api_key, carpeta_id):
    """
    Carga bills, OC, pickings y detecta etapas de una carpeta desde Odoo.
    Busca la OC por el campo partner_ref (Referencia de proveedor).
    """
    result = {"bills": [], "po": None, "pickings": [], "lc_ids": [],
              "stages": {}, "tc_oc": None, "error": None}
    try:
        m = xmlrpc.client.ServerProxy(models_url, allow_none=True)

        # 1. OC por partner_ref (campo "Referencia de proveedor")
        po_fields_ext = ["id", "name", "partner_id", "amount_total", "currency_id",
                         "state", "partner_ref", "x_studio_cotizacion_dolar"]
        po_fields_safe = ["id", "name", "partner_id", "amount_total", "currency_id",
                          "state", "partner_ref"]
        # Incluir draft para OCs auto-creadas que no pudieron confirmarse
        _po_states = ["draft", "sent", "purchase", "done"]
        try:
            pos = m.execute_kw(ODOO_DB, uid, api_key, "purchase.order", "search_read",
                [[("partner_ref", "ilike", carpeta_id), ("state", "in", _po_states)]],
                {"fields": po_fields_ext, "limit": 5})
        except Exception:
            pos = m.execute_kw(ODOO_DB, uid, api_key, "purchase.order", "search_read",
                [[("partner_ref", "ilike", carpeta_id), ("state", "in", _po_states)]],
                {"fields": po_fields_safe, "limit": 5})
        po = pos[0] if pos else None
        result["po"] = po

        # TC desde campo custom de la OC ("Cotización dólar")
        if po:
            tc_raw = po.get("x_studio_cotizacion_dolar")
            if tc_raw and float(tc_raw or 0) > 100:
                result["tc_oc"] = float(tc_raw)

        # 2. Bills por ref (carpeta_id como substring — incluye LUMI_291A etc.)
        bill_fields = ["id", "name", "ref", "partner_id", "invoice_date", "amount_total",
                       "amount_total_signed", "currency_id", "state", "invoice_currency_rate"]
        bills = m.execute_kw(ODOO_DB, uid, api_key, "account.move", "search_read",
            [[("move_type", "=", "in_invoice"), ("ref", "ilike", carpeta_id),
              ("state", "!=", "cancel")]],
            {"fields": bill_fields, "limit": 50})
        result["bills"] = bills

        # 3. Pickings vinculados a la OC
        pickings = []
        if po:
            pickings = m.execute_kw(ODOO_DB, uid, api_key, "stock.picking", "search_read",
                [[("purchase_id", "=", po["id"]), ("picking_type_code", "=", "incoming")]],
                {"fields": ["id", "name", "state", "date_done", "location_dest_id"]})
        result["pickings"] = pickings

        # 4. Landed Costs
        lc_ids = []
        if pickings:
            lcs = m.execute_kw(ODOO_DB, uid, api_key, "stock.landed.cost", "search_read",
                [[("picking_ids", "in", [p["id"] for p in pickings])]],
                {"fields": ["id", "name", "state"], "limit": 5})
            lc_ids = [lc["id"] for lc in lcs]
        result["lc_ids"] = lc_ids

        # 5. Detectar etapas automáticamente
        partner_ids = {b["partner_id"][0] for b in bills if b.get("partner_id")}
        stages = {k: False for k, *_ in ETAPAS_DEF}
        stages["0"]  = bool(po)
        stages["1"]  = 49328 in partner_ids
        stages["2"]  = 9 in partner_ids
        stages["2a"] = bool({48825, 48826, 48827, 48828} & partner_ids)
        stages["3"]  = any(p.get("state") == "done" for p in pickings)
        stages["4"]  = bool(lc_ids)
        result["stages"] = stages

    except Exception as e:
        result["error"] = str(e)
    return result


def parse_petdur_invoice_lines(text):
    """
    Parsea líneas de producto de una factura PETDUR (e-Ticket Uruguay).
    Formato por línea: Nro COD Descripcion Cantidad u Precio Monto
    Retorna lista de dicts: descripcion, cantidad, precio_unit, monto.
    """
    def _num(s):
        if "," in s:
            parts = s.rsplit(",", 1)
            return float(parts[0].replace(".", "").replace(",", "") + "." + parts[1])
        return float(s.replace(".", ""))

    lines = []
    pattern = re.compile(
        r"^\s*\d+\s+(.+?)\s+([\d.,]+)\s+u\s+([\d.,]+)\s+([\d.,]+)\s*$",
        re.MULTILINE)
    for m in pattern.finditer(text):
        desc_raw, cant_s, precio_s, monto_s = m.groups()
        desc_raw = desc_raw.strip()
        # Eliminar prefijo duplicado: "A10 PRO A10 PRO" → "A10 PRO"
        words = desc_raw.split()
        mid = len(words) // 2
        if mid >= 1 and words[:mid] == words[mid:mid + mid]:
            desc_raw = " ".join(words[mid:])
        try:
            lines.append({
                "descripcion": desc_raw,
                "cantidad":    _num(cant_s),
                "precio_unit": _num(precio_s),
                "monto":       _num(monto_s),
            })
        except Exception:
            pass
    return lines


def get_bill_lines(models_url, uid, api_key, bill_ids):
    """Trae líneas de producto de una lista de facturas. Retorna dict {bill_id: [lineas]}."""
    if not bill_ids:
        return {}
    try:
        m = xmlrpc.client.ServerProxy(models_url, allow_none=True)
        lines = m.execute_kw(
            ODOO_DB, uid, api_key, "account.move.line", "search_read",
            [[("move_id", "in", bill_ids),
              ("display_type", "=", False)]],
            {"fields": ["id", "move_id", "product_id", "name",
                        "quantity", "price_unit", "price_subtotal",
                        "price_total", "tax_ids"],
             "limit": 500})
        result = {}
        for ln in lines:
            mid = ln["move_id"][0] if isinstance(ln["move_id"], (list, tuple)) else ln["move_id"]
            result.setdefault(mid, []).append(ln)
        return result
    except Exception:
        return {}


def get_po_lines(models_url, uid, api_key, po_id):
    """Trae líneas de una OC (purchase.order.line). Retorna lista de dicts."""
    if not po_id:
        return []
    try:
        m = xmlrpc.client.ServerProxy(models_url, allow_none=True)
        lines = m.execute_kw(
            ODOO_DB, uid, api_key, "purchase.order.line", "search_read",
            [[("order_id", "=", po_id), ("product_id", "!=", False)]],
            {"fields": ["id", "product_id", "name",
                        "product_qty", "price_unit", "price_subtotal"],
             "limit": 200})
        return lines
    except Exception:
        return []


def _bill_currency(b):
    """Devuelve el nombre de moneda del bill ('ARS', 'USD', etc.)"""
    c = b.get("currency_id")
    if c and isinstance(c, (list, tuple)) and len(c) > 1:
        return str(c[1])
    return "ARS"

def _bill_ars_amount(b):
    """Monto del bill en ARS usando amount_total_signed (siempre ARS en Odoo)."""
    return abs(float(b.get("amount_total_signed") or b.get("amount_total") or 0))

def _calc_cost_breakdown(po_lines, bills, tc_usd):
    """
    Calcula costo estimado por producto en USD y ARS.
    Fuentes de costo:
      - FOB: líneas de la OC (precio del proveedor extranjero, en USD)
      - Nac: bills de TRICE/T4/MundoComex/SENASA (gastos locales, en ARS)
      - AFIP: DI AFIP (derechos + IVA aduanero, en ARS)
    El landeo = (total_nac_ARS + total_afip_ARS) / TC → expresado en USD por unidad.
    Coef. de landeo = landeo_unit_USD / FOB_unit_USD × 100%
    """
    if not po_lines or not tc_usd:
        return [], {}

    # ── Identificar bills por rol ──────────────────────────────────────────
    # PETDUR: proveedor extranjero (USD). Se usa solo para refinar el TC.
    petdur   = next((b for b in bills
                     if b.get("partner_id") and b["partner_id"][0] == 49328), None)

    # Nac: gastos locales en ARS. Filtramos SOLO bills en ARS para no
    # contaminar el total con el bill USD de PETDUR si el partner_id difiere.
    nac_bils = [b for b in bills
                if b.get("partner_id")
                and b["partner_id"][0] in {48825, 48826, 48827, 48828}
                and _bill_currency(b) != "USD"]

    # AFIP: derechos de importación (ARS)
    afip     = next((b for b in bills
                     if b.get("partner_id")
                     and b["partner_id"][0] == 9
                     and _bill_currency(b) != "USD"), None)

    # Refinar TC desde invoice_currency_rate del bill PETDUR
    if petdur:
        icr = petdur.get("invoice_currency_rate")
        if icr and icr is not False:
            tc_ref = _parse_odoo_rate({"rate": icr, "inverse_company_rate": None})
            if tc_ref:
                tc_usd = tc_ref

    # ── Totales de costos de nacionalización (todos en ARS) ───────────────
    total_nac_ars  = sum(_bill_ars_amount(b) for b in nac_bils)
    total_afip_ars = _bill_ars_amount(afip) if afip else 0
    # Landeo total en USD = costos ARS convertidos al TC
    total_landeo_usd = (total_nac_ars + total_afip_ars) / tc_usd if tc_usd > 0 else 0

    # ── FOB total (en USD, desde líneas OC) ───────────────────────────────
    total_fob_usd = sum(float(l.get("price_subtotal") or 0) for l in po_lines)
    if total_fob_usd == 0:
        return [], {}

    rows = []
    for ln in po_lines:
        prod   = ln["product_id"][1] if ln.get("product_id") else ln.get("name", "?")
        qty    = float(ln.get("product_qty") or 1)
        fob_total_usd = float(ln.get("price_subtotal") or 0)   # FOB total línea en USD
        fob_pu        = fob_total_usd / qty if qty > 0 else 0  # FOB por unidad en USD

        # Proporción de esta línea sobre el total FOB → distribuir landeo
        prop            = fob_total_usd / total_fob_usd
        landeo_line_usd = total_landeo_usd * prop               # landeo asignado a esta línea
        landeo_u_usd    = landeo_line_usd / qty if qty > 0 else 0  # landeo por unidad en USD

        # Total y coeficiente
        total_u_usd = fob_pu + landeo_u_usd                    # costo total por unidad en USD
        coef_pct    = (landeo_u_usd / fob_pu * 100) if fob_pu > 0 else 0
        cost_u_ars  = total_u_usd * tc_usd                     # costo por unidad en ARS
        rows.append({
            "Producto":           prod,
            "Cant.":              int(qty) if qty == int(qty) else qty,
            "FOB unit (u$s)":     fmt_usd(fob_pu),
            "Landeo unit (u$s)":  fmt_usd(landeo_u_usd),
            "Total unit (u$s)":   fmt_usd(total_u_usd),
            "Coef. landeo":       f"+{coef_pct:.1f}%",
            "Costo unit. (ARS)":  fmt_ars(cost_u_ars),
        })

    # Desglose de gastos de nac para el expander
    nac_detail = {}
    for b in nac_bils:
        pid   = b["partner_id"][0] if b.get("partner_id") else 0
        lbl   = PARTNER_TO_TIPO.get(pid, {}).get("label", "Otro")
        amt   = _bill_ars_amount(b)
        nac_detail[lbl] = nac_detail.get(lbl, 0) + amt
    if afip:
        nac_detail["AFIP"] = total_afip_ars

    total_landeo_ars = total_nac_ars + total_afip_ars
    summary = {
        "tc_usd":              tc_usd,
        "total_fob_usd":       total_fob_usd,
        "total_fob_ars":       total_fob_usd * tc_usd,
        "total_landeo_usd":    total_landeo_usd,
        "total_landeo_ars":    total_landeo_ars,
        "grand_total_usd":     total_fob_usd + total_landeo_usd,
        "grand_total_ars":     total_fob_usd * tc_usd + total_landeo_ars,
        "coef_landeo_total":   (total_landeo_usd / total_fob_usd * 100) if total_fob_usd > 0 else 0,
        "nac_detail":          nac_detail,
        # debug
        "_nac_bils_count":     len(nac_bils),
        "_afip_found":         afip is not None,
        "_petdur_found":       petdur is not None,
    }
    return rows, summary

def get_journal_purchase_account(models_url, uid, api_key, journal_id):
    """
    Devuelve un account_id de compras/gastos para líneas de factura de proveedor.
    Estrategias (en orden):
    1. Línea reciente de in_invoice en ese journal
    2. Línea reciente de in_invoice de cualquier journal con ese mismo partner (PETDUR=49328)
    3. default_account_id del journal
    4. Cualquier cuenta con tipo 'expense' o 'other' activa
    """
    try:
        m = xmlrpc.client.ServerProxy(models_url, allow_none=True)
        from collections import Counter

        def _most_common(rows):
            counts = Counter(l["account_id"][0] for l in rows if l.get("account_id"))
            return counts.most_common(1)[0][0] if counts else None

        # 1. Líneas recientes de facturas en el journal dado
        lines = m.execute_kw(
            ODOO_DB, uid, api_key, "account.move.line", "search_read",
            [[("journal_id", "=", journal_id),
              ("display_type", "=", False),
              ("account_id", "!=", False)]],
            {"fields": ["account_id"], "limit": 30, "order": "id desc"})
        if lines:
            result = _most_common(lines)
            if result:
                return result

        # 2. Líneas de facturas del partner PETDUR (cualquier journal)
        lines2 = m.execute_kw(
            ODOO_DB, uid, api_key, "account.move.line", "search_read",
            [[("partner_id", "=", 49328),
              ("display_type", "=", False),
              ("account_id", "!=", False)]],
            {"fields": ["account_id"], "limit": 30, "order": "id desc"})
        if lines2:
            result = _most_common(lines2)
            if result:
                return result

        # 3. default_account_id del journal
        jrows = m.execute_kw(
            ODOO_DB, uid, api_key, "account.journal", "read",
            [[journal_id]], {"fields": ["default_account_id"]})
        if jrows and jrows[0].get("default_account_id"):
            return jrows[0]["default_account_id"][0]

        # 4. Cualquier cuenta de gastos activa
        acc = m.execute_kw(
            ODOO_DB, uid, api_key, "account.account", "search_read",
            [[("account_type", "in", ["expense", "expense_direct_cost"]),
              ("deprecated", "=", False)]],
            {"fields": ["id"], "limit": 1, "order": "code asc"})
        if acc:
            return acc[0]["id"]
    except Exception:
        pass
    return None


def create_vendor_bill(models, uid, api_key, partner_id, ref, invoice_date,
                       filename, file_bytes, mimetype, journal_id=None, doc_type_id=None,
                       invoice_date_due=None, account_id=None, amount_neto=None,
                       currency_id=None, analytic_account_id=None, product_id=None,
                       l10n_latam_document_number=None, invoice_origin=None,
                       extra_lines=None, clear_taxes=False, line_name=None):
    """
    extra_lines: lista de dicts con keys opcionales:
        name, quantity, price_unit, account_id, product_id
    Si se pasa, reemplaza la lógica de línea única (account_id/amount_neto).
    """
    vals = {"move_type": "in_invoice"}
    if partner_id:       vals["partner_id"]   = partner_id
    if ref:              vals["ref"]          = ref
    if invoice_date:     vals["invoice_date"] = invoice_date
    if invoice_date_due: vals["invoice_date_due"] = invoice_date_due
    if journal_id:       vals["journal_id"]   = journal_id
    if doc_type_id:      vals["l10n_latam_document_type_id"] = doc_type_id
    if currency_id:      vals["currency_id"]  = currency_id
    if invoice_origin:   vals["invoice_origin"] = invoice_origin
    if l10n_latam_document_number:
        vals["l10n_latam_document_number"] = l10n_latam_document_number

    # Líneas explícitas (ej: de lineas_petdur)
    if extra_lines:
        bill_lines = []
        for ln in extra_lines:
            lv = {
                "name":       ln.get("name") or ref or "Artículo",
                "quantity":   float(ln.get("quantity") or ln.get("cantidad") or 1),
                "price_unit": float(ln.get("price_unit") or ln.get("precio_unit") or 0),
            }
            if ln.get("account_id"):
                lv["account_id"] = ln["account_id"]
            if ln.get("product_id"):
                lv["product_id"] = ln["product_id"]
            if analytic_account_id:
                lv["analytic_distribution"] = {str(analytic_account_id): 100}
            bill_lines.append((0, 0, lv))
        if bill_lines:
            vals["invoice_line_ids"] = bill_lines
    # Línea única legada (account_id/amount_neto)
    elif (account_id or product_id) and amount_neto:
        line_vals = {
            "name":       line_name or ref or "Factura proveedor",
            "price_unit": float(amount_neto),
            "quantity":   1,
        }
        if account_id:  line_vals["account_id"] = account_id
        if product_id:  line_vals["product_id"] = product_id
        if analytic_account_id:
            line_vals["analytic_distribution"] = {str(analytic_account_id): 100}
        if clear_taxes:
            line_vals["tax_ids"] = [(5, 0, 0)]   # limpiar impuestos (factura exenta)
        vals["invoice_line_ids"] = [(0, 0, line_vals)]

    move_id = call(models, uid, api_key, "account.move", "create", [vals])
    if file_bytes:
        attach_file(models, uid, api_key, "account.move", move_id, filename, file_bytes, mimetype)
    return move_id

def create_landed_cost(models, uid, api_key, picking_ids, cost_lines):
    vals = {
        "picking_ids": [(4, pid) for pid in picking_ids],
        "cost_lines": [(0, 0, {
            "product_id": line["product_id"],
            "price_unit": line["price_unit"],
            "split_method": "by_current_cost_price",
        }) for line in cost_lines],
    }
    return call(models, uid, api_key, "stock.landed.cost", "create", [vals])

def create_sale_order(models, uid, api_key, partner_id, note, lines, filename, file_bytes, mimetype,
                      client_order_ref=None, payment_term_id=None, date_order=None,
                      ejecutivo_field=None, ejecutivo_id=None):
    vals = {"partner_id": partner_id, "note": note or ""}
    if client_order_ref: vals["client_order_ref"] = client_order_ref
    if payment_term_id:  vals["payment_term_id"]  = payment_term_id
    if date_order:       vals["date_order"]        = date_order
    if ejecutivo_field and ejecutivo_id:
        vals[ejecutivo_field] = ejecutivo_id
    order_id = call(models, uid, api_key, "sale.order", "create", [vals])
    for ln in lines:
        line_vals = {
            "order_id":       order_id,
            "name":           ln.get("descripcion") or ln.get("producto") or "Sin descripción",
            "product_uom_qty": _to_float(ln.get("cantidad", 1)),
            "price_unit":     _to_float(ln.get("precio_unit") or ln.get("precio", 0)),
        }
        if ln.get("product_id"):
            _pid = ln["product_id"]
            # product_id puede ser product.template (estrategias 1&2) o product.product (EAN13)
            # Buscar variante por template_id; si no hay resultado ya es un variant ID correcto
            _vv = call(models, uid, api_key, "product.product", "search",
                       [[("product_tmpl_id", "=", _pid), ("active", "=", True)]], {"limit": 1})
            line_vals["product_id"] = _vv[0] if _vv else _pid
        elif ln.get("producto"):
            prod_ids = call(models, uid, api_key, "product.product", "search",
                            [[("name", "ilike", ln["producto"])]], {"limit": 1})
            if prod_ids:
                line_vals["product_id"] = prod_ids[0]
        call(models, uid, api_key, "sale.order.line", "create", [line_vals])
    if file_bytes:
        attach_file(models, uid, api_key, "sale.order", order_id, filename, file_bytes, mimetype)
    # Confirmar el pedido (draft → sale)
    try:
        call(models, uid, api_key, "sale.order", "action_confirm", [[order_id]])
    except Exception:
        pass  # si falla la confirmación el pedido queda en draft pero no bloquea
    # Cancelar los pickings de entrega auto-generados: el inventario se mueve con el remito
    try:
        _pick_ids = call(models, uid, api_key, "stock.picking", "search",
                         [[("sale_id", "=", order_id), ("picking_type_code", "=", "outgoing")]])
        if _pick_ids:
            call(models, uid, api_key, "stock.picking", "action_cancel", [_pick_ids])
    except Exception:
        pass  # no bloquear si falla la cancelación del picking
    return order_id

def _to_float(v):
    try:
        return float(str(v).replace(",", "."))
    except Exception:
        return 0.0

def parse_ar_date(raw):
    """
    Convierte fechas a ISO YYYY-MM-DD.
    Soporta: DD/MM/YYYY, DD-MM-YYYY, DD.MM.YYYY, YYYY-MM-DD, YYYY/MM/DD.
    """
    if not raw:
        return ""
    raw = raw.strip()
    # DD/MM/YYYY, DD-MM-YYYY, DD.MM.YYYY
    m = re.match(r"(\d{1,2})[/\-\.](\d{1,2})[/\-\.](\d{4})", raw)
    if m:
        d, mo, y = m.group(1).zfill(2), m.group(2).zfill(2), m.group(3)
        return f"{y}-{mo}-{d}"
    # YYYY-MM-DD, YYYY/MM/DD
    m2 = re.match(r"(\d{4})[/\-](\d{1,2})[/\-](\d{1,2})", raw)
    if m2:
        return raw[:10]
    return ""

def normalize_amount(raw):
    """
    Normaliza un número extraído de factura a float string con punto decimal.
    Soporta formato argentino (1.234,56) y formato US (1,234.56).
    """
    raw = raw.strip()
    last_comma = raw.rfind(",")
    last_dot   = raw.rfind(".")
    if last_comma > last_dot:
        # Formato argentino: último separador es coma → decimal
        return raw.replace(".", "").replace(",", ".")
    elif last_dot > last_comma:
        # Formato US o sin miles: último separador es punto → decimal
        return raw.replace(",", "")
    else:
        return raw.replace(",", ".")

_ODOO17_PATHS = {
    "account.move":       "odoo/accounting/vendor-bills",
    "purchase.order":     "odoo/purchase",
    "sale.order":         "odoo/sales",
    "stock.picking":      "odoo/inventory/receipts",
    "stock.landed.cost":  "odoo/inventory/landed-costs",
    "res.partner":        "odoo/contacts",
}

def odoo_url(model, record_id):
    """URL directa Odoo 17 para un registro. Funciona en Odoo 16+ también."""
    path = _ODOO17_PATHS.get(model)
    if path:
        return f"{ODOO_URL}/{path}/{record_id}"
    # fallback hash-URL por si el modelo no está mapeado
    return odoo_url("{model}", record_id)

def safe_float(v, default=0.0):
    """Convierte a float tolerando formato ARS (1.234,56), strings vacíos y None."""
    if v is None or v == "":
        return default
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("$", "").replace(" ", "")
    if "," in s and "." in s:
        # 1.234,56 → quitar punto de miles, coma como decimal
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    elif re.match(r"^\d+\.\d{3}$", s):
        # "77.896" → punto con exactamente 3 dígitos = separador de miles ARS → 77896
        s = s.replace(".", "")
    try:
        return float(s)
    except ValueError:
        return default


def fmt_ars(v):
    """Formatea número como moneda ARS: $ 1.234,56"""
    if not v:
        return ""
    try:
        s = "{:,.2f}".format(float(v))
        return "$ " + s.replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return str(v)

def fmt_usd(v):
    """Formatea número como moneda USD estilo Odoo: u$s 1.234,56"""
    try:
        s = "{:,.2f}".format(float(v))
        return "u$s " + s.replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return "u$s 0,00"

def parse_payment_terms(text):
    """
    Extrae días de pago de condiciones de venta.
    Retorna int con cantidad de días, o None si no se detecta.
    Ej: "CUENTA CORRIENTE A 10 DIAS" → 10
        "30 DIAS FECHA FACTURA"       → 30
        "A 15 DIAS FF"                → 15
        "CUENTA CORRIENTE"            → None
    """
    pats = [
        r"CUENTA\s+CORRIENTE\s+A\s+(\d+)\s+D[IÍ]AS?",
        r"\bA\s+(\d+)\s+D[IÍ]AS?\b",
        r"(\d+)\s+D[IÍ]AS?\s+(?:FECHA\s+FACTURA|FF|FV)",
        r"(?:Cond\.?\s*Vta\.?|Condici[oó]n(?:es)?\s+de\s+Venta)[:\s]+[^\n]*?(\d+)\s+D[IÍ]AS?",
        r"\b(\d{1,3})\s+D[IÍ]AS?\b",
    ]
    for pat in pats:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            days = int(m.group(1))
            if 1 <= days <= 365:   # sanidad: descartar años o CAE que puedan colar
                return days
    return None

def compute_vencimiento(fecha_iso, days):
    """
    Calcula vencimiento = fecha_iso + days días.
    fecha_iso: 'YYYY-MM-DD'. Retorna 'YYYY-MM-DD' o ''.
    """
    if not fecha_iso or days is None:
        return ""
    try:
        from datetime import date, timedelta
        y, mo, d = int(fecha_iso[:4]), int(fecha_iso[5:7]), int(fecha_iso[8:10])
        vto = date(y, mo, d) + timedelta(days=days)
        return vto.strftime("%Y-%m-%d")
    except Exception:
        return ""

@st.cache_data(ttl=120, show_spinner=False)
def search_partner_by_cuit(models_url, uid, api_key, cuit):
    """
    Busca un partner en Odoo por CUIT (campo vat).
    Retorna (partner_id, nombre) o None si no existe.
    """
    if not cuit:
        return None
    cuit_norm = re.sub(r"[^\d]", "", str(cuit))
    if len(cuit_norm) < 10:
        return None
    try:
        m = xmlrpc.client.ServerProxy(models_url, allow_none=True)
        # Intentar con formato estándar XX-XXXXXXXX-X y sin guiones
        variants = [cuit_norm]
        if len(cuit_norm) == 11:
            variants.append(f"{cuit_norm[:2]}-{cuit_norm[2:10]}-{cuit_norm[10:]}")
        for vat_val in variants:
            rows = m.execute_kw(ODOO_DB, uid, api_key, "res.partner", "search_read",
                [[("vat", "=", vat_val), ("active", "=", True)]],
                {"fields": ["id", "name", "vat"], "limit": 1})
            if rows:
                return (rows[0]["id"], rows[0]["name"])
        # Fallback: buscar por los últimos 8 dígitos del CUIT
        rows = m.execute_kw(ODOO_DB, uid, api_key, "res.partner", "search_read",
            [[("vat", "ilike", cuit_norm[2:10]), ("active", "=", True)]],
            {"fields": ["id", "name", "vat"], "limit": 3})
        if rows:
            return (rows[0]["id"], rows[0]["name"])
    except Exception:
        pass
    return None

def search_partner_by_cuit_or_name(models_url, uid, api_key, query, limit=8):
    """
    Busca partners en Odoo por CUIT o razón social (parcial).
    - Si el query tiene >= 10 dígitos → busca por VAT (exacto + ilike)
    - Si parece texto → busca por nombre (ilike)
    - Si tiene dígitos y texto → busca por ambos y une resultados
    Retorna lista de dicts con keys: id, name, vat
    """
    if not query or not query.strip():
        return []
    try:
        m   = xmlrpc.client.ServerProxy(models_url, allow_none=True)
        q   = query.strip()
        digits = re.sub(r"[^\d]", "", q)
        results = []
        seen_ids = set()

        def _fetch(domain):
            rows = m.execute_kw(
                ODOO_DB, uid, api_key, "res.partner", "search_read",
                [domain + [("active", "=", True)]],
                {"fields": ["id", "name", "vat"], "limit": limit, "order": "name asc"})
            return rows

        # Búsqueda por CUIT/VAT si hay suficientes dígitos
        if len(digits) >= 8:
            variants = [digits]
            if len(digits) == 11:
                variants.append(f"{digits[:2]}-{digits[2:10]}-{digits[10:]}")
            for vat_val in variants:
                for row in _fetch([("vat", "=", vat_val)]):
                    if row["id"] not in seen_ids:
                        results.append(row); seen_ids.add(row["id"])
            if not results:
                for row in _fetch([("vat", "ilike", digits[-8:])]):
                    if row["id"] not in seen_ids:
                        results.append(row); seen_ids.add(row["id"])

        # Búsqueda por nombre si el query tiene letras
        if re.search(r"[A-Za-záéíóúñÁÉÍÓÚÑ]", q):
            for row in _fetch([("name", "ilike", q)]):
                if row["id"] not in seen_ids:
                    results.append(row); seen_ids.add(row["id"])

        return results[:limit]
    except Exception:
        return []

@st.cache_data(ttl=30, show_spinner=False)
def check_invoice_exists(models_url, uid, api_key, ref):
    """
    Verifica si ya existe una factura de proveedor con esa referencia en Odoo.
    Retorna (True, move_id, name) si existe, (False, None, None) si no.
    """
    if not ref or not ref.strip():
        return False, None, None
    try:
        m = xmlrpc.client.ServerProxy(models_url, allow_none=True)
        rows = m.execute_kw(ODOO_DB, uid, api_key, "account.move", "search_read",
            [[("ref", "=", ref.strip()),
              ("move_type", "=", "in_invoice"),
              ("state", "!=", "cancel")]],
            {"fields": ["id", "name", "partner_id", "invoice_date", "amount_total"], "limit": 1})
        if rows:
            r = rows[0]
            label = r.get("name") or f"ID {r['id']}"
            return True, r["id"], label
        return False, None, None
    except Exception:
        return False, None, None

def _ai_extract_invoice_fields(text):
    """
    Extrae campos de factura usando Claude Haiku via API de Anthropic.
    Retorna un fields_dict con los mismos keys que extract_pdf_fields,
    o lanza excepción si falla (el caller hace fallback a regex).
    Requiere st.secrets["ANTHROPIC_API_KEY"].
    """
    import anthropic, json as _json

    _api_key = st.secrets.get("ANTHROPIC_API_KEY", "")
    if not _api_key:
        raise ValueError("ANTHROPIC_API_KEY no configurada en secrets")

    client = anthropic.Anthropic(api_key=_api_key)

    _prompt = """Extraé los campos de esta factura argentina. Respondé SOLO con JSON válido, sin texto extra ni bloques de código.

Formato esperado:
{
  "numero": "00004-00020659",
  "fecha": "DD/MM/YYYY",
  "proveedor": "Razón social del EMISOR (quien factura, no quien recibe)",
  "cuit": "30710058667",
  "total": 318124.13,
  "neto": 262912.50,
  "iva": 55211.63,
  "condiciones_venta": "A 7 dias FF",
  "tipo": "RI",
  "concepto": "Descripción breve del servicio o producto facturado"
}

Reglas:
- "numero" siempre en formato XXXXX-XXXXXXXX (5 dígitos, guión, 8 dígitos, con ceros a la izquierda)
- "cuit" del EMISOR (proveedor), sin guiones ni espacios
- "total" incluye todos los impuestos
- "neto" es la base imponible / subtotal gravado
- "iva" es la suma de todos los IVA (21%, 10.5%, 27%)
- "tipo": "RI" (Responsable Inscripto), "MONO" (Monotributo), "EX" (Exento)
- Si la factura es tipo C o el emisor es Monotributo: "iva" = null, "neto" = mismo valor que "total"
- "concepto": descripción breve del servicio/producto (1 línea, máx 100 chars). Buscá en el cuerpo de la factura: "concepto", "descripción", "detalle", o la primer línea del detalle de items. Si no hay, usar null
- Para campos no encontrados usar null
- Números como decimales sin símbolo de moneda (ej: 318124.13, no "$318.124,13")

Factura:
""" + text

    resp = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=512,
        messages=[{"role": "user", "content": _prompt}]
    )

    raw_json = resp.content[0].text.strip()
    # Limpiar por si el modelo wrapeó en ```json ... ```
    raw_json = re.sub(r"^```(?:json)?\s*|\s*```$", "", raw_json, flags=re.MULTILINE).strip()
    data = _json.loads(raw_json)

    # ── Construir fields_dict con el mismo schema que el parser regex ─────
    fields = {
        "numero":            str(data.get("numero") or "").strip(),
        "fecha":             str(data.get("fecha") or "").strip(),
        "fecha_iso":         "",
        "fecha_vencimiento": "",
        "fecha_vto_iso":     "",
        "proveedor":         str(data.get("proveedor") or "").strip()[:80],
        "cuit":              re.sub(r"[\s\-]", "", str(data.get("cuit") or "")),
        "total":             str(data.get("total") or "").replace(",", ".") if data.get("total") is not None else "",
        "neto":              str(data.get("neto") or "").replace(",", ".") if data.get("neto") is not None else "",
        "iva":               str(data.get("iva") or "").replace(",", ".") if data.get("iva") is not None else "",
        "condiciones_venta": str(data.get("condiciones_venta") or "").strip(),
        "concepto":          str(data.get("concepto") or "").strip()[:100],
        "dias_pago":         None,
    }

    # Convertir números a string limpio (sin notación científica)
    for _k in ("total", "neto", "iva"):
        try:
            if fields[_k]:
                fields[_k] = f"{float(fields[_k]):.2f}"
        except Exception:
            fields[_k] = ""

    # Fecha ISO y vencimiento (igual que el parser regex)
    if fields["fecha"]:
        fields["fecha_iso"] = parse_ar_date(fields["fecha"])
    cond_text = fields["condiciones_venta"] or text
    fields["dias_pago"] = parse_payment_terms(cond_text)
    if fields["dias_pago"] and fields["fecha_iso"]:
        fields["fecha_vencimiento"] = compute_vencimiento(fields["fecha_iso"], fields["dias_pago"])
        fields["fecha_vto_iso"]     = fields["fecha_vencimiento"]

    fields["_source"] = "ai"
    return fields


def _bot_extract(file_bytes: bytes, filename: str, mime_type: str, doc_type: str) -> dict:
    """
    Llama al endpoint /extract del bot Cloud Run usando Claude Sonnet con soporte nativo PDF.
    Retorna dict con campos extraídos, o {} si el bot no está configurado o falla.
    """
    import base64 as _b64
    import os
    _url   = os.getenv("BOT_URL", "").rstrip("/")
    _token = os.getenv("CHAT_TOKEN", "")
    if not _url or not _token:
        return {}
    try:
        r = requests.post(
            f"{_url}/extract",
            json={
                "file_b64":  _b64.b64encode(file_bytes).decode(),
                "file_name": filename,
                "file_mime": mime_type,
                "doc_type":  doc_type,
            },
            headers={"X-Chat-Token": _token},
            timeout=120,
        )
        r.raise_for_status()
        return r.json().get("fields", {})
    except Exception:
        return {}


def extract_pdf_fields(file_bytes):
    """
    Parser para facturas electrónicas argentinas.
    Intenta primero el bot (Claude Sonnet nativo PDF), luego IA local, luego regex.
    Retorna (fields_dict, raw_text).
    """
    # ── Bot Cloud Run (Claude Sonnet, soporte nativo PDF) ─────────────────
    _bot = _bot_extract(file_bytes, "factura.pdf", "application/pdf", "factura")
    if _bot.get("proveedor") or _bot.get("numero"):
        _bot["_source"] = "bot"
        try:
            import pdfplumber
            with pdfplumber.open(BytesIO(file_bytes)) as _pdf:
                _raw = "\n".join(p.extract_text() or "" for p in _pdf.pages)
        except Exception:
            _raw = ""
        return _bot, _raw

    try:
        import pdfplumber
        with pdfplumber.open(BytesIO(file_bytes)) as pdf:
            text = "\n".join(page.extract_text() or "" for page in pdf.pages)
    except Exception:
        return {}, ""
    if not text.strip():
        return {}, ""

    # ── Intentar extracción con IA local (Claude Haiku) ───────────────────
    try:
        _ai_fields = _ai_extract_invoice_fields(text)
        # Considerar exitoso si al menos tiene proveedor o número
        if _ai_fields.get("proveedor") or _ai_fields.get("numero"):
            return _ai_fields, text
    except Exception:
        pass  # silencioso: caer al parser regex

    # ── Fallback: parser regex original ──────────────────────────────────

    fields = {"numero": "", "fecha": "", "fecha_iso": "", "fecha_vencimiento": "",
              "fecha_vto_iso": "", "proveedor": "", "total": "", "neto": "", "iva": "",
              "cuit": "", "condiciones_venta": "", "dias_pago": None, "_source": "regex"}

    # ── NÚMERO DE COMPROBANTE ─────────────────────────────────────────────
    # Soporta:
    #   Formato AFIP estándar:     "Nro. Comp.: 00002-00013670"
    #   Formato con letra prefijo: "FACTURA A00005-00029174"
    # Primero: combinar "Punto de Venta: XXXX ... Comp. Nro: XXXXXXXX" (AFIP estándar)
    _m_pv = re.search(
        r"Punto\s+de\s+Venta[:\s]+(\d{4,5})[^\n]{0,60}?Comp\.?\s*Nro\.?[:\s]+(\d{6,8})",
        text, re.IGNORECASE)
    if _m_pv:
        fields["numero"] = f"{_m_pv.group(1).zfill(5)}-{_m_pv.group(2).zfill(8)}"

    if not fields["numero"]:
        num_pats = [
            r"(?:Nro\.?\s*Comp\.?(?:\s*\(Nro\.?\s*Orig\.?\))?|N[°º]\s*Comp\.?|Comprobante\s*N[°º]?)[:\s]*(\d{4,5}[-\s]\d{6,8})",
            r"(?:Punto\s+de\s+Venta[:\s]+\d+\s+)?(?:Comp\.?\s*Nro\.?|Nro\.)[:\s]+(\d{4,5}-\d{6,8})",
            r"(?:FACTURA|NOTA\s+DE\s+CR[EÉ]DITO|NOTA\s+DE\s+D[EÉ]BITO|RECIBO)\s+([A-Z]\d{4,5}-\d{6,8})",
            r"\b([A-Z]\d{4,5}-\d{6,8})\b",
            r"\b(\d{4,5}-\d{6,8})\b",
            # Patrón 6: requiere "Factura" o "Invoice" antes del N° para no capturar "CAE N°"
            r"(?:Factura|Invoice)\s*N[°º\.][:\s#]*([A-Z0-9\-]{5,20})",
        ]
        for pat in num_pats:
            m = re.search(pat, text, re.IGNORECASE)
            if m:
                fields["numero"] = m.group(1).strip()
                break

    # Formato "Nº0004 - 00020659" (Nº + espacios alrededor del guión, sin prefijo letra)
    if not fields["numero"]:
        m = re.search(r"N[°º]\s*(\d{4,5})\s*[-–]\s*(\d{6,8})", text)
        if m:
            fields["numero"] = f"{m.group(1).zfill(5)}-{m.group(2).zfill(8)}"

    # ── FECHA DE EMISIÓN ──────────────────────────────────────────────────
    emision_pats = [
        r"(?:Fecha\s+de\s+[Ee]misi[oó]n|Fecha\s+[Ee]mis\.?)[:\s]+(\d{1,2}/\d{1,2}/\d{4})",
        r"(?:^|\n|\s)(?:FECHA|Fecha)[:\s]+(\d{1,2}/\d{1,2}/\d{4})",
    ]
    for pat in emision_pats:
        m = re.search(pat, text, re.IGNORECASE | re.MULTILINE)
        if m:
            fields["fecha"] = m.group(1).strip()
            fields["fecha_iso"] = parse_ar_date(fields["fecha"])
            break
    if not fields["fecha"]:
        m = re.search(r"(\d{1,2}/\d{1,2}/\d{4})", text)
        if m:
            fields["fecha"] = m.group(1)
            fields["fecha_iso"] = parse_ar_date(fields["fecha"])

    # ── CONDICIONES DE VENTA ──────────────────────────────────────────────
    cond_pats = [
        r"(?:Condici[oó]n(?:es)?\s+de\s+Venta|Cond\.?\s*Vta\.?)[:\s]+([^\n]{3,80})",
    ]
    for pat in cond_pats:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            fields["condiciones_venta"] = m.group(1).strip()
            break

    # ── FECHA DE VENCIMIENTO DE PAGO ─────────────────────────────────────
    # Se calcula desde condiciones de venta (ej: "CUENTA CORRIENTE A 10 DIAS").
    # "Fecha de Vto." en facturas AFIP/CAE es el vencimiento del CAE, NO el de pago.
    cond_text = fields["condiciones_venta"] or text
    fields["dias_pago"] = parse_payment_terms(cond_text)
    if fields["dias_pago"] and fields["fecha_iso"]:
        fields["fecha_vencimiento"] = compute_vencimiento(fields["fecha_iso"], fields["dias_pago"])
        fields["fecha_vto_iso"]     = fields["fecha_vencimiento"]

    # ── IMPORTE TOTAL ─────────────────────────────────────────────────────
    total_pats = [
        r"(?:Importe\s+Total|Total\s+Factura|TOTAL\s+FACTURA)[:\s$]*\$?\s*([\d.,]+)",
        r"(?:^|\n|\s)TOTAL\s*:\s*\$?\s*([\d.,]+)",        # TOTAL: $amount
        r"(?:^|\n|\s)TOTAL\s+\$\s*([\d.,]+)",             # TOTAL $ amount
        r"(?:^|\n|\s)TOTAL\s+([\d.,]+)(?:\s|$)",          # TOTAL amount
        r"(?:^|\n|\s)PESOS\s+TOTAL[:\s$]*\$?\s*([\d.,]+)",
        r"Total\s+a\s+pagar[:\s$]*\$?\s*([\d.,]+)",
    ]
    for pat in total_pats:
        m = re.search(pat, text, re.IGNORECASE | re.MULTILINE)
        if m:
            fields["total"] = normalize_amount(m.group(1).strip())
            break

    # ── NETO GRAVADO ──────────────────────────────────────────────────────
    neto_pats = [
        # 1. Etiquetas explícitas (más seguras), incluyendo "Subt.Gravado" (abreviado)
        r"(?:Subt\.?\s*Gravado|Subtotal\s+Gravado|Neto\s+Gravado|Base\s+Imponible)[:\s$]*\$?\s*([\d.,]+)",
        r"(?:Gravado)\s*:\s*\$?\s*([\d.,]+)",
        # 2. "Subtotal:" con dos puntos
        r"(?:SUBTOTAL|Subtotal)\s*:\s*\$?\s*([\d.,]+)",
        # 3. "Subtotal" + newline opcional + "$" + monto (Odoo/columnas separadas en PDF)
        r"(?:SUBTOTAL|Subtotal)[^\n]*\n?\s*\$\s*([\d.,]+)",
        # 4. "Subtotal" + espacios + monto con separadores de miles (sin $)
        #    Exige X.XXX,XX o X,XXX.XX para no capturar "1,00" de tablas
        r"(?:SUBTOTAL|Subtotal)\s+([\d]{1,3}(?:[.,][\d]{3})+(?:[.,][\d]{2}))",
    ]
    for pat in neto_pats:
        m = re.search(pat, text, re.IGNORECASE | re.MULTILINE)
        if m:
            fields["neto"] = normalize_amount(m.group(1).strip())
            break

    # Factura C / Monotributo: sin IVA, neto = importe total
    if not fields["neto"] and fields["total"]:
        _is_monotrib = bool(re.search(
            r"(?:Responsable\s+Monotributo|MONOTRIBUTO|Factura\s+C\b|COD\.?\s*011)",
            text, re.IGNORECASE))
        if _is_monotrib:
            fields["neto"] = fields["total"]

    # ── IVA ───────────────────────────────────────────────────────────────
    iva_pats = [
        r"IVA\s+(?:21|10[.,]5|27)\s*%\s*\$?\s*([\d.,]+)",  # IVA 21 % $ amount
        r"I\.?V\.?A[^:\n]*(?:21|10\.5|27)[^:\n]*:[:\s$]*\$?\s*([\d.,]+)",
        r"I\.?V\.?A[:\s$%\d.]*:\s*([\d.,]+)",
        r"(?:Impuesto\s+)?IVA[:\s$]*\$?\s*([\d.,]+)",
    ]
    for pat in iva_pats:
        # Usar findall y tomar el ÚLTIMO match: las filas de tabla dan el IVA por línea
        # (primer match), el resumen al final da el IVA total (último match)
        matches = re.findall(pat, text, re.IGNORECASE)
        if matches:
            fields["iva"] = normalize_amount(matches[-1].strip())
            break

    # ── RESUMEN TABULAR (fila "T o t a l", común en facturas de servicios) ────
    # Header: "Subtotal ... I.V.A. 21% ... T o t a l"
    # Fila:   "$ 262,912.50  $ 262,912.50  $ 55,211.63  $ 318,124.13"
    # Mapeo:   [0]=neto       [1]=neto       [-2]=iva      [-1]=total
    if not fields["total"] or not fields["neto"] or not fields["iva"]:
        _m_sumrow = re.search(
            r"Subtotal[^\n]*T\s+o\s+t\s+a\s+l[^\n]*\n([^\n]+)",
            text, re.IGNORECASE)
        if _m_sumrow:
            _amounts = re.findall(r"\$\s*([\d.,]+)", _m_sumrow.group(1))
            if _amounts:
                if not fields["total"]:
                    fields["total"] = normalize_amount(_amounts[-1])
                if not fields["neto"] and len(_amounts) >= 1:
                    fields["neto"] = normalize_amount(_amounts[0])
                if not fields["iva"] and len(_amounts) >= 3:
                    # IVA es el penúltimo monto (antes del total)
                    fields["iva"] = normalize_amount(_amounts[-2])

    # ── RAZÓN SOCIAL / PROVEEDOR EMISOR ──────────────────────────────────
    razon_pats = [
        r"(?:Raz[oó]n\s+[Ss]ocial|Denominaci[oó]n)[:\s]+([^\n\d][^\n]{2,79})",
        r"(?:Apellido\s+y\s+Nombre\s+o\s+Raz[oó]n\s+[Ss]ocial|Nombre\s+y\s+Apellido)[:\s]+([^\n]{3,79})",
    ]
    for pat in razon_pats:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            name = m.group(1).strip()
            name = re.sub(r'\s*\d{2}-\d{8}-\d\s*', '', name).strip()
            if len(name) >= 3:
                fields["proveedor"] = name[:80]
                break

    # Fallback: primera línea significativa que no sea keyword AFIP
    if not fields["proveedor"]:
        skip = {"FACTURA", "NOTA", "RECIBO", "REMITO", "CUIT", "AFIP", "CAE",
                "PUNTO", "FECHA", "IMPORTE", "TOTAL", "VENCIMIENTO", "INGRESOS",
                "IVA", "MONOTRIBUTO", "RESPONSABLE", "INSCRIPTO", "ORIGINAL",
                "DUPLICADO", "TRIPLICADO", "CÓDIGO", "DOMICILIO", "PROVINCIA",
                "CODIGO", "SUBTOTAL", "DESCRIPCION", "CANTIDAD", "PRECIO",
                "CONDICION", "COMPROBANTE", "PESOS", "SON"}
        for line in (l.strip() for l in text.split("\n") if l.strip()):
            if len(line) < 4 or re.match(r'^[\d$.,/\s\-()]+$', line):
                continue
            upper = line.upper()
            if any(w in upper for w in skip):
                continue
            # Descartar líneas que son claramente el número de comprobante
            if re.match(r'^[A-Z]\d{4,5}-\d{6,8}$', line):
                continue
            fields["proveedor"] = line[:80]
            break

    # ── CUIT EMISOR ───────────────────────────────────────────────────────
    # Toma el primer CUIT encontrado en el documento (suele ser el emisor)
    cuit_m = re.search(r'(?:CUIT|C\.U\.I\.T)[:\s.]*(\d{2}[-\s]?\d{8}[-\s]?\d)', text, re.IGNORECASE)
    if cuit_m:
        fields["cuit"] = re.sub(r'[\s\-]', '', cuit_m.group(1))

    return fields, text

@st.cache_data(ttl=600, show_spinner=False)
def get_all_payment_terms(models_url, uid, api_key):
    """Retorna lista de (id, name) de plazos de pago activos en Odoo."""
    try:
        m = xmlrpc.client.ServerProxy(models_url, allow_none=True)
        rows = m.execute_kw(ODOO_DB, uid, api_key, "account.payment.term", "search_read",
            [[("active", "=", True)]],
            {"fields": ["id", "name"], "order": "name asc"})
        return [(r["id"], r["name"]) for r in rows]
    except Exception:
        return []

@st.cache_data(ttl=120, show_spinner=False)
def get_customer_payment_terms(models_url, uid, api_key, partner_id):
    """Retorna (payment_term_id, payment_term_name) del cliente, o (None, None)."""
    try:
        m = xmlrpc.client.ServerProxy(models_url, allow_none=True)
        rows = m.execute_kw(ODOO_DB, uid, api_key, "res.partner", "read",
            [[partner_id]],
            {"fields": ["property_payment_term_id"]})
        if rows:
            pt = rows[0].get("property_payment_term_id")
            if pt and isinstance(pt, (list, tuple)) and len(pt) == 2:
                return pt[0], pt[1]
    except Exception:
        pass
    return None, None

@st.cache_data(ttl=300, show_spinner=False)
def search_product_by_code_or_name(models_url, uid, api_key,
                                   code="", name_keywords="", limit=3, ean13=""):
    """
    Busca producto en Odoo. Prioridad:
      1. Código exacto/ilike en product.template  (ideal: LFANT00006, LCANO00022)
      2. Nombre en product.template — varias estrategias
      3. EAN13 barcode en product.product (último recurso)
    Siempre prefiere default_code que empiece con 'L' y mayor standard_price.
    """
    try:
        m   = xmlrpc.client.ServerProxy(models_url, allow_none=True)
        F   = ["id", "name", "default_code", "standard_price", "list_price"]

        def _best(rows):
            if not rows:
                return []
            l = [r for r in rows if str(r.get("default_code") or "").upper().startswith("L")]
            pool = l or rows
            return [max(pool, key=lambda r: float(r.get("standard_price") or 0))]

        def _tmpl(domain, lim=20):
            return m.execute_kw(ODOO_DB, uid, api_key,
                                "product.template", "search_read",
                                [domain], {"fields": F, "limit": lim})

        # ── 1. CÓDIGO exacto / ilike en default_code ──────────────────────────
        for c in dict.fromkeys([code.strip(), code.strip().lstrip("0")]):
            if not c:
                continue
            r = _best(_tmpl([("default_code", "=",     c), ("active", "=", True)], 5))
            if r: return r
            r = _best(_tmpl([("default_code", "ilike", c), ("active", "=", True)], 10))
            if r: return r

        # ── 1b. CÓDIGO como fragmento de nombre ────────────────────────────────
        # Útil para códigos de fabricante (GI-16, BH-1) que aparecen en el nombre.
        # NO generamos variante con espacio ("GI 16") porque es demasiado amplia.
        _code_s = code.strip()
        if _code_s and len(_code_s) >= 2:
            _code_variants = dict.fromkeys([
                _code_s,                                   # "GI-16" (con guión: específico)
                re.sub(r"[-_/ ]", "", _code_s),            # "GI16" (sin separadores)
            ])
            for cv in _code_variants:
                if not cv or len(cv) < 2:
                    continue
                # Código puramente numérico: demasiado genérico solo → combinarlo con
                # palabras significativas de la descripción (≥5 chars, no stopwords)
                if cv.isdigit():
                    _stops = {"para", "con", "por", "original", "canon", "pack",
                              "color", "negro", "plano", "escaner", "formato"}
                    _desc_words = [
                        w for w in re.sub(r"[^\w\s]", " ", name_keywords).split()
                        if len(w) >= 4 and w.lower() not in _stops
                    ]
                    for dw in _desc_words[:3]:
                        r = _best(_tmpl([("active", "=", True),
                                         ("name", "ilike", cv),
                                         ("name", "ilike", dw)], 10))
                        if r: return r
                else:
                    r = _best(_tmpl([("active", "=", True), ("name", "ilike", cv)], 10))
                    if r: return r

        # ── 2. NOMBRE por keywords ──────────────────────────────────────────────
        if name_keywords and name_keywords.strip():
            # Preservar guiones para capturar códigos como "GI-16", "BH-1", "LiDE-300"
            kw_hyphens = re.sub(r"[^\w\s\-]", " ", name_keywords)
            kw         = re.sub(r"[^\w\s]",   " ", name_keywords)
            words_hyph = kw_hyphens.split()
            words      = kw.split()
            # Tokens con guión Y dígitos: "GI-16", "BH-1", "LiDE-300" → tratar como modelo
            hyphen_model = [w for w in words_hyph
                            if "-" in w and re.search(r"\d", w) and len(w) >= 3]
            # Tokens con letras Y dígitos (sin guión): G1110, 190C, 190BK…
            model  = [w for w in words
                      if re.search(r"[A-Za-z]", w) and re.search(r"\d", w) and len(w) >= 4]
            # Tokens solo letras, cortos (contexto tipo "GI")
            short  = [w for w in words if w.isalpha() and len(w) == 2]
            # Tokens alfanuméricos cortos (E1, S1, V8, X9…) con palabra extra de contexto
            short_model = [w for w in words
                           if re.search(r"[A-Za-z]", w) and re.search(r"\d", w)
                           and 2 <= len(w) <= 3]

            # 2-extra: buscar por código con guión directamente en el nombre
            for hm in hyphen_model[:2]:
                r = _best(_tmpl([("active", "=", True), ("name", "ilike", hm)], 10))
                if r: return r

            # 2-extra-b: modelo corto + siguiente palabra como contexto ("E1 maxt", "V8 mate")
            for sm in short_model[:2]:
                ctx_words = [w for w in words if w != sm and len(w) >= 3]
                if ctx_words:
                    r = _best(_tmpl([("active", "=", True),
                                     ("name", "ilike", sm),
                                     ("name", "ilike", ctx_words[0])], 10))
                    if r: return r
                # Segundo intento: solo el modelo corto (más amplio)
                r = _best(_tmpl([("active", "=", True), ("name", "ilike", sm)], 10))
                if r: return r

            # 2a. Número de modelo literal ("G1110" → "PIXMA G1110")
            # Con contexto corto (ej: "GI") para evitar falsos positivos:
            # "190C" sin ctx matchea "190CM" en PILETA → con "GI" solo matchea GI-190 C
            if model:
                dom_2a = [("active", "=", True), ("name", "ilike", model[0])]
                if short:
                    dom_2a.append(("name", "ilike", short[0]))
                r = _best(_tmpl(dom_2a))
                if r: return r

            # 2b. Token alfanumérico separado con espacio + contexto 2-char
            # "190C" → partes ["190","C"] → busca "190 C" AND "GI" → "GI-190 C"
            # "190BK"→ partes ["190","BK"]→ "190 BK" no existe → AND("190","BK","GI")
            if model:
                parts = [p for p in
                         re.split(r"(?<=[A-Za-z])(?=\d)|(?<=\d)(?=[A-Za-z])", model[0])
                         if p]
                if len(parts) >= 2:
                    spaced = " ".join(parts)          # "190 C", "190 BK"
                    ctx    = short[:1]                # ["GI"] si existe

                    # 2b-i: spaced + contexto (más específico)
                    base = [("active", "=", True), ("name", "ilike", spaced)]
                    if ctx:
                        base.append(("name", "ilike", ctx[0]))
                    r = _best(_tmpl(base))
                    if r: return r

                    # 2b-ii: AND de partes + contexto (para "PGBK" q no tiene espacio)
                    base2 = ([("active", "=", True)]
                             + [("name", "ilike", p) for p in parts[:2]]
                             + ([("name", "ilike", ctx[0])] if ctx else []))
                    r = _best(_tmpl(base2))
                    if r: return r

            # ── 2c. Fallback texto puro: palabras largas, sin dígitos ──────────────
            # Cubre productos como "sillas gamer", "linterna recargable", etc.
            _stops_txt = {"para", "con", "por", "original", "pack", "color",
                          "negro", "unidad", "unidades", "precio", "total",
                          "costo", "envio", "marca", "modelo", "articulo",
                          "producto", "colores", "varios"}
            _txt_words = [
                w.lower() for w in re.sub(r"[^\w\s]", " ", name_keywords).split()
                if len(w) >= 5
                and w.lower() not in _stops_txt
                and not any(c.isdigit() for c in w)
            ]
            if _txt_words:
                # Dos palabras clave juntas (más específico)
                if len(_txt_words) >= 2:
                    r = _best(_tmpl([("active", "=", True),
                                     ("name", "ilike", _txt_words[0]),
                                     ("name", "ilike", _txt_words[1])], 10))
                    if r: return r
                # Una sola palabra clave
                r = _best(_tmpl([("active", "=", True),
                                 ("name", "ilike", _txt_words[0])], 15))
                if r: return r

        # ── 3. EAN13 ────────────────────────────────────────────────────────────
        if ean13 and len(str(ean13)) == 13 and str(ean13).isdigit():
            r = m.execute_kw(ODOO_DB, uid, api_key, "product.product", "search_read",
                             [[("barcode", "=", str(ean13)), ("active", "=", True)]],
                             {"fields": F, "limit": 1})
            if r: return r

    except Exception:
        pass
    return []


def get_ejecutivo_field(models_url, uid, api_key):
    """Detecta el nombre técnico y el modelo de relación del campo
    'Ejecutivo de cuenta' (o similar) en sale.order.
    Retorna (field_name, relation_model) o (None, None)."""
    try:
        _mx = xmlrpc.client.ServerProxy(models_url)
        fields = _mx.execute_kw(ODOO_DB, uid, api_key,
            "sale.order", "fields_get", [],
            {"attributes": ["string", "type", "relation"]})
        keywords = ["ejecutivo", "referido"]
        for fname, finfo in fields.items():
            if not fname.startswith("x_"):
                continue
            label = finfo.get("string", "").lower()
            if any(kw in label for kw in keywords):
                return fname, finfo.get("relation", "res.partner")
        return None, None
    except Exception:
        return None, None

def get_referidos(models_url, uid, api_key):
    """Devuelve lista de (id, nombre) de partners usados como Referido en Odoo."""
    try:
        _mx = xmlrpc.client.ServerProxy(models_url)
        groups = _mx.execute_kw(
            ODOO_DB, uid, api_key,
            "res.partner", "read_group",
            [[["x_studio_referido_1", "!=", False]],
             ["x_studio_referido_1"],
             ["x_studio_referido_1"]],
            {})
        result = []
        for g in groups:
            val = g.get("x_studio_referido_1")
            if val and isinstance(val, (list, tuple)) and len(val) == 2:
                result.append((val[0], val[1]))
        return sorted(result, key=lambda x: x[1])
    except Exception:
        return []

def create_partner(models, uid, api_key, name, vat, street="", phone="", email_addr=""):
    """Crea un nuevo cliente en Odoo y retorna su ID."""
    vals = {"name": name, "customer_rank": 1, "is_company": True}
    if vat:        vals["vat"]    = vat
    if street:     vals["street"] = street
    if phone:      vals["phone"]  = phone
    if email_addr: vals["email"]  = email_addr
    return call(models, uid, api_key, "res.partner", "create", [vals])

def create_vendor_partner(models, uid, api_key, name, vat, street="", phone="", email_addr=""):
    """Crea un nuevo proveedor en Odoo y retorna su ID."""
    vals = {"name": name, "supplier_rank": 1, "is_company": True}
    if vat:        vals["vat"]    = vat
    if street:     vals["street"] = street
    if phone:      vals["phone"]  = phone
    if email_addr: vals["email"]  = email_addr
    return call(models, uid, api_key, "res.partner", "create", [vals])


# ─────────────────────────────────────────────────────────────────────────────
# CONTACTOS — helpers ARCA + Odoo
# ─────────────────────────────────────────────────────────────────────────────

def extract_arca_fields(text):
    """
    Extrae campos de una Constancia de Inscripción ARCA (AFIP).
    Devuelve dict con: nombre, cuit, forma_juridica, street, city, zip_code,
    province_name, tipo_resp (RI/MONO/EX/otro), actividad_principal.
    """
    f = {
        "nombre": "", "cuit": "", "forma_juridica": "",
        "street": "", "city": "", "zip_code": "", "province_name": "",
        "tipo_resp": "RI",   # default: Responsable Inscripto
        "actividad_principal": "",
    }
    if not text:
        return f

    lines = text.splitlines()

    # Nombre y CUIT — primera línea con "CUIT:"
    for ln in lines[:10]:
        m = re.match(r"^(.+?)\s+CUIT:\s*([\d\-]+)", ln.strip())
        if m:
            f["nombre"] = m.group(1).strip()
            f["cuit"]   = re.sub(r"[\s\-]", "", m.group(2))
            break

    # Forma jurídica
    for ln in lines:
        m = re.match(r"Forma\s+Jur[íi]dica:\s*(.+)", ln, re.I)
        if m:
            f["forma_juridica"] = m.group(1).strip()
            break

    # Tipo responsabilidad AFIP
    text_up = text.upper()
    if "MONOTRIBUTO" in text_up or "RSOC " in text_up:
        f["tipo_resp"] = "MONO"
    elif "EXENTO" in text_up and "IVA" not in text_up:
        f["tipo_resp"] = "EX"
    else:
        f["tipo_resp"] = "RI"   # IVA registrado → Responsable Inscripto

    # Actividad principal
    for ln in lines:
        m = re.search(r"Actividad\s+principal:\s*\d+\s*(?:\(F-\d+\))?\s*(.+?)(?:\s+Mes de inicio|$)", ln, re.I)
        if m:
            f["actividad_principal"] = m.group(1).strip()[:120]
            break

    # Domicilio fiscal
    # Buscar el bloque después de "DOMICILIO FISCAL - ARCA"
    try:
        idx = next(i for i, l in enumerate(lines) if "DOMICILIO FISCAL" in l.upper() and "ARCA" in l.upper())
        addr_lines = [l.strip() for l in lines[idx+1:idx+5] if l.strip()]
        if addr_lines:
            f["street"] = addr_lines[0]
        if len(addr_lines) >= 2:
            f["city"] = addr_lines[1]
        if len(addr_lines) >= 3:
            # "5963-CORDOBA" → zip=5963, province=CORDOBA
            m_cp = re.match(r"(\d+)[\s\-]+(.+)", addr_lines[2])
            if m_cp:
                f["zip_code"]      = m_cp.group(1)
                f["province_name"] = m_cp.group(2).strip().title()
            else:
                f["province_name"] = addr_lines[2].strip().title()
    except StopIteration:
        pass

    return f


@st.cache_data(ttl=3600, show_spinner=False)
def get_ar_states(_models_url, uid, api_key):
    """Provincias argentinas: lista de (id, name)."""
    try:
        m = xmlrpc.client.ServerProxy(_models_url, allow_none=True)
        rows = m.execute_kw(ODOO_DB, uid, api_key, "res.country.state", "search_read",
            [[["country_id.code", "=", "AR"]]],
            {"fields": ["id", "name"], "order": "name asc"})
        return [(r["id"], r["name"]) for r in rows]
    except Exception:
        return []


@st.cache_data(ttl=3600, show_spinner=False)
def get_afip_resp_types(_models_url, uid, api_key):
    """Tipos de responsabilidad AFIP: lista de (id, name)."""
    try:
        m = xmlrpc.client.ServerProxy(_models_url, allow_none=True)
        rows = m.execute_kw(ODOO_DB, uid, api_key, "l10n_ar.afip.responsibility.type", "search_read",
            [[]], {"fields": ["id", "name"], "order": "sequence asc"})
        return [(r["id"], r["name"]) for r in rows]
    except Exception:
        return []


@st.cache_data(ttl=3600, show_spinner=False)
def get_cuit_id_type(_models_url, uid, api_key):
    """ID del tipo de identificación CUIT en Odoo."""
    try:
        m = xmlrpc.client.ServerProxy(_models_url, allow_none=True)
        rows = m.execute_kw(ODOO_DB, uid, api_key, "l10n_latam.identification.type", "search_read",
            [[["name", "ilike", "CUIT"]]],
            {"fields": ["id", "name"], "limit": 1})
        return rows[0]["id"] if rows else None
    except Exception:
        return None


@st.cache_data(ttl=3600, show_spinner=False)
def get_odoo_users(_models_url, uid, api_key):
    """Usuarios activos de Odoo: lista de (id, name)."""
    try:
        m = xmlrpc.client.ServerProxy(_models_url, allow_none=True)
        rows = m.execute_kw(ODOO_DB, uid, api_key, "res.users", "search_read",
            [[["active", "=", True], ["share", "=", False]]],
            {"fields": ["id", "name"], "order": "name asc"})
        return [(r["id"], r["name"]) for r in rows]
    except Exception:
        return []


@st.cache_data(ttl=3600, show_spinner=False)
def get_pricelists(_models_url, uid, api_key):
    """Listas de precios activas."""
    try:
        m = xmlrpc.client.ServerProxy(_models_url, allow_none=True)
        rows = m.execute_kw(ODOO_DB, uid, api_key, "product.pricelist", "search_read",
            [[["active", "=", True]]],
            {"fields": ["id", "name"], "order": "name asc"})
        return [(r["id"], r["name"]) for r in rows]
    except Exception:
        return []


@st.cache_data(ttl=3600, show_spinner=False)
def get_ar_accounts(_models_url, uid, api_key, account_type=None):
    """Cuentas contables filtradas por tipo (asset_receivable / liability_payable)."""
    try:
        m = xmlrpc.client.ServerProxy(_models_url, allow_none=True)
        domain = []
        if account_type:
            domain = [["account_type", "=", account_type]]
        rows = m.execute_kw(ODOO_DB, uid, api_key, "account.account", "search_read",
            [domain], {"fields": ["id", "name", "code"], "order": "code asc", "limit": 200})
        return [(r["id"], f"{r['code']} {r['name']}") for r in rows]
    except Exception:
        return []

def create_full_partner(models, uid, api_key, vals_dict):
    """
    Crea un res.partner completo en Odoo.
    vals_dict puede incluir cualquier campo válido de res.partner.
    Retorna partner_id.
    """
    return call(models, uid, api_key, "res.partner", "create", [vals_dict])


def match_ar_state(province_name, ar_states):
    """Busca el ID de la provincia argentina más parecida al nombre dado."""
    if not province_name or not ar_states:
        return None
    pn = province_name.strip().lower()
    # Exacto
    for sid, sname in ar_states:
        if sname.lower() == pn:
            return sid
    # Parcial (primera palabra significativa)
    pwords = [w for w in pn.split() if len(w) > 3]
    if pwords:
        for sid, sname in ar_states:
            if any(w in sname.lower() for w in pwords):
                return sid
    return None

def _image_to_ocr_text(file_bytes):
    """
    OCR de imagen con pytesseract.
    Devuelve (text, error_msg). Si tesseract no está instalado, error_msg != "".
    """
    try:
        import pytesseract
        from PIL import Image as _PILImage
        img = _PILImage.open(BytesIO(file_bytes))
        if img.mode in ("RGBA", "LA", "P"):
            img = img.convert("RGB")
        # Mejorar resolución para OCR (mínimo 300 DPI recomendado)
        w, h = img.size
        if w < 1000:
            factor = max(2, 1200 // w)
            img = img.resize((w * factor, h * factor), _PILImage.LANCZOS)
        try:
            text = pytesseract.image_to_string(img, lang="spa+eng",
                                               config="--psm 6 --oem 3")
        except Exception:
            text = pytesseract.image_to_string(img, lang="eng",
                                               config="--psm 6 --oem 3")
        return text, ""
    except Exception as e:
        return "", str(e)

def extract_image_fields(file_bytes):
    """OCR imagen → pipeline de facturas. Intenta bot primero, luego OCR."""
    _bot = _bot_extract(file_bytes, "factura.jpg", "image/jpeg", "factura")
    if _bot.get("proveedor") or _bot.get("numero"):
        _bot["_source"] = "bot"
        return _bot, ""
    try:
        import pytesseract
        from PIL import Image as _PILImage
        img = _PILImage.open(BytesIO(file_bytes))
        if img.mode in ("RGBA", "LA", "P"):
            img = img.convert("RGB")
        try:
            pdf = pytesseract.image_to_pdf_or_hocr(img, lang="spa+eng", extension="pdf")
        except Exception:
            pdf = pytesseract.image_to_pdf_or_hocr(img, lang="eng", extension="pdf")
        return extract_pdf_fields(pdf)
    except Exception:
        return {}, ""

def extract_image_oc_fields(file_bytes):
    """
    OCR imagen → extrae campos de Orden de Compra. Intenta bot primero, luego OCR.

    Formato observado en OCs Canon:
      Línea N-1: descripción completa del producto (viene antes de la línea del precio)
      Línea N  : [ítem#] [código] [desc parcial] $ [precio]
      Línea N+1: especificación técnica (a ignorar)
    """
    _bot = _bot_extract(file_bytes, "oc.jpg", "image/jpeg", "oc")
    if _bot.get("numero_oc") or _bot.get("cuit") or _bot.get("lineas"):
        _bot["_source"] = "bot"
        return _bot, {}, ""
    text, err = _image_to_ocr_text(file_bytes)
    if err or not text.strip():
        return {}, {}, err or ""

    result = {
        "cuit": "", "numero_oc": "", "fecha": "", "fecha_iso": "",
        "condiciones_pago": "", "dias_pago": None,
        "neto": "", "iva21": "", "iva105": "", "total": "",
        "lineas": [],
    }

    lines = [ln for ln in text.split("\n")]  # conservar índices

    # ── CUIT ──────────────────────────────────────────────────────────────
    _cuit_m = re.search(r"(?:CUIT|C\.U\.I\.T\.?)[:\s#]*(\d{2}[-\.\s]?\d{8}[-\.\s]?\d)", text, re.I)
    if _cuit_m:
        result["cuit"] = re.sub(r"[\s\.]", "-", _cuit_m.group(1).strip())

    # ── Número OC ─────────────────────────────────────────────────────────
    _oc_m = re.search(
        r"(?:orden\s+de\s+compra|N[°º#.]*\s*OC|OC\s*N[°º#.]?|OC)[:\s#]*([A-Z0-9][\-A-Z0-9]{1,20})",
        text, re.I)
    if _oc_m:
        result["numero_oc"] = _oc_m.group(1).strip()

    # ── Fecha ─────────────────────────────────────────────────────────────
    _fecha_m = re.search(r"(\d{1,2})[/\-\.](\d{1,2})[/\-\.](\d{2,4})", text)
    if _fecha_m:
        d, m2, y = _fecha_m.group(1), _fecha_m.group(2), _fecha_m.group(3)
        y = "20" + y if len(y) == 2 else y
        result["fecha"] = f"{d}/{m2}/{y}"
        try:
            result["fecha_iso"] = f"{y}-{int(m2):02d}-{int(d):02d}"
        except Exception:
            pass

    # ── Total ─────────────────────────────────────────────────────────────
    for _tp in [
        r"(?:total\s+(?:general|orden|a\s+pagar)?)[:\s$]*(\d[\d.,]+)",
        r"(?:^|\n)\s*TOTAL[:\s$]*(\d[\d.,]+)",
    ]:
        _tot_m = re.search(_tp, text, re.I | re.MULTILINE)
        if _tot_m:
            result["total"] = _tot_m.group(1).strip()
            break

    # ── Líneas de productos ───────────────────────────────────────────────
    # Patrón principal: línea que tiene $ seguido de número
    # Maneja OCR artifacts: ; en vez de , en precios (77.896; → 77.896,00)
    # Formato: [ítem] [código] [desc parcial] $ [precio]
    # IMPORTANTE: precio puede terminar en ; (OCR artifact) → no requerir dígito final
    _price_line_pat = re.compile(
        r"^\s*(?:(\d{1,2})\s+)?(.+?)\s*\$\s*([\d][\d.,;:]*)\s*$"
    )

    seen_prices = set()

    for i, line in enumerate(lines):
        m = _price_line_pat.match(line)
        if not m:
            continue

        item_s   = m.group(1)   # número de ítem (puede ser None)
        code_frag = m.group(2).strip()
        price_raw = m.group(3).strip()

        # Limpiar artefactos OCR en el precio
        # 1. Quitar trailing ; : , . (OCR los agrega al final cuando corta la línea)
        price_clean = price_raw.rstrip(";:., ")
        # 2. Reemplazar ; y : internos por , (OCR confunde coma con punto y coma)
        price_clean = re.sub(r"[;:]", ",", price_clean)

        # Validar que es un precio razonable (> 100 ARS)
        price_val = safe_float(price_clean)
        if price_val < 100:
            continue

        # Evitar duplicados
        if price_clean in seen_prices:
            continue
        seen_prices.add(price_clean)

        # ── Extraer código: primer token del fragmento en la línea del precio
        parts = code_frag.split(None, 1)
        codigo    = parts[0] if parts else ""
        desc_frag = parts[1] if len(parts) > 1 else ""

        # ── Buscar descripción completa en la línea ANTERIOR (no vacía)
        desc_prev = ""
        for j in range(i - 1, max(i - 4, -1), -1):
            prev = lines[j].strip()
            # Ignorar líneas vacías o que son solo números/encabezados
            if prev and not re.match(r"^\d+$", prev) and len(prev) > 3:
                desc_prev = prev
                break

        # Usar descripción previa como descripción principal; si no hay, usar fragmento
        descripcion = desc_prev if desc_prev else desc_frag or code_frag

        # Limpiar artefactos OCR comunes en la descripción
        descripcion = re.sub(r"^[A-Z]\s+", "", descripcion)  # letra suelta al inicio (ej: "E Cabezal...")
        descripcion = descripcion.strip(" ,-—")

        result["lineas"].append({
            "codigo":      codigo,
            "descripcion": descripcion,
            "cantidad":    1,
            "precio_unit": price_clean,
            "subtotal":    price_clean,
            "iva_pct":     21,
        })

    return result, {}, text



def _claude_api_extract_oc(file_bytes: bytes, mime_type: str = "application/pdf") -> dict:
    """
    Fallback: usa Claude API directamente para extraer campos de un pedido/OC.
    Se usa cuando BOT_URL no esta configurado o el bot falla.
    """
    import base64 as _b64c2, os as _os2, json as _jc2
    _ant_key2 = _os2.getenv("ANTHROPIC_API_KEY", "")
    if not _ant_key2:
        return {}
    try:
        import anthropic as _ac2
        _client2 = _ac2.Anthropic(api_key=_ant_key2)
        _prompt2 = (
            "Analiza este documento (puede ser un pedido, orden de compra, relacion de pedidos, "
            "presupuesto o cualquier formato de solicitud de productos). "
            "Extrae los datos y devuelve SOLO un JSON con exactamente estos campos:\n"
            "{\n"
            '  "cuit": "CUIT del cliente/empresa que hace el pedido (solo digitos, sin guiones), vacio si no aparece",\n'
            '  "cliente_nombre": "nombre o razon social del cliente que hace el pedido",\n'
            '  "numero_oc": "numero de pedido, orden de compra u orden",\n'
            '  "fecha": "fecha del documento en formato DD/MM/YYYY",\n'
            '  "condiciones_pago": "texto de condicion o plazo de pago",\n'
            '  "dias_pago": numero entero de dias de pago o null,\n'
            '  "total": "monto total como texto (ej: 17.690.000)",\n'
            '  "lineas": [\n'
            '    {\n'
            '      "codigo": "codigo interno del producto",\n'
            '      "descripcion": "articulo o descripcion",\n'
            '      "marca": "marca",\n'
            '      "modelo": "modelo",\n'
            '      "cantidad": cantidad como numero entero,\n'
            '      "precio_unit": precio unitario como numero flotante,\n'
            '      "subtotal": total de la linea como numero flotante\n'
            '    }\n'
            '  ]\n'
            "}"
        )
        _resp2 = _client2.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=2048,
            messages=[{"role": "user", "content": [
                {"type": "document", "source": {
                    "type": "base64", "media_type": mime_type,
                    "data": _b64c2.b64encode(file_bytes).decode()}},
                {"type": "text", "text": _prompt2}
            ]}]
        )
        _text2 = _resp2.content[0].text.strip()
        _js2 = re.search(r'\{[\s\S]*\}', _text2)
        if not _js2:
            return {}
        _data2 = _jc2.loads(_js2.group())
        result2 = {
            "cuit": str(_data2.get("cuit", "") or "").replace("-","").replace(" ",""),
            "cliente_nombre": str(_data2.get("cliente_nombre", "") or ""),
            "numero_oc": str(_data2.get("numero_oc", "") or ""),
            "fecha": str(_data2.get("fecha", "") or ""),
            "fecha_iso": "",
            "condiciones_pago": str(_data2.get("condiciones_pago", "") or ""),
            "dias_pago": _data2.get("dias_pago"),
            "total": str(_data2.get("total", "") or ""),
            "lineas": [],
            "_source": "claude_api",
        }
        # parse fecha_iso
        _fm2 = re.match(r"(\d{1,2})/(\d{1,2})/(\d{4})", result2["fecha"])
        if _fm2:
            result2["fecha_iso"] = f"{_fm2.group(3)}-{_fm2.group(2).zfill(2)}-{_fm2.group(1).zfill(2)}"
        for _ln2 in (_data2.get("lineas") or []):
            result2["lineas"].append({
                "codigo":      str(_ln2.get("codigo", "") or ""),
                "descripcion": str(_ln2.get("descripcion", "") or ""),
                "marca":       str(_ln2.get("marca", "") or ""),
                "modelo":      str(_ln2.get("modelo", "") or ""),
                "cantidad":    float(_ln2.get("cantidad") or 0),
                "precio_unit": float(_ln2.get("precio_unit") or 0),
                "subtotal":    float(_ln2.get("subtotal") or 0),
                "iva_pct":     21.0,
            })
        return result2
    except Exception:
        return {}

def extract_oc_fields(file_bytes):
    """
    Parser para Órdenes de Compra de clientes (formato heterogéneo).
    Intenta bot primero (Claude Sonnet nativo PDF), luego pdfplumber+regex.
    Retorna (fields_dict, all_tables, raw_text).
    """
    _bot = _bot_extract(file_bytes, "oc.pdf", "application/pdf", "oc")
    if _bot.get("numero_oc") or _bot.get("cuit") or _bot.get("lineas"):
        _bot["_source"] = "bot"
        try:
            import pdfplumber
            with pdfplumber.open(BytesIO(file_bytes)) as _pdf:
                _raw = "\n".join(p.extract_text() or "" for p in _pdf.pages)
        except Exception:
            _raw = ""
        return _bot, [], _raw

    # Fallback: Claude API directo (cuando BOT_URL no esta configurado)
    _claude_r = _claude_api_extract_oc(file_bytes)
    if _claude_r.get("numero_oc") or _claude_r.get("lineas"):
        try:
            import pdfplumber
            with pdfplumber.open(BytesIO(file_bytes)) as _pdf:
                _raw = "\n".join(p.extract_text() or "" for p in _pdf.pages)
        except Exception:
            _raw = ""
        return _claude_r, [], _raw

    try:
        import pdfplumber
        with pdfplumber.open(BytesIO(file_bytes)) as pdf:
            pages_text = [page.extract_text() or "" for page in pdf.pages]
            text = "\n".join(pages_text)
            all_tables = []
            for page in pdf.pages:
                tbls = page.extract_tables()
                if tbls:
                    all_tables.extend(tbls)
    except Exception:
        return {}, [], ""
    if not text.strip():
        return {}, [], ""

    fields = {
        "cuit": "", "numero_oc": "", "fecha": "", "fecha_iso": "",
        "condiciones_pago": "", "dias_pago": None,
        "lineas": [],
        "subtotal_neto": "", "iva_21": "", "iva_105": "", "total": "",
    }

    # ── CUIT emisor ──────────────────────────────────────────────────────
    cuits_found = re.findall(r'\b(\d{2}-\d{8}-\d)\b', text)
    if not cuits_found:
        cuits_found = re.findall(r'\bCUIT[:\s.]*(\d{11})\b', text, re.IGNORECASE)
    if cuits_found:
        fields["cuit"] = re.sub(r'[-\s]', '', cuits_found[0])

    # ── Número de OC ─────────────────────────────────────────────────────
    oc_pats = [
        # CASTILLO: "Orden de Compra N 0001-0118,667"
        r"(?:Orden\s+de\s+[Cc]ompra|O\.?C\.?\s*N[°o]?|ORDEN\s+DE\s+COMPRA\s*N?\s*)[:\s]*([0-9]{4}[-/][0-9,]{4,})",
        # Carsa/MUSIMUNDO: "Orden Definitiva de Provisión Número: 4501653808"
        r"(?:Orden\s+Definitiva|Orden\s+de\s+Provisi[oó]n)\b.{0,40}N[úu]mero[:\s]+(\d{6,})",
        r"(?:N[°º]\s*(?:de\s+)?[Oo]rden|Pedido\s+N[°º]|N[°º]\s*[Pp]edido)[:\s]*([0-9]{4}[-/][0-9,]{4,}|\d{6,})",
        r"\b(0{4}[-/][0-9,]{4,})\b",
        # La Anónima / genérico: "Número: 22620313"
        r"\bN[úu]mero[:\s]+(\d{5,})\b",
    ]
    for pat in oc_pats:
        mo = re.search(pat, text, re.IGNORECASE)
        if mo:
            fields["numero_oc"] = mo.group(1).strip().replace(",", "")
            break

    # ── Fecha ─────────────────────────────────────────────────────────────
    # Soporta DD/MM/YYYY, DD.MM.YYYY, DD-MM-YYYY
    date_pats = [
        r"(?:Fecha\s+[Ee]misi[oó]n|Fecha\s+OC|Fecha)[:\s]+(\d{1,2}[/.\-]\d{1,2}[/.\-]\d{4})",
        r"(?:^|\s)(\d{1,2}[/.\-]\d{1,2}[/.\-]\d{4})(?:\s|$)",
    ]
    for pat in date_pats:
        mo = re.search(pat, text, re.IGNORECASE | re.MULTILINE)
        if mo:
            fields["fecha"] = mo.group(1).strip()
            fields["fecha_iso"] = parse_ar_date(fields["fecha"])
            break

    # ── Condiciones de pago ───────────────────────────────────────────────
    cond_pats = [
        r"(?:Condici[oó]n(?:es)?\s+de\s+[Pp]ago|Forma\s+de\s+[Pp]ago)[:\s]+([^\n]{3,80})",
        # Carsa: "Condición: 0016 - 60 Dias"
        r"(?:^|\n)\s*Condici[oó]n[:\s]+([^\n]{3,60})",
        r"(CUENTA\s+CORRIENTE[^\n]{0,60})",
    ]
    for pat in cond_pats:
        mo = re.search(pat, text, re.IGNORECASE | re.MULTILINE)
        if mo:
            fields["condiciones_pago"] = mo.group(1).strip()
            break

    # Días: buscar "Intervalo: 30", "60 Dias", "LOS XX DIAS"
    intervalo_mo = re.search(r"[Ii]ntervalo[:\s]+(\d+)", text)
    if intervalo_mo:
        fields["dias_pago"] = int(intervalo_mo.group(1))
    else:
        fields["dias_pago"] = parse_payment_terms(fields["condiciones_pago"] or text)

    # ── Totales ───────────────────────────────────────────────────────────
    mo = re.search(r"(?:Sub[-\s]?[Tt]otal\s+[Nn]eto|SUBTOTAL\s+NETO|Neto\s+Gravado)[:\s$]*\$?\s*([\d.,]+)", text, re.IGNORECASE)
    if mo:
        fields["subtotal_neto"] = normalize_amount(mo.group(1))
    # Solo capturar IVA si hay un monto en la MISMA línea (no cruzar newline)
    # Esto evita capturar EAN13 del producto siguiente como monto de IVA
    mo = re.search(r"IVA\s+21\s*%[ \t:$]*\$?[ \t]*([\d.,]+)", text, re.IGNORECASE)
    if mo:
        _iva21_raw = mo.group(1)
        # Descartar si parece un EAN13 u otro código de barras (≥12 dígitos sin coma/punto)
        if not re.match(r'^\d{12,}$', _iva21_raw.replace(',','').replace('.','')):
            fields["iva_21"] = normalize_amount(_iva21_raw)
    mo = re.search(r"IVA\s+10[.,]5\s*%[ \t:$]*\$?[ \t]*([\d.,]+)", text, re.IGNORECASE)
    if mo:
        _iva105_raw = mo.group(1)
        if not re.match(r'^\d{12,}$', _iva105_raw.replace(',','').replace('.','')):
            fields["iva_105"] = normalize_amount(_iva105_raw)
    # Carsa: "TOTAL PEDIDO DE COMPRAS ARS 35.923.359,50"
    total_pats = [
        r"(?:Total\s+OC|TOTAL\s+OC|Total\s+[Oo]rden)[:\s$]*\$?\s*([\d.,]+)",
        r"TOTAL\s+PEDIDO[^\d]+([\d.,]+)",
        r"(?:^|\n)\s*TOTAL[:\s$]*\$?\s*([\d.,]+)\s*(?:$|\n)",
    ]
    for _tp in total_pats:
        mo = re.search(_tp, text, re.IGNORECASE | re.MULTILINE)
        if mo:
            fields["total"] = normalize_amount(mo.group(1))
            break

    # ── Líneas de productos (desde tablas pdfplumber) ─────────────────────
    def _try_num(s):
        s = str(s or "").strip()
        if not s or not re.search(r'\d', s):
            return None
        try:
            return float(normalize_amount(s))
        except Exception:
            return None

    for table in all_tables:
        if not table or len(table) < 2:
            continue
        header = [str(c or "").strip().lower() for c in table[0]]
        col_map = {}
        for i, h in enumerate(header):
            if re.search(r"c[oó]d|sku|art[ií]culo|codigo", h) and "codigo" not in col_map:
                col_map["codigo"] = i
            elif re.search(r"cant|qty|cantidad", h) and "cantidad" not in col_map:
                col_map["cantidad"] = i
            elif re.search(r"detal|descri|product|nombre|item", h) and "descripcion" not in col_map:
                col_map["descripcion"] = i
            elif re.search(r"\bneto\b|p\.?\s*unit|precio\s*unit|unitario|unit\s*price", h) and "precio_unit" not in col_map:
                col_map["precio_unit"] = i
            elif re.search(r"\biva\b|tax|%\s*iva|alicuota", h) and "iva_pct" not in col_map:
                col_map["iva_pct"] = i
            elif re.search(r"sub.?total|importe|c\.?final|precio\s*final", h) and "subtotal" not in col_map:
                col_map["subtotal"] = i
        if not col_map or ("descripcion" not in col_map and "codigo" not in col_map):
            continue
        for row in table[1:]:
            if not row or all(not c for c in row):
                continue
            def _gcol(key, default=""):
                idx = col_map.get(key)
                if idx is None or idx >= len(row):
                    return default
                return str(row[idx] or "").strip()
            desc = _gcol("descripcion")
            cod  = _gcol("codigo")
            if not desc and not cod:
                continue
            if re.match(r'^(detalle|descripci[oó]n|producto|item|total|subtotal)$', desc, re.IGNORECASE):
                continue
            qty   = _try_num(_gcol("cantidad"))
            price = _try_num(_gcol("precio_unit"))
            sub   = _try_num(_gcol("subtotal"))
            iva   = _try_num(_gcol("iva_pct"))
            if qty is None and price is None and sub is None:
                continue
            fields["lineas"].append({
                "codigo":      cod,
                "descripcion": desc,
                "cantidad":    qty   if qty   is not None else 0,
                "precio_unit": price if price is not None else 0,
                "iva_pct":     iva   if iva   is not None else 21.0,
                "subtotal":    sub   if sub   is not None else (
                                   (qty * price) if (qty and price) else 0),
            })

    # ── Fallback EAN13: formato Carsa/MUSIMUNDO ──────────────────────────
    # Líneas: {EAN13} {INTCODE-DESCRIPCION} {M3} {QTY} UN {PRECIO} {SUBTOTAL}
    # pdfplumber puede wrappear la línea; combinamos hasta 3 líneas siguientes
    # para armar el registro completo antes de aplicar el regex.
    # IVA en línea posterior ("IVA 21%" / "IVA 10,5%")
    if not fields["lineas"] and text:
        _ean_lines = text.split("\n")
        _ean_pat   = re.compile(
            r'^(\d{13})\s+(.+?)\s+UN\s+([\d.]+,\d{2})\s+([\d.]+,\d{2})\s*$')
        _ean_stop  = re.compile(
            r'^(?:\d{13}\s|Entregar:|TOTAL|Vencimiento|P[aá]gina)',
            re.IGNORECASE)
        for _ei, _eln in enumerate(_ean_lines):
            _strip = _eln.strip()
            if not re.match(r'^\d{13}\b', _strip):
                continue
            # Combinar con las siguientes líneas SOLO si la línea actual no matchea completa
            # (evita agregar el modelo del producto siguiente que rompe el regex)
            _combined = _strip
            if not _ean_pat.match(_combined.strip()):
                for _fwd_idx in range(_ei + 1, min(_ei + 4, len(_ean_lines))):
                    _nxt = _ean_lines[_fwd_idx].strip()
                    if _ean_stop.match(_nxt):
                        break
                    _combined += ' ' + _nxt
                    if _ean_pat.match(_combined.strip()):
                        break
            _em = _ean_pat.match(_combined.strip())
            if not _em:
                continue
            _ean       = _em.group(1)
            _rest_ean  = _em.group(2).strip()
            _price_raw = _em.group(3)
            _sub_raw   = _em.group(4)

            # Extraer qty (último entero antes de UN — ya consumido por regex previo)
            # qty está al final de _rest_ean: "... 3,540 60"
            _qty_m = re.search(r'\s+(\d+)\s*$', _rest_ean)
            _qty   = int(_qty_m.group(1)) if _qty_m else 0
            _rest_ean = (_rest_ean[:_qty_m.start()].strip() if _qty_m else _rest_ean)

            # Remover M3 (decimal con coma como separador decimal: "3,540")
            _m3_m = re.search(r'\s+([\d]+,\d{3})\s*$', _rest_ean)
            if _m3_m:
                _rest_ean = _rest_ean[:_m3_m.start()].strip()

            # Extraer código interno: "176270-DESCRIPCION"
            _ic_m = re.match(r'^(\d{4,8})-(.+)$', _rest_ean)
            _int_code = _ic_m.group(1) if _ic_m else _ean
            _desc_ean = (_ic_m.group(2).strip() if _ic_m else _rest_ean.strip())

            # Líneas de continuación: absorber hasta encontrar "Entregar:" o otro EAN13
            _ei2 = _ei + 1
            while _ei2 < len(_ean_lines):
                _nl2 = _ean_lines[_ei2].strip()
                if (re.match(r'^\d{13}\s', _nl2)
                        or re.match(r'^Entregar:', _nl2, re.IGNORECASE)
                        or re.match(r'^TOTAL|^Vencimiento|^P[aá]gina', _nl2, re.IGNORECASE)):
                    break
                if _nl2 and re.search(r'[A-Za-z0-9]', _nl2) and not re.match(r'^IVA', _nl2, re.IGNORECASE):
                    _desc_ean += " " + _nl2
                _ei2 += 1

            # Buscar IVA en las líneas siguientes (después de Entregar:)
            _iva_ean = 21.0
            for _fwd in range(_ei + 1, min(_ei + 6, len(_ean_lines))):
                _fwd_ln = _ean_lines[_fwd].strip()
                _iva_m2 = re.match(r'IVA\s+([\d,.]+)\s*%', _fwd_ln, re.IGNORECASE)
                if _iva_m2:
                    try:
                        _iva_ean = float(normalize_amount(_iva_m2.group(1)))
                    except Exception:
                        pass
                    break

            _price_f = float(normalize_amount(_price_raw))
            _sub_f   = float(normalize_amount(_sub_raw))
            _desc_ean = re.sub(r'\s+', ' ', _desc_ean).strip()

            fields["lineas"].append({
                "codigo":      _int_code,
                "ean13":       _ean,        # código de barras EAN13 original
                "descripcion": _desc_ean,
                "cantidad":    float(_qty),
                "precio_unit": _price_f,
                "iva_pct":     _iva_ean,
                "subtotal":    _sub_f,
            })

    # ── Fallback: parser de texto cuando no hay tablas ───────────────────
    # Detecta líneas de producto del tipo:
    #   CODIGO  QTY  DESCRIPCION  NETO  IVA%  IT  CFINAL  SUBTOTAL
    # Maneja números fusionados (artefacto de pdfplumber): "297,004.1329,700,413.00"
    if not fields["lineas"] and text:
        _lines = text.split("\n")
        _hdr_idx = None
        for _i, _ln in enumerate(_lines):
            if re.search(r'\bC[oó]digo\b.{0,30}\bCant\b', _ln, re.IGNORECASE):
                _hdr_idx = _i
                break

        if _hdr_idx is not None:
            _stop = re.compile(
                r'^(?:Sub[-\s]?[Tt]otal|Totales|TOTALES|Sub-Totales|Observaciones|IMPORTANTE)',
                re.IGNORECASE)
            _num_re = r'\d{1,3}(?:[.,]\d{3})*[.,]\d{2}|\d+[.,]\d{2}'

            def _parse_num(s):
                s = str(s).strip()
                lc, ld = s.rfind(","), s.rfind(".")
                if lc > ld:
                    return float(s.replace(".", "").replace(",", "."))
                return float(s.replace(",", ""))

            _j = _hdr_idx + 1
            while _j < len(_lines):
                _raw = _lines[_j].strip()
                if _stop.match(_raw):
                    break
                _cm = re.match(r'^(\d{6,12})\s+(\d+)\s+(.+)$', _raw)
                if _cm:
                    _cod   = _cm.group(1)
                    _qty_s = _cm.group(2)
                    _rest  = _cm.group(3).strip()

                    # Absorber línea siguiente si es continuación de la descripción
                    if _j + 1 < len(_lines):
                        _nl = _lines[_j + 1].strip()
                        if (_nl and not re.match(r'^\d{6,12}\s', _nl)
                                and not _stop.match(_nl)
                                and re.search(r'[A-Za-z]', _nl)):
                            _rest += " " + _nl
                            _j += 1

                    # Separar números fusionados: ".13" seguido de dígito → ".13 "
                    _rest_fixed = re.sub(r'(\.\d{2})(\d)', r'\1 \2', _rest)

                    _raw_nums = re.findall(_num_re, _rest_fixed)
                    _nums = []
                    for _rn in _raw_nums:
                        try:
                            _nums.append(_parse_num(_rn))
                        except Exception:
                            pass

                    if len(_nums) >= 3:
                        _qty    = float(_qty_s)
                        # Estructura esperada (de izquierda a derecha):
                        #   Neto  [%Desc]  IVA%  IT  C.Final  Sub-Total
                        # El último siempre es Sub-Total, el primero es Neto
                        _sub    = _nums[-1]
                        _neto   = _nums[0]
                        # IVA%: buscar el valor típico 21.0 o 10.5 en las posiciones centrales
                        _iva    = 21.0
                        for _n in _nums[1:-1]:
                            if abs(_n - 21.0) < 0.6:
                                _iva = 21.0; break
                            if abs(_n - 10.5) < 0.6:
                                _iva = 10.5; break

                        # Descripción: texto antes del primer número + texto después del último
                        # Esto captura el modelo que queda al final de la línea (ej: G2110, G3110)
                        _fst_m  = re.search(_num_re, _rest_fixed)
                        _all_ms = list(re.finditer(_num_re, _rest_fixed))
                        _pre  = (_rest_fixed[:_fst_m.start()].strip()
                                 if _fst_m else _rest_fixed.strip())
                        _post = (_rest_fixed[_all_ms[-1].end():].strip()
                                 if _all_ms else "")
                        _desc = ((_pre + " " + _post).strip() if _post else _pre)
                        _desc = re.sub(r'\s+', ' ', _desc).strip()

                        fields["lineas"].append({
                            "codigo":      _cod,
                            "descripcion": _desc,
                            "cantidad":    _qty,
                            "precio_unit": round(_neto, 2),
                            "iva_pct":     _iva,
                            "subtotal":    round(_sub, 2),
                        })
                _j += 1

    # ── Fallback: La Anónima / tabular texto plano (Cod.Art.Prov.) ──────────
    # Header: "Cod.Art. Cod.Art.Prov. Descripción Marca Bto. Cont. U/M Cant. Costo % Bonif. % Iva Total"
    # Línea:  "2383809 LCANO00015 BOTELLA GL-190 CYA CANON 1 1 CU 25 17098.3 0.00 21.00 427458.00"
    # Números en formato US (punto como decimal): 17098.3  0.00  21.00  427458.00
    if not fields["lineas"] and text:
        _la_lines = text.split("\n")
        _la_hdr = None
        for _li, _ll in enumerate(_la_lines):
            if re.search(r'Cod\.Art\.Prov', _ll, re.IGNORECASE):
                _la_hdr = _li
                break
        if _la_hdr is not None:
            _la_stop = re.compile(
                r'^(?:Sub[-\s]?[Tt]otal|Total\s|Totales|Bonificaci[oó]n|Observaciones|'
                r'Sr\.?\s+Proveedor|Toda\s+Orden|RESERVAR)',
                re.IGNORECASE)
            for _ll in _la_lines[_la_hdr + 1:]:
                _lraw = _ll.strip()
                if not _lraw or _la_stop.match(_lraw):
                    break
                # INT_CODE PROV_CODE ... resto
                _lcm = re.match(r'^(\d{5,10})\s+([A-Z][A-Z0-9]{2,})\s+(.+)', _lraw)
                if not _lcm:
                    continue
                _lint_code  = _lcm.group(1)
                _lprov_code = _lcm.group(2)
                _lrest      = _lcm.group(3).strip()
                # Extraer los últimos 5 tokens numéricos (qty costo bonif iva% total)
                # La Anónima usa formato US: solo dígitos y punto decimal
                _ltoks     = _lrest.split()
                _lnum_toks = []
                for _lt in reversed(_ltoks):
                    if re.match(r'^\d+(?:\.\d+)?$', _lt) and len(_lnum_toks) < 5:
                        _lnum_toks.insert(0, _lt)
                    elif _lnum_toks:
                        break
                if len(_lnum_toks) < 4:
                    continue
                _lqty   = float(_lnum_toks[0])
                _lcosto = float(_lnum_toks[1])
                _liva   = float(_lnum_toks[3]) if len(_lnum_toks) > 3 else 21.0
                _ltotal = float(_lnum_toks[4]) if len(_lnum_toks) > 4 else _lqty * _lcosto
                # Descripción: tokens antes de los numéricos; limpiar U/M y Bto/Cont al final
                _lnum_start = len(_ltoks) - len(_lnum_toks)
                _ldesc = " ".join(_ltoks[:_lnum_start]).strip()
                _ldesc = re.sub(r'\s+\d+\s+\d+\s+[A-Z]{1,3}\s*$', '', _ldesc).strip()
                fields["lineas"].append({
                    "codigo":      _lprov_code,
                    "descripcion": _ldesc,
                    "cantidad":    _lqty,
                    "precio_unit": _lcosto,
                    "iva_pct":     _liva,
                    "subtotal":    _ltotal,
                })

    return fields, all_tables, text


def extract_excel_oc_fields(file_bytes, filename=""):
    """
    Parser flexible para pedidos en Excel (.xls / .xlsx).
    Soporta múltiples formatos de cliente: CANT./CANTIDAD/QTY, IMP.UNIT./PRECIO, etc.
    También extrae metadata (CUIT, razón social, fecha) de las filas de encabezado.
    """
    # ── intentar bot primero (solo xlsx) ─────────────────────────────────
    _fname = (filename or "oc.xlsx").lower()
    if _fname.endswith(".xlsx"):
        _bot = _bot_extract(
            file_bytes, "oc.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "oc")
        if _bot.get("lineas") or _bot.get("numero_oc"):
            _bot["_source"] = "bot"
            return _bot

    fields = {
        "cuit": "", "cliente": "", "numero_oc": "", "fecha": "", "fecha_iso": "",
        "condiciones_pago": "", "dias_pago": None,
        "lineas": [],
        "subtotal_neto": "", "iva_21": "", "iva_105": "", "total": "",
        "fuente": "excel",
    }

    # ── cargar el workbook (xls o xlsx) ──────────────────────────────────
    all_rows = []
    try:
        if _fname.endswith(".xls"):
            import xlrd
            wb = xlrd.open_workbook(file_contents=file_bytes)
            ws = wb.sheet_by_index(0)
            # Convertir a lista de tuplas igual que openpyxl
            for r in range(ws.nrows):
                row = []
                for c in range(ws.ncols):
                    cell = ws.cell(r, c)
                    # xlrd type 0=empty 1=text 2=number 3=date 4=bool 5=error
                    if cell.ctype == 0:
                        row.append(None)
                    elif cell.ctype == 3:
                        # Fecha serial Excel → string
                        try:
                            import datetime
                            dt = xlrd.xldate_as_datetime(cell.value, wb.datemode)
                            row.append(dt.strftime("%d/%m/%Y"))
                        except Exception:
                            row.append(cell.value)
                    else:
                        row.append(cell.value)
                all_rows.append(tuple(row))
        else:
            from openpyxl import load_workbook
            wb2 = load_workbook(BytesIO(file_bytes), data_only=True)
            ws2 = wb2.active
            all_rows = list(ws2.iter_rows(values_only=True))
    except Exception:
        return fields

    # ── Paso 0: extraer metadata de filas pre-header ──────────────────────
    # Buscar CUIT, nombre de cliente, fecha, NRO PEDIDO en filas libres
    _cuit_re = re.compile(r"(\d{2}-\d{8}-\d|\d{11})")
    _fecha_re = re.compile(r"(\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4})")
    for row in all_rows[:20]:
        vals = [str(v or "").strip() for v in row]
        joined = " ".join(vals)
        # CUIT
        if not fields["cuit"]:
            _cm = _cuit_re.search(joined)
            if _cm:
                fields["cuit"] = re.sub(r"[\s\-]", "", _cm.group(1))
        # Fecha
        if not fields["fecha"]:
            _fm = _fecha_re.search(joined)
            if _fm:
                fields["fecha"] = _fm.group(1)
        # Cliente: buscar patrón "CLIENTE <valor>" en celdas adyacentes
        for ci, v in enumerate(vals):
            vl = v.lower()
            if vl in ("cliente", "razon social", "razón social", "cliente:") and ci + 1 < len(vals):
                _cname = vals[ci + 1].strip()
                if _cname and not fields["cliente"]:
                    fields["cliente"] = _cname
            if vl in ("fecha", "fecha:", "date") and ci + 1 < len(vals):
                _fv = vals[ci + 1].strip()
                if _fv and not fields["fecha"]:
                    fields["fecha"] = _fv
            if vl.startswith("nro") and "pedido" in vl and ci + 1 < len(vals):
                _nv = vals[ci + 1].strip()
                if _nv:
                    fields["numero_oc"] = _nv
        # Si la fecha quedó como float (Excel serial) convertirla
        if fields["fecha"] and not fields["fecha_iso"]:
            try:
                _ffl = float(str(fields["fecha"]))
                import xlrd as _xlrd2
                import datetime as _dt2
                _fdt = _xlrd2.xldate_as_datetime(_ffl, 0)
                fields["fecha"] = _fdt.strftime("%d/%m/%Y")
                fields["fecha_iso"] = _fdt.strftime("%Y-%m-%d")
            except Exception:
                pass

    # ── Paso 1: detectar fila de encabezado de productos ─────────────────
    # Aliases de columnas — normalizamos quitando puntos, espacios y acentos
    def _norm(s):
        s = str(s or "").lower().strip()
        s = s.replace(".", "").replace("á","a").replace("é","e").replace("í","i")
        s  = s.replace("ó","o").replace("ú","u").replace("  "," ")
        return s

    # SKU / código
    _SKU_KW    = {"cod", "code", "codigo", "sku", "ref", "referencia", "art", "articulo"}
    # Descripción / modelo / producto
    _PROD_KW   = {"producto", "productos", "descripcion", "descripcion", "detalle",
                  "modelo", "model", "nombre", "item", "articulo", "art", "denominacion"}
    # Cantidad pedida
    _QTY_KW    = {"cant", "cantidad", "qty", "pedido", "unidades", "u", "ctd",
                  "cant pedida", "cantidad pedida", "order qty", "pedir"}
    # Precio unitario
    _PRICE_KW  = {"imp unit", "precio unit", "precio unitario", "p unit", "unitario",
                  "precio", "pvp", "price", "valor unit", "valor unitario",
                  "imp unit s/iva", "prec unit"}
    # Subtotal / importe total
    _TOTAL_KW  = {"imp total", "total", "subtotal", "importe", "monto",
                  "imp tot", "total linea", "subtot"}
    # Observaciones
    _OBS_KW    = {"obs", "observaciones", "observacion", "nota", "notas", "comment"}

    col_map = {}
    hdr_row_idx = None
    for ri, row in enumerate(all_rows[:30]):
        norms = [_norm(c) for c in row]
        hits = sum(1 for n in norms if (
            n in _SKU_KW or n in _PROD_KW or n in _QTY_KW or
            n in _PRICE_KW or n in _TOTAL_KW
        ))
        if hits >= 2:
            hdr_row_idx = ri
            for ci, n in enumerate(norms):
                if n in _SKU_KW and "sku" not in col_map:
                    col_map["sku"] = ci
                if n in _PROD_KW and "modelo" not in col_map:
                    col_map["modelo"] = ci
                if n in _QTY_KW and "pedido" not in col_map:
                    col_map["pedido"] = ci
                if n in _PRICE_KW and "precio_unit" not in col_map:
                    col_map["precio_unit"] = ci
                if n in _TOTAL_KW and "subtotal" not in col_map:
                    col_map["subtotal"] = ci
                if n in _OBS_KW and "obs" not in col_map:
                    col_map["obs"] = ci
            break

    if "pedido" not in col_map and "modelo" not in col_map:
        return fields

    # Si no hay columna de cantidad pero hay modelo + precio, tomar cantidad = 1
    _has_qty = "pedido" in col_map

    # ── Paso 2: leer filas de datos ───────────────────────────────────────
    _skip_norms = _SKU_KW | _PROD_KW | _QTY_KW | _PRICE_KW | _TOTAL_KW | _OBS_KW
    for row in all_rows[hdr_row_idx + 1:]:
        if not any(c not in (None, "", 0, 0.0) for c in row):
            continue

        def _gc(ci):
            return row[ci] if ci is not None and ci < len(row) else None

        sku_val    = str(_gc(col_map.get("sku"))    or "").strip()
        modelo_val = str(_gc(col_map.get("modelo")) or "").strip()
        obs_val    = str(_gc(col_map.get("obs"))    or "").strip()

        _ident = sku_val or modelo_val
        if not _ident:
            continue
        # Saltar si es otra fila de totales o sub-encabezado
        if _norm(_ident) in _skip_norms:
            continue
        # Saltar filas de totales/descuentos (sin ident numérico ni alfanumérico de producto)
        _ident_up = _ident.upper()
        if any(kw in _ident_up for kw in ("TOTAL", "SUBTOTAL", "DESCUENTO", "DESC.", "IVA", "PLAZO", "EXPRESO", "TEL:", "CALLE")):
            continue

        # Cantidad
        if _has_qty:
            try:
                qty = float(str(_gc(col_map["pedido"]) or "0").replace(",", ".").strip())
            except Exception:
                qty = 0.0
            if qty <= 0:
                continue
        else:
            qty = 1.0

        # Precio unitario
        try:
            precio_unit = float(str(_gc(col_map.get("precio_unit")) or "0").replace(",", ".").strip())
        except Exception:
            precio_unit = 0.0

        # Subtotal
        try:
            subtotal = float(str(_gc(col_map.get("subtotal")) or "0").replace(",", ".").strip())
            if subtotal <= 0:
                subtotal = precio_unit * qty
        except Exception:
            subtotal = precio_unit * qty

        fields["lineas"].append({
            "codigo":      sku_val,
            "modelo":      modelo_val,
            "descripcion": (modelo_val or sku_val)[:200],
            "ean":         "",
            "cantidad":    qty,
            "precio_unit": precio_unit,
            "iva_pct":     21.0,
            "subtotal":    subtotal,
            "obs":         obs_val,
        })

    return fields

def classify_document(text, carpeta_id=""):
    """
    Clasifica un documento de importación.
    Prioridad: sin-texto → no-aplica → CUIT → keyword.
    Retorna dict con: tipo, label, partner_id, journal_id, doc_type,
                      no_aplica (bool), mismatch (bool), extracted (dict).
    """
    _other = {"tipo":"other","label":"Otro comprobante","partner_id":None,
              "journal_id":10,"doc_type":None,
              "no_aplica":False,"mismatch":False,"extracted":{}}

    # ── Sin texto ─────────────────────────────────────────────────────────
    if not text.strip():
        return {**_other, "label":"Sin texto — no aplica", "no_aplica":True}

    tu = text.upper()
    extracted = {}

    # ── Extracción de campos comunes ──────────────────────────────────────
    # Fecha (DD/MM/YYYY, DD-MM-YYYY, DD.MM.YYYY)
    _fm = re.search(r'(\d{2}[/\-\.]\d{2}[/\-\.]\d{4})', text)
    if _fm:
        extracted["fecha"] = parse_ar_date(_fm.group(1))

    # CUIT (con o sin guiones): captura el primero que aparece
    _cm = re.search(r'(\d{2})[-\s]?(\d{8})[-\s]?(\d)\b', text)
    if _cm:
        extracted["cuit"]      = f"{_cm.group(1)}-{_cm.group(2)}-{_cm.group(3)}"
        extracted["cuit_norm"] = _cm.group(1) + _cm.group(2) + _cm.group(3)

    # Monto total
    _am = re.search(r'TOTAL[^\d$]*([\d\.]+,[\d]{2})', tu)
    if _am:
        extracted["monto"] = normalize_amount(_am.group(1))

    # TC desde texto: "Tipo de cambio USD 1 = ARG 1.505,18000" o similar
    _tc_m = re.search(
        r'(?:TIPO\s+DE\s+CAMBIO|T\.?\s*C\.?)\s*.*?(?:USD\s*1\s*=\s*(?:ARG\s*)?)?([\d\.]+,[\d]+)',
        tu, re.DOTALL)
    if _tc_m:
        extracted["tc_pdf"] = normalize_amount(_tc_m.group(1))

    # N° comprobante — patrón argentino: letra + 4 dígitos + guion + 8 dígitos
    _nr = re.search(r'\b([A-Z]\d{4}-\d{8})\b', text)
    if _nr:
        extracted["nro_comp"] = _nr.group(1)
    else:
        _nr2 = re.search(r'N[°º]?\s*(?:COMP\.?|FACTURA|COMPROBANTE)?[:\s]+([A-Z0-9\-]{5,20})', tu)
        if _nr2:
            _nr2_val = _nr2.group(1).strip()
            # Descartar si no tiene ningún dígito (serían palabras como "MONTO", "TOTAL"…)
            if re.search(r'\d', _nr2_val):
                extracted["nro_comp"] = _nr2_val

    # ── No aplica ─────────────────────────────────────────────────────────
    if "VOLANTE ELECTRONICO DE PAGO" in tu or (
            re.search(r'\bVEP\b', tu) and ("PAGO" in tu or "AFIP" in tu)):
        return {**_other, "label":"VEP — no aplica", "no_aplica":True, "extracted":extracted}

    if re.search(r'\bPRESUPUESTO\b', tu) and not re.search(r'\bFACTURA\b', tu):
        return {**_other, "label":"Presupuesto — no aplica", "no_aplica":True, "extracted":extracted}

    if "BILL OF LADING" in tu or ("CONOCIMIENTO" in tu and "EMBARQUE" in tu):
        return {**_other, "label":"Bill of Lading — no aplica", "no_aplica":True, "extracted":extracted}

    # ── Mismatch de carpeta ────────────────────────────────────────────────
    mismatch = False
    if carpeta_id:
        _carp_norm = re.sub(r'[_\s]', '_', carpeta_id.strip().upper())
        _refs = re.findall(r'LUMI[_\s]?\d+[A-Z]?', tu)
        for _r in _refs:
            _r_norm = re.sub(r'[_\s]', '_', _r.strip())
            if _r_norm != _carp_norm:
                mismatch = True
                extracted["mismatch_ref"] = _r_norm
                break

    # ── Líneas de producto PETDUR + N° factura uruguaya ──────────────────────
    _cuit_no_sep = tu.replace("-", "").replace(" ", "")
    if "PETDUR" in tu or "217016440010" in _cuit_no_sep:
        _p_lns = parse_petdur_invoice_lines(text)
        if _p_lns:
            extracted["lineas_petdur"] = _p_lns
        # Número de factura uruguaya: "Factura A 873", "e-Ticket A 873"
        if not extracted.get("nro_comp"):
            # Patrón directo contextual: 'FACTURA A 873' o 'E-TICKET A 873'
            _uy_direct = re.search(r'(?:FACTURA|E-TICKET)\s+([A-Z])\s+(\d+)', tu)
            if _uy_direct:
                extracted["nro_comp"] = f"{_uy_direct.group(1)}{int(_uy_direct.group(2)):04d}"
            else:
                # Fallback: letra sola al inicio de palabra + espacio + dígitos
                _uy2 = re.search(r'(?<![A-Z])([A-Z]) (\d{3,8})(?!\d)', tu)
                if _uy2:
                    extracted["nro_comp"] = f"{_uy2.group(1)}{int(_uy2.group(2)):04d}"

    # ── Clasificación por CUIT ─────────────────────────────────────────────
    cuit_norm = extracted.get("cuit_norm", "")
    # Buscar en todo el texto (por si el PDF tiene CUITs sin guiones)
    _tu_no_sep = tu.replace("-","").replace(" ","")
    for _ck, _cfg in CUIT_TO_PARTNER.items():
        if cuit_norm == _ck or _ck in _tu_no_sep:
            return {**_cfg, "no_aplica":False, "mismatch":mismatch, "extracted":extracted}

    # ── Fallback por keyword ───────────────────────────────────────────────
    _kw = [
        ("PETDUR",                {"tipo":"petdur", "label":"Bill PETDUR (Etapa 1)",  "partner_id":49328,"journal_id":71,"doc_type":None}),
        ("217016440010",          {"tipo":"petdur", "label":"Bill PETDUR (Etapa 1)",  "partner_id":49328,"journal_id":71,"doc_type":None}),
        ("26001IC",               {"tipo":"di_afip","label":"DI AFIP (Etapa 2)",      "partner_id":9,    "journal_id":10,"doc_type":66}),
        ("DECLARACION DE IMPORT", {"tipo":"di_afip","label":"DI AFIP (Etapa 2)",      "partner_id":9,    "journal_id":10,"doc_type":66}),
        ("SUBREGIMEN",            {"tipo":"di_afip","label":"DI AFIP (Etapa 2)",      "partner_id":9,    "journal_id":10,"doc_type":66}),
        ("SENASA",                {"tipo":"nac",    "label":"Bill SENASA (Etapa 2a)", "partner_id":48827,"journal_id":10,"doc_type":None}),
        ("TRICE",                 {"tipo":"nac",    "label":"Bill TRICE Transport (Etapa 2a)", "partner_id":48825,"journal_id":10,"doc_type":None}),
        ("TERMINAL 4",            {"tipo":"nac",    "label":"Bill Terminal 4 SA (Etapa 2a)",   "partner_id":48828,"journal_id":10,"doc_type":None}),
        ("MUNDO COMEX",           {"tipo":"nac",    "label":"Bill Mundo Comex (Etapa 2a)",     "partner_id":48826,"journal_id":10,"doc_type":None}),
    ]
    for _kword, _cfg in _kw:
        if _kword in tu:
            return {**_cfg, "no_aplica":False, "mismatch":mismatch, "extracted":extracted}

    return {**_other, "mismatch":mismatch, "extracted":extracted}


# ───────────────────────────────────────────────────
# SESSION STATE
# ───────────────────────────────────────────────────
DEFAULTS = {
    "logged_in":    False,
    "user_email":   "",
    "odoo_uid":     None,   # uid personal del usuario logueado en Odoo
    "odoo_password": "",    # contraseña Odoo (usada en cada llamada XML-RPC)
    "history": [],
    "carpeta_id": "", "carpeta_po": None, "carpeta_bills": [], "carpeta_lc_id": None,
    "etapas": {k: False for k, *_ in ETAPAS_DEF},
}
for k, v in DEFAULTS.items():
    if k not in st.session_state:
        st.session_state[k] = v


# ─── Entorno activo (producción / testing) ───────────────────────────────
if "odoo_env" not in st.session_state:
    st.session_state["odoo_env"] = "prod"
if st.session_state["odoo_env"] == "test":
    ODOO_URL = TEST_ODOO_URL
    ODOO_DB  = TEST_ODOO_DB

# ═══════════════════════════════════════════════════
# SIDEBAR
# ═══════════════════════════════════════════════════
with st.sidebar:
    import os
    if os.path.exists("logo.png"):
        st.image("logo.png", width=180)
    else:
        st.markdown("""<div class="lumi-sidebar-logo">
  <span class="lumi-logo-dot">🛒</span>
  <span class="lumi-logo-text">LUMINATEC</span>
</div>""", unsafe_allow_html=True)
    st.markdown("---")

    # ── Selector de entorno (solo ivarela@luminatec.com) ────────────────
    _dev_email = "ivarela@luminatec.com"
    _current_email = st.session_state.get("user_email", "")
    # Mostrar el toggle solo si ya está logueado como dev, o si aún no hay sesión
    # pero el email ingresado coincide — para eso lo mostramos siempre antes del login
    # y lo ocultamos post-login si no es el dev.
    _show_env_toggle = (
        not st.session_state.get("logged_in")          # antes del login: visible
        or _current_email == _dev_email                 # o es el dev
    )
    if _show_env_toggle:
        _is_test = st.toggle(
            "🧪 Modo testing",
            value=(st.session_state.get("odoo_env") == "test"),
            help="Entorno de prueba — solo disponible para el administrador",
        )
        if _is_test and st.session_state.get("odoo_env") != "test":
            st.session_state["odoo_env"] = "test"
            st.session_state.logged_in    = False
            st.session_state.odoo_uid     = None
            st.session_state.odoo_password = ""
            st.rerun()
        elif not _is_test and st.session_state.get("odoo_env") != "prod":
            st.session_state["odoo_env"] = "prod"
            st.session_state.logged_in    = False
            st.session_state.odoo_uid     = None
            st.session_state.odoo_password = ""
            st.rerun()
        if st.session_state.get("odoo_env") == "test":
            st.warning("🧪 **Entorno de TESTING**")
    else:
        # Usuario no-dev logueado: forzar producción silenciosamente
        if st.session_state.get("odoo_env") == "test":
            st.session_state["odoo_env"] = "prod"
            st.session_state.logged_in    = False
            st.session_state.odoo_uid     = None
            st.session_state.odoo_password = ""
            st.rerun()
    st.markdown("---")

    if not st.session_state.logged_in:
        st.markdown("### 🔐 Iniciar sesión")
        st.caption(f"`{ODOO_URL}`")
        with st.form("login_form"):
            email_in = st.text_input("Email", placeholder="tu@luminatec.com")
            pass_in  = st.text_input("Contraseña", type="password", placeholder="••••••••")
            login_btn = st.form_submit_button("Ingresar", use_container_width=True)
        if login_btn:
            if email_in and pass_in:
                with st.spinner("Verificando en Odoo..."):
                    _uid, _err = odoo_authenticate(email_in, pass_in)
                if _uid:
                    st.session_state.logged_in    = True
                    st.session_state.user_email   = email_in.strip().lower()
                    st.session_state.odoo_uid      = _uid
                    st.session_state.odoo_password = pass_in
                    st.rerun()
                else:
                    st.error(_err or "Email o contraseña incorrectos.")
            else:
                st.warning("Completá email y contraseña.")
    else:
        is_admin = st.session_state.user_email in ADMIN_EMAILS
        st.success("✅ Sesión activa")
        st.markdown(
            '<span class="user-chip">👤 ' + st.session_state.user_email + '</span>',
            unsafe_allow_html=True
        )
        if is_admin:
            st.markdown('<span class="admin-badge">🛳️ Importaciones habilitado</span>',
                        unsafe_allow_html=True)
            st.caption(f"Carpeta: **{st.session_state.carpeta_id or 'sin selección'}**")
        if st.button("🔓 Cerrar sesión", use_container_width=True):
            st.session_state.logged_in    = False
            st.session_state.user_email   = ""
            st.session_state.odoo_uid      = None
            st.session_state.odoo_password = ""
            st.session_state.history       = []
            st.session_state.carpeta_id    = ""
            st.session_state.carpeta_po    = None
            st.session_state.carpeta_bills = []
            st.session_state.carpeta_lc_id = None
            st.session_state.etapas = {k: False for k, *_ in ETAPAS_DEF}
            st.rerun()

    st.divider()
    _env_label = "🟢 Producción" if st.session_state.get("odoo_env","prod") == "prod" else "🧪 Testing"
    st.caption(f"{_env_label} · `{ODOO_DB}`")


# ═══════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════
st.markdown('<h1 class="main-title">🛒 <span>LUMINA</span>TEC · Carga Odoo</h1>', unsafe_allow_html=True)
st.caption("Facturas de proveedores, pedidos de clientes e importaciones — todo en un lugar.")

if not st.session_state.logged_in:
    st.info("👈 Iniciá sesión desde el panel lateral para empezar.")
    st.stop()

# ── Keepalive: evita que Streamlit cierre la sesión por inactividad ───────────
import streamlit.components.v1 as _stc
_stc.html("""
<script>
(function() {
    // Cada 2 minutos hace un fetch al healthcheck de Streamlit para mantener
    // el WebSocket activo. Funciona desde iframe (mismo origen).
    function ping() {
        fetch('/_stcore/health', {method: 'GET', cache: 'no-store'})
            .catch(function() {
                // Fallback: intentar raíz si /_stcore/health no responde
                fetch(window.location.origin + '/', {method: 'GET',
                    mode: 'no-cors', cache: 'no-store'}).catch(function(){});
            });
    }
    ping(); // ping inicial
    setInterval(ping, 120000); // luego cada 2 minutos
})();
</script>
""", height=0, scrolling=False)

# Conexión usando las credenciales del usuario logueado
uid        = st.session_state.odoo_uid
api_key    = st.session_state.odoo_password   # en XML-RPC, password == api_key
models     = get_models_proxy()
models_url = f"{ODOO_URL}/xmlrpc/2/object"

if not uid or not api_key:
    st.error("⚠️ Sesión inválida. Por favor cerrá sesión y volvé a ingresar.")
    st.stop()

is_admin = st.session_state.user_email in ADMIN_EMAILS


# ───────────────────────────────────────────────────────────────────────────
# ORDENES DE PAGO — helpers
# ───────────────────────────────────────────────────────────────────────────
@st.cache_data(ttl=300, show_spinner=False)
def get_pending_bills(models_url, uid, api_key):
    """Todas las FAs de proveedor confirmadas y con saldo pendiente."""
    try:
        m = xmlrpc.client.ServerProxy(models_url, allow_none=True)
        rows = m.execute_kw(ODOO_DB, uid, api_key, "account.move", "search_read",
            [[("move_type",     "=",  "in_invoice"),
              ("state",         "=",  "posted"),
              ("payment_state", "in", ["not_paid", "partial"])]],
            {"fields": ["id", "name", "ref", "partner_id", "invoice_date",
                        "invoice_date_due", "amount_total", "amount_residual",
                        "currency_id", "payment_state", "journal_id"],
             "order":  "invoice_date asc",
             "limit":  300})
        return rows
    except Exception as e:
        return []

@st.cache_data(ttl=300, show_spinner=False)
def get_payment_journals(models_url, uid, api_key):
    """
    Diarios banco/caja activos en ARS + MercadoPago.
    Excluye diarios con moneda extranjera explícita (USD, EUR, etc.).
    Incluye diarios sin moneda (= ARS de la empresa) y los que dicen ARS/PESO.
    """
    try:
        m = xmlrpc.client.ServerProxy(models_url, allow_none=True)
        rows = m.execute_kw(ODOO_DB, uid, api_key, "account.journal", "search_read",
            [[]],
            {"fields": ["id", "name", "currency_id", "type", "active"], "order": "name asc",
             "context": {"active_test": False}})
        result = []
        for r in rows:
            if r.get("type") not in ("bank", "cash"):
                continue
            if not r.get("active", True):
                continue
            cur       = r.get("currency_id")
            has_cur   = bool(cur and isinstance(cur, (list, tuple)) and cur[0])
            cur_name  = cur[1] if has_cur else "ARS"
            cur_upper = cur_name.upper()
            name_upper = r["name"].upper()
            is_mp      = "MERCADOPAGO" in name_upper or "MERCADO PAGO" in name_upper
            is_foreign = has_cur and not any(k in cur_upper for k in ("ARS", "PESO"))
            if is_foreign and not is_mp:
                continue
            result.append((r["id"], r["name"], cur_name))
        return result   # list of (id, label, currency_name)
    except Exception as e:
        return []


@st.cache_data(ttl=3600, show_spinner=False)
def get_all_banks(_models_url, uid, api_key):
    """Lista de (id, name) de res.bank para matchear bancos de cheques."""
    try:
        m = xmlrpc.client.ServerProxy(_models_url, allow_none=True)
        rows = m.execute_kw(ODOO_DB, uid, api_key, "res.bank", "search_read",
            [[]], {"fields": ["id", "name"], "limit": 500, "order": "name asc"})
        return [(r["id"], r["name"]) for r in rows]
    except Exception:
        return []


def match_bank_id(bank_name, all_banks):
    """Nombre de banco del home banking -> res.bank.id o None."""
    if not bank_name or not all_banks:
        return None
    clean = re.sub(r"\s*-\s*\d{5,}\s*$", "", bank_name.strip()).upper()
    stop = {"BANCO","DEL","DE","LA","LAS","LOS","EL","Y","SA","SRL"}
    words = [w for w in re.sub(r"[^\w\s]","",clean).split() if len(w)>=4 and w not in stop]
    kw = words[0].lower() if words else clean[:15].lower()
    for bid, bname in all_banks:
        if kw in bname.lower(): return bid
    return None


def register_payment_wizard(models, uid, api_key, move_ids, payment_date, journal_id):
    """
    Genera un pago via el wizard account.payment.register.
    Funciona para uno o varios bills del mismo proveedor/moneda.
    """
    ctx = {
        "active_model": "account.move",
        "active_ids":   move_ids,
        "active_id":    move_ids[0] if move_ids else None,
    }
    try:
        wiz_id = models.execute_kw(ODOO_DB, uid, api_key,
            "account.payment.register", "create",
            [{"payment_date": payment_date, "journal_id": journal_id}],
            {"context": ctx})
        result = models.execute_kw(ODOO_DB, uid, api_key,
            "account.payment.register", "action_create_payments",
            [[wiz_id]], {"context": ctx})
        return True, result
    except Exception as e:
        return False, str(e)

def _get_payment_doc_type(models, uid, api_key, payment_type, partner_type, journal_id):
    """Devuelve el l10n_latam_document_type_id correcto para un pago usando default_get de Odoo.
    Esto garantiza que el pago aparezca en Clientes>Recibos o Proveedores>Órdenes de pago."""
    try:
        ctx = {
            "default_payment_type": payment_type,
            "default_partner_type": partner_type,
            "default_journal_id":   journal_id,
        }
        defaults = models.execute_kw(ODOO_DB, uid, api_key, "account.payment", "default_get",
            [["l10n_latam_document_type_id"]], {"context": ctx})
        return defaults.get("l10n_latam_document_type_id") or False
    except Exception:
        return False

def create_advance_payment(models, uid, api_key, partner_id, amount,
                           currency_id, payment_date, journal_id, memo=""):
    """Pago a cuenta: crea y confirma un pago SIN vincular a ninguna FA.
    Setea l10n_latam_document_type_id para que aparezca en Proveedores > Órdenes de pago."""
    vals = {
        "payment_type": "outbound",
        "partner_type": "supplier",
        "partner_id":   partner_id,
        "amount":       float(amount),
        "currency_id":  currency_id,
        "date":         payment_date,
        "journal_id":   journal_id,
        "ref":          memo or "",
    }
    _doc_type = _get_payment_doc_type(models, uid, api_key, "outbound", "supplier", journal_id)
    if _doc_type:
        vals["l10n_latam_document_type_id"] = _doc_type
    pay_id = models.execute_kw(ODOO_DB, uid, api_key, "account.payment", "create", [vals])
    models.execute_kw(ODOO_DB, uid, api_key, "account.payment", "action_post", [[pay_id]])
    return pay_id

@st.cache_data(ttl=300, show_spinner=False)
def get_pending_expense_sheets(models_url, uid, api_key):
    """Notas de gastos aprobadas pendientes de pago (hr.expense.sheet state=post)."""
    try:
        m = xmlrpc.client.ServerProxy(models_url, allow_none=True)
        rows = m.execute_kw(ODOO_DB, uid, api_key, "hr.expense.sheet", "search_read",
            [[("state", "=", "post")]],
            {"fields": ["id", "name", "employee_id", "total_amount",
                        "currency_id", "payment_state"],
             "order": "id asc", "limit": 100})
        return rows
    except Exception:
        return []

def register_expense_payment(models, uid, api_key, sheet_id, payment_date, journal_id):
    """Registra el pago de una nota de gastos aprobada via el wizard de Odoo."""
    ctx = {
        "active_model": "hr.expense.sheet",
        "active_id":    sheet_id,
        "active_ids":   [sheet_id],
    }
    try:
        wiz_id = models.execute_kw(ODOO_DB, uid, api_key,
            "account.payment.register", "create",
            [{"payment_date": payment_date, "journal_id": journal_id}],
            {"context": ctx})
        result = models.execute_kw(ODOO_DB, uid, api_key,
            "account.payment.register", "action_create_payments",
            [[wiz_id]], {"context": ctx})
        return True, result
    except Exception as e:
        return False, str(e)


@st.cache_data(ttl=180, show_spinner=False)
def search_partners_by_cuits(models_url, uid, api_key, cuits_tuple):
    """Busca socios en Odoo por tupla de CUITs. Retorna dict {cuit_sin_guiones: (id, name)}.
    Maneja que Odoo puede guardar el VAT con guiones (30-71189948-7) o sin (30711899487)."""
    try:
        m = xmlrpc.client.ServerProxy(models_url, allow_none=True)

        def _cuit_variants(c):
            """Devuelve variantes de un CUIT para cubrir todos los formatos de Odoo:
            - sin guiones:        30711899487
            - con guiones:        30-71189948-7
            - con prefijo AR:     AR30711899487   (Odoo l10n_ar)
            - con prefijo AR+gu:  AR30-71189948-7
            """
            clean = str(c).replace("-", "").replace(" ", "").replace(".", "").strip()
            # Si ya viene con prefijo AR, quitarlo para la base
            if clean.upper().startswith("AR"):
                clean = clean[2:]
            variants = [clean]
            if len(clean) == 11 and clean.isdigit():
                fmt = f"{clean[:2]}-{clean[2:10]}-{clean[10]}"
                variants.append(fmt)
                variants.append(f"AR{clean}")
                variants.append(f"AR{fmt}")
            return variants

        norm_set = set()   # CUITs sin guiones para lookup final
        all_variants = []  # todas las variantes para el query
        for c in cuits_tuple:
            vs = _cuit_variants(c)
            norm_set.add(vs[0])
            all_variants.extend(vs)

        rows = m.execute_kw(ODOO_DB, uid, api_key, "res.partner", "search_read",
            [[("vat", "in", all_variants), ("active", "=", True)]],
            {"fields": ["id", "name", "vat", "is_company", "parent_id",
                        "customer_rank", "type"], "limit": 300})

        def _partner_score(r):
            """Mayor puntaje = registro más apropiado para asociar un cobro.
            Prioriza: cliente activo > empresa raíz > sin tipo especial."""
            score = 0
            if (r.get("customer_rank") or 0) > 0: score += 100
            if r.get("is_company"):                score += 50
            if not r.get("parent_id"):             score += 30
            if r.get("type") in (False, "contact", "other"): score += 10
            return score

        # Agrupar por CUIT normalizado y quedarse con el de mayor score
        _candidates = {}   # cuit_clean → list of rows
        for r in rows:
            vat_raw   = (r.get("vat") or "").strip().upper()
            if vat_raw.startswith("AR"):
                vat_raw = vat_raw[2:]
            vat_clean = vat_raw.replace("-", "").replace(" ", "")
            if vat_clean in norm_set:
                _candidates.setdefault(vat_clean, []).append(r)

        result = {}
        for vat_clean, candidates in _candidates.items():
            best = max(candidates, key=_partner_score)
            result[vat_clean] = (best["id"], best["name"])
        return result
    except Exception:
        return {}

@st.cache_data(ttl=180, show_spinner=False)
def get_customer_unpaid_invoices(models_url, uid, api_key, partner_ids_tuple):
    """Facturas de cliente sin pagar (posted, not_paid/partial/in_payment).
    Usa child_of para incluir contactos secundarios del mismo socio.
    Incluye in_payment porque en Odoo AR muchas FAs quedan en ese estado
    cuando el pago está registrado pero sin conciliar con extracto bancario."""
    try:
        m = xmlrpc.client.ServerProxy(models_url, allow_none=True)
        rows = m.execute_kw(ODOO_DB, uid, api_key, "account.move", "search_read",
            [[("move_type", "=", "out_invoice"),
              ("state", "=", "posted"),
              ("payment_state", "in", ["not_paid", "partial", "in_payment"]),
              ("partner_id", "child_of", list(partner_ids_tuple))]],
            {"fields": ["id", "name", "invoice_date", "invoice_date_due",
                        "amount_total", "amount_residual", "currency_id", "partner_id"],
             "order": "invoice_date asc", "limit": 300})
        return rows
    except Exception as _e:
        st.warning(f"⚠️ Error al cargar facturas pendientes: {_e}")
        return []

def register_customer_payment(models, uid, api_key,
                               partner_id, amount, currency_id,
                               payment_date, journal_id,
                               move_ids=None, memo="", cheques=None,
                               withholdings=None):
    """Registra un recibo de cobro de cliente usando account.payment.group.
    Flujo correcto para Odoo AR (módulo account_payments_group):
      1. Crear account.payment.group con payment_type='receivable'
      2. Crear account.payment vinculado al grupo (payment_group_id)
      3. Opcionalmente linkear facturas (move_line_ids)
      4. Llamar post() en el grupo para confirmar
    Esto garantiza que el recibo aparezca en Clientes > Recibos."""
    try:
        # 1. Obtener las líneas receivable de las facturas seleccionadas
        inv_line_ids = []
        if move_ids:
            inv_lines = models.execute_kw(ODOO_DB, uid, api_key,
                "account.move.line", "search_read",
                [[("move_id", "in", move_ids),
                  ("account_id.account_type", "=", "asset_receivable"),
                  ("reconciled", "=", False)]],
                {"fields": ["id"], "limit": 100})
            inv_line_ids = [l["id"] for l in inv_lines]

        # 2. Crear el grupo (solo campos del grupo, sin pagos inline)
        group_vals = {
            "payment_type": "receivable",
            "partner_id":   partner_id,
            "currency_id":  currency_id,
            "date":         payment_date,
        }
        if inv_line_ids:
            group_vals["move_line_ids"] = [(6, 0, inv_line_ids)]
        group_id = models.execute_kw(ODOO_DB, uid, api_key,
            "account.payment.group", "create", [group_vals])

        # 3. Crear el payment vinculado al grupo
        #    Nota: en esta instalación el campo se llama "memo" (no "ref")

        # 3a. Buscar la linea de metodo de pago "Cheque de Terceros Existente"
        #     para el journal seleccionado (code: out_third_party_checks = Existing Third Party Checks)
        try:
            pml_lines = models.execute_kw(ODOO_DB, uid, api_key,
                "account.payment.method.line", "search_read",
                [[["journal_id", "=", journal_id],
                  ["payment_method_id.code", "=", "new_third_party_checks"]]],
                {"fields": ["id"], "limit": 1})
            pml_id = pml_lines[0]["id"] if pml_lines else None
        except Exception:
            pml_id = None

        # Si hay retenciones Y cheques, el pago principal debe ser el monto TOTAL
        # de los cheques. Odoo AR valida: payment.amount == sum(cheques.amount).
        # Las retenciones se crean como pagos adicionales en el mismo grupo.
        _cheque_total = sum(float(ch.get("amount") or 0) for ch in (cheques or []))
        _pay_amount = _cheque_total if (cheques and withholdings) else float(amount)

        pay_vals = {
            "payment_type":   "inbound",
            "partner_type":   "customer",
            "partner_id":     partner_id,
            "amount":         _pay_amount,
            "currency_id":    currency_id,
            "date":           payment_date,
            "journal_id":     journal_id,
            "memo":           memo or "",
            "payment_group_id": group_id,
        }
        if pml_id:
            pay_vals["payment_method_line_id"] = pml_id

        # 3b. Adjuntar cheques inline (l10n_latam_new_check_ids)
        if cheques:
            check_lines = []
            for ch in cheques:
                ch_vals = {
                    "payment_date": ch.get("payment_date") or payment_date,
                    "amount":       float(ch.get("amount") or amount),
                }
                if ch.get("nro"):       ch_vals["name"] = str(ch["nro"])
                if ch.get("bank_id"):   ch_vals["bank_id"] = ch["bank_id"]
                if ch.get("issuer_vat"):ch_vals["issuer_vat"] = str(ch["issuer_vat"])
                if pml_id:             ch_vals["payment_method_line_id"] = pml_id
                check_lines.append((0, 0, ch_vals))
            pay_vals["l10n_latam_new_check_ids"] = check_lines

        models.execute_kw(ODOO_DB, uid, api_key,
            "account.payment", "create", [pay_vals])

        # 3c. Pagos adicionales por retenciones (en el mismo grupo)
        if withholdings:
            # Cargar todos los diarios generales/cash una sola vez
            _all_jrnls = []
            try:
                _all_jrnls = models.execute_kw(ODOO_DB, uid, api_key,
                    "account.journal", "search_read",
                    [[("type", "in", ["general", "cash"])]],
                    {"fields": ["id", "name"], "limit": 100})
            except Exception:
                pass

            def _find_journal_for_wh(concepto_str):
                """Elige el diario más específico para la retención dada su descripción."""
                _clow = (concepto_str or "").lower()
                # Provincias argentinas para matchear en nombre de diario
                _provs = [
                    "misiones", "buenos aires", "cordoba", "córdoba",
                    "santa fe", "mendoza", "corrientes", "chaco",
                    "formosa", "salta", "jujuy", "tucuman", "tucumán",
                    "catamarca", "rioja", "san juan", "san luis",
                    "neuquen", "neuquén", "rio negro", "chubut",
                    "santa cruz", "entre rios", "la pampa", "tierra del fuego",
                ]
                # 1. Diario que coincide en provincia + "retenc"/"iibb"
                for _prov in _provs:
                    if _prov in _clow:
                        for _j in _all_jrnls:
                            _jlow = _j["name"].lower()
                            if _prov in _jlow and ("retenc" in _jlow or "iibb" in _jlow):
                                return _j["id"]
                # 2. Diario que coincide solo en provincia
                for _prov in _provs:
                    if _prov in _clow:
                        for _j in _all_jrnls:
                            if _prov in _j["name"].lower():
                                return _j["id"]
                # 3. Cualquier diario con "retenc" o "iibb"
                for _kw in ["retenc", "retencion", "iibb"]:
                    for _j in _all_jrnls:
                        if _kw in _j["name"].lower():
                            return _j["id"]
                # 4. Fallback misc/general
                for _kw in ["misc", "varios", "general", "op"]:
                    for _j in _all_jrnls:
                        if _kw in _j["name"].lower():
                            return _j["id"]
                return _all_jrnls[0]["id"] if _all_jrnls else None

            _wh_errors = []
            for _wh in withholdings:
                _wh_amt = float(_wh.get("monto") or _wh.get("amount") or 0)
                if _wh_amt <= 0:
                    continue
                # Diario específico según provincia/concepto de esta retención
                _ret_journal_id = _find_journal_for_wh(_wh.get("concepto", ""))
                if not _ret_journal_id:
                    _wh_errors.append(f"Sin diario para retención: {_wh.get('concepto','')}")
                    continue
                # payment_method_line_id para este diario
                _wh_pml_id = None
                try:
                    _wh_pml_lines = models.execute_kw(ODOO_DB, uid, api_key,
                        "account.payment.method.line", "search_read",
                        [[["journal_id", "=", _ret_journal_id]]],
                        {"fields": ["id"], "limit": 1})
                    _wh_pml_id = _wh_pml_lines[0]["id"] if _wh_pml_lines else None
                except Exception:
                    pass
                # Pago outbound en el grupo: reduce el neto aplicado a la factura
                _wh_vals = {
                    "payment_type":     "outbound",
                    "partner_type":     "customer",
                    "partner_id":       partner_id,
                    "amount":           _wh_amt,
                    "currency_id":      currency_id,
                    "date":             payment_date,
                    "journal_id":       _ret_journal_id,
                    "memo":             _wh.get("concepto") or "Retención",
                    "payment_group_id": group_id,
                }
                if _wh_pml_id:
                    _wh_vals["payment_method_line_id"] = _wh_pml_id
                try:
                    models.execute_kw(ODOO_DB, uid, api_key,
                        "account.payment", "create", [_wh_vals])
                except Exception as _whe:
                    _wh_errors.append(str(_whe))
            if _wh_errors:
                # Retornar los errores como advertencia (el recibo principal ya quedó)
                return True, f"__WH_WARN__{'|'.join(_wh_errors)}"

        # 4. Confirmar el grupo
        # Nota: post() retorna None en esta instalacion, lo que causa un error
        # de marshalling en XML-RPC. Se ignora ese error especifico y se verifica
        # el estado real del grupo para confirmar que se confirmo correctamente.
        try:
            models.execute_kw(ODOO_DB, uid, api_key,
                "account.payment.group", "post", [[group_id]])
        except Exception as post_err:
            err_str = str(post_err)
            if "marshal" in err_str.lower() or "none" in err_str.lower() or "nil" in err_str.lower():
                # post() retorno None -> XML-RPC no puede serializarlo,
                # pero la accion se ejecuto. Verificar estado real.
                try:
                    grp_check = models.execute_kw(ODOO_DB, uid, api_key,
                        "account.payment.group", "read",
                        [[group_id]], {"fields": ["state", "name"]})
                    if grp_check and grp_check[0].get("state") == "posted":
                        return True, group_id  # Confirmado OK a pesar del error XML-RPC
                except Exception:
                    pass
            raise  # Re-lanzar si no es el error esperado

        return True, group_id
    except Exception as e:
        return False, str(e)


_tabs = ["🧾 Facturas prov.", "📦 Pedidos", "🏦 Órdenes de Pago", "💰 Recibos de Cobro", "👥 Contactos"]
if is_admin:
    _tabs.append("🛳️ Importaciones")
_tabs.append("🤖 Asistente")
_tabs.append("📋 Historial")
_tab_objs = st.tabs(_tabs)
if is_admin:
    tab_bills, tab_orders, tab_op, tab_recibos, tab_contacts, tab_import, tab_chat, tab_history = _tab_objs
else:
    tab_bills, tab_orders, tab_op, tab_recibos, tab_contacts, tab_chat, tab_history = _tab_objs
    tab_import = None
    tab_import = None


# ═══════════════════════════════════════════════════
# TAB 1 — FACTURAS DE PROVEEDORES
# ═══════════════════════════════════════════════════
with tab_bills:
    st.subheader("Facturas de proveedores")
    files = st.file_uploader("Arrastrá o elegí archivos (PDF, JPG, PNG, XLSX)",
        type=["pdf","jpg","jpeg","png","xlsx","xls"], accept_multiple_files=True, key="bills_upload")
    if not files:
        st.caption("Subí uno o más archivos para empezar.")
    _total_upfiles = len(files) if files else 0
    if _total_upfiles > 1:
        st.caption(f"📂 {_total_upfiles} archivo(s) cargados — procesando uno por uno.")
    for _uf_idx, uf in enumerate(files or []):
        st.divider()
        ext        = uf.name.rsplit(".", 1)[-1].lower()
        file_bytes = uf.read()
        mimetype   = MIMETYPES.get(ext, "application/octet-stream")
        _file_lbl = f"({_uf_idx + 1}/{_total_upfiles}) " if _total_upfiles > 1 else ""
        st.markdown(f"**📎 {_file_lbl}{uf.name}**  `{ext.upper()}`  ({len(file_bytes)//1024} KB)")
        if ext in ("xlsx", "xls"):
            # ── Detección temprana: ¿es un Excel de pedido de cliente? ───────
            _oc_check = extract_excel_oc_fields(file_bytes)
            if _oc_check.get("lineas"):
                st.warning(
                    f"⚠️ **Este Excel parece un pedido de cliente** — se detectaron "
                    f"{len(_oc_check['lineas'])} producto(s) con columna de cantidad/pedido.\n\n"
                    "Este archivo **no debe cargarse como factura de proveedor**. "
                    "Por favor subilo en la pestaña **📦 Pedidos** para crear el pedido en Ventas."
                )
                continue
            try:
                df = pd.read_excel(BytesIO(file_bytes), dtype=str).fillna("")
                df.columns = [c.strip() for c in df.columns]
            except Exception as e:
                st.error(f"No se pudo leer el Excel: {e}"); continue
            st.caption(f"📊 {len(df)} filas · Columnas: {', '.join(df.columns)}")
            st.dataframe(df.head(10), use_container_width=True, height=180)
            cols_opts = ["(ninguna)"] + list(df.columns)
            c1, c2, c3, c4 = st.columns(4)
            col_prov  = c1.selectbox("Proveedor",    cols_opts, key=f"bp_{uf.name}")
            col_fecha = c2.selectbox("Fecha",         cols_opts, key=f"bf_{uf.name}")
            col_ref   = c3.selectbox("N° Factura",   cols_opts, key=f"br_{uf.name}")
            col_total = c4.selectbox("Total (info)",  cols_opts, key=f"bt_{uf.name}")
            if st.button(f"⬆️ Cargar {len(df)} facturas en Odoo", key=f"load_bills_xls_{uf.name}"):
                bar = st.progress(0)
                ok, errs = 0, []
                for i, row in df.iterrows():
                    try:
                        prov_name  = row.get(col_prov, "")  if col_prov  != "(ninguna)" else ""
                        fecha      = row.get(col_fecha, "") if col_fecha != "(ninguna)" else ""
                        ref        = row.get(col_ref, "")   if col_ref   != "(ninguna)" else ""
                        partner_id = False
                        if prov_name:
                            m2 = search_partners(models_url, uid, api_key, prov_name, limit=1)
                            partner_id = m2[0][0] if m2 else False
                        move_id = create_vendor_bill(models, uid, api_key,
                            partner_id=partner_id, ref=str(ref),
                            invoice_date=str(fecha) if fecha else False,
                            filename=f"{uf.name}_fila{i+1}.pdf",
                            file_bytes=file_bytes, mimetype=mimetype)
                        ok += 1
                        url = odoo_url("account.move", move_id)
                        st.session_state.history.append({"tipo":"Factura proveedor",
                            "archivo":f"{uf.name}·fila{i+1}","id":move_id,"url":url,"estado":"✅","hora":_dt_now.now().strftime("%H:%M")})
                    except Exception as e:
                        errs.append(f"Fila {i+1}: {str(e)[:100]}")
                    bar.progress((i+1)/len(df))
                if ok: st.toast(f"{ok} de {len(df)} facturas creadas en Odoo.", icon="✅")
                for err in errs: st.warning(err)
        else:
            extracted, raw_text = {}, ""
            if ext == "pdf":
                with st.spinner(f"Analizando PDF con IA... {_file_lbl}"):
                    extracted, raw_text = extract_pdf_fields(file_bytes)
                _src_tag = "✨ IA" if extracted.get("_source") == "ai" else "🔣 Regex"
                if extracted.get("proveedor") or extracted.get("numero"):
                    st.caption(f"🤖 Datos detectados [{_src_tag}] — revisá antes de confirmar.")
                else:
                    st.caption("ℹ️ PDF sin texto extraíble. Completá los datos a mano.")
            elif ext in ("jpg","jpeg","png"):
                st.image(file_bytes, caption="Vista previa", width=420)
                with st.spinner(f"Leyendo imagen con OCR... {_file_lbl}"):
                    extracted, raw_text = extract_image_fields(file_bytes)
                st.caption("🤖 Datos detectados por OCR — revisá antes de confirmar." if extracted.get("proveedor")
                           else "ℹ️ OCR no detectó datos. Completá los campos a mano.")

            # Cargar cuentas contables (cacheado)
            _bill_accounts = get_all_accounts(models_url, uid, api_key)
            _acct_labels   = ["— Sin cuenta —"] + [lbl for _, lbl in _bill_accounts]

            # ── CUIT editable en tiempo real (fuera del form) ────────────────────
            _cuit_raw    = extracted.get("cuit", "")
            _cond_venta  = extracted.get("condiciones_venta", "")
            _dias_pago   = extracted.get("dias_pago")
            _vto_auto    = extracted.get("fecha_vencimiento", "")

            _cuit_edit_key = f"bill_cuit_edit_{uf.name}"
            if _cuit_edit_key not in st.session_state:
                st.session_state[_cuit_edit_key] = _cuit_raw

            _cuit_effective = st.text_input(
                "CUIT del proveedor",
                key=_cuit_edit_key,
                placeholder="30-12345678-9",
                help="Detectado automáticamente del PDF. Corregilo si es necesario.",
            )
            _cuit_raw  = _cuit_effective.strip() if _cuit_effective else ""
            _cuit_norm = _cuit_raw.replace("-","").replace(" ","").strip()
            _ss_vendor_key = f"vendor_created_{_cuit_norm}"

            # Buscar primero en session state (recién creado), luego en Odoo
            _partner_preloaded = st.session_state.get(_ss_vendor_key)
            if not _partner_preloaded and _cuit_raw:
                _partner_preloaded = search_partner_by_cuit(models_url, uid, api_key, _cuit_raw)

            # Pre-selección de cuenta contable según proveedor
            _default_acct_idx = 0
            if _partner_preloaded:
                _def_acct = get_partner_default_account(models_url, uid, api_key, _partner_preloaded[0])
                if _def_acct:
                    _def_acct_label = _def_acct[1]
                    for _i, _lbl in enumerate(_acct_labels):
                        if _lbl == _def_acct_label:
                            _default_acct_idx = _i
                            break

            # Cargar productos de gasto y calcular default del proveedor
            _expense_products = get_expense_products(models_url, uid, api_key)
            _prod_labels = ["— Sin producto —"] + [lbl for _, lbl in _expense_products]
            _default_prod_idx = 0
            if _partner_preloaded:
                _def_prod = get_partner_default_product(models_url, uid, api_key, _partner_preloaded[0])
                if _def_prod:
                    _def_prod_label = _def_prod[1]
                    for _pi, _plbl in enumerate(_prod_labels):
                        if _plbl == _def_prod_label:
                            _default_prod_idx = _pi
                            break

            # Cargar cuentas analíticas (Centros de Costo)
            _analytic_accounts = get_analytic_accounts(models_url, uid, api_key)
            _analytic_labels   = ["— Sin centro de costo —"] + [lbl for _, lbl in _analytic_accounts]

            # ── Estado del proveedor + opción de crear nuevo ──────────────────────
            _create_new_vend_key = f"bill_create_new_vend_{uf.name}"
            # Inicializar siempre en False para evitar que quede marcado de sesiones anteriores
            if _create_new_vend_key not in st.session_state:
                st.session_state[_create_new_vend_key] = False

            if _partner_preloaded:
                st.info(f"🏢 Proveedor detectado por CUIT: **{_partner_preloaded[1]}**")
            elif _cuit_raw:
                # CUIT ingresado pero no encontrado → ofrecer crear
                st.warning(f"⚠️ CUIT **{_cuit_raw}** no encontrado en Odoo.")
                st.checkbox("➕ Crear nuevo proveedor en Odoo", key=_create_new_vend_key)
            else:
                # Sin CUIT → permitir buscar por nombre o CUIT, o crear nuevo proveedor
                _search_query = st.text_input(
                    "🔍 Buscar proveedor por nombre o CUIT",
                    placeholder="Ej: Acme SA  ó  30-12345678-9",
                    key=f"vend_search_{uf.name}",
                    help="Buscá en Odoo por nombre o CUIT del proveedor",
                )
                if _search_query and _search_query.strip():
                    _srch_results = search_partners(
                        models_url, uid, api_key, _search_query.strip(), limit=6)
                    if _srch_results:
                        _srch_opts = {f"{r[1]} ({r[2]})": r[0] for r in _srch_results}
                        _chosen_lbl = st.selectbox(
                            "Proveedor encontrado",
                            options=list(_srch_opts.keys()),
                            key=f"vend_sel_{uf.name}",
                        )
                        if _chosen_lbl:
                            st.session_state[f"partner_override_{uf.name}"] = _srch_opts[_chosen_lbl]
                            st.session_state[f"partner_name_{uf.name}"]     = _chosen_lbl
                    else:
                        st.info("No se encontraron proveedores. Podés crear uno nuevo.")
                st.checkbox("➕ Crear nuevo proveedor en Odoo", key=_create_new_vend_key)

            if st.session_state.get(_create_new_vend_key):
                with st.expander("📝 Datos del nuevo proveedor", expanded=True):
                    with st.form(key=f"new_vendor_form_{uf.name}"):
                        st.markdown("Completá los datos mínimos para dar de alta el proveedor:")
                        _nv_c1, _nv_c2 = st.columns(2)
                        _nv_name  = _nv_c1.text_input("Razón social *",
                            value=extracted.get("proveedor","")[:80],
                            placeholder="Nombre en Odoo")
                        _nv_cuit  = _nv_c2.text_input("CUIT *",
                            value=_cuit_raw,
                            placeholder="30-12345678-9")
                        _nv_street = _nv_c1.text_input("Dirección", placeholder="Av. Corrientes 1234")
                        _nv_phone  = _nv_c2.text_input("Teléfono", placeholder="+54 11 4xxx-xxxx")
                        _nv_email  = st.text_input("E-mail", placeholder="proveedor@empresa.com")
                        _nv_go = st.form_submit_button("Crear proveedor en Odoo", use_container_width=True)
                    if _nv_go:
                        if not _nv_name.strip() or not _nv_cuit.strip():
                            st.error("Razón social y CUIT son obligatorios.")
                        else:
                            try:
                                _nv_pid = create_vendor_partner(
                                    models, uid, api_key,
                                    name=_nv_name.strip(),
                                    vat=_nv_cuit.strip().replace("-",""),
                                    street=_nv_street.strip(),
                                    phone=_nv_phone.strip(),
                                    email_addr=_nv_email.strip())
                                _cuit_for_key = _nv_cuit.strip().replace("-","")
                                st.session_state[f"vendor_created_{_cuit_for_key}"] = (_nv_pid, _nv_name.strip())
                                st.session_state[_create_new_vend_key] = False
                                st.success(f"✅ Proveedor **{_nv_name}** creado (ID {_nv_pid}). Recargando...")
                                st.rerun()
                            except Exception as _nv_e:
                                st.error(f"Error al crear proveedor: {_nv_e}")

            if _cond_venta:
                if _dias_pago:
                    st.caption(f"📅 Condición de venta: **{_cond_venta}** → vencimiento calculado: `{_vto_auto}`")
                else:
                    st.caption(f"📅 Condición de venta: **{_cond_venta}** (sin días detectados — completá el vencimiento a mano)")

            # ── Chequeo de duplicado en tiempo real ───────────────────────────────
            _num_raw = extracted.get("numero","")
            _dup_exists, _dup_id, _dup_name = False, None, None
            if _num_raw:
                _dup_exists, _dup_id, _dup_name = check_invoice_exists(models_url, uid, api_key, _num_raw)
            if _dup_exists:
                _dup_url = odoo_url("account.move", _dup_id)
                st.error(
                    f"🚫 Esta factura **ya fue cargada** en Odoo ({_dup_name}). "
                    f"[Ver factura existente]({_dup_url})"
                )

            with st.form(key=f"bill_form_{uf.name}"):
                # CUIT ya está fuera del form para lookup en tiempo real
                cuit_i = _cuit_raw
                c1, c2 = st.columns(2)
                ref_i   = c2.text_input("N° de factura",
                            value=extracted.get("numero",""))
                fecha_i = c1.text_input("Fecha emisión (AAAA-MM-DD)",
                            value=extracted.get("fecha_iso",""),
                            placeholder="2026-05-12")
                fecha_vto_i = c2.text_input("Fecha vencimiento (AAAA-MM-DD)",
                            value=_vto_auto,
                            placeholder="2026-05-20",
                            help="Se calcula automáticamente si se detectan días en las condiciones de venta")

                concepto_i = st.text_input(
                    "📋 Concepto / Descripción",
                    value=extracted.get("concepto", ""),
                    placeholder="Descripción del servicio o producto",
                    help="Se usa como descripción de la línea en Odoo",
                )

                # Proveedor por nombre como fallback si no hay CUIT
                prov_i = st.text_input("Nombre del proveedor (fallback si no hay CUIT)",
                            value=(_partner_preloaded[1] if _partner_preloaded else extracted.get("proveedor","")[:60]),
                            placeholder="Nombre exacto en Odoo",
                            help="Se usa solo si el CUIT no resuelve a ningún proveedor")

                # Monto a cargar (editable; por defecto: neto gravado / total si exenta)
                _ca, _cb = st.columns([2, 1])
                _total_ref = extracted.get("total") or ""
                _neto_ref  = extracted.get("neto")  or ""
                _iva_ref   = extracted.get("iva")   or ""
                # Default: neto (base imponible); Odoo aplica impuestos encima.
                # Si no hay neto usa total (caso exenta: neto == total).
                _default_amount = safe_float(_neto_ref) or safe_float(_total_ref)
                amount_i = _ca.number_input(
                    "💰 Importe neto (base) *",
                    min_value=0.0,
                    value=_default_amount,
                    step=0.01,
                    format="%.2f",
                    key=f"amount_i_{uf.name}",
                    help="Base imponible (sin IVA). Odoo calcula los impuestos encima. "
                         "Para facturas exentas, ingresá el total.",
                )
                _ca.caption(
                    f"Extraído → Neto: {fmt_ars(_neto_ref)}  |  "
                    f"IVA: {fmt_ars(_iva_ref)}  |  "
                    f"Total: {fmt_ars(_total_ref)}"
                )
                # Exenta: sin IVA extraído → pre-marcar
                _exenta_default = not bool(safe_float(_iva_ref))
                exenta_i = st.checkbox(
                    "🔒 Factura exenta / Monotributo (sin impuestos)",
                    value=_exenta_default,
                    key=f"exenta_{uf.name}",
                    help="Si está marcado, se crea la línea sin ningún impuesto en Odoo.",
                )

                st.text_area("Notas internas", height=55, key=f"notas_{uf.name}")

                # Producto, Cuenta contable y Centro de Costo
                st.markdown("##### 📦 Producto / Servicio y Contabilidad")
                prod_sel = st.selectbox(
                    "Producto / Servicio",
                    options=_prod_labels,
                    index=_default_prod_idx,
                    key=f"prod_g_{uf.name}",
                    help="Producto de Odoo que se asigna a la línea de factura. "
                         "Se pre-selecciona según el proveedor.",
                )
                # Cuenta contable: auto-detectada en background (no se muestra en UI)
                # cuenta_sel sigue disponible para el asiento estimado y fallback
                cuenta_sel = _acct_labels[_default_acct_idx] if _acct_labels and _default_acct_idx < len(_acct_labels) else None
                analytic_sel = st.selectbox(
                    "Centro de Costo",
                    options=_analytic_labels,
                    index=0,
                    key=f"cc_g_{uf.name}",
                    help="Centro de costo (cuenta analítica) que absorbe el gasto. Opcional.",
                )

                _btn_label = "⬆️ Cargar en Odoo"
                if _dup_exists:
                    _btn_label = "⚠️ Ya existe — Cargar igual"
                go = st.form_submit_button(_btn_label, use_container_width=True)

            # Asiento estimado — visible siempre que haya montos
            _neto_f = extracted.get("neto","")
            _iva_f  = extracted.get("iva","")
            _tot_f  = extracted.get("total","")
            if _neto_f or _tot_f:
                _cuenta_disp = (cuenta_sel if (cuenta_sel and cuenta_sel != "— Sin cuenta —")
                                else "*(cuenta de gasto — seleccioná arriba)*")
                st.markdown("**📒 Asiento estimado en Odoo:**")
                st.markdown(
                    f"| Cuenta | Debe | Haber |\n"
                    f"|---|---|---|\n"
                    f"| {_cuenta_disp} | {fmt_ars(_neto_f)} | |\n"
                    f"| IVA Crédito Fiscal (si aplica) | {fmt_ars(_iva_f)} | |\n"
                    f"| Proveedor (por pagar) | | {fmt_ars(_tot_f)} |"
                )

            if go:
                with st.spinner("Procesando..."):
                    try:
                        partner_id = False
                        # 0. Proveedor seleccionado vía búsqueda por nombre/CUIT (sin-CUIT branch)
                        _ov_id = st.session_state.get(f"partner_override_{uf.name}")
                        if _ov_id:
                            partner_id = _ov_id
                            st.caption(f"Proveedor seleccionado: {st.session_state.get(f'partner_name_{uf.name}','')}")
                        # 1. Buscar por CUIT ingresado en el form
                        if cuit_i and cuit_i.strip():
                            _found = search_partner_by_cuit(models_url, uid, api_key, cuit_i.strip())
                            if _found:
                                partner_id = _found[0]
                                st.caption(f"Proveedor por CUIT: {_found[1]}")
                        # 2. Fallback: buscar por nombre
                        if not partner_id and prov_i and prov_i.strip():
                            m2 = search_partners(models_url, uid, api_key, prov_i.strip(), limit=3)
                            if m2:
                                partner_id = m2[0][0]
                                st.caption(f"Proveedor por nombre: {m2[0][1]}")
                            else:
                                st.warning(f"'{prov_i}' no encontrado — se creará sin proveedor asignado.")

                        # 3. Chequeo de duplicado en el submit (segunda línea de defensa)
                        if ref_i and ref_i.strip():
                            _dup2, _dup2_id, _dup2_name = check_invoice_exists(
                                models_url, uid, api_key, ref_i.strip())
                            if _dup2:
                                _dup2_url = odoo_url("account.move", _dup2_id)
                                st.error(
                                    f"🚫 La factura **{ref_i}** ya existe en Odoo ({_dup2_name}). "
                                    f"[Ver factura existente]({_dup2_url})"
                                )
                                continue

                        # 4. Resolver cuenta seleccionada
                        account_id_sel = None
                        if cuenta_sel and cuenta_sel != "— Sin cuenta —":
                            for _aid, _albl in _bill_accounts:
                                if _albl == cuenta_sel:
                                    account_id_sel = _aid
                                    break

                        # 5. Resolver Centro de Costo
                        analytic_id_sel = None
                        if analytic_sel and analytic_sel != "— Sin centro de costo —":
                            for _anid, _anlbl in _analytic_accounts:
                                if _anlbl == analytic_sel:
                                    analytic_id_sel = _anid
                                    break

                        # 6. Resolver Producto seleccionado
                        product_id_sel = None
                        if prod_sel and prod_sel != "— Sin producto —":
                            for _epid, _eplbl in _expense_products:
                                if _eplbl == prod_sel:
                                    product_id_sel = _epid
                                    break

                        # 7. l10n_latam_document_number (número sin prefijo de letra)
                        _latam_num = (ref_i or "").strip()
                        # Si el número extraído tiene prefijo de letra, quitarlo
                        if _latam_num and re.match(r"^[A-Za-z]\d", _latam_num):
                            _latam_num = _latam_num[1:]

                        move_id = create_vendor_bill(models, uid, api_key,
                            partner_id=partner_id, ref=concepto_i.strip() or ref_i,
                            invoice_date=fecha_i or False,
                            invoice_date_due=fecha_vto_i or None,
                            filename=uf.name, file_bytes=file_bytes, mimetype=mimetype,
                            account_id=account_id_sel,
                            amount_neto=amount_i if amount_i else None,
                            analytic_account_id=analytic_id_sel,
                            product_id=product_id_sel,
                            l10n_latam_document_number=_latam_num or None,
                            clear_taxes=exenta_i,
                            line_name=concepto_i.strip() or None)
                        url = odoo_url("account.move", move_id)
                        st.toast("Factura creada en Odoo", icon="✅")
                        st.markdown(f"📎 [Abrir en Odoo]({url})")
                        st.session_state.history.append({"tipo":"Factura proveedor",
                            "archivo":uf.name,"id":move_id,"url":url,"estado":"✅","hora":_dt_now.now().strftime("%H:%M")})
                    except Exception as e:
                        st.error(f"❌ {e}")


# ═══════════════════════════════════════════════════
# TAB 2 — PEDIDOS DE CLIENTES
# ═══════════════════════════════════════════════════
with tab_orders:
    st.subheader("Pedidos de clientes")
    files_o = st.file_uploader("Arrastrá o elegí archivos (PDF, JPG, PNG, XLSX)",
        type=["pdf","jpg","jpeg","png","xlsx","xls"], accept_multiple_files=True, key="orders_upload")
    if not files_o:
        st.caption("Subí uno o más archivos para empezar.")
    for uf in (files_o or []):
        st.divider()
        ext        = uf.name.rsplit(".", 1)[-1].lower()
        file_bytes = uf.read()
        mimetype   = MIMETYPES.get(ext, "application/octet-stream")
        st.markdown(f"**📎 {uf.name}**  `{ext.upper()}`  ({len(file_bytes)//1024} KB)")
        if ext in ("xlsx","xls"):
            # ── Parseo inteligente de Excel de pedido ─────────────────────
            with st.spinner("Leyendo Excel..."):
                oc_fields_xl = extract_excel_oc_fields(file_bytes, filename=uf.name)
            _lineas_xl = oc_fields_xl.get("lineas", [])
            if _lineas_xl:
                st.caption(f"✅ {len(_lineas_xl)} productos con pedido > 0 detectados automáticamente.")
            else:
                st.warning("No se detectaron productos con cantidad pedida. Revisá el archivo.")

            # ── Cliente: CUIT editable + auto-lookup ──────────────────────
            st.markdown("##### 🏢 Cliente")
            _cuit_key_xl = f"xl_cuit_{uf.name}"
            _pid_key_xl  = f"xl_pid_{uf.name}"
            _pnm_key_xl  = f"xl_pnm_{uf.name}"
            for _k, _dv in [(_cuit_key_xl,""), (_pid_key_xl, None), (_pnm_key_xl,"")]:
                if _k not in st.session_state:
                    st.session_state[_k] = _dv
            # Pre-llenar CUIT si lo extrajo el parser
            if not st.session_state[_cuit_key_xl] and oc_fields_xl.get("cuit"):
                st.session_state[_cuit_key_xl] = oc_fields_xl["cuit"]

            # ── Búsqueda de cliente por CUIT o razón social ───────────────
            _xl_pid = st.session_state[_pid_key_xl]
            _xl_pnm = st.session_state[_pnm_key_xl]

            # Si ya hay cliente asignado mostrar con opción de cambiar
            if _xl_pid:
                _cc1, _cc2 = st.columns([5, 1])
                _cc1.success(f"✅ Cliente: **{_xl_pnm}**")
                if _cc2.button("✏️ Cambiar", key=f"xl_change_{uf.name}"):
                    st.session_state[_pid_key_xl] = None
                    st.session_state[_pnm_key_xl] = ""
                    st.session_state[_cuit_key_xl] = ""
                    st.rerun()
            else:
                _xl_q = st.text_input(
                    "Buscar cliente por CUIT o razón social",
                    key=_cuit_key_xl,
                    placeholder="ej: 30-12345678-9  o  Brant  o  Fanttik",
                    help="Escribí el CUIT completo o parte de la razón social y presioná Enter",
                )
                _xl_q_val = (_xl_q or "").strip()
                if _xl_q_val and len(_xl_q_val) >= 3:
                    _xl_cands = search_partner_by_cuit_or_name(
                        models_url, uid, api_key, _xl_q_val, limit=8)
                    if len(_xl_cands) == 1:
                        # Match único: asignar directo
                        st.session_state[_pid_key_xl] = _xl_cands[0]["id"]
                        st.session_state[_pnm_key_xl] = _xl_cands[0]["name"]
                        st.rerun()
                    elif len(_xl_cands) > 1:
                        # Múltiples resultados: selectbox de elección
                        _xl_opt_map = {
                            f"{r['name']}  [{r.get('vat') or '—'}]": r
                            for r in _xl_cands}
                        _xc1, _xc2 = st.columns([4, 1])
                        with _xc1:
                            _xl_sel_lbl = st.selectbox(
                                "Resultados", list(_xl_opt_map.keys()),
                                key=f"xl_sel_{uf.name}",
                                label_visibility="collapsed")
                        with _xc2:
                            st.write("")
                            if st.button("✅ Usar", key=f"xl_use_{uf.name}"):
                                _ch = _xl_opt_map[_xl_sel_lbl]
                                st.session_state[_pid_key_xl] = _ch["id"]
                                st.session_state[_pnm_key_xl] = _ch["name"]
                                st.rerun()
                    else:
                        st.warning(f"⚠️ Sin resultados para «{_xl_q_val}».")
                        with st.expander("➕ Crear nuevo cliente", expanded=False):
                            _xnc1, _xnc2 = st.columns(2)
                            _xnc_name   = _xnc1.text_input("Razón social *", key=f"xl_nc_name_{uf.name}")
                            _xnc_cuit_v = _xnc2.text_input("CUIT *", key=f"xl_nc_cuit_{uf.name}",
                                value=_xl_q_val if re.sub(r"[^\d]","",_xl_q_val).__len__() >= 10 else "",
                                placeholder="30-12345678-9")
                            _xnc_street = _xnc1.text_input("Dirección", key=f"xl_nc_st_{uf.name}")
                            _xnc_phone  = _xnc2.text_input("Teléfono",  key=f"xl_nc_ph_{uf.name}")
                            _xnc_email  = st.text_input("Email", key=f"xl_nc_em_{uf.name}")
                            if st.button("Crear cliente", key=f"xl_btn_nc_{uf.name}"):
                                if _xnc_name and _xnc_cuit_v:
                                    try:
                                        _new_xl_pid = create_partner(models, uid, api_key,
                                            _xnc_name, _xnc_cuit_v,
                                            _xnc_street, _xnc_phone, _xnc_email)
                                        st.session_state[_pid_key_xl] = _new_xl_pid
                                        st.session_state[_pnm_key_xl] = _xnc_name
                                        st.rerun()
                                    except Exception as _xe:
                                        st.error(f"❌ {_xe}")
                                else:
                                    st.warning("Razón social y CUIT son obligatorios.")
                elif _xl_q_val:
                    st.caption("Escribí al menos 3 caracteres para buscar.")
                else:
                    st.info("📌 Ingresá el CUIT o razón social del cliente para continuar.")

            # ── Plazo de pago (siempre visible, opciones del sistema) ─────
            st.markdown("##### 📅 Plazo de pago")
            _all_pts_xl  = get_all_payment_terms(models_url, uid, api_key)
            _pt_opts_xl  = {name: pid for pid, name in _all_pts_xl}
            _xl_pt_def   = 0
            _xl_cpt_name = ""
            if _xl_pid and _pt_opts_xl:
                _xl_cpt_id, _xl_cpt_name = get_customer_payment_terms(models_url, uid, api_key, _xl_pid)
                if _xl_cpt_name and _xl_cpt_name in _pt_opts_xl:
                    _xl_pt_def = list(_pt_opts_xl.keys()).index(_xl_cpt_name)
            _xl_pt_names = list(_pt_opts_xl.keys()) if _pt_opts_xl else ["(sin opciones)"]
            _xl_pt_sel   = st.selectbox("Plazo de pago a usar",
                options=_xl_pt_names, index=_xl_pt_def, key=f"xl_pt_{uf.name}",
                help=f"Plazo cargado en Odoo para este cliente: {_xl_cpt_name or 'no configurado'}")
            _xl_pt_id = _pt_opts_xl.get(_xl_pt_sel)

            # ── Productos ─────────────────────────────────────────────────
            st.markdown("##### 📦 Productos")
            _xl_enriched = []
            if _lineas_xl:
                for _i_ln, _ln in enumerate(_lineas_xl):
                    _ov_key    = f"prod_ov_{uf.name}_{_i_ln}"
                    _cache_key = f"prod_cache_{uf.name}_{_i_ln}"
                    if _cache_key not in st.session_state:
                        _prods = search_product_by_code_or_name(
                            models_url, uid, api_key,
                            code=_ln.get("modelo","") or _ln.get("codigo",""),
                            name_keywords=(_ln.get("descripcion","") or _ln.get("modelo","")).strip(),
                            limit=1)
                        if _prods:
                            st.session_state[_cache_key] = _prods[0]
                    _op_auto = st.session_state.get(_cache_key)
                    _op      = st.session_state[_ov_key] if _ov_key in st.session_state else _op_auto
                    _cost    = float(_op["standard_price"]) if _op else 0.0
                    _price   = safe_float(_ln.get("precio_unit", 0))
                    _margin  = ((_price - _cost) / _price * 100) if _price > 0 else 0.0
                    _xl_enriched.append({**_ln, "odoo_product": _op,
                                         "cost": _cost, "margin_pct": _margin,
                                         "_ov_key": _ov_key})

                # ── Tabla de productos (Producto Odoo editable inline) ──────
                _tbl_rows = []
                for _i_el, _el in enumerate(_xl_enriched):
                    _op = _el.get("odoo_product")
                    _tbl_rows.append({
                        "":            "✅" if _op else "⚠️",
                        "Modelo":      _el.get("modelo") or "",
                        "Descripción": (_el.get("descripcion") or "")[:50],
                        "Cant.":       int(_el.get("cantidad", 0)),
                        "P. Unit.":    float(_el.get("precio_unit", 0)),
                        "IVA %":       int(_el.get("iva_pct", 21)),
                        "Costo":       float(_el.get("cost", 0)),
                        "Margen %":    round(float(_el.get("margin_pct", 0)), 1),
                        "Producto Odoo": _op["name"] if _op else "",
                    })
                _tbl_cfg = {
                    "":              st.column_config.TextColumn("", width="small"),
                    "Cant.":         st.column_config.NumberColumn("Cant.", format="%d"),
                    "P. Unit.":      st.column_config.NumberColumn("P. Unit.", format="$ %.2f"),
                    "IVA %":         st.column_config.NumberColumn("IVA %", format="%d%%"),
                    "Costo":         st.column_config.NumberColumn("Costo", format="$ %.2f"),
                    "Margen %":      st.column_config.NumberColumn("Margen %", format="%.1f%%"),
                    "Producto Odoo": st.column_config.TextColumn(
                        "Producto Odoo",
                        help="✏️ Escribí nombre o código (parcial) y presioná Enter para buscar"),
                }
                _tbl_disabled_cols = ["", "Modelo", "Descripción", "Cant.",
                                      "P. Unit.", "IVA %", "Costo", "Margen %"]
                # Versión del key: cambia tras cada match exitoso para limpiar el estado del editor
                _tbl_ver   = st.session_state.get(f"_tbl_ver_{uf.name}", 0)
                _tbl_edited = st.data_editor(
                    pd.DataFrame(_tbl_rows), column_config=_tbl_cfg,
                    disabled=_tbl_disabled_cols,
                    use_container_width=True, hide_index=True,
                    key=f"tbl_ped_{uf.name}_{_tbl_ver}")

                # ── Procesar ediciones inline de Producto Odoo ─────────────
                _tbl_updated  = False
                _tbl_no_match = []
                for _i_ed, _row_ed in _tbl_edited.iterrows():
                    _new_q   = (_row_ed.get("Producto Odoo") or "").strip()
                    _old_op  = _xl_enriched[_i_ed].get("odoo_product")
                    _old_name = _old_op["name"] if _old_op else ""
                    if not _new_q or _new_q == _old_name:
                        continue  # sin cambio
                    _ck_ed = _xl_enriched[_i_ed].get("_ov_key")
                    _res = search_product_by_code_or_name(
                        models_url, uid, api_key,
                        code=_new_q, name_keywords=_new_q, limit=5)
                    if len(_res) == 1:
                        # Match único: asignar directo
                        if _ck_ed:
                            st.session_state[_ck_ed] = _res[0]
                        _xl_enriched[_i_ed]["odoo_product"] = _res[0]
                        _tbl_updated = True
                    elif len(_res) > 1:
                        # Múltiples resultados: guardar para mostrar selectbox
                        st.session_state[f"_dis_{uf.name}_{_i_ed}"] = (
                            _new_q, _res, _ck_ed, _i_ed)
                    else:
                        _tbl_no_match.append(
                            f"⚠️ Sin resultados para «{_new_q}» (línea {_i_ed+1})")

                if _tbl_updated:
                    st.session_state[f"_tbl_ver_{uf.name}"] = _tbl_ver + 1
                    st.rerun()

                # ── Desambiguación: mostrar opciones cuando hay varios resultados ──
                for _i_ed in range(len(_xl_enriched)):
                    _dis_key = f"_dis_{uf.name}_{_i_ed}"
                    if _dis_key not in st.session_state:
                        continue
                    _dq, _dres, _dck, _di = st.session_state[_dis_key]
                    _opts_map = {
                        f"{r['name']}  [{r.get('default_code') or '—'}]": r
                        for r in _dres}
                    _dc1, _dc2 = st.columns([4, 1])
                    with _dc1:
                        _mod_lbl = (_xl_enriched[_di].get("modelo")
                                    or _xl_enriched[_di].get("descripcion")
                                    or f"Línea {_di+1}")
                        _dis_sel = st.selectbox(
                            f"**{_mod_lbl}** — elegí el producto:",
                            list(_opts_map.keys()),
                            key=f"dis_sel_{uf.name}_{_i_ed}")
                    with _dc2:
                        st.write("")
                        st.write("")
                        if st.button("✅ Usar", key=f"dis_use_{uf.name}_{_i_ed}"):
                            _chosen = _opts_map[_dis_sel]
                            if _dck:
                                st.session_state[_dck] = _chosen
                            _xl_enriched[_di]["odoo_product"] = _chosen
                            del st.session_state[_dis_key]
                            st.session_state[f"_tbl_ver_{uf.name}"] = (
                                st.session_state.get(f"_tbl_ver_{uf.name}", 0) + 1)
                            st.rerun()

                for _nm_msg in _tbl_no_match:
                    st.caption(_nm_msg)
            else:
                st.info("No se detectaron productos con cantidad pedida.")

            # ── Resumen financiero ─────────────────────────────────────────
            _xl_neto  = sum(safe_float(_el.get("subtotal",0)) for _el in _xl_enriched)
            _xl_iva   = sum(
                safe_float(_el.get("subtotal",0)) * safe_float(_el.get("iva_pct",21)) / 100
                for _el in _xl_enriched)
            _xl_total = _xl_neto + _xl_iva
            if _xl_enriched:
                st.markdown("##### 💰 Resumen financiero")
                def _fin_card_xl(lbl, val):
                    return (f'<div style="border-left:3px solid #e63946;padding:5px 10px;'
                            f'background:#fff8f8;border-radius:4px">'
                            f'<div style="font-size:11px;color:#888;margin-bottom:2px">{lbl}</div>'
                            f'<div style="font-size:15px;font-weight:700;color:#e63946">{fmt_ars(val)}</div></div>')
                _xlrf1, _xlrf2, _xlrf3 = st.columns(3)
                _xlrf1.markdown(_fin_card_xl("Neto s/IVA",  _xl_neto), unsafe_allow_html=True)
                _xlrf2.markdown(_fin_card_xl("IVA",         _xl_iva),  unsafe_allow_html=True)
                _xlrf3.markdown(_fin_card_xl("Total c/IVA", _xl_total),unsafe_allow_html=True)

            # ── Crear pedido ───────────────────────────────────────────────
            st.markdown("---")
            _xl_btn_disabled = not bool(_xl_pid)
            if _xl_btn_disabled:
                st.caption("🔒 Identificá el cliente para habilitar la creación del pedido.")
            if st.button("⬆️ Crear pedido en Odoo", key=f"btn_xl_order_{uf.name}",
                         type="primary", disabled=_xl_btn_disabled):
                with st.spinner("Creando pedido..."):
                    try:
                        _xl_order_lines = [{
                            "product_id":  _el["odoo_product"]["id"] if _el.get("odoo_product") else None,
                            "descripcion": _el.get("descripcion",""),
                            "cantidad":    _el.get("cantidad", 1),
                            "precio_unit": _el.get("precio_unit", 0),
                        } for _el in _xl_enriched]
                        _xl_order_id = create_sale_order(
                            models, uid, api_key,
                            partner_id      = _xl_pid,
                            note            = f"Importado desde {uf.name}",
                            lines           = _xl_order_lines,
                            filename        = uf.name,
                            file_bytes      = file_bytes,
                            mimetype        = mimetype,
                            payment_term_id = _xl_pt_id or None,
                        )
                        url = odoo_url("sale.order", _xl_order_id)
                        st.toast("Pedido creado en Odoo", icon="✅")
                        st.markdown(f"📎 [Abrir en Odoo]({url})")
                        st.session_state.history.append({"tipo":"Pedido cliente",
                            "archivo":uf.name,"id":_xl_order_id,"url":url,"estado":"✅","hora":_dt_now.now().strftime("%H:%M")})
                    except Exception as _xe:
                        st.error(f"❌ {_xe}")
        else:
            # ── Parseo del PDF de OC ──────────────────────────────────────
            oc_fields, _oc_tables, _oc_raw = {}, [], ""
            if ext == "pdf":
                with st.spinner("Leyendo OC..."):
                    oc_fields, _oc_tables, _oc_raw = extract_oc_fields(file_bytes)
            elif ext in ("jpg","jpeg","png"):
                st.image(file_bytes, caption="Vista previa", width=420)
                with st.spinner("Leyendo imagen con OCR..."):
                    oc_fields, _oc_tables, _oc_raw = extract_image_oc_fields(file_bytes)
                _n_lineas = len(oc_fields.get("lineas", []))
                if not _oc_raw:
                    st.error("❌ OCR falló — Tesseract puede no estar instalado aún. Intentá en 2 minutos o subí un PDF.")
                else:
                    if _n_lineas == 0:
                        st.warning("⚠️ OCR leyó el texto pero no detectó líneas de productos. Revisá el texto crudo y completá a mano si es necesario.")
                    else:
                        st.caption(f"🤖 OCR detectó {_n_lineas} línea(s). Revisá antes de confirmar.")
                    with st.expander("🔍 Ver texto bruto detectado por OCR"):
                        st.code(_oc_raw if _oc_raw else "(vacío)")

            # ── Session state para partner de esta OC ─────────────────────
            _ss_pid   = f"oc_pid_{uf.name}"
            _ss_pname = f"oc_pname_{uf.name}"
            if _ss_pid not in st.session_state:
                st.session_state[_ss_pid] = None
            if _ss_pname not in st.session_state:
                st.session_state[_ss_pname] = ""

            # ── SECCIÓN 1: CLIENTE ────────────────────────────────────────
            st.markdown("##### 🏢 Cliente")

            # CUIT editable en tiempo real (fuera del form), igual que en Facturas
            _oc_cuit_key = f"oc_cuit_edit_{uf.name}"
            _oc_cuit_detected = oc_fields.get("cuit", "")
            _oc_name_detected = oc_fields.get("cliente_nombre", "")
            if _oc_cuit_key not in st.session_state:
                # Pre-llenar con nombre si no hay CUIT
                st.session_state[_oc_cuit_key] = _oc_cuit_detected or _oc_name_detected
            _oc_cuit = st.text_input(
                "CUIT o razon social del cliente",
                key=_oc_cuit_key,
                placeholder="30-12345678-9  o  COPPEL S.A.",
                help="Detectado del documento. Podés buscar por CUIT o nombre.",
            ).strip()
            _oc_cuit_norm = _oc_cuit.replace("-","").replace(" ","")

            # Lookup por CUIT o nombre si todavía no tenemos partner resuelto
            if _oc_cuit and not st.session_state[_ss_pid]:
                # Intentar por CUIT primero
                _partner_oc = search_partner_by_cuit(models_url, uid, api_key, _oc_cuit) if len(_oc_cuit_norm) >= 10 else None
                if _partner_oc:
                    st.session_state[_ss_pid]   = _partner_oc[0]
                    st.session_state[_ss_pname] = _partner_oc[1]
                elif len(_oc_cuit) >= 3:
                    # Fallback: buscar por nombre
                    _oc_cands = search_partner_by_cuit_or_name(models_url, uid, api_key, _oc_cuit, limit=6)
                    if len(_oc_cands) == 1:
                        st.session_state[_ss_pid]   = _oc_cands[0]["id"]
                        st.session_state[_ss_pname] = _oc_cands[0]["name"]
                    elif len(_oc_cands) > 1:
                        _oc_opt_map = {f"{r['name']}  [{r.get('vat') or '—'}]": r for r in _oc_cands}
                        _occ1, _occ2 = st.columns([4, 1])
                        with _occ1:
                            _oc_sel_lbl = st.selectbox("Resultados", list(_oc_opt_map.keys()),
                                key=f"oc_sel_{uf.name}", label_visibility="collapsed")
                        with _occ2:
                            st.write("")
                            if st.button("Usar", key=f"oc_use_{uf.name}"):
                                _ch = _oc_opt_map[_oc_sel_lbl]
                                st.session_state[_ss_pid]   = _ch["id"]
                                st.session_state[_ss_pname] = _ch["name"]
                                st.rerun()

            _partner_id_oc   = st.session_state[_ss_pid]
            _partner_name_oc = st.session_state[_ss_pname]

            # Checkbox crear nuevo (solo aparece cuando CUIT ingresado y no encontrado)
            _oc_create_key = f"oc_create_new_{uf.name}"
            if _oc_create_key not in st.session_state:
                st.session_state[_oc_create_key] = False

            if _partner_id_oc:
                st.success(f"✅ Cliente: **{_partner_name_oc}** (CUIT {_oc_cuit})")
                _pt_id, _pt_name = get_customer_payment_terms(models_url, uid, api_key, _partner_id_oc)
            elif _oc_cuit:
                st.warning(f"⚠️ CUIT **{_oc_cuit}** no encontrado en Odoo.")
                st.checkbox("➕ Crear nuevo cliente en Odoo", key=_oc_create_key)
                _pt_id, _pt_name = None, None
            else:
                st.info("ℹ️ Ingresá el CUIT del cliente para buscarlo en Odoo.")
                _pt_id, _pt_name = None, None

            if st.session_state.get(_oc_create_key):
                with st.expander("📝 Datos del nuevo cliente", expanded=True):
                    _nc1, _nc2 = st.columns(2)
                    nc_name   = _nc1.text_input("Razón social *", key=f"nc_name_{uf.name}")
                    nc_cuit   = _nc2.text_input("CUIT *", value=_oc_cuit,
                                    key=f"nc_cuit_{uf.name}", placeholder="30-12345678-9")
                    nc_street = _nc1.text_input("Dirección", key=f"nc_street_{uf.name}")
                    nc_phone  = _nc2.text_input("Teléfono", key=f"nc_phone_{uf.name}")
                    nc_email  = st.text_input("Email", key=f"nc_email_{uf.name}")
                    if st.button("Crear cliente en Odoo", key=f"btn_nc_{uf.name}"):
                        if nc_name and nc_cuit:
                            try:
                                _new_pid = create_partner(models, uid, api_key,
                                    nc_name, nc_cuit, nc_street, nc_phone, nc_email)
                                st.session_state[_ss_pid]   = _new_pid
                                st.session_state[_ss_pname] = nc_name
                                st.session_state[_oc_create_key] = False
                                st.success(f"✅ Cliente **{nc_name}** creado en Odoo (ID {_new_pid})")
                                st.rerun()
                            except Exception as _e:
                                st.error(f"❌ {_e}")
                        else:
                            st.warning("Razón social y CUIT son obligatorios.")

            # ── SECCIÓN 2: DATOS DE LA OC ─────────────────────────────────
            # Forzar sesión con datos nuevos si antes estaba vacío
            for _oc_fk, _oc_fv in [
                (f"ocnum_{uf.name}",  oc_fields.get("numero_oc", "")),
                (f"ocfec_{uf.name}",  oc_fields.get("fecha_iso", "") or oc_fields.get("fecha", "")),
                (f"occond_{uf.name}", oc_fields.get("condiciones_pago", "")),
            ]:
                if _oc_fv and not st.session_state.get(_oc_fk):
                    st.session_state[_oc_fk] = _oc_fv
            st.markdown("##### 📋 Datos de la Orden de Compra")
            _oc1, _oc2, _oc3 = st.columns(3)
            _oc_num_i  = _oc1.text_input("N° OC",   value=oc_fields.get("numero_oc",""),
                                          key=f"ocnum_{uf.name}")
            _oc_fec_i  = _oc2.text_input("Fecha",   value=oc_fields.get("fecha_iso","") or oc_fields.get("fecha",""),
                                          key=f"ocfec_{uf.name}", placeholder="AAAA-MM-DD")
            _oc_cond_i = _oc3.text_input("Condición de pago",
                                          value=oc_fields.get("condiciones_pago",""),
                                          key=f"occond_{uf.name}")

            # ── SECCIÓN 3: PRODUCTOS ──────────────────────────────────────
            st.markdown("##### 📦 Productos")
            _lineas_oc = oc_fields.get("lineas", [])
            _enriched  = []

            # Session state para overrides manuales de producto
            _ss_overrides = f"prod_ov_{uf.name}"
            if _ss_overrides not in st.session_state:
                st.session_state[_ss_overrides] = {}

            if _lineas_oc:
                # Build enriched list from current overrides/cache
                for _li, _ln in enumerate(_lineas_oc):
                    _cands_key = f"oc_cands_{uf.name}_{_li}"
                    if _cands_key not in st.session_state:
                        _cands = search_product_by_code_or_name(
                            models_url, uid, api_key,
                            code=_ln.get("codigo",""),
                            name_keywords=(
                                (_ln.get("descripcion","") + " " +
                                 _ln.get("marca","") + " " +
                                 _ln.get("modelo","")).strip()
                            ),
                            ean13=_ln.get("ean13",""),
                            limit=1,
                        )
                        st.session_state[_cands_key] = _cands
                    _cands = st.session_state.get(_cands_key) or []
                    _override = st.session_state[_ss_overrides].get(_li)
                    _op = _override if _override is not None else (_cands[0] if _cands else None)
                    _cost  = float(_op["standard_price"]) if _op else 0.0
                    _price = safe_float(_ln.get("precio_unit", 0))
                    _margin = ((_price - _cost) / _price * 100) if _price > 0 else 0.0
                    _enriched.append({**_ln, "odoo_product": _op, "cost": _cost, "margin_pct": _margin})

                # Table data
                _df_rows = []
                for _el in _enriched:
                    _desc_full = " · ".join(x for x in [
                        _el.get("descripcion",""), _el.get("marca",""), _el.get("modelo","")
                    ] if x)
                    _match_txt = (_el["odoo_product"]["name"] if _el.get("odoo_product") else "")
                    _df_rows.append({
                        "Código":       _el.get("codigo",""),
                        "Descripción":  _desc_full,
                        "Cant.":        int(_el.get("cantidad",0)),
                        "Precio unit.": fmt_ars(_el.get("precio_unit",0)),
                        "Subtotal":     fmt_ars(_el.get("subtotal",0)),
                        "Margen %":     f"{_el.get('margin_pct',0):.1f}%",
                        "Match Odoo":   _match_txt,
                    })
                _prev_matches = [r["Match Odoo"] for r in _df_rows]

                _edited_df = st.data_editor(
                    pd.DataFrame(_df_rows),
                    column_config={
                        "Código":       st.column_config.TextColumn(disabled=True, width="small"),
                        "Descripción":  st.column_config.TextColumn(disabled=True),
                        "Cant.":        st.column_config.NumberColumn(disabled=True, width="small"),
                        "Precio unit.": st.column_config.TextColumn(disabled=True, width="medium"),
                        "Subtotal":     st.column_config.TextColumn(disabled=True, width="medium"),
                        "Margen %":     st.column_config.TextColumn(disabled=True, width="small"),
                        "Match Odoo":   st.column_config.TextColumn(
                            "Match Odoo ✏️",
                            help="Escribí nombre o código Odoo para reasignar",
                            width="large",
                        ),
                    },
                    use_container_width=True,
                    hide_index=True,
                    key=f"oc_editor_{uf.name}",
                )

                # Detect Match Odoo changes → search & update overrides
                _needs_rerun = False
                for _idx, _erow in _edited_df.iterrows():
                    _new_txt = str(_erow.get("Match Odoo","") or "").strip()
                    _old_txt = _prev_matches[_idx] if _idx < len(_prev_matches) else ""
                    if _new_txt == _old_txt:
                        continue
                    if not _new_txt:
                        st.session_state[_ss_overrides][_idx] = None
                        st.session_state.pop(f"oc_cands_{uf.name}_{_idx}", None)
                        _needs_rerun = True
                    else:
                        _sr = search_product_by_code_or_name(
                            models_url, uid, api_key,
                            code=_new_txt, name_keywords=_new_txt, limit=6)
                        if len(_sr) == 1:
                            st.session_state[_ss_overrides][_idx] = _sr[0]
                            st.session_state[f"oc_cands_{uf.name}_{_idx}"] = _sr
                            _needs_rerun = True
                        elif len(_sr) > 1:
                            _res_opts = {
                                f"{r['name']}  [{r.get('default_code','')}]": r for r in _sr}
                            st.warning(f"Fila {_idx+1} — múltiples resultados para «{_new_txt}»:")
                            _chosen_lbl2 = st.selectbox(
                                "Seleccioná el producto correcto",
                                list(_res_opts.keys()),
                                key=f"oc_res_{uf.name}_{_idx}",
                                label_visibility="collapsed",
                            )
                            if st.button("✅ Confirmar", key=f"oc_res_btn_{uf.name}_{_idx}"):
                                st.session_state[_ss_overrides][_idx] = _res_opts[_chosen_lbl2]
                                st.session_state[f"oc_cands_{uf.name}_{_idx}"] = _sr
                                _needs_rerun = True
                        else:
                            st.warning(f"Fila {_idx+1}: sin resultados para «{_new_txt}» — intentá otro término.")
                if _needs_rerun:
                    st.rerun()
            else:
                st.info("No se detectaron líneas de productos automáticamente.")

            # Cálculo de totales desde líneas enriquecidas
            _calc_neto = sum(safe_float(_el.get("subtotal",0)) for _el in _enriched)
            _calc_iva21 = sum(
                safe_float(_el.get("subtotal",0)) * 0.21
                for _el in _enriched if abs(safe_float(_el.get("iva_pct",21)) - 21) < 1)
            _calc_iva105 = sum(
                safe_float(_el.get("subtotal",0)) * 0.105
                for _el in _enriched if abs(safe_float(_el.get("iva_pct",21)) - 10.5) < 1)
            _calc_iva   = _calc_iva21 + _calc_iva105
            _calc_total = _calc_neto + _calc_iva
            _calc_costo = sum(safe_float(_el.get("cost",0)) * safe_float(_el.get("cantidad",1))
                              for _el in _enriched)
            _margin_total = ((_calc_neto - _calc_costo) / _calc_neto * 100
                             if _calc_neto > 0 else 0.0)

            # Cuando hay líneas enriquecidas, usar valores calculados (más exactos).
            # Solo caer en los valores del PDF si no hay líneas detectadas.
            if _enriched:
                _show_neto   = _calc_neto
                _show_iva21  = _calc_iva21
                _show_iva105 = _calc_iva105
                _show_iva    = _calc_iva
                _show_total  = _calc_neto + _calc_iva
            else:
                _show_neto   = float(oc_fields.get("subtotal_neto") or 0)
                _show_iva21  = float(oc_fields.get("iva_21")  or 0)
                _show_iva105 = float(oc_fields.get("iva_105") or 0)
                _show_iva    = _show_iva21 + _show_iva105
                _show_total  = float(oc_fields.get("total") or 0) or (_show_neto + _show_iva)

            # ── SECCIÓN 4: RESUMEN FINANCIERO ────────────────────────────
            st.markdown("##### 💰 Resumen financiero")
            def _fin_card(lbl, val, pct=False):
                color = "#e63946"
                fval = f"{val:.1f}%" if pct else fmt_ars(val)
                return (f'<div style="border-left:3px solid {color};padding:5px 10px;'
                        f'background:#fff8f8;border-radius:4px">'
                        f'<div style="font-size:11px;color:#888;margin-bottom:2px">{lbl}</div>'
                        f'<div style="font-size:15px;font-weight:700;color:{color}">{fval}</div></div>')
            _rf1, _rf2, _rf3, _rf4 = st.columns(4)
            _rf1.markdown(_fin_card("Total Neto",  _show_neto), unsafe_allow_html=True)
            _rf2.markdown(_fin_card("IVA",         _show_iva),  unsafe_allow_html=True)
            _rf3.markdown(_fin_card("Total c/IVA", _show_total),unsafe_allow_html=True)
            _rf4.markdown(_fin_card("Margen total",_margin_total, pct=True), unsafe_allow_html=True)

            # ── SECCIÓN 5: PLAZO DE PAGO ──────────────────────────────────
            st.markdown("##### 📅 Plazo de pago")
            _oc_dias     = oc_fields.get("dias_pago")
            _oc_cond_str = oc_fields.get("condiciones_pago","")
            _pt_choice_id = _pt_id  # default: plazo del cliente en Odoo

            _all_pts_pdf = get_all_payment_terms(models_url, uid, api_key)
            _pt_opts_pdf = {name: pid for pid, name in _all_pts_pdf}

            if _pt_id and _pt_name:
                _odoo_dias_est = parse_payment_terms(_pt_name)
                _hay_disc = (
                    _oc_dias is not None
                    and _odoo_dias_est is not None
                    and abs(_odoo_dias_est - _oc_dias) > 3
                )
                if _hay_disc:
                    st.warning(
                        f"⚠️ Discrepancia: la OC indica **{_oc_dias} días** "
                        f"({_oc_cond_str}), pero el cliente tiene **{_pt_name}** en Odoo."
                    )
                # Selectbox con todas las opciones del sistema
                _pt_def_idx = 0
                if _pt_name in _pt_opts_pdf:
                    _pt_def_idx = list(_pt_opts_pdf.keys()).index(_pt_name)
                _oc_hint = f" | OC: {_oc_cond_str}" if _oc_cond_str else (f" | OC: {_oc_dias} días" if _oc_dias else "")
                _pt_sel_pdf = st.selectbox(
                    "Plazo de pago a usar",
                    options=list(_pt_opts_pdf.keys()),
                    index=_pt_def_idx,
                    key=f"pt_sel_{uf.name}",
                    help=f"Plazo del cliente en Odoo: {_pt_name}{_oc_hint}",
                )
                _pt_choice_id = _pt_opts_pdf.get(_pt_sel_pdf)
                if not _hay_disc:
                    st.caption(f"✅ Plazo del cliente en Odoo: **{_pt_name}**"
                               + (f" — OC: {_oc_cond_str}" if _oc_cond_str else ""))
            elif _oc_dias:
                st.info(f"📅 OC indica **{_oc_dias} días** ({_oc_cond_str}) "
                        f"— cliente sin plazo configurado en Odoo.")
            elif _pt_id:
                st.info(f"📅 Plazo del cliente en Odoo: **{_pt_name}**")

            # ── SECCIÓN 6: ASIENTO ESTIMADO ───────────────────────────────
            st.markdown("##### 📒 Asiento estimado en Odoo")
            _iva_rows_md = ""
            if _show_iva105 > 0:
                _iva_rows_md += f"| IVA Débito Fiscal 10,5% | | {fmt_ars(_show_iva105)} |\n"
            if _show_iva21 > 0:
                _iva_rows_md += f"| IVA Débito Fiscal 21% | | {fmt_ars(_show_iva21)} |"
            elif _show_iva > 0 and not _iva_rows_md:
                _iva_rows_md += f"| IVA Débito Fiscal | | {fmt_ars(_show_iva)} |"
            st.markdown(
                f"| Cuenta | Debe | Haber |\n"
                f"|---|---|---|\n"
                f"| Cuentas por Cobrar (Clientes) | {fmt_ars(_show_total)} | |\n"
                f"| Ventas / Ingresos | | {fmt_ars(_show_neto)} |\n"
                + _iva_rows_md
            )

            # ── SECCIÓN 7: CREAR PEDIDO ───────────────────────────────────
            st.markdown("---")

            # Campo Referido / Ejecutivo de cuenta
            _oc_ejecutivo_field, _oc_ejecutivo_relation = get_ejecutivo_field(models_url, uid, api_key)
            _oc_referidos  = get_referidos(models_url, uid, api_key)
            _oc_ref_map    = {n: i for i, n in _oc_referidos}
            _oc_ref_opts   = ["— Sin referido —"] + list(_oc_ref_map.keys())
            _oc_ref_default = 0
            if _partner_id_oc and _oc_ref_map:
                try:
                    _oc_pdata = models.execute_kw(ODOO_DB, uid, api_key,
                        "res.partner", "read", [[_partner_id_oc]],
                        {"fields": ["x_studio_referido_1"]})[0]
                    _oc_existing = _oc_pdata.get("x_studio_referido_1")
                    if _oc_existing and isinstance(_oc_existing, (list, tuple)):
                        _rname = _oc_existing[1]
                        if _rname in _oc_ref_opts:
                            _oc_ref_default = _oc_ref_opts.index(_rname)
                except Exception:
                    pass
            _oc_ref_sel = st.selectbox(
                "Referido", _oc_ref_opts, index=_oc_ref_default,
                key=f"oc_ref_{uf.name}",
                help="Quién refirió a este cliente",
            )
            # DEBUG temporal — mostrar campo detectado para Ejecutivo de cuenta
            if _oc_ejecutivo_field:
                st.caption(f"🔍 Debug: campo detectado = `{_oc_ejecutivo_field}` · relation = `{_oc_ejecutivo_relation}`")
            else:
                st.caption("🔍 Debug: campo 'Ejecutivo de cuenta' NO detectado en sale.order")

            _btn_disabled = not bool(_partner_id_oc)
            if _btn_disabled:
                st.caption("🔒 Identificá o creá el cliente para habilitar la creación del pedido.")
            if st.button("⬆️ Crear pedido en Odoo", key=f"btn_order_{uf.name}",
                         type="primary", disabled=_btn_disabled):
                with st.spinner("Creando pedido en Odoo..."):
                    try:
                        _order_lines = []
                        for _el in _enriched:
                            _order_lines.append({
                                "product_id":  _el["odoo_product"]["id"] if _el.get("odoo_product") else None,
                                "descripcion": _el.get("descripcion",""),
                                "cantidad":    _el.get("cantidad", 1),
                                "precio_unit": _el.get("precio_unit", 0),
                            })
                        _ref_oc = _oc_num_i or oc_fields.get("numero_oc","")
                        _fec_oc = _oc_fec_i or oc_fields.get("fecha_iso","") or None
                        if _oc_ref_sel != "— Sin referido —" and _oc_ref_sel in _oc_ref_map:
                            try:
                                models.execute_kw(ODOO_DB, uid, api_key,
                                    "res.partner", "write",
                                    [[_partner_id_oc],
                                     {"x_studio_referido_1": _oc_ref_map[_oc_ref_sel]}])
                            except Exception:
                                pass
                        _ref_partner_id = _oc_ref_map.get(_oc_ref_sel) if _oc_ref_sel != "— Sin referido —" else None
                        # Si el campo espera res.users, buscar el usuario cuyo partner coincide
                        _ref_id_oc = _ref_partner_id
                        if _ref_partner_id and _oc_ejecutivo_relation == "res.users":
                            try:
                                _usr = models.execute_kw(ODOO_DB, uid, api_key,
                                    "res.users", "search_read",
                                    [[("partner_id", "=", _ref_partner_id)]],
                                    {"fields": ["id"], "limit": 1})
                                if _usr:
                                    _ref_id_oc = _usr[0]["id"]
                            except Exception:
                                pass
                        order_id = create_sale_order(
                            models, uid, api_key,
                            partner_id       = _partner_id_oc,
                            note             = f"OC {_ref_oc}" if _ref_oc else "",
                            lines            = _order_lines,
                            filename         = uf.name,
                            file_bytes       = file_bytes,
                            mimetype         = mimetype,
                            client_order_ref = _ref_oc or None,
                            payment_term_id  = _pt_choice_id or None,
                            date_order       = _fec_oc,
                            ejecutivo_field  = _oc_ejecutivo_field,
                            ejecutivo_id     = _ref_id_oc,
                        )
                        url = odoo_url("sale.order", order_id)
                        st.toast("Pedido creado en Odoo", icon="✅")
                        st.markdown(f"📎 [Abrir en Odoo]({url})")
                        st.session_state.history.append({"tipo":"Pedido cliente",
                            "archivo":uf.name,"id":order_id,"url":url,"estado":"✅","hora":_dt_now.now().strftime("%H:%M")})
                    except Exception as _e:
                        st.error(f"❌ {_e}")


# ═══════════════════════════════════════════════════
# TAB 3 — IMPORTACIONES (ADMIN)
# ═══════════════════════════════════════════════════

# ═══════════════════════════════════════════════════
# TAB — CONTACTOS (clientes / proveedores via ARCA)
# ═══════════════════════════════════════════════════
with tab_contacts:
    st.subheader("Alta de Contactos")
    st.caption("Subí la constancia de ARCA para pre-completar los datos, o completá a mano.")

    _ct_file = st.file_uploader(
        "Constancia de inscripción ARCA (PDF) — opcional",
        type=["pdf"], key="ct_arca_upload",
        accept_multiple_files=False,
    )

    # ── Extraer datos del PDF si se subió ─────────────────────────────────
    _arca = {}
    if _ct_file:
        _ct_bytes = _ct_file.read()
        with st.spinner("Leyendo constancia ARCA..."):
            try:
                import pdfplumber
                with pdfplumber.open(BytesIO(_ct_bytes)) as _pdf:
                    _ct_text = "\n".join(p.extract_text() or "" for p in _pdf.pages)
                _arca = extract_arca_fields(_ct_text)
                st.success(f"✅ {_arca.get('nombre','?')} · CUIT {_arca.get('cuit','?')}")
            except Exception as _ce:
                st.warning(f"No se pudo leer el PDF: {_ce}")

    # ── Cargar catálogos Odoo ──────────────────────────────────────────────
    _ct_states   = get_ar_states(models_url, uid, api_key)
    _ct_afip     = get_afip_resp_types(models_url, uid, api_key)
    _ct_users    = get_odoo_users(models_url, uid, api_key)
    _ct_pts      = get_all_payment_terms(models_url, uid, api_key)   # [(id,name)]
    _ct_plists   = get_pricelists(models_url, uid, api_key)
    _ct_cuit_tid  = get_cuit_id_type(models_url, uid, api_key)
    _ct_referidos = get_referidos(models_url, uid, api_key)

    # helpers: map name→id
    _state_map  = {n: i for i, n in _ct_states}
    _afip_map   = {n: i for i, n in _ct_afip}
    _user_map   = {n: i for i, n in _ct_users}
    _pt_map     = {n: i for i, n in _ct_pts}
    _plist_map    = {n: i for i, n in _ct_plists}
    _referido_map = {n: i for i, n in _ct_referidos}

    # Default provincia
    _def_prov = ""
    if _arca.get("province_name"):
        _sid = match_ar_state(_arca["province_name"], _ct_states)
        if _sid:
            _def_prov = next((n for i, n in _ct_states if i == _sid), "")

    # Default tipo responsabilidad
    _afip_names = list(_afip_map.keys())
    _def_afip_idx = 0
    _tr = _arca.get("tipo_resp", "RI")
    for _ai, _an in enumerate(_afip_names):
        _an_up = _an.upper()
        if _tr == "RI" and ("RESPONSABLE INSCRIPTO" in _an_up or "IVA" in _an_up):
            _def_afip_idx = _ai; break
        if _tr == "MONO" and "MONOTRIBUTO" in _an_up:
            _def_afip_idx = _ai; break
        if _tr == "EX" and "EXENTO" in _an_up:
            _def_afip_idx = _ai; break

    # Pre-cargar cuentas contables
    _ct_accts_rec = get_ar_accounts(models_url, uid, api_key, "asset_receivable")
    _ct_accts_pay = get_ar_accounts(models_url, uid, api_key, "liability_payable")
    _ct_acct_rec_map = {n: i for i, n in _ct_accts_rec}
    _ct_acct_pay_map = {n: i for i, n in _ct_accts_pay}


    with st.form("ct_form"):
        # ── Persona / Empresa ──────────────────────────────────────────────
        _ct_company_type = st.radio(
            "Tipo de entidad", ["🏢 Empresa", "👤 Persona"],
            horizontal=True, index=0,
        )
        _ct_is_company_val = (_ct_company_type == "🏢 Empresa")

        # ── Tipo de contacto ───────────────────────────────────────────────
        st.markdown("##### 🏷️ Rol")
        _ct_c1, _ct_c2 = st.columns(2)
        _ct_is_customer = _ct_c1.checkbox("Es cliente", value=True)
        _ct_is_supplier = _ct_c2.checkbox("Es proveedor", value=False)

        # ── Datos básicos ──────────────────────────────────────────────────
        st.markdown("##### 🏢 Datos básicos")
        _ct_b1, _ct_b2 = st.columns(2)
        _ct_name  = _ct_b1.text_input("Razón social *",
            value=_arca.get("nombre", ""),
            placeholder="ACME S.A.")
        _ct_cuit  = _ct_b2.text_input("CUIT *",
            value=_arca.get("cuit", ""),
            placeholder="30-12345678-9")
        _ct_phone = _ct_b1.text_input("Teléfono",
            placeholder="+54 351 xxx-xxxx")
        _ct_email = _ct_b2.text_input("Correo electrónico",
            placeholder="contacto@empresa.com")
        _ct_web   = st.text_input("Sitio web",
            placeholder="https://www.empresa.com")

        # ── Dirección ──────────────────────────────────────────────────────
        st.markdown("##### 📍 Dirección fiscal")
        _ct_d1, _ct_d2, _ct_d3 = st.columns([3, 2, 1])
        _ct_street = _ct_d1.text_input("Calle y número",
            value=_arca.get("street", ""))
        _ct_city   = _ct_d2.text_input("Ciudad / Localidad",
            value=_arca.get("city", ""))
        _ct_zip    = _ct_d3.text_input("C.P.",
            value=_arca.get("zip_code", ""))
        _ct_state_opts = ["— Seleccionar —"] + list(_state_map.keys())
        _ct_state_def  = _ct_state_opts.index(_def_prov) if _def_prov in _ct_state_opts else 0
        _ct_state_sel  = st.selectbox("Provincia", _ct_state_opts, index=_ct_state_def)

        # ── AFIP ───────────────────────────────────────────────────────────
        st.markdown("##### 🏛️ Fiscal AFIP")
        _ct_f1, _ct_f2 = st.columns(2)
        _ct_afip_sel = _ct_f1.selectbox(
            "Tipo de responsabilidad AFIP",
            _afip_names if _afip_names else ["(no disponible)"],
            index=_def_afip_idx,
        )
        _ct_ref = _ct_f2.text_input("Referencia interna", placeholder="Ej: Canal Online, Zona Norte")

        # ── Ventas ─────────────────────────────────────────────────────────
        st.markdown("##### 💼 Ventas")
        _ct_v1, _ct_v2, _ct_v3 = st.columns(3)
        _ct_user_opts = ["— Sin vendedor —"] + list(_user_map.keys())
        _ct_user_sel  = _ct_v1.selectbox("Vendedor", _ct_user_opts)
        _ct_pt_opts   = ["— Sin plazo —"] + list(_pt_map.keys())
        _ct_pt_sel    = _ct_v2.selectbox("Términos de pago (ventas)", _ct_pt_opts)
        _ct_pl_opts   = ["— Predeterminado —"] + list(_plist_map.keys())
        _ct_pl_sel    = _ct_v3.selectbox("Lista de precios", _ct_pl_opts)

        _ct_ref_opts  = ["— Sin referido —"] + list(_referido_map.keys())
        _ct_ref_sel   = st.selectbox(
            "Referido",
            _ct_ref_opts,
            help="Quien refirio a este contacto",
        )

        # ── Compras ────────────────────────────────────────────────────────
        st.markdown("##### 🛒 Compras")
        _ct_p1, _ct_p2 = st.columns(2)
        _ct_pt_purch_sel = _ct_p1.selectbox("Términos de pago (compras)", _ct_pt_opts,
            key="ct_pt_purch")
        _ct_ref_purch = _ct_p2.text_input("Referencia del proveedor",
            placeholder="Código que el proveedor nos asigna")

        # ── Contabilidad ───────────────────────────────────────────────────
        st.markdown("##### 📊 Contabilidad")
        _ct_acct_c1, _ct_acct_c2 = st.columns(2)
        _ct_rec_opts = ["— Predeterminada —"] + list(_ct_acct_rec_map.keys())
        _ct_pay_opts = ["— Predeterminada —"] + list(_ct_acct_pay_map.keys())
        # Default: primera cuenta de cada tipo
        _ct_rec_def = 1 if _ct_accts_rec else 0
        _ct_pay_def = 1 if _ct_accts_pay else 0
        _ct_acct_rec_sel = _ct_acct_c1.selectbox(
            "Cuenta por cobrar", _ct_rec_opts, index=_ct_rec_def,
            help="Cuenta contable para créditos por ventas")
        _ct_acct_pay_sel = _ct_acct_c2.selectbox(
            "Cuenta por pagar", _ct_pay_opts, index=_ct_pay_def,
            help="Cuenta contable para proveedores")
        _ct_credit_limit = st.number_input(
            "Límite de crédito", min_value=0.0, value=0.0,
            step=1000.0, format="%.2f",
            help="0 = sin límite")

        # ── Notas ──────────────────────────────────────────────────────────
        st.markdown("##### 📝 Notas internas")
        _ct_notes = st.text_area("Notas", height=60, label_visibility="collapsed",
            value=_arca.get("actividad_principal", ""))

        _ct_go = st.form_submit_button("💾 Crear en Odoo", use_container_width=True, type="primary")

    if _ct_go:
        if not _ct_name.strip():
            st.error("La razón social es obligatoria.")
        elif not _ct_cuit.strip():
            st.error("El CUIT es obligatorio.")
        else:
            with st.spinner("Creando contacto en Odoo..."):
                try:
                    _ct_vals = {
                        "name":          _ct_name.strip(),
                        "is_company":    _ct_is_company_val,
                        "company_type":  "company" if _ct_is_company_val else "person",
                        "customer_rank": 1 if _ct_is_customer else 0,
                        "supplier_rank": 1 if _ct_is_supplier else 0,
                    }
                    # CUIT
                    _vat_clean = re.sub(r"[\s\-]", "", _ct_cuit.strip())
                    _ct_vals["vat"] = _vat_clean
                    if _ct_cuit_tid:
                        _ct_vals["l10n_latam_identification_type_id"] = _ct_cuit_tid

                    # Contacto
                    if _ct_phone.strip(): _ct_vals["phone"]   = _ct_phone.strip()
                    if _ct_email.strip(): _ct_vals["email"]   = _ct_email.strip()
                    if _ct_web.strip():   _ct_vals["website"] = _ct_web.strip()

                    # Dirección
                    if _ct_street.strip(): _ct_vals["street"] = _ct_street.strip()
                    if _ct_city.strip():   _ct_vals["city"]   = _ct_city.strip()
                    if _ct_zip.strip():    _ct_vals["zip"]    = _ct_zip.strip()
                    if _ct_state_sel and _ct_state_sel != "— Seleccionar —":
                        _ct_vals["state_id"]   = _state_map[_ct_state_sel]
                        _ct_vals["country_id"] = models.execute_kw(
                            ODOO_DB, uid, api_key, "res.country", "search",
                            [[["code", "=", "AR"]]], {"limit": 1})[0]

                    # AFIP
                    if _ct_afip_sel and _ct_afip_sel in _afip_map:
                        _ct_vals["l10n_ar_afip_responsibility_type_id"] = _afip_map[_ct_afip_sel]

                    # Referencia
                    if _ct_ref.strip(): _ct_vals["ref"] = _ct_ref.strip()

                    # Ventas
                    if _ct_user_sel != "— Sin vendedor —":
                        _ct_vals["user_id"] = _user_map[_ct_user_sel]
                    if _ct_pt_sel != "— Sin plazo —":
                        _ct_vals["property_payment_term_id"] = _pt_map[_ct_pt_sel]
                    if _ct_pl_sel != "— Predeterminado —":
                        _ct_vals["property_product_pricelist"] = _plist_map[_ct_pl_sel]
                    if _ct_ref_sel != "— Sin referido —" and _ct_ref_sel in _referido_map:
                        _ct_vals["x_studio_referido_1"] = _referido_map[_ct_ref_sel]

                    # Compras
                    if _ct_pt_purch_sel != "— Sin plazo —":
                        _ct_vals["property_supplier_payment_term_id"] = _pt_map[_ct_pt_purch_sel]
                    if _ct_ref_purch.strip():
                        _ct_vals["ref"] = _ct_ref_purch.strip()

                    # Notas
                    if _ct_notes.strip():
                        _ct_vals["comment"] = _ct_notes.strip()


                    # Cuentas contables
                    if _ct_acct_rec_sel != "— Predeterminada —" and _ct_acct_rec_sel in _ct_acct_rec_map:
                        _ct_vals["property_account_receivable_id"] = _ct_acct_rec_map[_ct_acct_rec_sel]
                    if _ct_acct_pay_sel != "— Predeterminada —" and _ct_acct_pay_sel in _ct_acct_pay_map:
                        _ct_vals["property_account_payable_id"] = _ct_acct_pay_map[_ct_acct_pay_sel]

                    # Límite de crédito
                    if _ct_credit_limit > 0:
                        _ct_vals["credit_limit"] = _ct_credit_limit

                    _new_pid = create_full_partner(models, uid, api_key, _ct_vals)
                    _new_url = odoo_url("res.partner", _new_pid)
                    st.toast("Contacto creado en Odoo", icon="✅")
                    st.markdown(f"🎉 **{_ct_name}** creado · [Abrir en Odoo]({_new_url})")
                    st.session_state.history.append({
                        "tipo": "Contacto",
                        "archivo": _ct_name,
                        "id": _new_pid,
                        "url": _new_url,
                        "estado": "✅",
                        "hora": _dt_now.now().strftime("%H:%M"),
                    })
                except Exception as _cte:
                    st.error(f"❌ {_cte}")

if tab_import is not None:
    with tab_import:
        st.subheader("🛳️ Importaciones — Modo Claude")

        # ── Input carpeta ─────────────────────────────────────────
        col_ci, col_btns = st.columns([3, 2])
        carp_in = col_ci.text_input("Carpeta", value=st.session_state.carpeta_id,
            placeholder="LUMI_293", key="input_carpeta", label_visibility="collapsed")
        with col_btns:
            _b1, _b2, _b3 = st.columns(3)
            load_btn   = _b1.button("🔍 Cargar",   key="btn_load_carp",   use_container_width=True)
            reset_btn  = _b2.button("🔄 Nueva",    key="btn_reset_carp",  use_container_width=True)
            cancel_btn = _b3.button("❌ Cancelar", key="btn_cancel_carp", use_container_width=True)

        if reset_btn or cancel_btn:
            for k in ["carpeta_id", "carpeta_po", "carpeta_bills", "carpeta_lc_id"]:
                st.session_state[k] = DEFAULTS.get(k, "")
            st.session_state.etapas = {k: False for k, *_ in ETAPAS_DEF}
            # Borrar el estado interno del widget text_input (no se puede setear, solo borrar)
            for _sk in ["input_carpeta", "_imp_preview_open", "carp_data"]:
                st.session_state.pop(_sk, None)
            st.rerun()

        if carp_in != st.session_state.carpeta_id:
            st.session_state.carpeta_id = carp_in
            if "carp_data" in st.session_state:
                del st.session_state["carp_data"]

        if load_btn and st.session_state.carpeta_id:
            load_carpeta_full.clear()
            with st.spinner(f"Cargando {st.session_state.carpeta_id} desde Odoo..."):
                cdata = load_carpeta_full(models_url, uid, api_key, st.session_state.carpeta_id)
            st.session_state["carp_data"] = cdata
            # Siempre resetear carpeta_po y carpeta_bills desde Odoo (evitar valor "sucio" de otra carpeta)
            st.session_state.carpeta_po    = cdata.get("po")   # None si no hay OC
            st.session_state.carpeta_bills = [b["id"] for b in cdata.get("bills", [])]
            st.session_state.etapas        = {k: False for k, *_ in ETAPAS_DEF}
            for k, v in cdata.get("stages", {}).items():
                if v:
                    st.session_state.etapas[k] = True
            st.rerun()

        if not st.session_state.carpeta_id:
            st.info("Ingresá el número de carpeta y presioná **🔍 Cargar** para traer los datos de Odoo.")

        if st.session_state.carpeta_id:
            carp_data = st.session_state.get("carp_data")


            st.divider()

            # ── Subir comprobantes ────────────────────────────────────
            st.markdown("#### ⬆️ Subir comprobantes")
            # ── Resultado de la última carga (persiste tras rerun) ─────────
            if "_imp_create_result" in st.session_state:
                _res = st.session_state.pop("_imp_create_result")
                if _res.get("ok", 0):
                    st.success(f"✅ {_res['ok']} registro(s) creados para **{_res['carp']}**.")
                    for _itm in _res.get("items", []):
                        st.markdown(f"  📎 [{_itm['file']}]({_itm['url']}) → ID {_itm['id']}")
                for _rerr in _res.get("errs", []):
                    st.error(_rerr)

            st.info("💡 Podés subir documentos de cualquier etapa en cualquier orden. "
                    "seleccioná múltiples archivos con **Ctrl+A** en el explorador, "
                    "o arrastrá y soltá varios archivos al mismo tiempo.")

            TIPO_OPTIONS_IMP = {
                "Bill PETDUR (Etapa 1)":           {"tipo":"petdur",  "partner_id":49328,"journal_id":71, "doc_type":None},
                "DI AFIP (Etapa 2)":               {"tipo":"di_afip", "partner_id":9,    "journal_id":10, "doc_type":66},
                "Bill TRICE Transport (Etapa 2a)": {"tipo":"nac",     "partner_id":48825,"journal_id":10, "doc_type":None},
                "Bill Terminal 4 SA (Etapa 2a)":   {"tipo":"nac",     "partner_id":48828,"journal_id":10, "doc_type":None},
                "Bill Mundo Comex (Etapa 2a)":     {"tipo":"nac",     "partner_id":48826,"journal_id":10, "doc_type":None},
                "Bill SENASA (Etapa 2a)":          {"tipo":"nac",     "partner_id":48827,"journal_id":10, "doc_type":None},
                "Otro comprobante":                {"tipo":"other",   "partner_id":None, "journal_id":10, "doc_type":None},
            }

            imp_files = st.file_uploader(
                f"Documentos de {st.session_state.carpeta_id} — seleccioná uno o todos a la vez",
                type=["pdf","jpg","jpeg","png"], accept_multiple_files=True, key="import_uploader")

            classified_docs = []
            if imp_files:
                st.markdown("**Clasificación automática — revisá y ajustá si hace falta**")
                for uf in imp_files:
                    ext        = uf.name.rsplit(".", 1)[-1].lower()
                    file_bytes = uf.read()
                    mimetype   = MIMETYPES.get(ext, "application/octet-stream")
                    raw_text   = ""
                    if ext == "pdf":
                        _, raw_text = extract_pdf_fields(file_bytes)
                    auto        = classify_document(raw_text, st.session_state.carpeta_id)
                    default_lbl = auto["label"] if auto["label"] in TIPO_OPTIONS_IMP else "Otro comprobante"
                    _ext_info   = auto.get("extracted", {})
                    _icon       = "🚫" if auto.get("no_aplica") else ("⚠️" if auto.get("mismatch") else "📎")
                    with st.expander(f"{_icon} {uf.name} — {auto['label']}", expanded=True):
                        # Warnings de no-aplica y mismatch
                        if auto.get("no_aplica"):
                            st.warning(f"🚫 **Este documento no aplica** ({auto['label']}) — "
                                       "no se cargará en Odoo. Podés ignorarlo.")
                        elif auto.get("mismatch"):
                            st.warning(f"⚠️ **Carpeta mismatch** — este doc referencia "
                                       f"**{_ext_info.get('mismatch_ref','')}** "
                                       f"pero la carpeta activa es **{st.session_state.carpeta_id}**. "
                                       "Verificá que sea el documento correcto.")

                        # ── Info extraída del PDF ──────────────────────────────
                        _imp_fields = [
                            ("N° Comp.",  _ext_info.get("nro_comp")),
                            ("Fecha",     _ext_info.get("fecha")),
                            ("CUIT",      _ext_info.get("cuit")),
                            ("Monto",     _ext_info.get("monto")),
                            ("TC (PDF)",  _ext_info.get("tc_pdf")),
                        ]
                        _vis_fields = [(k, v) for k, v in _imp_fields if v]
                        if _vis_fields:
                            _info_html = "  &nbsp;·&nbsp;  ".join(
                                f"<b>{k}:</b>&nbsp;<code>{v}</code>"
                                for k, v in _vis_fields)
                            st.markdown(
                                f'<div style="font-size:0.92rem; background:#f0f2f6; '
                                f'padding:7px 14px; border-radius:7px; margin:4px 0; '
                                f'line-height:1.7; color:#262730;">{_info_html}</div>',
                                unsafe_allow_html=True)
                        # Líneas de productos (PETDUR)
                        _pet_lns = _ext_info.get("lineas_petdur", [])
                        if _pet_lns:
                            st.markdown("**Líneas de la factura:**")
                            _pet_df = pd.DataFrame([{
                                "Descripción": ln["descripcion"],
                                "Cant.":       ln["cantidad"],
                                "P. Unit.":    ln["precio_unit"],
                                "Monto":       ln["monto"],
                            } for ln in _pet_lns])
                            _pet_cfg = {
                                "Cant.":    st.column_config.NumberColumn("Cant.", format="%.2f"),
                                "P. Unit.": st.column_config.NumberColumn("P. Unit.", format="%.2f"),
                                "Monto":    st.column_config.NumberColumn("Monto", format="%.2f"),
                            }
                            st.dataframe(_pet_df, column_config=_pet_cfg,
                                         use_container_width=True, hide_index=True)

                        if not auto.get("no_aplica"):
                            ct1, ct2, ct3, ct4 = st.columns([3, 2, 2, 1])
                            tipo_sel  = ct1.selectbox("Tipo", list(TIPO_OPTIONS_IMP.keys()),
                                index=list(TIPO_OPTIONS_IMP.keys()).index(default_lbl), key=f"tipo_{uf.name}")
                            # Pre-poblar session_state desde extracción si el campo está vacío
                            _ref_key = f"ref_d_{uf.name}"
                            _fec_key = f"fec_d_{uf.name}"
                            if not st.session_state.get(_ref_key) and _ext_info.get("nro_comp"):
                                st.session_state[_ref_key] = _ext_info["nro_comp"]
                            if not st.session_state.get(_fec_key) and _ext_info.get("fecha"):
                                st.session_state[_fec_key] = _ext_info["fecha"]
                            ref_doc   = ct2.text_input("N° comprobante", key=_ref_key)
                            fecha_doc = ct3.text_input("Fecha (AAAA-MM-DD)", key=_fec_key,
                                placeholder="2026-05-12")
                            _def_mon  = "USD" if auto.get("tipo") == "petdur" else "ARS"
                            moneda    = ct4.selectbox("Moneda", ["ARS","USD"],
                                index=["ARS","USD"].index(_def_mon), key=f"cur_{uf.name}")
                            if moneda == "USD":
                                _rd = fecha_doc or pd.Timestamp.today().strftime("%Y-%m-%d")
                                _tc_up, _dt_up = get_usd_rate_odoo(models_url, uid, api_key, _rd)
                                st.caption(f"TC Odoo para {_dt_up or _rd}: **$ {_tc_up:,.2f}**"
                                           if _tc_up else "TC no encontrado en Odoo para esa fecha")
                            classified_docs.append({
                                "filename":  uf.name, "file_bytes": file_bytes, "mimetype": mimetype,
                                "tipo_cfg":  TIPO_OPTIONS_IMP[tipo_sel],
                                "ref":       ref_doc, "fecha": fecha_doc, "moneda": moneda,
                                "extracted": _ext_info,
                            })

                if classified_docs:
                    # ── Vista previa antes de crear ────────────────────────────
                    _prev_btn, _create_btn = st.columns([1, 2])
                    _show_prev = _prev_btn.button("👁️ Vista previa", key="btn_preview_imp",
                                                  use_container_width=True)
                    if _show_prev or st.session_state.get("_imp_preview_open"):
                        st.session_state["_imp_preview_open"] = True
                        st.markdown("**Resumen de lo que se va a crear en Odoo:**")
                        _prev_rows = []
                        for _d in classified_docs:
                            _cfg = _d["tipo_cfg"]
                            _pid = _cfg.get("partner_id")
                            _pname = PARTNER_TO_TIPO.get(_pid, {}).get("label", "—") if _pid else "Sin asignar"
                            _tc_prev = "—"
                            if _d.get("moneda") == "USD":
                                _ref_date = _d.get("fecha") or pd.Timestamp.today().strftime("%Y-%m-%d")
                                _tv, _td = get_usd_rate_odoo(models_url, uid, api_key, _ref_date)
                                _tc_prev = f"$ {_tv:,.0f}" if _tv else "sin TC"
                            _prev_rows.append({
                                "Archivo":    _d["filename"],
                                "Tipo":       _cfg.get("label","—")[:30],
                                "Proveedor":  _pname,
                                "Ref.":       (f"{st.session_state.carpeta_id} / {_d['ref']}"
                                              if _d.get("ref") else st.session_state.carpeta_id),
                                "Fecha":      _d.get("fecha") or "—",
                                "Moneda":     _d.get("moneda","ARS"),
                                "TC ARS/USD": _tc_prev,
                            })
                        st.dataframe(pd.DataFrame(_prev_rows), use_container_width=True, hide_index=True)
                        st.caption("Revisá los datos antes de confirmar. Podés modificar cualquier campo arriba.")

                    if _create_btn.button(f"⬆️ Confirmar y crear {len(classified_docs)} registro(s) en Odoo",
                                 type="primary", key="btn_create_all_imp", use_container_width=True):
                        st.session_state["_imp_preview_open"] = False
                        _prog = st.progress(0)
                        _ok, _errs = 0, []
                        _created_items = []
                        _carp = st.session_state.carpeta_id
                        _usd_id = get_currency_id(models_url, uid, api_key, "USD")
                        # TC del día para x_studio_cotizacion_dolar (campo requerido por Odoo)
                        _today_str = pd.Timestamp.today().strftime("%Y-%m-%d")
                        _tc_hoy, _  = get_usd_rate_odoo(models_url, uid, api_key, _today_str)

                        # ── Auto-crear OC PETDUR (Etapa 0) si no existe y hay Bill PETDUR ──
                        _tiene_petdur = any(d["tipo_cfg"]["tipo"] == "petdur" for d in classified_docs)
                        _petdur_doc   = next((d for d in classified_docs
                                              if d["tipo_cfg"]["tipo"] == "petdur"), None)
                        _existing_po  = st.session_state.get("carpeta_po")
                        _has_oc       = bool(_existing_po)
                        _new_po_id    = None
                        _oc_error     = None
                        if _tiene_petdur and not _has_oc:
                            try:
                                _po_id = create_purchase_order_petdur(
                                    models, uid, api_key, _carp,
                                    currency_id = _usd_id if _usd_id else None,
                                    tc_usd      = _tc_hoy,   # Cotización Dólar requerida por Odoo
                                    filename    = (_petdur_doc or {}).get("filename"),
                                    file_bytes  = (_petdur_doc or {}).get("file_bytes"),
                                    mimetype    = (_petdur_doc or {}).get("mimetype"),
                                    lineas      = ((_petdur_doc or {}).get("extracted") or {}).get("lineas_petdur") or [],
                                )
                                _po_url    = odoo_url("purchase.order", _po_id)
                                _new_po_id = _po_id
                                st.session_state.etapas["0"] = True
                                st.session_state.carpeta_po  = {"id": _po_id, "name": _carp}
                                _created_items.append({
                                    "file": f"🏭 OC {_carp} (Etapa 0 auto-creada)",
                                    "id": _po_id, "url": _po_url
                                })
                                _ok += 1
                            except Exception as _e_po:
                                _oc_error = str(_e_po)[:300]
                                _errs.append(f"❌ No se pudo crear OC: {_oc_error}")
                        elif _tiene_petdur and _has_oc:
                            # OC ya existe — agregar al resumen para que el usuario la vea
                            _ex_po_id  = (_existing_po or {}).get("id")
                            _ex_po_url = odoo_url("purchase.order", _ex_po_id) if _ex_po_id else "#"
                            _created_items.append({
                                "file": f"🏭 OC ya existente ({_carp})",
                                "id":   _ex_po_id, "url": _ex_po_url
                            })

                        # Pre-cargar cuenta contable del journal USD para líneas PETDUR
                        _petdur_jid     = 71
                        _petdur_account = get_journal_purchase_account(
                            models_url, uid, api_key, _petdur_jid)

                        for _i, _doc in enumerate(classified_docs):
                            try:
                                _tipo     = _doc["tipo_cfg"]["tipo"]
                                _ext      = _doc.get("extracted", {})
                                _full_ref = f"{_carp} / {_doc['ref']}" if _doc.get("ref") else _carp
                                _cur_id   = _usd_id if _doc.get("moneda", "ARS") == "USD" else None

                                # Si es PETDUR y la OC falló, no crear la factura tampoco
                                if _tipo == "petdur" and _tiene_petdur and _oc_error:
                                    _errs.append(
                                        f"⏭️ {_doc['filename']}: factura omitida porque la OC no se pudo crear")
                                    continue

                                # Construir líneas para facturas PETDUR
                                _bill_lines = None
                                if _tipo == "petdur" and _petdur_account:
                                    _raw_lns = _ext.get("lineas_petdur", [])
                                    if _raw_lns:
                                        _bill_lines = [
                                            {
                                                "name":       ln["descripcion"],
                                                "quantity":   ln["cantidad"],
                                                "price_unit": ln["precio_unit"],
                                                "account_id": _petdur_account,
                                            }
                                            for ln in _raw_lns
                                        ]
                                    elif _ext.get("monto"):
                                        _bill_lines = [{
                                            "name":       _full_ref,
                                            "quantity":   1,
                                            "price_unit": float(_ext["monto"]),
                                            "account_id": _petdur_account,
                                        }]

                                # N° comprobante: primero lo que ingresó el usuario, luego extracted
                                _nro_comp = (_doc.get("ref") or _ext.get("nro_comp") or "").strip()

                                _move_id = create_vendor_bill(models, uid, api_key,
                                    partner_id      = _doc["tipo_cfg"]["partner_id"],
                                    ref             = _full_ref,
                                    invoice_date    = _doc["fecha"] or False,
                                    filename        = _doc["filename"],
                                    file_bytes      = _doc["file_bytes"],
                                    mimetype        = _doc["mimetype"],
                                    journal_id      = _doc["tipo_cfg"]["journal_id"],
                                    doc_type_id     = _doc["tipo_cfg"]["doc_type"],
                                    currency_id     = _cur_id,
                                    invoice_origin  = _carp,
                                    l10n_latam_document_number = _nro_comp or None,
                                    extra_lines     = _bill_lines,
                                )
                                _url = odoo_url("account.move", _move_id)
                                _created_items.append({"file": _doc["filename"], "id": _move_id, "url": _url})
                                _tipo = _doc["tipo_cfg"]["tipo"]
                                if _tipo == "petdur":    st.session_state.etapas["1"]  = True
                                elif _tipo == "di_afip": st.session_state.etapas["2"]  = True
                                elif _tipo == "nac":     st.session_state.etapas["2a"] = True
                                if _move_id not in st.session_state.carpeta_bills:
                                    st.session_state.carpeta_bills.append(_move_id)
                                st.session_state.history.append({
                                    "tipo":   f"Importación {_carp}",
                                    "archivo":_doc["filename"], "id":_move_id,
                                    "url":_url, "estado":"✅",
                                    "hora": _dt_now.now().strftime("%H:%M")
                                })
                                _ok += 1
                            except Exception as _e:
                                _errs.append(f"❌ {_doc['filename']}: {str(_e)[:120]}")
                            _prog.progress((_i + 1) / len(classified_docs))
                        st.session_state["_imp_create_result"] = {
                            "ok": _ok, "carp": _carp,
                            "items": _created_items, "errs": _errs
                        }
                        if _ok:
                            load_carpeta_full.clear()
                        st.rerun()

            st.divider()

            st.divider()

            # ── Resumen de carpeta ────────────────────────────────────
            if carp_data:
                if carp_data.get("error"):
                    st.error(f"Error al cargar desde Odoo: {carp_data['error']}")
                po       = carp_data.get("po")
                bills    = carp_data.get("bills", [])
                pickings = carp_data.get("pickings", [])
                tc_oc    = carp_data.get("tc_oc")
                if po:
                    pname    = po["name"]
                    ppartner = po.get("partner_id", [0, "—"])[1]
                    ptotal   = po.get("amount_total", 0)
                    pcur     = po.get("currency_id", [0, "USD"])[1] if po.get("currency_id") else "USD"
                    po_url   = odoo_url("purchase.order", po['id'])
                    ci1, ci2, ci3 = st.columns([3, 2, 1])
                    ci1.success(f"📦 **{st.session_state.carpeta_id}** — OC {pname} · {ppartner}")
                    ci2.info(f"💵 {pcur} {ptotal:,.2f}" + (f" · TC: **$ {tc_oc:,.0f}**" if tc_oc else ""))
                    ci3.markdown(f"[🔗 Ver OC]({po_url})")
                    done_picks = [p for p in pickings if p.get("state") == "done"]
                    if pickings:
                        st.caption(f"📥 {len(pickings)} picking(s) · {len(done_picks)} recibido(s) en depósito")
                elif bills:
                    st.warning(f"⚠️ {len(bills)} comprobante(s) encontrados pero sin OC vinculada.")
                else:
                    st.info(f"**{st.session_state.carpeta_id}** no encontrada en Odoo. "
                            f"Subí los documentos para crear los registros.")

            # ── Progreso de etapas ────────────────────────────────────
            st.markdown("#### 📋 Estado de la carpeta")
            cols_et = st.columns(len(ETAPAS_DEF))
            for i, (key, label, desc) in enumerate(ETAPAS_DEF):
                done = st.session_state.etapas.get(key, False)
                icon = "✅" if done else "⏳"
                cols_et[i].markdown(f"**{icon}**  \n<small>{label}</small>",
                                    unsafe_allow_html=True, help=desc)
                _auto_etapas = {"0", "1", "2", "2a"}  # se setean solo al subir docs
                if not done and st.session_state.carpeta_id:
                    if key in _auto_etapas:
                        cols_et[i].caption("auto")
                    else:
                        if cols_et[i].button("✓", key=f"et_{key}", help=f"Marcar {label}"):
                            _conf_key = f"_et_conf_{key}"
                            st.session_state[_conf_key] = True
                            st.rerun()
            completadas = sum(1 for v in st.session_state.etapas.values() if v)
            st.progress(completadas / len(ETAPAS_DEF),
                        text=f"{completadas}/{len(ETAPAS_DEF)} etapas completadas")

            # ── Confirmación de marcado manual de etapa ───────────────────
            for _ck, _clbl, _cdesc in ETAPAS_DEF:
                _conf_key = f"_et_conf_{_ck}"
                if st.session_state.get(_conf_key):
                    st.warning(f"⚠️ ¿Confirmás que **{_clbl}** está completada en Odoo? "
                               "Esta acción no se puede deshacer automáticamente.")
                    _cc1, _cc2 = st.columns(2)
                    if _cc1.button(f"✅ Sí, marcar {_clbl}", key=f"et_conf_yes_{_ck}"):
                        st.session_state.etapas[_ck] = True
                        st.session_state.pop(_conf_key)
                        st.rerun()
                    if _cc2.button("Cancelar", key=f"et_conf_no_{_ck}"):
                        st.session_state.pop(_conf_key)
                        st.rerun()

            st.divider()

            # ── Comprobantes cargados en Odoo ─────────────────────────────────
            if carp_data and carp_data.get("bills"):
                st.markdown("#### 📄 Comprobantes en Odoo")

                _bill_ids_all = [b["id"] for b in carp_data["bills"]]
                _all_lines    = get_bill_lines(models_url, uid, api_key, _bill_ids_all)
                _estado_map   = {
                    "draft":  "Borrador 📝",
                    "posted": "Confirmado ✅",
                    "cancel": "Cancelado ❌",
                }
                _total_ars_bills = 0.0

                for b in carp_data["bills"]:
                    pid      = b["partner_id"][0] if b.get("partner_id") else 0
                    tipo_inf = PARTNER_TO_TIPO.get(pid,
                        {"etapa": "—", "label": b["partner_id"][1]
                         if b.get("partner_id") else "Otro"})
                    cur_name = b["currency_id"][1] if b.get("currency_id") else "ARS"

                    tc_val = None
                    if cur_name == "USD":
                        icr = b.get("invoice_currency_rate")
                        if icr and icr is not False:
                            tc_val = _parse_odoo_rate({"rate": icr, "inverse_company_rate": None})
                        if not tc_val and carp_data.get("tc_oc"):
                            tc_val = carp_data["tc_oc"]

                    amt_orig = float(b.get("amount_total") or 0)
                    amt_ars  = abs(float(b.get("amount_total_signed") or amt_orig or 0))
                    _total_ars_bills += amt_ars
                    bill_url   = odoo_url("account.move", b["id"])
                    estado_lbl = _estado_map.get(b.get("state", ""), b.get("state", ""))
                    _b_name    = b.get("name") or f"ID {b['id']}"
                    _b_lines   = _all_lines.get(b["id"], [])

                    _exp_title = (
                        f"Etapa {tipo_inf['etapa']} · {tipo_inf['label']} · "
                        f"{_b_name} · {cur_name} {amt_orig:,.2f} · {estado_lbl}"
                    )
                    with st.expander(_exp_title, expanded=False):
                        _bc1, _bc2, _bc3, _bc4 = st.columns([2, 2, 3, 1])
                        _bc1.markdown(f"**Fecha:** {b.get('invoice_date') or '—'}")
                        _bc2.markdown(
                            f"**Moneda:** {cur_name}"
                            + (f"  ·  TC: ${tc_val:,.0f}" if tc_val else ""))
                        _bc3.markdown(
                            f"**Total:** {cur_name} {amt_orig:,.2f}"
                            + (f"  =  ARS {fmt_ars(amt_ars)}" if cur_name == "USD" else ""))
                        _bc4.markdown(f"[🔗 Odoo]({bill_url})")

                        if _b_lines:
                            _ln_rows = []
                            for ln in _b_lines:
                                _pname = (ln["product_id"][1]
                                          if ln.get("product_id") else ln.get("name", "—"))
                                _ln_rows.append({
                                    "Producto":    _pname,
                                    "Cant.":       float(ln.get("quantity") or 0),
                                    "P. Unit.":    float(ln.get("price_unit") or 0),
                                    "Subtotal":    float(ln.get("price_subtotal") or 0),
                                    "Total c/imp": float(ln.get("price_total") or 0),
                                })
                            _ln_cfg = {
                                "Cant.":       st.column_config.NumberColumn("Cant.", format="%.2f"),
                                "P. Unit.":    st.column_config.NumberColumn("P. Unit.", format="%.4f"),
                                "Subtotal":    st.column_config.NumberColumn("Subtotal", format="%.2f"),
                                "Total c/imp": st.column_config.NumberColumn("Total c/imp", format="%.2f"),
                            }
                            st.dataframe(pd.DataFrame(_ln_rows), column_config=_ln_cfg,
                                         use_container_width=True, hide_index=True)
                            _sum_sub = sum(r["Subtotal"]    for r in _ln_rows)
                            _sum_tot = sum(r["Total c/imp"] for r in _ln_rows)
                            _pie = [
                                f"<b>Neto s/imp:</b>&nbsp;<code>{cur_name} {_sum_sub:,.2f}</code>",
                                f"<b>Total c/imp:</b>&nbsp;<code>{cur_name} {_sum_tot:,.2f}</code>",
                            ]
                            if tc_val:
                                _pie.append(
                                    f"<b>ARS equiv.:</b>&nbsp;"
                                    f"<code>{fmt_ars(_sum_tot * tc_val)}</code>")
                            st.markdown(
                                '<div style="font-size:0.9rem;background:#f0f2f6;'
                                'padding:6px 12px;border-radius:6px;margin:4px 0;">'
                                + "&nbsp;&nbsp;·&nbsp;&nbsp;".join(_pie)
                                + "</div>",
                                unsafe_allow_html=True)
                        else:
                            # Para PETDUR: mostrar líneas de la OC como referencia
                            _is_petdur = (pid == 49328)
                            _po_ref = carp_data.get("po") if carp_data else None
                            if _is_petdur and _po_ref:
                                _po_lns = get_po_lines(models_url, uid, api_key, _po_ref["id"])
                                if _po_lns:
                                    st.caption("📋 Sin líneas en la factura — mostrando líneas de la OC como referencia:")
                                    _po_rows = []
                                    for pl in _po_lns:
                                        _po_rows.append({
                                            "Producto":  pl["product_id"][1] if pl.get("product_id") else pl.get("name","—"),
                                            "Cant.":     float(pl.get("product_qty") or 0),
                                            "P. Unit.":  float(pl.get("price_unit") or 0),
                                            "Subtotal":  float(pl.get("price_subtotal") or 0),
                                        })
                                    _po_cfg = {
                                        "Cant.":    st.column_config.NumberColumn("Cant.", format="%.2f"),
                                        "P. Unit.": st.column_config.NumberColumn("P. Unit.", format="%.4f"),
                                        "Subtotal": st.column_config.NumberColumn("Subtotal", format="%.2f"),
                                    }
                                    st.dataframe(pd.DataFrame(_po_rows), column_config=_po_cfg,
                                                 use_container_width=True, hide_index=True)
                                    _oc_sub = sum(r["Subtotal"] for r in _po_rows)
                                    st.markdown(
                                        '<div style="font-size:0.9rem;background:#f0f2f6;'
                                        'padding:6px 12px;border-radius:6px;margin:4px 0;">'
                                        f"<b>Subtotal OC:</b>&nbsp;<code>{cur_name} {_oc_sub:,.2f}</code>"
                                        "</div>", unsafe_allow_html=True)
                                else:
                                    st.caption("Sin líneas en la factura ni en la OC.")
                            else:
                                st.caption("Sin líneas de producto en esta factura.")

                st.markdown(
                    '<div style="font-size:0.9rem;background:#e8f4fd;'
                    'padding:6px 14px;border-radius:6px;margin:6px 0;">'
                    f"<b>{len(carp_data['bills'])} comprobante(s)</b>"
                    "&nbsp;&nbsp;·&nbsp;&nbsp;"
                    f"<b>Equiv. ARS total:</b>&nbsp;<code>{fmt_ars(_total_ars_bills)}</code>"
                    "</div>",
                    unsafe_allow_html=True)

            st.divider()

            # ── Desglose de costos por producto ───────────────────────
            _po_cost    = st.session_state.get("carpeta_po") or (carp_data.get("po") if carp_data else None)
            _bills_cost = carp_data.get("bills", []) if carp_data else []
            _tc_oc      = carp_data.get("tc_oc")       if carp_data else None

            if _po_cost and _bills_cost:
                st.markdown("#### 💰 Desglose de costos por producto")

                # Determinar TC: OC primero, luego Odoo currency rate
                _tc_src = None
                if not _tc_oc:
                    _petdur_b = next((b for b in _bills_cost
                                      if b.get("partner_id") and b["partner_id"][0] == 49328), None)
                    _ref_date = ((_petdur_b.get("invoice_date") if _petdur_b else None)
                                 or pd.Timestamp.today().strftime("%Y-%m-%d"))
                    _tc_oc, _tc_dt = get_usd_rate_odoo(models_url, uid, api_key, _ref_date)
                    _tc_src = f"Odoo ({_tc_dt})" if _tc_oc else None
                else:
                    _tc_src = f"OC {_po_cost.get('name', '')} (campo Cotización dólar)"

                if not _tc_oc:
                    st.warning("⚠️ No se encontró TC USD/ARS. Verificá el campo 'Cotización dólar' en "
                               "la OC o que esté cargado en Odoo (Contabilidad → Divisas → USD).")
                else:
                    with st.spinner("Calculando costos..."):
                        _po_lns    = get_po_lines(models_url, uid, api_key, _po_cost["id"])
                        _cost_rows, _summary = _calc_cost_breakdown(_po_lns, _bills_cost, _tc_oc)

                    if _cost_rows and _summary:
                        st.caption(
                            f"TC USD/ARS: **$ {_summary['tc_usd']:,.2f}** — Fuente: {_tc_src}  ·  "
                            f"Coef. landeo total: **+{_summary['coef_landeo_total']:.1f}%**")
                        c1, c2, c3, c4 = st.columns(4)
                        c1.metric("FOB total (u$s)",    fmt_usd(_summary["total_fob_usd"]))
                        c2.metric("Landeo total (u$s)",  fmt_usd(_summary["total_landeo_usd"]))
                        c3.metric("Total (u$s)",         fmt_usd(_summary["grand_total_usd"]))
                        c4.metric("Total (ARS)",         fmt_ars(_summary["grand_total_ars"]))

                        with st.expander("📊 Detalle gastos de nacionalización (ARS)", expanded=False):
                            _tot_nac = _summary["total_landeo_ars"]
                            for _lbl, _amt in _summary.get("nac_detail", {}).items():
                                _pct = _amt / _tot_nac * 100 if _tot_nac > 0 else 0
                                st.caption(f"• **{_lbl}**: {fmt_ars(_amt)}  ({_pct:.1f}%)")
                            st.caption(
                                f"Bills nac encontradas: {_summary['_nac_bils_count']}  ·  "
                                f"AFIP: {'✅' if _summary['_afip_found'] else '—'}  ·  "
                                f"PETDUR: {'✅' if _summary['_petdur_found'] else '—'}")

                        st.dataframe(pd.DataFrame(_cost_rows), use_container_width=True, hide_index=True)

                        # ── Asientos que va a generar el Landed Cost ──────────
                        with st.expander("📒 Asientos estimados del Landed Cost", expanded=False):
                            st.caption(
                                "Basado en `split_method: by_current_cost_price`. "
                                "Los montos son estimados — el LC de Odoo los calcula al validar.")
                            _tc_ae   = _summary["tc_usd"]
                            _nac_ars = _summary["total_landeo_ars"]
                            _nac_usd = _summary["total_landeo_usd"]

                            # Asiento 1: bills nac ya registradas (DR Cuenta Puente / CR Proveedor)
                            st.markdown("**Asiento 1 — Bills de nacionalización (ya creadas)**")
                            _ae1 = []
                            for _lbl, _amt in _summary.get("nac_detail", {}).items():
                                _ae1.append({"Cuenta (DR)": "3284 · Cuenta Puente Recepciones",
                                             "Cuenta (CR)": f"Proveedores · {_lbl}",
                                             "Monto (ARS)": fmt_ars(_amt)})
                            if _ae1:
                                st.dataframe(pd.DataFrame(_ae1), use_container_width=True, hide_index=True)
                            else:
                                st.caption("Sin bills de nac cargadas aún.")

                            st.markdown("**Asiento 2 — Landed Cost (al validar)**")
                            st.caption("DR: Valuación de inventario por producto · CR: 3284 Cuenta Puente Recepciones")
                            _ae2 = []
                            for _row in _cost_rows:
                                # Reconstruct amounts from the formatted strings is hard; recalculate
                                pass
                            # Recalculate LC distribution from po_lines
                            _ae2 = []
                            for _ln in _po_lns:
                                _prod_name = _ln["product_id"][1] if _ln.get("product_id") else _ln.get("name","?")
                                _qty       = float(_ln.get("product_qty") or 1)
                                _fob_line  = float(_ln.get("price_subtotal") or 0)
                                _prop      = _fob_line / _summary["total_fob_usd"] if _summary["total_fob_usd"] > 0 else 0

                                _nac_usd_ln   = _nac_usd * _prop
                                _lc_unit_usd  = (_nac_usd_ln / _qty) if _qty > 0 else 0
                                _lc_total_ars = _nac_usd_ln * _tc_ae
                                _ae2.append({
                                    "Producto":       _prod_name[:40],
                                    "Cant.":          int(_qty),
                                    "LC dist. (u$s)": fmt_usd(_nac_usd_ln),
                                    "LC dist. (ARS)": fmt_ars(_lc_total_ars),
                                    "costo/u (u$s)":  fmt_usd(_lc_unit_usd),
                                })
                            if _ae2:
                                st.dataframe(pd.DataFrame(_ae2),
                                             use_container_width=True, hide_index=True)
                                st.caption(
                                    "Total CR 3284: "
                                    + fmt_ars(_nac_ars)
                                    + "  ·  Total LC en u$s: "
                                    + fmt_usd(_nac_usd))
                            else:
                                st.caption("Sin datos de OC para calcular el Landed Cost.")

            st.divider()

            # ── Crear Landed Cost ──────────────────────────────────────────────
            st.markdown("#### ⚓ Crear Landed Cost")
            _done_picks = [p for p in (carp_data.get("pickings", []) if carp_data else [])
                           if p.get("state") == "done"]
            _lc_ids     = carp_data.get("lc_ids", []) if carp_data else []

            if not _done_picks:
                st.info("Sin pickings recibidos (estado 'done') para esta carpeta.")
            elif _lc_ids:
                st.success(f"✅ Ya existe {len(_lc_ids)} Landed Cost(s) para esta carpeta.")
                for _lc_id in _lc_ids:
                    st.markdown(f"[\U0001f517 Ver LC en Odoo]({odoo_url('stock.landed.cost', _lc_id)})")
            else:
                with st.form("form_lc"):
                    st.caption("Picking(s): " + ", ".join(p["name"] for p in _done_picks))
                    _lc_lines = []
                    for _lc_pid, _lc_pname in LC_PRODUCTS.items():
                        _lc_c1, _lc_c2 = st.columns([3, 1])
                        _lc_c1.caption(_lc_pname)
                        _lc_amt = _lc_c2.number_input(
                            "Monto ARS", min_value=0.0, value=0.0,
                            key=f"lc_amt_{_lc_pid}", label_visibility="collapsed")
                        if _lc_amt > 0:
                            _lc_lines.append({"product_id": _lc_pid, "price_unit": _lc_amt})
                    _lc_submit = st.form_submit_button("⚓ Crear Landed Cost", type="primary")
                    if _lc_submit:
                        if not _lc_lines:
                            st.warning("Completá al menos un monto antes de crear el LC.")
                        else:
                            try:
                                _new_lc  = create_landed_cost(models, uid, api_key,
                                               [p["id"] for p in _done_picks], _lc_lines)
                                _lc_url2 = odoo_url("stock.landed.cost", _new_lc)
                                st.success(f"✅ Landed Cost ID {_new_lc}")
                                st.markdown(f"[\U0001f517 Ver en Odoo]({_lc_url2})")
                                st.session_state.etapas["4"] = True
                                load_carpeta_full.clear()
                            except Exception as _lce:
                                st.error(f"Error al crear LC: {_lce}")

            st.divider()

            # ── Decálogo CFO ──────────────────────────────────────────────────
            if st.session_state.etapas.get("4"):
                with st.expander("✅ Checklist Decálogo CFO (Etapa 6)", expanded=not st.session_state.etapas.get("6")):
                    for _di, _ditem in enumerate(DECALOGO):
                        st.checkbox(_ditem, key=f"dec_{_di}")

                    st.divider()
                    # Validación de cierre
                    _etapas_requeridas = ["0","1","2","2a","3","4","5"]
                    _et_faltantes = [lbl for k, lbl, _ in ETAPAS_DEF
                                     if k in _etapas_requeridas
                                     and not st.session_state.etapas.get(k)]
                    _dec_faltantes = [DECALOGO[i] for i in range(len(DECALOGO))
                                      if not st.session_state.get(f"dec_{i}", False)]
                    _puede_cerrar = not _et_faltantes and not _dec_faltantes

                    if st.session_state.etapas.get("6"):
                        st.success("🔒 Importación cerrada — Acta CFO firmada.")
                    else:
                        if not _puede_cerrar:
                            if _et_faltantes:
                                st.warning("⚠️ Etapas pendientes: " + ", ".join(_et_faltantes))
                            if _dec_faltantes:
                                st.warning(f"⚠️ {len(_dec_faltantes)} check(s) del Decálogo sin marcar.")
                        _cerrar_disabled = not _puede_cerrar
                        if st.button("🔒 Cerrar importación y firmar Acta CFO",
                                     type="primary", use_container_width=True,
                                     key="btn_cerrar_carpeta",
                                     disabled=_cerrar_disabled):
                            st.session_state.etapas["6"] = True
                            st.toast(f"Importación {st.session_state.carpeta_id} cerrada.", icon="🔒")
                            st.rerun()



    # ─────────────────────────────────────────────────────────────────────────────
    # TAB — ÓRDENES DE PAGO
    # ─────────────────────────────────────────────────────────────────────────────
with tab_op:
    st.subheader("🏦 Órdenes de Pago")

    _op_tipo = st.radio(
        "Tipo de orden",
        ["💳 Pago de factura", "📤 Pago a cuenta", "🧾 Gastos / VEPs"],
        horizontal=True, key="op_tipo",
    )
    st.divider()

    from datetime import date as _date_cls

    _jours = get_payment_journals(models_url, uid, api_key)
    if _jours:
        _jour_opts = {label: jid for jid, label, _ in _jours}
        _jour_cur  = {label: cur for _,   label, cur in _jours}
    else:
        _jour_opts, _jour_cur = {}, {}

    # =========================================================================
    # MODO 1 — PAGO DE FACTURA CONTABILIZADA
    # =========================================================================
    if _op_tipo == "💳 Pago de factura":

        _op_c1, _op_c2, _op_c3, _op_c4 = st.columns([2, 2, 1, 1])
        _op_partner_filter = _op_c1.text_input(
            "Filtrar proveedor", key="op_filt_partner", placeholder="Nombre del proveedor")
        _op_cur_filter    = _op_c2.selectbox(
            "Moneda", ["Todas", "ARS", "USD"], key="op_filt_cur")
        _op_only_vencidas = _op_c3.checkbox("Solo vencidas", key="op_vencidas")
        _op_refresh       = _op_c4.button("🔄 Actualizar", key="op_refresh_btn")

        if _op_refresh:
            get_pending_bills.clear()
            st.session_state.pop("_op_bills_ok", None)

        if not st.session_state.get("_op_bills_ok"):
            st.info("Presioná **Actualizar** para cargar las facturas pendientes.")
            _all_pending = []
        else:
            with st.spinner("Cargando facturas pendientes..."):
                _all_pending = get_pending_bills(models_url, uid, api_key)

        # Marcar como cargado la primera vez que el usuario presiona Actualizar
        if _op_refresh:
            st.session_state["_op_bills_ok"] = True
            with st.spinner("Cargando facturas pendientes..."):
                _all_pending = get_pending_bills(models_url, uid, api_key)

        if not _all_pending:
            st.info("No hay facturas de proveedor pendientes de pago.")

        if _all_pending:
            _today = _date_cls.today().isoformat()
            _filtered = _all_pending
            if _op_partner_filter:
                _pf = _op_partner_filter.lower()
                _filtered = [b for b in _filtered
                             if _pf in (b.get("partner_id") or [0, ""])[1].lower()]
            if _op_cur_filter != "Todas":
                _filtered = [b for b in _filtered
                             if (b.get("currency_id") or [0, "ARS"])[1] == _op_cur_filter]
            if _op_only_vencidas:
                _filtered = [b for b in _filtered
                             if b.get("invoice_date_due") and b["invoice_date_due"] < _today]

            _tot_ars_filt = sum(float(b.get("amount_residual") or 0) for b in _filtered
                                if (b.get("currency_id") or [0,"ARS"])[1] == "ARS")
            _tot_usd_filt = sum(float(b.get("amount_residual") or 0) for b in _filtered
                                if (b.get("currency_id") or [0,"ARS"])[1] == "USD")
            _tot_ars_all  = sum(float(b.get("amount_residual") or 0) for b in _all_pending
                                if (b.get("currency_id") or [0,"ARS"])[1] == "ARS")
            _tot_usd_all  = sum(float(b.get("amount_residual") or 0) for b in _all_pending
                                if (b.get("currency_id") or [0,"ARS"])[1] == "USD")

            _hay_filtro = len(_filtered) != len(_all_pending)
            _res_parts = []
            if _tot_ars_filt > 0: _res_parts.append(f"**ARS:** {fmt_ars(_tot_ars_filt)}")
            if _tot_usd_filt > 0: _res_parts.append(f"**USD:** {fmt_usd(_tot_usd_filt)}")
            _res_parts.append(
                f"**{len(_filtered)}** factura(s)"
                + (f" de {len(_all_pending)} totales" if _hay_filtro else ""))
            st.info("  ·  ".join(_res_parts) if _res_parts else "Sin deuda pendiente.")
            if _hay_filtro and (_tot_ars_all != _tot_ars_filt or _tot_usd_all != _tot_usd_filt):
                _sf = []
                if _tot_ars_all > 0: _sf.append(f"ARS {fmt_ars(_tot_ars_all)}")
                if _tot_usd_all > 0: _sf.append(f"USD {fmt_usd(_tot_usd_all)}")
                st.caption("Sin filtro: " + "  ·  ".join(_sf))

            if not _filtered:
                st.warning("Ninguna factura cumple los filtros aplicados.")

            if _filtered:
                _op_rows = []
                for _b in _filtered:
                    _cur   = (_b.get("currency_id") or [0,"ARS"])[1]
                    _resid = float(_b.get("amount_residual") or 0)
                    _total = float(_b.get("amount_total") or 0)
                    _due   = _b.get("invoice_date_due") or ""
                    _venc_flag = "⚠️" if (_due and _due < _today) else ""
                    _pstate_map = {"not_paid": "Sin pagar", "partial": "Parcial"}
                    _op_rows.append({
                        "Sel": False, "Venc.": _venc_flag,
                        "Proveedor":   (_b.get("partner_id") or [0, "—"])[1],
                        "Comprobante": _b.get("name") or f"ID {_b['id']}",
                        "Ref.":        (_b.get("ref") or "")[:30],
                        "Fecha FA":    _b.get("invoice_date") or None,
                        "Vto. pago":   _due or None,
                        "Moneda": _cur, "Total": _total, "Pendiente": _resid,
                        "Estado pago": _pstate_map.get(_b.get("payment_state",""),
                                                       _b.get("payment_state","")),
                        "_id": _b["id"],
                        "_partner_id":  (_b.get("partner_id") or [0])[0],
                        "_currency_id": (_b.get("currency_id") or [0])[0],
                    })

                _df_op = pd.DataFrame(_op_rows)
                _col_cfg_op = {
                    "Sel":       st.column_config.CheckboxColumn("✓", width="small"),
                    "Venc.":     st.column_config.TextColumn("", width="small"),
                    "Fecha FA":  st.column_config.DateColumn("Fecha FA",  format="DD/MM/YYYY"),
                    "Vto. pago": st.column_config.DateColumn("Vto. pago", format="DD/MM/YYYY"),
                    "Total":     st.column_config.NumberColumn("Total",    format="%.2f"),
                    "Pendiente": st.column_config.NumberColumn("Pendiente",format="%.2f"),
                    "_id": None, "_partner_id": None, "_currency_id": None,
                }
                _display_cols = ["Sel","Venc.","Proveedor","Comprobante","Ref.",
                                 "Fecha FA","Vto. pago","Moneda","Total","Pendiente","Estado pago"]

                st.markdown("**Seleccioná las facturas a pagar:**")
                _edited_op = st.data_editor(
                    _df_op[_display_cols + ["_id","_partner_id","_currency_id"]],
                    column_config=_col_cfg_op, column_order=_display_cols,
                    use_container_width=True, hide_index=True, key="op_data_editor",
                    disabled=[c for c in _display_cols if c != "Sel"],
                )

                _selected_op = _edited_op[_edited_op["Sel"] == True]
                n_sel = len(_selected_op)

                if n_sel > 0:
                    st.divider()
                    _total_ars_sel = _selected_op[_selected_op["Moneda"]=="ARS"]["Pendiente"].sum()
                    _total_usd_sel = _selected_op[_selected_op["Moneda"]=="USD"]["Pendiente"].sum()
                    _rs_parts = [f"{n_sel} factura(s) seleccionada(s)"]
                    if _total_ars_sel > 0: _rs_parts.append(f"ARS {fmt_ars(_total_ars_sel)}")
                    if _total_usd_sel > 0: _rs_parts.append(f"USD {fmt_usd(_total_usd_sel)}")
                    st.info("  ·  ".join(_rs_parts))

                    with st.expander("Ver detalle de seleccionadas", expanded=False):
                        for _, _sr in _selected_op.iterrows():
                            st.caption(
                                f"• **{_sr['Proveedor']}** — {_sr['Comprobante']}"
                                f" — {_sr['Moneda']} {_sr['Pendiente']:,.2f}"
                                + (f" (venc. {_sr['Vto. pago']})" if _sr.get("Vto. pago") else ""))

                    st.markdown("#### Datos del pago")
                    if not _jour_opts:
                        st.error("No se encontraron diarios de pago en Odoo.")
                    else:
                        _fp_c1, _fp_c2 = st.columns(2)
                        _pay_journal    = _fp_c1.selectbox(
                            "Diario de pago", list(_jour_opts.keys()), key="op_journal")
                        _pay_date       = _fp_c2.date_input(
                            "Fecha de pago", value=_date_cls.today(), key="op_pay_date")
                        _pay_journal_id  = _jour_opts[_pay_journal]
                        _pay_journal_cur = _jour_cur[_pay_journal]

                        with st.expander("📒 Asientos que generará cada pago", expanded=True):
                            _ae_rows = []
                            for _, _sr in _selected_op.iterrows():
                                _cur_op   = _sr["Moneda"]
                                _monto_op = _sr["Pendiente"]
                                _fmt_m    = (fmt_usd(_monto_op) if _cur_op == "USD"
                                             else fmt_ars(_monto_op))
                                _ae_rows.append({
                                    "Tipo": "DR", "Cuenta": "Proveedores",
                                    "Descripción": (f"{_sr['Proveedor']}"
                                                    f" · {_sr['Comprobante']}"),
                                    "Moneda": _cur_op, "Monto": _fmt_m,
                                })
                                _ae_rows.append({
                                    "Tipo": "  CR", "Cuenta": _pay_journal,
                                    "Descripción": (f"Pago {_sr['Comprobante']}"
                                                    f" — {_pay_date}"),
                                    "Moneda": _pay_journal_cur, "Monto": _fmt_m,
                                })
                            if _ae_rows:
                                st.dataframe(pd.DataFrame(_ae_rows),
                                             use_container_width=True, hide_index=True)
                                st.caption(
                                    "Estimado. Odoo calcula los importes exactos al confirmar.")

                        st.caption(
                            f"Se generará **una OP por factura** ({n_sel} OP en total).")

                        _op_btn = st.button(
                            f"💸 Generar {n_sel} Orden(es) de Pago en Odoo",
                            type="primary", key="btn_gen_op")

                        if _op_btn:
                            _op_ok, _op_errs = 0, []
                            _pay_date_str = _pay_date.strftime("%Y-%m-%d")
                            _prog_op = st.progress(0)
                            for _op_i, (_, _sr) in enumerate(_selected_op.iterrows()):
                                _mid   = int(_sr["_id"])
                                _cname = _sr["Comprobante"]
                                _pname = _sr["Proveedor"]
                                _curx  = _sr["Moneda"]
                                _mont  = _sr["Pendiente"]
                                try:
                                    _ok, _res = register_payment_wizard(
                                        models, uid, api_key,
                                        [_mid], _pay_date_str, _pay_journal_id)
                                    if _ok:
                                        st.toast(
                                            f"OP generada — {_pname} · {_cname} · {_curx} {_mont:,.2f}", icon="✅")
                                        _op_ok += 1
                                    else:
                                        _op_errs.append(f"❌ {_cname} ({_pname}): {_res}")
                                except Exception as _ope:
                                    _op_errs.append(
                                        f"❌ {_cname} ({_pname}): {str(_ope)[:150]}")
                                _prog_op.progress((_op_i + 1) / n_sel)
                            for _oe in _op_errs:
                                st.error(_oe)
                            if _op_ok:
                                st.toast(f"{_op_ok} OP(s) generada(s) correctamente.", icon="✅")
                                get_pending_bills.clear()
                                st.info(
                                    "Presioná 🔄 Actualizar para ver el nuevo estado.")
                else:
                    st.info(
                        "Marcá el ✓ en la columna izquierda para seleccionar facturas.")

            # =========================================================================
            # MODO 2 — PAGO A CUENTA (sin factura previa)
            # =========================================================================
    elif _op_tipo == "📤 Pago a cuenta":
        st.info(
            "📤 **Pago a cuenta:** el pago se registra sin vincularlo a ninguna "
            "factura existente. Queda como crédito en la cuenta del proveedor y "
            "se imputa a una factura futura desde Odoo (Contabilidad → Proveedores "
            "→ Pagos)."
        )
        with st.form("form_pago_a_cuenta"):
            _pac_c1, _pac_c2 = st.columns(2)
            _pac_prov_name = _pac_c1.text_input(
                "Proveedor (nombre)", placeholder="Nombre o razón social")
            _pac_prov_cuit = _pac_c2.text_input(
                "CUIT (opcional)", placeholder="Sin guiones")
            _pac_c3, _pac_c4, _pac_c5 = st.columns(3)
            _pac_amount   = _pac_c3.number_input(
                "Monto", min_value=0.01, value=1.0, step=0.01)
            _pac_currency = _pac_c4.selectbox(
                "Moneda", ["ARS", "USD"], key="pac_cur")
            _pac_date     = _pac_c5.date_input(
                "Fecha de pago", value=_date_cls.today(), key="pac_date")
            _pac_c6, _pac_c7 = st.columns(2)
            _pac_journal = (
                _pac_c6.selectbox(
                    "Diario de pago", list(_jour_opts.keys()), key="pac_journal")
                if _jour_opts else None
            )
            _pac_memo = _pac_c7.text_input(
                "Referencia / Concepto",
                placeholder="Ej: Anticipo FA-0001-00000123")

            if _pac_journal and _pac_amount > 0:
                _pac_j_cur = _jour_cur.get(_pac_journal, "ARS")
                _pac_fmt_m = (fmt_usd(_pac_amount) if _pac_currency == "USD"
                              else fmt_ars(_pac_amount))
                st.markdown("**Asiento estimado:**")
                st.dataframe(pd.DataFrame([
                    {"Tipo": "DR",
                     "Cuenta": "Proveedores",
                     "Descripción": _pac_prov_name or "(proveedor)",
                     "Monto": _pac_fmt_m},
                    {"Tipo": "  CR",
                     "Cuenta": _pac_journal,
                     "Descripción": _pac_memo or "Pago a cuenta",
                     "Monto": _pac_fmt_m},
                ]), use_container_width=True, hide_index=True)

            _pac_submit = st.form_submit_button(
                "📤 Registrar Pago a Cuenta", type="primary")

            if _pac_submit:
                if not _pac_prov_name and not _pac_prov_cuit:
                    st.warning("Ingresá el nombre o CUIT del proveedor.")
                elif not _pac_journal:
                    st.warning("Seleccioná un diario de pago.")
                else:
                    _pac_pid = None
                    if _pac_prov_cuit:
                        _found = search_partner_by_cuit(
                            models_url, uid, api_key, _pac_prov_cuit)
                        if _found:
                            _pac_pid = _found[0]
                    if not _pac_pid and _pac_prov_name:
                        _res2 = search_partners(
                            models_url, uid, api_key, _pac_prov_name, limit=1)
                        if _res2:
                            _pac_pid = _res2[0][0]
                    if not _pac_pid:
                        st.error(
                            "No se encontró el proveedor en Odoo. "
                            "Verificá el nombre o CUIT.")
                    else:
                        _pac_cur_id = get_currency_id(
                            models_url, uid, api_key, _pac_currency)
                        if not _pac_cur_id:
                            st.error("No se encontró la moneda en Odoo.")
                        else:
                            try:
                                _pac_pay_id = create_advance_payment(
                                    models, uid, api_key,
                                    partner_id   = _pac_pid,
                                    amount       = _pac_amount,
                                    currency_id  = _pac_cur_id,
                                    payment_date = _pac_date.strftime("%Y-%m-%d"),
                                    journal_id   = _jour_opts[_pac_journal],
                                    memo         = _pac_memo or "",
                                )
                                _pac_url = (
                                    f"{ODOO_URL}/odoo/accounting"
                                    f"/payments/{_pac_pay_id}"
                                )
                                st.toast(f"Pago a cuenta registrado — ID {_pac_pay_id}", icon="✅")
                                st.markdown(f"[🔗 Ver en Odoo]({_pac_url})")
                            except Exception as _pace:
                                st.error(
                                    f"Error al registrar el pago: {_pace}")

    # =========================================================================
    # MODO 3 — GASTOS / VEPs
    # =========================================================================
    else:
        _gasto_refresh = st.button("🔄 Actualizar", key="gasto_refresh_btn")
        if _gasto_refresh:
            get_pending_expense_sheets.clear()
            st.session_state.pop("_op_sheets_ok", None)

        if not st.session_state.get("_op_sheets_ok"):
            st.info("Presioná **Actualizar** para cargar las notas de gastos pendientes.")
            _sheets = []
        else:
            with st.spinner("Cargando notas de gastos pendientes..."):
                _sheets = get_pending_expense_sheets(models_url, uid, api_key)

        if _gasto_refresh:
            st.session_state["_op_sheets_ok"] = True
            with st.spinner("Cargando notas de gastos pendientes..."):
                _sheets = get_pending_expense_sheets(models_url, uid, api_key)

        if not _sheets:
            st.info(
                "No hay notas de gastos aprobadas pendientes de pago en Odoo.  \n"
                "Si esperabas ver algo acá, verificá que el estado de la nota "
                "sea **'Publicado'** (aprobada y con asiento contable generado)."
            )
        else:
            st.info(
                f"**{len(_sheets)}** nota(s) de gastos aprobada(s) pendiente(s) de pago.")

            # Tabla de notas de gastos
            _sheet_rows = []
            for _s in _sheets:
                _cur  = (_s.get("currency_id") or [0, "ARS"])[1]
                _monto = float(_s.get("total_amount") or 0)
                _emp  = (_s.get("employee_id") or [0, "—"])[1]
                _pmap = {"not_paid": "Sin pagar", "partial": "Parcial", "paid": "Pagado"}
                _sheet_rows.append({
                    "Sel":        False,
                    "Empleado":   _emp,
                    "Nota":       _s.get("name") or f"ID {_s['id']}",
                    "Moneda":     _cur,
                    "Total":      _monto,
                    "Pago":       _pmap.get(_s.get("payment_state",""),
                                            _s.get("payment_state","")),
                    "_id":        _s["id"],
                })

            _df_sheets = pd.DataFrame(_sheet_rows)
            _col_cfg_sheets = {
                "Sel":   st.column_config.CheckboxColumn("✓", width="small"),
                "Total": st.column_config.NumberColumn("Total", format="%.2f"),
                "_id":   None,
            }
            _disp_sheets = ["Sel", "Empleado", "Nota", "Moneda", "Total", "Pago"]

            st.markdown("**Seleccioná las notas a pagar:**")
            _edited_sheets = st.data_editor(
                _df_sheets[_disp_sheets + ["_id"]],
                column_config=_col_cfg_sheets,
                column_order=_disp_sheets,
                use_container_width=True, hide_index=True,
                key="gasto_data_editor",
                disabled=[c for c in _disp_sheets if c != "Sel"],
            )

            _sel_sheets = _edited_sheets[_edited_sheets["Sel"] == True]
            n_sel_s = len(_sel_sheets)

            if n_sel_s > 0:
                st.divider()
                _tot_ars_s = _sel_sheets[_sel_sheets["Moneda"]=="ARS"]["Total"].sum()
                _tot_usd_s = _sel_sheets[_sel_sheets["Moneda"]=="USD"]["Total"].sum()
                _gs_parts  = [f"{n_sel_s} nota(s) seleccionada(s)"]
                if _tot_ars_s > 0: _gs_parts.append(f"ARS {fmt_ars(_tot_ars_s)}")
                if _tot_usd_s > 0: _gs_parts.append(f"USD {fmt_usd(_tot_usd_s)}")
                st.info("  ·  ".join(_gs_parts))

                st.markdown("#### Datos del pago")
                if not _jour_opts:
                    st.error("No se encontraron diarios de pago en Odoo.")
                else:
                    _gs_c1, _gs_c2 = st.columns(2)
                    _gs_journal = _gs_c1.selectbox(
                        "Diario de pago", list(_jour_opts.keys()), key="gs_journal")
                    _gs_date    = _gs_c2.date_input(
                        "Fecha de pago", value=_date_cls.today(), key="gs_date")
                    _gs_jour_id  = _jour_opts[_gs_journal]
                    _gs_jour_cur = _jour_cur[_gs_journal]

                    with st.expander("📒 Asientos que generará cada pago", expanded=True):
                        _gs_ae = []
                        for _, _sr in _sel_sheets.iterrows():
                            _cur_g  = _sr["Moneda"]
                            _mont_g = _sr["Total"]
                            _fmt_g  = (fmt_usd(_mont_g) if _cur_g == "USD"
                                       else fmt_ars(_mont_g))
                            _gs_ae.append({
                                "Tipo": "DR",
                                "Cuenta": "Gastos / Empleado",
                                "Descripción": f"{_sr['Empleado']} · {_sr['Nota']}",
                                "Moneda": _cur_g, "Monto": _fmt_g,
                            })
                            _gs_ae.append({
                                "Tipo": "  CR",
                                "Cuenta": _gs_journal,
                                "Descripción": f"Pago {_sr['Nota']} — {_gs_date}",
                                "Moneda": _gs_jour_cur, "Monto": _fmt_g,
                            })
                        if _gs_ae:
                            st.dataframe(pd.DataFrame(_gs_ae),
                                         use_container_width=True, hide_index=True)
                            st.caption(
                                "Estimado. Odoo calcula los importes exactos al confirmar.")

                    st.caption(
                        f"Se generará **una OP por nota de gastos** "
                        f"({n_sel_s} OP en total).")

                    _gs_btn = st.button(
                        f"💸 Registrar {n_sel_s} Pago(s) de Gastos en Odoo",
                        type="primary", key="btn_gen_gs")

                    if _gs_btn:
                        _gs_ok, _gs_errs = 0, []
                        _gs_date_str = _gs_date.strftime("%Y-%m-%d")
                        _prog_gs = st.progress(0)
                        for _gi, (_, _sr) in enumerate(_sel_sheets.iterrows()):
                            _sid   = int(_sr["_id"])
                            _snota = _sr["Nota"]
                            _semp  = _sr["Empleado"]
                            try:
                                _ok, _res = register_expense_payment(
                                    models, uid, api_key,
                                    _sid, _gs_date_str, _gs_jour_id)
                                if _ok:
                                    st.toast(f"Pago registrado — {_semp} · {_snota}", icon="✅")
                                    _gs_ok += 1
                                else:
                                    _gs_errs.append(f"❌ {_snota} ({_semp}): {_res}")
                            except Exception as _gse:
                                _gs_errs.append(
                                    f"❌ {_snota} ({_semp}): {str(_gse)[:150]}")
                            _prog_gs.progress((_gi + 1) / n_sel_s)
                        for _ge in _gs_errs:
                            st.error(_ge)
                        if _gs_ok:
                            st.toast(f"{_gs_ok} pago(s) registrado(s).", icon="✅")
                            get_pending_expense_sheets.clear()
                            st.info(
                                "Presioná 🔄 Actualizar para ver el nuevo estado.")
            else:
                st.info(
                    "Marcá el ✓ en la columna izquierda para seleccionar "
                    "notas de gastos.")

        st.divider()
        st.caption(
            f"Para cargar nuevos gastos o aprobar notas pendientes, "
            f"usá el [módulo de Gastos en Odoo]"
            f"({ODOO_URL}/odoo/expenses) directamente.")




# ─────────────────────────────────────────────────────────────────────────────
# TAB — RECIBOS DE COBRO
# ─────────────────────────────────────────────────────────────────────────────
with tab_recibos:
    from datetime import date as _rc_date_cls

    st.subheader("💰 Recibos de Cobro")

    # ── helpers locales ─────────────────────────────────────────────────────
    def _rc_parse_monto(val):
        """Convierte importe de home banking a float.
        Soporta formato AR (1.234,56), US/Excel (1234.56) y sin centavos (1234)."""
        s = str(val).replace("$", "").replace("\xa0", "").replace(" ", "").strip()
        if not s or s.lower() in ("nan", "none", ""):
            return 0.0
        try:
            return float(normalize_amount(s))
        except Exception:
            return 0.0

    def _rc_parse_retencion(file_bytes):
        """Parsea PDF de retención de IIBB y retorna dict con los datos clave."""
        try:
            import pdfplumber
            text = ""
            with pdfplumber.open(BytesIO(file_bytes)) as pdf:
                text = (pdf.pages[0].extract_text() or "") if pdf.pages else ""
            if not text:
                return None
            result = {}
            # Nombre emisor: primera línea significativa
            lines = [l.strip() for l in text.splitlines() if l.strip()]
            result["nombre"] = lines[0] if lines else ""
            # CUIT emisor: primer CUIT en el texto (antes de "Razón social")
            _m = re.search(r"C\.U\.I\.T\.:\s*([\d\-]+)", text)
            if _m:
                result["cuit"] = _m.group(1).replace("-", "").replace(" ", "").strip()
            # Fecha
            _m = re.search(r"Fecha:\s*(\d{2}/\d{2}/\d{4})", text)
            if _m:
                _p = _m.group(1).split("/")
                result["fecha"] = _m.group(1)
                result["fecha_iso"] = f"{_p[2]}-{_p[1]}-{_p[0]}"
            # Nro certificado
            _m = re.search(r"Nro\.\s*de\s*certificado:\s*(.+)", text)
            if _m: result["nro_certificado"] = _m.group(1).strip()
            # Nro pago
            _m = re.search(r"Pago:\s*(.+)", text)
            if _m: result["nro_pago"] = _m.group(1).strip()
            # Concepto
            _m = re.search(r"Concepto del pago:\s*(.+)", text)
            if _m: result["concepto"] = _m.group(1).strip()
            # Importe retenido
            _m = re.search(r"Importe retenido:\s*([\d.,]+)", text)
            if _m: result["importe"] = _rc_parse_monto(_m.group(1))
            # Importe sujeto
            _m = re.search(r"Importe pagado sujeto a retenci[oó]n:\s*([\d.,]+)", text)
            if _m: result["importe_sujeto"] = _rc_parse_monto(_m.group(1))
            return result if result.get("cuit") and result.get("importe") else None
        except Exception as _e:
            return None

    def _rc_parse_cheques(file_bytes, filename):
        fname = filename.lower()
        if fname.endswith(".csv"):
            try:
                text = file_bytes.decode("latin-1")
            except Exception:
                text = file_bytes.decode("utf-8", errors="replace")
            lines = [l for l in text.splitlines() if l.strip()]
            if len(lines) < 3:
                return []
            rows = []
            for line in lines[2:]:
                parts = [p.strip() for p in line.split(";")]
                while len(parts) < 40:
                    parts.append("")
                estado = parts[7].lower()
                # Col D (idx 3) = CUIT del que nos lo entrego (directo)
                # Col N (idx 13) = CUIT del beneficiario original (endosado)
                if "endoso" in estado:
                    cuit   = parts[13]
                    nombre = parts[12]
                else:
                    cuit   = parts[3]
                    nombre = parts[2]
                cuit = str(cuit).replace("-", "").replace(" ", "").strip()
                if not cuit or len(cuit) < 8:
                    continue
                rows.append({
                    "nro":       parts[0],
                    "nombre":    nombre.strip(),
                    "cuit":      cuit,
                    "fecha_pago": parts[4],
                    "importe":   _rc_parse_monto(parts[6]),
                    "estado":    parts[7],
                    "banco":     parts[8],
                })
            return rows
        elif fname.endswith((".xlsx", ".xls")):
            try:
                _dfx = pd.read_excel(BytesIO(file_bytes), header=1, dtype=str).fillna("")
            except Exception as _xe:
                st.error(f"No se pudo leer el Excel: {_xe}")
                return []
            rows = []
            for _, _xr in _dfx.iterrows():
                parts = list(_xr.values)
                while len(parts) < 40:
                    parts.append("")
                estado = str(parts[7]).lower()
                if "endoso" in estado:
                    cuit   = str(parts[13])
                    nombre = str(parts[12])
                else:
                    cuit   = str(parts[3])
                    nombre = str(parts[2])
                cuit = cuit.replace("-", "").replace(" ", "").strip()
                if not cuit or len(cuit) < 8:
                    continue
                # Normalizar fecha: el xlsx trae '2026-06-19 00:00:12' → tomar solo YYYY-MM-DD
                _fp_raw = str(parts[4]).strip()
                _fp = _fp_raw[:10] if len(_fp_raw) >= 10 else _fp_raw
                rows.append({
                    "nro":       str(parts[0]),
                    "nombre":    nombre.strip(),
                    "cuit":      cuit,
                    "fecha_pago": _fp,
                    "importe":   _rc_parse_monto(str(parts[6])),
                    "estado":    str(parts[7]),
                    "banco":     str(parts[8]).strip(),
                })
            return rows
        else:
            st.error("Formato no soportado. Subí un archivo CSV o XLSX del home banking.")
            return []

    # ── Journal fijo: Cheques a depositar (id=73) ─────────────────────────────
    _RC_JOURNAL_ID = 73
    _rc_all_banks  = get_all_banks(models_url, uid, api_key)

    # ── Uploaders ────────────────────────────────────────────────────────────
    _rcu_col1, _rcu_col2 = st.columns([3, 2])
    _rc_file = _rcu_col1.file_uploader(
        "📄 Cheques (CSV o XLSX del home banking)",
        type=["csv", "xlsx", "xls"], key="rc_file_uploader",
        help="Descargalo desde tu home banking y subilo sin modificar.")
    _rc_ret_files = _rcu_col2.file_uploader(
        "🔖 Retenciones (PDF)",
        type=["pdf"], key="rc_ret_uploader",
        accept_multiple_files=True,
        help="Subí uno o más comprobantes de retención de IIBB o Ganancias.")

    # ── Parsear retenciones subidas ────────────────────────────────────────
    # Guardamos en session_state para que no se pierdan al reruns
    if "rc_retenciones" not in st.session_state:
        st.session_state["rc_retenciones"] = {}   # {cuit: [ret_dict, ...]}
    if _rc_ret_files:
        for _rf in _rc_ret_files:
            _rb = _rf.read()
            _ret = _rc_parse_retencion(_rb)
            if _ret and _ret.get("cuit"):
                _rcuit_ret = _ret["cuit"]
                _ret["_filename"] = _rf.name
                # Evitar duplicados por nombre de archivo
                existing = st.session_state["rc_retenciones"].get(_rcuit_ret, [])
                if not any(r.get("_filename") == _rf.name for r in existing):
                    existing.append(_ret)
                    st.session_state["rc_retenciones"][_rcuit_ret] = existing
            else:
                st.warning(f"No se pudo parsear la retención en {_rf.name}.")

    # Mostrar retenciones cargadas
    if st.session_state["rc_retenciones"]:
        _all_rets = [r for rl in st.session_state["rc_retenciones"].values() for r in rl]
        st.caption(
            f"✅ {len(_all_rets)} retención(es) cargada(s): "
            + ", ".join(f"{r['nombre']} — ARS {fmt_ars(r['importe'])}" for r in _all_rets)
            + "  ·  [Limpiar]" if _all_rets else "")
        if st.button("🗑️ Limpiar retenciones cargadas", key="rc_clear_rets"):
            st.session_state["rc_retenciones"] = {}
            st.rerun()

    if _rc_file is None:
        st.caption(
            "Subí el archivo del home banking para procesar los cheques recibidos "
            "y generar los recibos de cobro en Odoo.")
    else:
        _rc_bytes   = _rc_file.read()
        _rc_cheques = _rc_parse_cheques(_rc_bytes, _rc_file.name)

        if not _rc_cheques:
            st.error(
                "No se encontraron cheques en el archivo. "
                "Verificá que sea el export correcto del home banking.")
        else:
            # Group by CUIT, preserving order of first appearance
            _rc_groups = {}
            for _ch in _rc_cheques:
                _rc_groups.setdefault(_ch["cuit"], []).append(_ch)

            _rc_total_ars = sum(c["importe"] for c in _rc_cheques)
            st.info(
                f"**{len(_rc_cheques)} cheque(s)** de **{len(_rc_groups)} cliente(s)** "
                f"— Total ARS {fmt_ars(_rc_total_ars)}")

            # ── Cargar datos de Odoo para todos los CUITs a la vez ──────────
            _rc_all_cuits = tuple(sorted(_rc_groups.keys()))
            with st.spinner("Buscando clientes en Odoo..."):
                _rc_partner_map = search_partners_by_cuits(
                    models_url, uid, api_key, _rc_all_cuits)

            _rc_pids_found = tuple(pid for pid, _ in _rc_partner_map.values())
            if _rc_pids_found:
                with st.spinner("Cargando facturas pendientes de cobro..."):
                    _rc_all_inv = get_customer_unpaid_invoices(
                        models_url, uid, api_key, _rc_pids_found)
            else:
                _rc_all_inv = []

            # Construir mapa inverso: id_hijo → id_padre (para agrupar facturas de contactos)
            # Una factura puede estar a nombre de un contacto (hijo) del socio principal
            _rc_child_to_parent = {}
            if _rc_pids_found:
                try:
                    _all_contacts = models.execute_kw(
                        ODOO_DB, uid, api_key, "res.partner", "search_read",
                        [[("parent_id", "in", list(_rc_pids_found))]],
                        {"fields": ["id", "parent_id"], "limit": 500})
                    for _ct in _all_contacts:
                        _parent = (_ct.get("parent_id") or [0])[0]
                        if _parent:
                            _rc_child_to_parent[_ct["id"]] = _parent
                except Exception:
                    pass

            # Index invoices by partner_id (normalizando hijos al padre)
            _rc_inv_by_pid = {}
            for _rci in _rc_all_inv:
                _rpid = (_rci.get("partner_id") or [0])[0]
                # Si es un contacto hijo, agrupar bajo el padre
                _rpid_norm = _rc_child_to_parent.get(_rpid, _rpid)
                _rc_inv_by_pid.setdefault(_rpid_norm, []).append(_rci)

            if st.button("🔄 Actualizar desde Odoo", key="rc_refresh"):
                search_partners_by_cuits.clear()
                get_customer_unpaid_invoices.clear()
                st.rerun()

            st.divider()

            # ── Un expander por cliente ──────────────────────────────────────
            for _rcuit, _rchs in _rc_groups.items():
                _rcp_data  = _rc_partner_map.get(_rcuit)
                _rctotal   = sum(c["importe"] for c in _rchs)
                _rc_tag    = "" if _rcp_data else "  ⚠️ no encontrado en Odoo"
                _rc_exp_lbl = (
                    f"👤 {_rchs[0]['nombre']} — CUIT {_rcuit} — "
                    f"{len(_rchs)} cheque(s) · ARS {fmt_ars(_rctotal)}{_rc_tag}")

                with st.expander(_rc_exp_lbl, expanded=(len(_rc_groups) == 1)):

                    # Tabla de cheques
                    _rch_rows = [
                        {"Nro": c["nro"], "Banco": c["banco"],
                         "Fecha cobro": c["fecha_pago"],
                         "Estado": c["estado"],
                         "Importe ARS": fmt_ars(c["importe"])}
                        for c in _rchs
                    ]
                    st.dataframe(
                        pd.DataFrame(_rch_rows),
                        use_container_width=True, hide_index=True)

                    if not _rcp_data:
                        st.warning(
                            f"CUIT **{_rcuit}** no encontrado en Odoo. "
                            "Verificá que el cliente esté registrado con ese CUIT en el campo VAT.")
                        continue

                    _rc_pid_cuit, _rc_pname_cuit = _rcp_data

                    # ── Detectar mismatch de nombre ───────────────────────────
                    _rc_nombre_excel = _rchs[0].get("nombre", "").strip().upper()
                    _rc_nombre_odoo  = _rc_pname_cuit.strip().upper()
                    # Limpiar puntuación de las keywords (ej: "LAURET," → "LAURET")
                    _rc_kws = [
                        re.sub(r"[^\w]", "", w)
                        for w in _rc_nombre_excel.split()
                        if len(re.sub(r"[^\w]", "", w)) >= 5
                    ]
                    _rc_match_ok = not _rc_kws or any(
                        kw in _rc_nombre_odoo for kw in _rc_kws)

                    if not _rc_match_ok:
                        # Hay mismatch: dejar elegir entre el partner del CUIT y buscar por nombre
                        st.warning(
                            f"⚠️ El CUIT **{_rcuit}** está asignado en Odoo a "
                            f"**{_rc_pname_cuit}** (ID {_rc_pid_cuit}), "
                            f"pero el cheque fue emitido por **{_rchs[0].get('nombre','')}**.")

                        _rc_sel_key  = f"rc_who_{_rcuit}"
                        _rc_srch_key = f"rc_srch_{_rcuit}"
                        _rc_choice = st.radio(
                            "¿A qué cliente asignar este cobro?",
                            options=["cuit", "nombre"],
                            format_func=lambda x: (
                                f"Usar {_rc_pname_cuit} (CUIT coincide en Odoo)"
                                if x == "cuit"
                                else f"Buscar {_rchs[0].get('nombre','')} por nombre"
                            ),
                            key=_rc_sel_key, horizontal=True)

                        if _rc_choice == "cuit":
                            _rc_pid   = _rc_pid_cuit
                            _rc_pname = _rc_pname_cuit
                        else:
                            # Buscar por nombre en Odoo
                            _rc_name_q = st.text_input(
                                "Nombre a buscar en Odoo",
                                value=_rchs[0].get("nombre", ""),
                                key=_rc_srch_key)
                            _rc_name_results = []
                            if _rc_name_q and len(_rc_name_q) >= 3:
                                try:
                                    _rc_name_results = models.execute_kw(
                                        ODOO_DB, uid, api_key, "res.partner", "search_read",
                                        [[("name", "ilike", _rc_name_q),
                                          ("customer_rank", ">", 0),
                                          ("active", "=", True)]],
                                        {"fields": ["id", "name", "vat"], "limit": 10})
                                except Exception:
                                    pass
                            if not _rc_name_results:
                                st.info("Ingresá al menos 3 caracteres para buscar.")
                                continue
                            _rc_name_opts = {
                                f"{r['name']} (CUIT {r.get('vat','?')})": (r["id"], r["name"])
                                for r in _rc_name_results}
                            _rc_name_sel = st.selectbox(
                                "Seleccioná el cliente correcto",
                                list(_rc_name_opts.keys()),
                                key=f"rc_namesel_{_rcuit}")
                            _rc_pid, _rc_pname = _rc_name_opts[_rc_name_sel]
                    else:
                        _rc_pid   = _rc_pid_cuit
                        _rc_pname = _rc_pname_cuit
                        st.markdown(f"**Cliente Odoo:** {_rc_pname} (ID {_rc_pid})")

                    _rc_invs = _rc_inv_by_pid.get(_rc_pid, [])

                    # ── Selector de facturas ─────────────────────────────────
                    _rcsel_ids  = []
                    _rcsel_saldo = 0.0
                    if not _rc_invs:
                        st.info(
                            "No hay facturas pendientes para este cliente. "
                            "El cobro se registrará como pago a cuenta.")
                    else:
                        st.markdown("**Seleccioná las facturas a cobrar:**")
                        _rci_rows = []
                        for _rci in _rc_invs:
                            _rcic  = (_rci.get("currency_id") or [0, "ARS"])[1]
                            _rcres = float(_rci.get("amount_residual") or 0)
                            _rcto  = float(_rci.get("amount_total") or 0)
                            _vence_raw = str(_rci.get("invoice_date_due") or "")
                            _today_str = str(_rc_date_cls.today())
                            _vence_disp = (f"⚠️ {_vence_raw}" if _vence_raw and _vence_raw < _today_str else _vence_raw)
                            _rci_rows.append({
                                "Sel":        False,
                                "Factura":    _rci.get("name") or f"ID {_rci['id']}",
                                "Fecha":      str(_rci.get("invoice_date") or ""),
                                "Vence":      _vence_disp,
                                "Moneda":     _rcic,
                                "Total":      fmt_ars(_rcto)  if _rcic == "ARS" else f"{_rcic} {_rcto:,.2f}",
                                "Saldo":      fmt_ars(_rcres) if _rcic == "ARS" else f"{_rcic} {_rcres:,.2f}",
                                "_saldo_num": _rcres,   # columna numérica oculta para cálculos
                                "_id":        _rci["id"],
                            })
                        _rci_df  = pd.DataFrame(_rci_rows)
                        _rci_cfg = {
                            "Sel":        st.column_config.CheckboxColumn("✓", width="small"),
                            "Total":      st.column_config.TextColumn("Total"),
                            "Saldo":      st.column_config.TextColumn("Saldo"),
                            "_saldo_num": None,   # oculta
                            "_id":        None,
                        }
                        _rci_disp = ["Sel", "Factura", "Fecha", "Vence",
                                     "Moneda", "Total", "Saldo"]
                        _rci_edited = st.data_editor(
                            _rci_df[_rci_disp + ["_saldo_num", "_id"]],
                            column_config=_rci_cfg,
                            column_order=_rci_disp,
                            use_container_width=True, hide_index=True,
                            key=f"rc_inv_{_rcuit}",
                            disabled=[c for c in _rci_disp if c != "Sel"],
                        )
                        _rcsel       = _rci_edited[_rci_edited["Sel"] == True]
                        _rcsel_ids   = [int(r) for r in _rcsel["_id"].tolist()]
                        _rcsel_saldo = float(_rcsel["_saldo_num"].sum())

                    # ── Formulario de pago ────────────────────────────────────
                    st.markdown("#### Datos del recibo")
                    _rc_jour_id = _RC_JOURNAL_ID
                    _rcc2, _rcc3 = st.columns([1, 1])
                    _rc_date = _rcc2.date_input(
                        "Fecha de cobro",
                        value=_rc_date_cls.today(),
                        key=f"rc_date_{_rcuit}")
                    _rc_amount = _rcc3.number_input(
                        "Importe cobrado (ARS)",
                        min_value=0.0,
                        value=float(_rctotal),
                        step=0.01, format="%.2f",
                        key=f"rc_amt_{_rcuit}",
                        help="Pre-completado con el total de cheques. "
                             "Ajustá si hay retenciones o NC.")

                    # ── Deducciones dinámicas (Retenciones / NC) ────────────
                    _ded_key     = f"rc_deds_{_rcuit}"
                    _ded_cnt_key = f"rc_deds_cnt_{_rcuit}"
                    if _ded_key     not in st.session_state: st.session_state[_ded_key]     = []
                    if _ded_cnt_key not in st.session_state: st.session_state[_ded_cnt_key] = 0

                    # ── Auto-agregar retenciones del mismo cliente ─────────
                    _cuit_rets = st.session_state.get("rc_retenciones", {}).get(_rcuit, [])
                    for _cret in _cuit_rets:
                        _cret_fname = _cret.get("_filename", "")
                        _already = any(
                            d.get("_ret_filename") == _cret_fname
                            for d in st.session_state[_ded_key])
                        if not _already and _cret.get("importe", 0) > 0:
                            _new_uid = st.session_state[_ded_cnt_key] + 1
                            st.session_state[_ded_cnt_key] = _new_uid
                            _cret_concepto_pdf = _cret.get("concepto", "")
                            # Construir lista de cuentas igual que el widget para calcular índice
                            _rc_accts_pre = get_all_accounts(models_url, uid, api_key)
                            _rc_accts_pre_s = sorted(
                                _rc_accts_pre,
                                key=lambda x: (1 if "(copia)" in x[1].lower() else 0, x[1]))
                            _rc_acct_opts_pre = ["— Seleccionar concepto —"] + [lbl for _, lbl in _rc_accts_pre_s]
                            # Buscar cuenta por "ingresos brutos" / "iibb" / "retenc"
                            _ret_acct_id  = None
                            _ret_acct_lbl = ""
                            _ret_acct_idx = 0
                            _kws_ret = ["ingresos brutos", "iibb", "retenc"]
                            # Primero intentar matchear con el concepto del PDF
                            _concepto_low = _cret_concepto_pdf.lower()
                            for _aid, _albl in _rc_accts_pre_s:
                                _albl_low = _albl.lower()
                                if "(copia)" in _albl_low:
                                    continue
                                # Match exacto con palabras del concepto PDF
                                if any(kw in _concepto_low and kw in _albl_low for kw in _kws_ret):
                                    _ret_acct_id  = _aid
                                    _ret_acct_lbl = _albl
                                    break
                            # Fallback: cualquier cuenta de retención
                            if not _ret_acct_id:
                                for _aid, _albl in _rc_accts_pre_s:
                                    _albl_low = _albl.lower()
                                    if any(kw in _albl_low for kw in _kws_ret) and "(copia)" not in _albl_low:
                                        _ret_acct_id  = _aid
                                        _ret_acct_lbl = _albl
                                        break
                            if _ret_acct_lbl and _ret_acct_lbl in _rc_acct_opts_pre:
                                _ret_acct_idx = _rc_acct_opts_pre.index(_ret_acct_lbl)
                            st.session_state[_ded_key].append({
                                "uid":           _new_uid,
                                "monto":         float(_cret["importe"]),
                                "concepto_idx":  _ret_acct_idx,
                                "concepto":      _ret_acct_lbl,
                                "account_id":    _ret_acct_id,
                                "_ret_filename": _cret_fname,
                                "_ret_auto":     True,
                            })
                            st.toast(
                                f"Retención de {_cret.get('nombre','')} "
                                f"(ARS {fmt_ars(_cret['importe'])}) agregada automáticamente.",
                                icon="🔖")

                    # Cargar cuentas para conceptos (reutiliza cache de facturas)
                    _rc_accts_raw = get_all_accounts(models_url, uid, api_key)
                    # Poner cuentas "(copia)" al final para no confundir con las originales
                    _rc_accts = sorted(
                        _rc_accts_raw,
                        key=lambda x: (1 if "(copia)" in x[1].lower() else 0, x[1]))
                    _rc_acct_opts = ["— Seleccionar concepto —"] + [lbl for _, lbl in _rc_accts]

                    _deds = st.session_state[_ded_key]
                    if _deds:
                        st.markdown("**Deducciones / Retenciones / NC:**")
                    _to_remove = None
                    for _ded in _deds:
                        _uid = _ded["uid"]
                        _dc1, _dc2, _dc3 = st.columns([5, 2, 1])
                        _cpt_key = f"rc_ded_cpt_{_rcuit}_{_uid}"
                        _mnt_key = f"rc_ded_mnt_{_rcuit}_{_uid}"
                        _cur_idx = _ded.get("concepto_idx", 0)
                        _new_cpt = _dc1.selectbox(
                            "", options=_rc_acct_opts, index=_cur_idx,
                            key=_cpt_key, label_visibility="collapsed")
                        _new_mnt = _dc2.number_input(
                            "", value=float(_ded.get("monto", 0.0)),
                            min_value=0.0, step=0.01, format="%.2f",
                            key=_mnt_key, label_visibility="collapsed")
                        if _dc3.button("✕", key=f"rc_rm_{_rcuit}_{_uid}"):
                            _to_remove = _uid
                        _cpt_idx = _rc_acct_opts.index(_new_cpt) if _new_cpt in _rc_acct_opts else 0
                        _acct_id = next((aid for aid, albl in _rc_accts if albl == _new_cpt), None)
                        _ded.update({"concepto": _new_cpt, "concepto_idx": _cpt_idx,
                                     "account_id": _acct_id, "monto": _new_mnt})

                    if _to_remove is not None:
                        st.session_state[_ded_key] = [d for d in _deds if d["uid"] != _to_remove]
                        st.rerun()

                    if st.button("➕ Agregar retención / NC", key=f"rc_add_ded_{_rcuit}"):
                        _new_uid = st.session_state[_ded_cnt_key] + 1
                        st.session_state[_ded_cnt_key] = _new_uid
                        st.session_state[_ded_key].append(
                            {"uid": _new_uid, "monto": 0.0, "concepto_idx": 0,
                             "concepto": "", "account_id": None})
                        st.rerun()

                    _rc_ajuste_total = sum(
                        d.get("monto", 0.0) for d in st.session_state[_ded_key])

                    _rc_memo = st.text_input(
                        "Referencia / Memo",
                        value=f"Recibo cheques — {_rchs[0]['nombre']}",
                        key=f"rc_memo_{_rcuit}")

                    _rc_neto = _rc_amount - _rc_ajuste_total
                    _rc_info = f"**Importe neto:** ARS {fmt_ars(_rc_neto)}"
                    if _rc_ajuste_total > 0:
                        _rc_info += f"  ·  Deducciones: ARS {fmt_ars(_rc_ajuste_total)}"
                    if _rcsel_ids:
                        _rc_info += (
                            f"  ·  {len(_rcsel_ids)} factura(s) seleccionada(s) "
                            f"(saldo ARS {fmt_ars(_rcsel_saldo)})")
                    else:
                        _rc_info += "  ·  Sin facturas → se registra como pago a cuenta"
                    st.info(_rc_info)

                    _dup_pending = st.session_state.get(f"rc_confirm_dup_{_rcuit}", False)
                    if _dup_pending:
                        st.button("↩️ Cancelar", key=f"rc_cancel_dup_{_rcuit}",
                                  on_click=lambda: st.session_state.pop(
                                      f"rc_confirm_dup_{_rcuit}", None))
                        _rc_reg_btn = st.button(
                            "⚠️ Registrar igual (puede ser duplicado)",
                            type="secondary", key=f"rc_btn_{_rcuit}")
                    else:
                        _rc_reg_btn = st.button(
                            f"💵 Registrar Recibo en Odoo",
                            type="primary", key=f"rc_btn_{_rcuit}")

                    if _rc_reg_btn:
                        if _rc_neto <= 0:
                            st.error("El importe neto debe ser mayor a cero.")
                        else:
                            _rc_date_str = _rc_date.strftime("%Y-%m-%d")
                            # Obtener currency_id: primero buscar ARS dinámicamente
                            _rc_cur_id = None
                            try:
                                _ars_cur = models.execute_kw(ODOO_DB, uid, api_key,
                                    "res.currency", "search_read",
                                    [[("name", "=", "ARS"), ("active", "in", [True, False])]],
                                    {"fields": ["id"], "limit": 1})
                                if _ars_cur:
                                    _rc_cur_id = _ars_cur[0]["id"]
                            except Exception:
                                pass
                            if not _rc_cur_id:
                                _rc_cur_id = 1  # fallback
                            if _rcsel_ids and _rc_invs:
                                _rci_first = next(
                                    (i for i in _rc_invs
                                     if i["id"] == _rcsel_ids[0]), None)
                                if _rci_first:
                                    _rcurr = _rci_first.get("currency_id")
                                    if _rcurr and isinstance(_rcurr, (list, tuple)):
                                        _rc_cur_id = _rcurr[0]

                            # ── Validación de duplicados ──────────────────
                            # Busca pagos ya registrados con mismo cliente,
                            # monto, fecha y journal en estado confirmado.
                            _dup_key = f"rc_confirm_dup_{_rcuit}"
                            _dup_existing = []
                            try:
                                _dup_existing = models.execute_kw(
                                    ODOO_DB, uid, api_key,
                                    "account.payment", "search_read",
                                    [[
                                        ("partner_id",   "=",  _rc_pid),
                                        ("amount",       "=",  _rc_neto),
                                        ("date",         "=",  _rc_date_str),
                                        ("journal_id",   "=",  _rc_jour_id),
                                        ("payment_type", "=",  "inbound"),
                                        ("state",        "in", ["posted", "reconciled"]),
                                    ]],
                                    {"fields": ["id", "name", "amount", "date"], "limit": 3})
                            except Exception:
                                pass

                            if _dup_existing and not st.session_state.get(_dup_key):
                                _dup_names = ", ".join(
                                    d.get("name","?") for d in _dup_existing)
                                st.warning(
                                    f"⚠️ Ya existe un cobro registrado con el mismo cliente, "
                                    f"monto y fecha: **{_dup_names}**. "
                                    f"Si es un cobro diferente, confirmá para continuar.")
                                st.session_state[_dup_key] = True
                                st.rerun()
                            else:
                                # Limpiar flag de confirmación para próxima vez
                                st.session_state.pop(_dup_key, None)
                                _rc_cheque_vals = []
                                for _rch in _rchs:
                                    _bk = match_bank_id(_rch.get("banco",""),_rc_all_banks)
                                    _dt = _rch.get("fecha_pago","") or ""
                                    try:
                                        if "/" in _dt:
                                            _p = _dt.split("/")
                                            if len(_p)==3:
                                                _dt = f"{_p[2]}-{_p[1].zfill(2)}-{_p[0].zfill(2)}"
                                    except Exception:
                                        _dt = _rc_date_str
                                    if not _dt or len(_dt)<8: _dt = _rc_date_str
                                    _rc_cheque_vals.append({
                                        "nro":          _rch.get("nro",""),
                                        "bank_id":      _bk,
                                        "issuer_vat":   _rch.get("cuit",""),
                                        "payment_date": _dt,
                                        "amount":       float(_rch.get("importe") or 0),
                                    })
                                with st.spinner("Registrando cobro en Odoo..."):
                                    # Pasar retenciones como withholdings para que
                                    # register_customer_payment cree pagos adicionales
                                    # en el grupo y evite el error de monto vs cheques
                                    _rc_withholdings = [
                                        d for d in st.session_state.get(_ded_key, [])
                                        if float(d.get("monto", 0)) > 0
                                    ] if _rc_cheque_vals else None
                                    _rc_ok, _rc_res = register_customer_payment(
                                        models, uid, api_key,
                                        _rc_pid, _rc_neto, _rc_cur_id,
                                        _rc_date_str, _rc_jour_id,
                                        move_ids=_rcsel_ids if _rcsel_ids else None,
                                        memo=_rc_memo,
                                        cheques=_rc_cheque_vals if _rc_cheque_vals else None,
                                        withholdings=_rc_withholdings)
                                if _rc_ok:
                                    _wh_warn = isinstance(_rc_res, str) and _rc_res.startswith("__WH_WARN__")
                                    st.toast(
                                        f"Recibo registrado para {_rc_pname} — ARS {fmt_ars(_rc_neto)}", icon="✅")
                                    search_partners_by_cuits.clear()
                                    get_customer_unpaid_invoices.clear()
                                    if _wh_warn:
                                        _wh_detail = _rc_res.replace("__WH_WARN__", "")
                                        st.warning(
                                            f"⚠️ Recibo registrado, pero **no se pudo crear el pago de retención** "
                                            f"en Odoo. Registralo manualmente.\n\n"
                                            f"Detalle: `{_wh_detail}`")
                                    else:
                                        st.info(
                                            "Presioná 🔄 Actualizar para ver "
                                            "el estado actualizado.")
                                else:
                                    st.error(f"Error al registrar en Odoo: {_rc_res}")

    st.divider()
    st.caption(
        f"Para emitir facturas de venta o gestionar cobros manualmente, "
        f"usá [Odoo Ventas]({ODOO_URL}/odoo/accounting/customers/invoices) directamente.")



# ═══════════════════════════════════════════════════
# TAB ASISTENTE — Chat Claude-Odoo
# ═══════════════════════════════════════════════════
with tab_chat:
    import json   as _jc
    import base64 as _b64c
    import io     as _ioc
    import datetime as _dtc

    st.subheader("Asistente Luminatec")
    st.caption("Pregunta sobre facturas, saldos, socios. Pedi PDFs o exportes Excel. Adjunta archivos para analizarlos.")

    _ant_key = st.secrets.get("ANTHROPIC_API_KEY", "")
    if not _ant_key:
        st.warning("Falta ANTHROPIC_API_KEY en los Secrets de Streamlit Cloud.")
        st.stop()

    # Session state
    if "chat_msgs" not in st.session_state: st.session_state.chat_msgs = []
    if "chat_dl"   not in st.session_state: st.session_state.chat_dl   = []
    if "chat_qr"   not in st.session_state: st.session_state.chat_qr   = []  # quick replies

    def _blk_to_dict(b):
        if isinstance(b, dict): return b
        t = getattr(b, "type", None)
        if t == "text":        return {"type": "text", "text": b.text}
        if t == "tool_use":    return {"type": "tool_use", "id": b.id, "name": b.name, "input": b.input}
        return {"type": "text", "text": str(b)}

    def _odoo_pdf(doc_id, report="account.report_invoice_with_payments"):
        import base64 as _b64p
        # Metodo 1: buscar adjunto PDF ya generado en Odoo (ir.attachment)
        try:
            _atts = models.execute_kw(ODOO_DB, uid, api_key,
                "ir.attachment", "search_read",
                [[["res_model", "=", "account.move"],
                  ["res_id", "=", doc_id],
                  ["mimetype", "=", "application/pdf"]]],
                {"fields": ["id", "name", "datas"], "limit": 1,
                 "order": "id desc"})
            if _atts and _atts[0].get("datas"):
                return _b64p.b64decode(_atts[0]["datas"]), None
        except Exception:
            pass
        # Metodo 2: sesion HTTP (funciona si el usuario ingreso con password real)
        try:
            import requests as _rq
            _s = _rq.Session()
            _auth = _s.post(
                f"{ODOO_URL}/web/session/authenticate",
                json={"jsonrpc": "2.0", "method": "call", "id": 1,
                      "params": {"db": ODOO_DB,
                                 "login": st.session_state.get("user_email", ""),
                                 "password": api_key}},
                timeout=15)
            _uid_r = (_auth.json().get("result") or {}).get("uid")
            if _uid_r:
                _r = _s.get(f"{ODOO_URL}/report/pdf/{report}/{doc_id}", timeout=90)
                if _r.status_code == 200 and _r.content[:4] == b"%PDF":
                    return _r.content, None
        except Exception:
            pass
        # Metodo 3: devolver link directo a Odoo para que el usuario lo descargue
        _odoo_link = f"{ODOO_URL}/odoo/accounting/customer-invoices/{doc_id}"
        return None, f"PDF_LINK:{_odoo_link}"

    def _odoo_xlsx(model, domain, fields, filename="export.xlsx"):
        try:
            import pandas as _pd
            recs = models.execute_kw(ODOO_DB, uid, api_key, model, "search_read",
                [domain], {"fields": fields, "limit": 1000, "order": "id desc"})
            df = _pd.DataFrame(recs)
            buf = _ioc.BytesIO()
            with _pd.ExcelWriter(buf, engine="openpyxl") as _w:
                df.to_excel(_w, index=False)
            return buf.getvalue(), filename, None
        except Exception as _e:
            return None, filename, str(_e)

    _tools = [
        {
            "name": "odoo_search",
            "description": (
                "Busca registros en Odoo. Modelos disponibles:\n"
                "- account.move: facturas (campos: id, name, partner_id, invoice_date, amount_total, "
                "state, move_type, payment_state, invoice_origin, ref). "
                "move_type: out_invoice=FC venta, in_invoice=FC compra, out_refund=NC venta, in_refund=NC compra. "
                "state=posted para confirmadas. payment_state: not_paid, partial, paid, in_payment.\n"
                "- account.move.line: lineas de factura (campos: id, move_id, name, product_id, "
                "quantity, price_unit, price_subtotal, account_id). "
                "Usar para buscar por producto dentro de facturas.\n"
                "- res.partner: socios/clientes/proveedores (campos: id, name, vat, email, phone, "
                "customer_rank, supplier_rank, is_company, category_id).\n"
                "- account.payment: pagos (campos: id, name, partner_id, amount, date, "
                "payment_type, state, journal_id).\n"
                "- product.product: variantes de producto (campos: id, name, default_code, "
                "list_price, standard_price, categ_id, active).\n"
                "- product.template: plantillas de producto (campos: id, name, default_code, "
                "list_price, type, categ_id, active).\n"
                "- purchase.order: ordenes de compra (campos: id, name, partner_id, date_order, "
                "amount_total, state).\n"
                "- stock.move: movimientos de stock.\n"
                "Para saldo pendiente: account.move con payment_state in ['not_paid','partial']."
            ),
            "input_schema": {
                "type": "object",
                "properties": {
                    "model":  {"type": "string"},
                    "domain": {"type": "array", "description": "Condiciones Odoo, ej: [['move_type','=','out_invoice'],['state','=','posted']]"},
                    "fields": {"type": "array", "items": {"type": "string"}},
                    "limit":  {"type": "integer", "default": 10},
                    "order":  {"type": "string", "default": ""}
                },
                "required": ["model", "domain", "fields"]
            }
        },
        {
            "name": "odoo_get_pdf",
            "description": "Genera el PDF de una factura de Odoo y lo deja disponible para descargar.",
            "input_schema": {
                "type": "object",
                "properties": {
                    "doc_id":   {"type": "integer", "description": "ID del documento en Odoo"},
                    "doc_name": {"type": "string",  "description": "Nombre para el archivo, ej: FCE-A_00011-00000779"}
                },
                "required": ["doc_id"]
            }
        },
        {
            "name": "odoo_export_xlsx",
            "description": "Exporta registros de Odoo a Excel descargable.",
            "input_schema": {
                "type": "object",
                "properties": {
                    "model":    {"type": "string"},
                    "domain":   {"type": "array"},
                    "fields":   {"type": "array", "items": {"type": "string"}},
                    "filename": {"type": "string"}
                },
                "required": ["model", "domain", "fields", "filename"]
            }
        },
        {
            "name": "odoo_aggregate",
            "description": (
                "Agrega datos de Odoo con SUM/COUNT/AVG directamente en el servidor (read_group). "
                "Ideal para totales: cuanto se facturo en un periodo, cantidad de facturas, "
                "suma de pagos, etc. Devuelve un solo registro con los totales calculados. "
                "Ejemplo: para total facturado en Mayo usar model='account.move', "
                "domain=[['move_type','=','out_invoice'],['state','=','posted'],"
                "['invoice_date','>=','2026-05-01'],['invoice_date','<=','2026-05-31']], "
                "aggregate_fields=['amount_total:sum','id:count']."
            ),
            "input_schema": {
                "type": "object",
                "properties": {
                    "model":            {"type": "string"},
                    "domain":           {"type": "array"},
                    "aggregate_fields": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "Campos con funcion: 'amount_total:sum', 'id:count', 'amount_total:avg'"
                    },
                    "group_by": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "Opcional: agrupar por campo, ej: ['partner_id'] o ['invoice_date:month']"
                    }
                },
                "required": ["model", "domain", "aggregate_fields"]
            }
        }
    ]

    def _exec_tool(name, inp):
        if name == "odoo_search":
            try:
                recs = models.execute_kw(ODOO_DB, uid, api_key,
                    inp["model"], "search_read",
                    [inp["domain"]],
                    {"fields": inp["fields"], "limit": inp.get("limit", 10), "order": inp.get("order", "")})
                return _jc.dumps(recs, ensure_ascii=False, default=str)
            except Exception as _e:
                return f"Error busqueda: {_e}"
        elif name == "odoo_get_pdf":
            _did   = inp["doc_id"]
            _dname = inp.get("doc_name", f"documento_{_did}").replace("/", "-").replace(" ", "_")
            _pdf, _err = _odoo_pdf(_did)
            if _pdf:
                _fname = _dname if _dname.endswith(".pdf") else f"{_dname}.pdf"
                st.session_state.chat_dl.append({"name": _fname, "data": _pdf, "mime": "application/pdf"})
                return f"PDF '{_fname}' generado ({len(_pdf):,} bytes). Disponible para descargar."
            if _err and _err.startswith("PDF_LINK:"):
                _link = _err.replace("PDF_LINK:", "")
                return f"No pude generar el PDF automaticamente (requiere password real, no API key). Podés descargarlo directamente desde Odoo: {_link}"
            return f"No se pudo generar el PDF: {_err}"
        elif name == "odoo_export_xlsx":
            _xb, _fn, _err = _odoo_xlsx(inp["model"], inp["domain"], inp["fields"],
                                         filename=inp.get("filename", "export.xlsx"))
            if _xb:
                st.session_state.chat_dl.append({
                    "name": _fn, "data": _xb,
                    "mime": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"})
                return f"Excel '{_fn}' generado. Disponible para descargar."
            return f"No se pudo generar el Excel: {_err}"
        elif name == "odoo_aggregate":
            try:
                _agg_fields = inp.get("aggregate_fields", [])
                _group_by   = inp.get("group_by", [])
                # Parsear campos de agregacion: "amount_total:sum" → ("amount_total", "sum")
                _spec_fields = []
                _spec_agg    = {}
                for _af in _agg_fields:
                    if ":" in _af:
                        _fn, _func = _af.rsplit(":", 1)
                        _spec_fields.append(_fn)
                        _spec_agg[_fn] = _func
                    else:
                        _spec_fields.append(_af)
                _all_fields = list(set(_spec_fields + _group_by))
                _rg = models.execute_kw(ODOO_DB, uid, api_key,
                    inp["model"], "read_group",
                    [inp["domain"], _all_fields, _group_by or []],
                    {"lazy": False})
                # Simplificar resultado
                _out_rows = []
                for _row in _rg:
                    _out_row = {}
                    for _fn in _spec_fields:
                        _out_row[_fn] = _row.get(_fn)
                        if _fn + "_count" in _row:
                            _out_row[_fn + "_count"] = _row[_fn + "_count"]
                    for _gb in _group_by:
                        _gb_base = _gb.split(":")[0]
                        if _gb_base in _row:
                            _out_row[_gb_base] = _row[_gb_base]
                    _out_row["__count"] = _row.get("__count", 0)
                    _out_rows.append(_out_row)
                return _jc.dumps(_out_rows, ensure_ascii=False, default=str)
            except Exception as _ae:
                return f"Error agregacion: {_ae}"
        return "Herramienta desconocida"

    def _run_agent(user_text, file_blocks=None):
        try:
         _run_agent_inner(user_text, file_blocks)
        except Exception as _outer_err:
            st.session_state.chat_msgs.append({
                "role": "assistant",
                "content": [{"type": "text", "text": f"⚠️ Error inesperado: {_outer_err}"}]
            })

    def _run_agent_inner(user_text, file_blocks=None):
        import anthropic as _ac2
        _client = _ac2.Anthropic(api_key=_ant_key)
        _today = _dtc.date.today().isoformat()
        _system = "\n".join([
            "Sos el asistente inteligente de Luminatec, conectado a Odoo en tiempo real.",
            "Podés buscar facturas, socios, pagos, productos y cualquier registro.",
            "Podés generar PDFs de facturas y exportar listas a Excel.",
            "",
            "== ESTRATEGIA DE BUSQUEDA — SIEMPRE SEGUILA ==",
            "",
            "BUSQUEDA POR CLIENTE/PROVEEDOR:",
            "1. Buscá en res.partner: domain [['name','ilike','NOMBRE']], fields ['id','name','customer_rank','vat'], order='customer_rank desc', limit 10.",
            "2. Si hay UNO, usalo directamente. Si hay VARIOS, listalos con formato:",
            "   - NOMBRE EMPRESA (ID XXXXX) — uno por línea. Pedí que confirmen.",
            "3. Si no encontrás nada, probá variantes: truncá el nombre, sacá palabras cortas,",
            "   o buscá por CUIT si lo tenés.",
            "4. Con el partner_id confirmado, buscá en account.move con",
            "   [['partner_id','=',ID],['state','=','posted']].",
            "",
            "BUSQUEDA POR PRODUCTO (código o nombre como LFANT, PERUG, etc.):",
            "1. Primero buscá el producto: product.product domain [['default_code','ilike','COD']]",
            "   o [['name','ilike','NOMBRE']], fields ['id','name','default_code'], limit 10.",
            "2. Si no encontrás por default_code, probá buscar en account.move.line:",
            "   domain [['name','ilike','COD'],['move_id.state','=','posted']],",
            "   fields ['id','move_id','name','product_id','quantity','price_subtotal'], limit 20.",
            "3. También podés buscar: [['product_id.default_code','ilike','COD']]",
            "   o [['product_id.name','ilike','NOMBRE']].",
            "4. Una vez que tenés los move_id de las líneas, podés buscar las facturas completas.",
            "",
            "BUSQUEDA POR NUMERO DE FACTURA:",
            "1. Buscá en account.move: domain [['name','ilike','NUMERO'],['state','=','posted']],",
            "   fields ['id','name','partner_id','invoice_date','amount_total','payment_state'].",
            "",
            "REGLAS GENERALES:",
            "- Para preguntas de TOTALES o CANTIDADES (cuanto se facturo, cuantas facturas, etc.)",
            "  SIEMPRE usa odoo_aggregate con read_group en lugar de odoo_search. Es mas rapido y exacto.",
            "- NUNCA te rindas en el primer intento fallido. Probá al menos 2 estrategias distintas.",
            "- Si una búsqueda da vacío, ajustá el domain y reintentá automáticamente.",
            "- Usá ilike para búsquedas parciales, nunca '=' para nombres.",
            "- Para la última factura: order='invoice_date desc', limit 1.",
            "- Para facturas pendientes: payment_state in ['not_paid','partial'].",
            "- Cuando mostrás facturas, incluí: número, fecha, monto total, estado de pago.",
            "- Si el usuario pide el PDF de una factura específica, usá odoo_get_pdf con su ID.",
            "- Si pide exportar a Excel, usá odoo_export_xlsx.",
            "- Respondé siempre en castellano. Sé conciso pero completo.",
            "",
            "== CUANDO RECIBES UN DOCUMENTO ADJUNTO (PDF o imagen) ==",
            "1. Analizalo INMEDIATAMENTE. Identificá qué tipo es:",
            "   pedido de compra, factura, remito, nota de crédito, presupuesto, etc.",
            "2. Extraé TODOS los datos clave:",
            "   - Encabezado: empresa emisora, destinatario, fecha, número de doc, condición de pago.",
            "   - Si tiene productos: código, descripción, marca, modelo, cantidad, precio unit, total.",
            "   - Totales: subtotal, IVA, total general.",
            "3. Presentá los datos en una tabla markdown clara.",
            "4. ACCIÓN AUTOMÁTICA: buscá en Odoo los datos relevantes sin esperar que te lo pidan.",
            "   - Si es un pedido de un cliente: buscá ese cliente en res.partner.",
            "   - Si tiene códigos de producto: buscá cada uno en product.product por default_code.",
            "   - Si es una factura de proveedor: buscá el proveedor en res.partner.",
            "5. Ofrecé acciones concretas según el tipo de doc (crear pedido, buscar facturas, etc.).",
            f"Fecha hoy: {_today}",
        ])
        _ublocks = []
        if file_blocks:
            _ublocks.extend(file_blocks)
        _ublocks.append({"type": "text", "text": user_text})
        st.session_state.chat_msgs.append({"role": "user", "content": _ublocks})
        _msgs = [{"role": m["role"], "content": m["content"]}
                 for m in st.session_state.chat_msgs]
        for _ in range(10):
            _resp = _client.messages.create(
                model="claude-sonnet-4-6",
                max_tokens=4096,
                system=_system,
                tools=_tools,
                messages=_msgs,
            )
            _msgs.append({"role": "assistant", "content": _resp.content})
            if _resp.stop_reason != "tool_use":
                break
            _results = []
            for _blk in _resp.content:
                if _blk.type == "tool_use":
                    _out = _exec_tool(_blk.name, _blk.input)
                    _results.append({"type": "tool_result", "tool_use_id": _blk.id, "content": _out})
            _msgs.append({"role": "user", "content": _results})
        st.session_state.chat_msgs = [
            {"role": m["role"],
             "content": [_blk_to_dict(b) for b in (
                 m["content"] if isinstance(m["content"], list)
                 else [{"type": "text", "text": str(m["content"])}])]}
            for m in _msgs
        ]
        # Extraer opciones clickeables de la ultima respuesta del asistente
        import re as _re
        st.session_state.chat_qr = []
        for _lm in reversed(st.session_state.chat_msgs):
            if _lm.get("role") == "assistant":
                _ltexts = [b["text"] for b in _lm.get("content", [])
                           if isinstance(b, dict) and b.get("type") == "text"]
                _ltext = " ".join(_ltexts)
                _qr_matches = _re.findall(r"[-*]?\s*(.+?)\s+\(ID\s+(\d+)\)", _ltext)
                if len(_qr_matches) > 1:
                    st.session_state.chat_qr = [{"label": n.strip(), "id": int(i)} for n, i in _qr_matches]
                break


    # Downloads pendientes
    if st.session_state.chat_dl:
        _dl_cols = st.columns(min(len(st.session_state.chat_dl), 4))
        for _i, _dl in enumerate(st.session_state.chat_dl):
            _dl_cols[_i % 4].download_button(
                label=f"Descargar {_dl['name']}",
                data=_dl["data"],
                file_name=_dl["name"],
                mime=_dl["mime"],
                key=f"dl_{_i}_{_dl['name']}"
            )
        st.divider()

    # Historial de mensajes
    for _m in st.session_state.chat_msgs:
        _role    = _m.get("role", "assistant")
        _content = _m.get("content", [])
        if not isinstance(_content, list):
            _content = [{"type": "text", "text": str(_content)}]
        _texts      = [b["text"] for b in _content
                       if isinstance(b, dict) and b.get("type") == "text" and b.get("text", "").strip()]
        _tool_calls = [b for b in _content
                       if isinstance(b, dict) and b.get("type") == "tool_use"]
        if not _texts and not _tool_calls:
            continue
        if _role == "user":
            if _texts:
                with st.chat_message("user"):
                    st.markdown("\n".join(_texts))
        else:
            with st.chat_message("assistant", avatar="🤖"):
                for _tc in _tool_calls:
                    with st.expander(f"Tool: {_tc.get('name','tool')}", expanded=False):
                        st.json(_tc.get("input", {}))
                if _texts:
                    st.markdown("\n".join(_texts))

    # Quick reply buttons (opciones clickeables)
    if st.session_state.chat_qr:
        st.markdown("**Selecciona uno:**")
        _qr_cols = st.columns(min(len(st.session_state.chat_qr), 3))
        for _qi, _qr in enumerate(st.session_state.chat_qr):
            if _qr_cols[_qi % 3].button(
                _qr["label"], key=f"qr_{_qi}_{_qr['id']}", use_container_width=True
            ):
                _sel_msg = f"{_qr['label']} (ID {_qr['id']})"
                st.session_state.chat_qr = []
                with st.spinner("Pensando..."):
                    _run_agent(_sel_msg)
                st.rerun()

    # Input y upload
    st.divider()
    _cu1, _cu2 = st.columns([5, 1])
    with _cu1:
        _chat_upload = st.file_uploader(
            "Adjuntar", type=["pdf", "xlsx", "xls", "png", "jpg", "jpeg"],
            key=f"chat_up_{len(st.session_state.chat_msgs)}",
            label_visibility="collapsed")
    with _cu2:
        if st.button("Nueva", key="chat_new_btn", use_container_width=True):
            st.session_state.chat_msgs = []
            st.session_state.chat_dl   = []
            st.rerun()

    _chat_in = st.chat_input("Pregunta algo, ej: Facturas pendientes de PETDUR / Descargame la ultima factura de Castillo")

    if _chat_in or _chat_upload:
        _fblocks = []
        if _chat_upload:
            _fb = _chat_upload.read()
            _fn = _chat_upload.name.lower()
            if _fn.endswith(".pdf"):
                _fblocks.append({"type": "document", "source": {
                    "type": "base64", "media_type": "application/pdf",
                    "data": _b64c.b64encode(_fb).decode()}})
            elif _fn.endswith((".png", ".jpg", ".jpeg")):
                _mime2 = "image/png" if _fn.endswith(".png") else "image/jpeg"
                _fblocks.append({"type": "image", "source": {
                    "type": "base64", "media_type": _mime2,
                    "data": _b64c.b64encode(_fb).decode()}})
            else:
                _fblocks.append({"type": "text", "text": f"[Archivo adjunto: {_chat_upload.name}]"})
        _user_text = _chat_in if _chat_in else (
            f"Analizá este documento adjunto ({_chat_upload.name}) y extrae toda la informacion relevante."
            if _chat_upload else "")
        if not _user_text:
            st.rerun()
        _spin_msg = "Analizando documento..." if (not _chat_in and _chat_upload) else "Pensando..."
        with st.spinner(_spin_msg):
            _run_agent(_user_text, _fblocks or None)
        st.rerun()


# ═══════════════════════════════════════════════════
# TAB HISTORIAL
# ═══════════════════════════════════════════════════
with tab_history:
    st.subheader("📋 Historial de esta sesión")
    _hist = st.session_state.get("history", [])
    if not _hist:
        st.caption("Todavía no se procesó ningún documento en esta sesión.")
    else:
        import pandas as _pd_hist
        _hdf = _pd_hist.DataFrame(_hist)
        _hcols = [c for c in ["hora","tipo","archivo","estado","id","url"] if c in _hdf.columns]
        _hdf_disp = _hdf[_hcols].copy()
        if "url" in _hdf_disp.columns:
            _hdf_disp["url"] = _hdf_disp["url"].apply(
                lambda u: f"[Abrir]({u})" if u else "")
        st.dataframe(_hdf_disp, use_container_width=True, hide_index=True)


