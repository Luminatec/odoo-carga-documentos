"""
Luminatec · Odoo — Parsers de documentos
PDF (facturas), imágenes, Excel (órdenes de compra), ARCA (constancias), DOCX (formularios).
"""
import re
import base64
import streamlit as st
from io import BytesIO
from datetime import datetime as _dt_now
import pandas as pd
import config as _cfg
from odoo_client import (
    normalize_amount,
    parse_payment_terms,
    compute_vencimiento,
    parse_petdur_invoice_lines,
    safe_float,
)


def parse_ar_date(raw):
    """Convierte fechas a ISO YYYY-MM-DD.
    Soporta: DD/MM/YYYY, DD-MM-YYYY, DD.MM.YYYY, YYYY-MM-DD, YYYY/MM/DD."""
    if not raw:
        return ""
    raw = raw.strip()
    m = re.match(r"(\d{1,2})[/\-\.](\d{1,2})[/\-\.](\d{4})", raw)
    if m:
        d, mo, y = m.group(1).zfill(2), m.group(2).zfill(2), m.group(3)
        return f"{y}-{mo}-{d}"
    m2 = re.match(r"(\d{4})[/\-](\d{1,2})[/\-](\d{1,2})", raw)
    if m2:
        return raw[:10]
    return ""



def parse_ar_date(raw):
    """Convierte fechas a ISO YYYY-MM-DD.
    Soporta: DD/MM/YYYY, DD-MM-YYYY, DD.MM.YYYY, YYYY-MM-DD, YYYY/MM/DD."""
    if not raw:
        return ""
    raw = raw.strip()
    m = re.match(r"(\d{1,2})[/\-\.](\d{1,2})[/\-\.](\d{4})", raw)
    if m:
        d, mo, y = m.group(1).zfill(2), m.group(2).zfill(2), m.group(3)
        return f"{y}-{mo}-{d}"
    m2 = re.match(r"(\d{4})[/\-](\d{1,2})[/\-](\d{1,2})", raw)
    if m2:
        return raw[:10]
    return ""

def _ai_extract_invoice_fields(text):
    """
    Extrae campos de factura usando Claude Haiku via API de Anthropic.
    Retorna un fields_dict con los mismos keys que extract_pdf_fields,
    o lanza excepción si falla (el caller hace fallback a regex).
    Requiere st.secrets["ANTHROPIC_API_KEY"].
    """
    import anthropic, json as _json

    _api_key = st.secrets.get("ANTHROPIC_API_KEY", "")
    if not _api_key:
        raise ValueError("ANTHROPIC_API_KEY no configurada en secrets")

    client = anthropic.Anthropic(api_key=_api_key)

    _prompt = """Extraé los campos de esta factura argentina. Respondé SOLO con JSON válido, sin texto extra ni bloques de código.

Formato esperado:
{
  "numero": "00004-00020659",
  "fecha": "DD/MM/YYYY",
  "proveedor": "Razón social del EMISOR (quien factura, no quien recibe)",
  "cuit": "30710058667",
  "total": 318124.13,
  "neto": 262912.50,
  "iva": 55211.63,
  "condiciones_venta": "A 7 dias FF",
  "tipo": "RI",
  "concepto": "Descripción breve del servicio o producto facturado"
}

Reglas:
- "numero" siempre en formato XXXXX-XXXXXXXX (5 dígitos, guión, 8 dígitos, con ceros a la izquierda)
- "cuit" del EMISOR (proveedor), sin guiones ni espacios
- "total" incluye todos los impuestos
- "neto" es la base imponible / subtotal gravado
- "iva" es la suma de todos los IVA (21%, 10.5%, 27%)
- "tipo": "RI" (Responsable Inscripto), "MONO" (Monotributo), "EX" (Exento)
- Si la factura es tipo C o el emisor es Monotributo: "iva" = null, "neto" = mismo valor que "total"
- "concepto": descripción breve del servicio/producto (1 línea, máx 100 chars). Buscá en el cuerpo de la factura: "concepto", "descripción", "detalle", o la primer línea del detalle de items. Si no hay, usar null
- Para campos no encontrados usar null
- Números como decimales sin símbolo de moneda (ej: 318124.13, no "$318.124,13")

Factura:
""" + text[:3000]

    resp = client.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=300,
        messages=[{"role": "user", "content": _prompt}]
    )

    raw_json = resp.content[0].text.strip()
    # Limpiar por si el modelo wrapeó en ```json ... ```
    raw_json = re.sub(r"^```(?:json)?\s*|\s*```$", "", raw_json, flags=re.MULTILINE).strip()
    data = _json.loads(raw_json)

    # ── Construir fields_dict con el mismo schema que el parser regex ─────
    fields = {
        "numero":            str(data.get("numero") or "").strip(),
        "fecha":             str(data.get("fecha") or "").strip(),
        "fecha_iso":         "",
        "fecha_vencimiento": "",
        "fecha_vto_iso":     "",
        "proveedor":         str(data.get("proveedor") or "").strip()[:80],
        "cuit":              re.sub(r"[\s\-]", "", str(data.get("cuit") or "")),
        "total":             str(data.get("total") or "").replace(",", ".") if data.get("total") is not None else "",
        "neto":              str(data.get("neto") or "").replace(",", ".") if data.get("neto") is not None else "",
        "iva":               str(data.get("iva") or "").replace(",", ".") if data.get("iva") is not None else "",
        "condiciones_venta": str(data.get("condiciones_venta") or "").strip(),
        "concepto":          str(data.get("concepto") or "").strip()[:100],
        "dias_pago":         None,
    }

    # Convertir números a string limpio (sin notación científica)
    for _k in ("total", "neto", "iva"):
        try:
            if fields[_k]:
                fields[_k] = f"{float(fields[_k]):.2f}"
        except Exception:
            fields[_k] = ""

    # Fecha ISO y vencimiento (igual que el parser regex)
    if fields["fecha"]:
        fields["fecha_iso"] = parse_ar_date(fields["fecha"])
    cond_text = fields["condiciones_venta"] or text
    fields["dias_pago"] = parse_payment_terms(cond_text)
    if fields["dias_pago"] and fields["fecha_iso"]:
        fields["fecha_vencimiento"] = compute_vencimiento(fields["fecha_iso"], fields["dias_pago"])
        fields["fecha_vto_iso"]     = fields["fecha_vencimiento"]

    fields["_source"] = "ai"
    return fields



def _extract_percepcion_iva(text):
    """Extrae Percepciones de IVA (ej: RG2408/08) — formato Telecom y similares.
    Retorna (total_float, [{"label": str, "importe": float}]).
    """
    try:
        import re as _re
        from odoo_client import normalize_amount as _norm
        def _pn(s):
            try: return float(_norm(str(s)))
            except Exception: return 0.0

        detalle = []
        # Restringir al bloque resumen para evitar duplicados de páginas de detalle
        _bloque_iva = text
        _sm2 = _re.search(
            r"SUBTOTAL\s*SIN\s*IMPUESTOS.*?TOTAL\s*CARGOS\s*DEL\s*MES",
            text, _re.IGNORECASE | _re.DOTALL)
        if _sm2:
            _bloque_iva = _sm2.group(0)
        _seen_iva = set()
        for m in _re.finditer(
                r"Percepci[oó]n\s*(?:de\s*)?IVA\s*([A-Za-z0-9/]*)\s*[\d.,]+%\s+([\d.,]+)",
                _bloque_iva, _re.IGNORECASE):
            label = ("Percepción IVA " + m.group(1).strip()).strip()
            imp   = _pn(m.group(2))
            if imp > 0 and label not in _seen_iva:
                _seen_iva.add(label)
                detalle.append({"label": label, "importe": imp})

        if detalle:
            return sum(d["importe"] for d in detalle), detalle
    except Exception:
        pass
    return 0.0, []

def _bot_extract(file_bytes: bytes, filename: str, mime_type: str, doc_type: str) -> dict:
    """
    Llama al endpoint /extract del bot Cloud Run usando Claude Sonnet con soporte nativo PDF.
    Retorna dict con campos extraídos, o {} si el bot no está configurado o falla.
    """
    import base64 as _b64
    import os
    _url   = os.getenv("BOT_URL", "").rstrip("/")
    _token = os.getenv("CHAT_TOKEN", "")
    if not _url or not _token:
        return {}
    try:
        r = requests.post(
            f"{_url}/extract",
            json={
                "file_b64":  _b64.b64encode(file_bytes).decode(),
                "file_name": filename,
                "file_mime": mime_type,
                "doc_type":  doc_type,
            },
            headers={"X-Chat-Token": _token},
            timeout=120,
        )
        r.raise_for_status()
        return r.json().get("fields", {})
    except Exception:
        return {}


def _extract_percepciones_iibb(text):
    """Extrae percepciones IIBB con detalle por provincia.
    Retorna (total_float, [{"provincia": str, "importe": float}, ...]).
    """
    try:
        import re as _re
        from odoo_client import normalize_amount as _norm
        def _pn(s):
            try: return float(_norm(str(s)))
            except Exception: return 0.0

        detalle = []

        # Formato 1: filas "base  Per. IIBB PROVINCIA  tasa  importe"
        for m2 in _re.finditer(
                "[\d.,]+\s+Per\.?\s*IIBB\s+([A-Za-z\s]+?)\s+[\d.,]+\s+([\d.,]+)\s*$",
                text, _re.IGNORECASE | _re.MULTILINE):
            prov = m2.group(1).strip()
            imp  = _pn(m2.group(2))
            if imp > 0:
                detalle.append({"provincia": prov, "importe": imp})

        # Formato 2: Telecom/Personal — "PercepciónIIBBCABA5% 1.882,20"
        # Solo del bloque resumen (entre SUBTOTALSINIMPUESTOS y TOTALCARGOSDELMES)
        if not detalle:
            import re as _re2
            _bloque = text
            _sm = _re2.search(
                r"SUBTOTAL\s*SIN\s*IMPUESTOS.*?TOTAL\s*CARGOS\s*DEL\s*MES",
                text, _re2.IGNORECASE | _re2.DOTALL)
            if _sm:
                _bloque = _sm.group(0)
            _seen_provs = set()
            for m in _re2.finditer(
                    r"Percepci[oó]n\s*IIBB\s*([A-Za-záéíóúÁÉÍÓÚ]+(?:\s+[A-Za-záéíóúÁÉÍÓÚ]+)*?)\s*[\d.,]+%\s+([\d.,]+)",
                    _bloque, _re2.IGNORECASE):
                prov = m.group(1).strip()
                imp  = _pn(m.group(2))
                if imp > 0 and prov not in _seen_provs:
                    _seen_provs.add(prov)
                    detalle.append({"provincia": prov, "importe": imp})

        if detalle:
            return sum(d["importe"] for d in detalle), detalle

        # Formato 2b: línea de tabla con IIBB / Ingresos Brutos como línea de producto
        # Cubre: "IIBB BUENOS AIRES 0.00 0.00 3,918.00", "Ingresos Brutos CABA 5% 638.97", etc.
        if not detalle:
            _pat2b = (r"^(?:(?:Perc[ep.]*\s+)?(?:IIBB|Ing\.?\s*Brutos?|Ingresos\s+Brutos?|"
                      r"Retenci[oó]n\s+Ing\.?\s*Brutos?)\s*((?:[A-Za-z\xc1\xe1\xc9\xe9"
                      r"\xcd\xed\xd3\xf3\xda\xfa\s]+?))(?:\s+\([^)]*(?:PERCEP|RETEN)[^)]*\))?"
                      r"\s+(?:[\d.,]+\s+)*?([\d.,]+))\s*$")
            for m2b in _re.finditer(_pat2b, text, _re.IGNORECASE | _re.MULTILINE):
                prov = _re.sub(r"[\d.,%\s]+$", "", m2b.group(1)).strip()
                prov = _re.sub(r"\s+", " ", prov).strip()
                imp  = _pn(m2b.group(2))
                if imp > 0 and prov and prov not in [d["provincia"] for d in detalle]:
                    detalle.append({"provincia": prov, "importe": imp})
        if detalle:
            return sum(d["importe"] for d in detalle), detalle

        # Formato 3: resumen Andreani (solo total)
        _p1 = "Subtotal[^\n]*IIBB[^\n]*\n\s*([\d.,]+)\s+([\d.,]+)\s+[\d.,]+\s+([\d.,]+)"
        m1 = _re.search(_p1, text, _re.IGNORECASE)
        if m1:
            total = _pn(m1.group(2))
            return total, [{"provincia": "IIBB", "importe": total}]

        # Formato 4: etiqueta
        m3 = _re.search("Percepci[^\n]*IIBB[\s:]+([\d.,]+)", text, _re.IGNORECASE)
        if m3:
            total = _pn(m3.group(1))
            return total, [{"provincia": "IIBB", "importe": total}]

    except Exception:
        pass
    return 0.0, []


def extract_pdf_fields(file_bytes):
    """
    Parser para facturas electrónicas argentinas.
    Orden: 1) Bot Cloud Run  2) Regex  3) Haiku IA (solo si regex falla en campos críticos)
    Retorna (fields_dict, raw_text).
    """
    # ── 1. Bot Cloud Run (Claude Sonnet nativo PDF) ───────────────────────
    _bot = _bot_extract(file_bytes, "factura.pdf", "application/pdf", "factura")
    if _bot.get("proveedor") or _bot.get("numero"):
        _bot["_source"] = "bot"
        try:
            import pdfplumber
            with pdfplumber.open(BytesIO(file_bytes)) as _pdf:
                _raw = "\n".join(p.extract_text() or "" for p in _pdf.pages)
        except Exception:
            _raw = ""
        _p, _pd = _extract_percepciones_iibb(_raw)
        if _p: _bot["percepcion_iibb"] = _p
        if _pd: _bot["percepcion_iibb_detalle"] = _pd
        _piva, _pivad = _extract_percepcion_iva(_raw)
        if _piva: _bot["percepcion_iva"] = _piva
        if _pivad: _bot["percepcion_iva_detalle"] = _pivad
        import re as _re_iva27
        _m27 = _re_iva27.search(r"IVA\s*27\s*%\s*([\d.,]+)", _raw, _re_iva27.IGNORECASE)
        if _m27:
            try:
                from odoo_client import normalize_amount as _na27
                _bot["iva_27"] = float(_na27(_m27.group(1)))
            except Exception: pass
        _m21b = _re_iva27.search(r"IVA\s*21\s*%\s*([\d.,]+)", _raw, _re_iva27.IGNORECASE)
        if _m21b:
            try:
                from odoo_client import normalize_amount as _na21b
                _v21b = float(_na21b(_m21b.group(1)))
                if _v21b > 100:
                    _bot["iva_21"] = _v21b
                    _bot["iva"] = f"{(_v21b + float(_bot.get('iva_27') or 0)):.2f}"
            except Exception: pass
        _bn = str(_bot.get("numero") or "").strip()
        if not re.match(r"^\d{4,5}-\d{6,8}$", _bn):
            for _bnp in [r"Nro\.Comprobante[:\s]*(\d{4,5}-\d{6,8})", r"\b(\d{4,5}-\d{6,8})\b"]:
                _bnm = re.search(_bnp, _raw, re.IGNORECASE)
                if _bnm:
                    _bparts = _bnm.group(1).split("-")
                    _bot["numero"] = f"{_bparts[0].zfill(5)}-{_bparts[1].zfill(8)}" if len(_bparts)==2 else _bnm.group(1)
                    break
        return _bot, _raw

    # ── 2. Extraer texto del PDF ──────────────────────────────────────────
    try:
        import pdfplumber
        with pdfplumber.open(BytesIO(file_bytes)) as pdf:
            text = "\n".join(page.extract_text() or "" for page in pdf.pages)
    except Exception:
        return {}, ""
    if not text.strip():
        return {}, ""

    # ── 3. Parser regex ───────────────────────────────────────────────────

    fields = {"numero": "", "fecha": "", "fecha_iso": "", "fecha_vencimiento": "",
              "fecha_vto_iso": "", "proveedor": "", "total": "", "neto": "", "iva": "",
              "cuit": "", "condiciones_venta": "", "dias_pago": None, "_source": "regex"}

    # ── NÚMERO DE COMPROBANTE ─────────────────────────────────────────────
    # Soporta:
    #   Formato AFIP estándar:     "Nro. Comp.: 00002-00013670"
    #   Formato con letra prefijo: "FACTURA A00005-00029174"
    # Primero: combinar "Punto de Venta: XXXX ... Comp. Nro: XXXXXXXX" (AFIP estándar)
    _m_pv = re.search(
        r"Punto\s+de\s+Venta[:\s]+(\d{4,5})[^\n]{0,60}?Comp\.?\s*Nro\.?[:\s]+(\d{6,8})",
        text, re.IGNORECASE)
    if _m_pv:
        fields["numero"] = f"{_m_pv.group(1).zfill(5)}-{_m_pv.group(2).zfill(8)}"

    if not fields["numero"]:
        num_pats = [
            r"(?:Nro\.?\s*Comp\.?(?:\s*\(Nro\.?\s*Orig\.?\))?|N[°º]\s*Comp\.?|Comprobante\s*N[°º]?)[:\s]*(\d{4,5}[-\s]\d{6,8})",
            r"(?:Punto\s+de\s+Venta[:\s]+\d+\s+)?(?:Comp\.?\s*Nro\.?|Nro\.)[:\s]+(\d{4,5}-\d{6,8})",
            r"(?:FACTURA|NOTA\s+DE\s+CR[EÉ]DITO|NOTA\s+DE\s+D[EÉ]BITO|RECIBO)\s+([A-Z]\d{4,5}-\d{6,8})",
            r"\b([A-Z]\d{4,5}-\d{6,8})\b",
            r"\b(\d{4,5}-\d{6,8})\b",
            # Patrón 6: requiere "Factura" o "Invoice" antes del N° para no capturar "CAE N°"
            r"Nro\.Comprobante[:\s]*(\d{4,5}-\d{6,8})",
            r"(?:Factura|Invoice)\s*N[°º\.][:\s#]*([A-Z0-9\-]{5,20})",
        ]
        for pat in num_pats:
            m = re.search(pat, text, re.IGNORECASE)
            if m:
                _raw_num = m.group(1).strip()
                _np = _raw_num.split("-")
                if len(_np) == 2 and _np[0].isdigit() and _np[1].isdigit():
                    fields["numero"] = f"{_np[0].zfill(5)}-{_np[1].zfill(8)}"
                else:
                    fields["numero"] = _raw_num
                break

    # Formato "Nº0004 - 00020659" (Nº + espacios alrededor del guión, sin prefijo letra)
    if not fields["numero"]:
        m = re.search(r"N[°º]\s*(\d{4,5})\s*[-–]\s*(\d{6,8})", text)
        if m:
            fields["numero"] = f"{m.group(1).zfill(5)}-{m.group(2).zfill(8)}"

    # ── FECHA DE EMISIÓN ──────────────────────────────────────────────────
    emision_pats = [
        r"(?:Fecha\s+de\s+[Ee]misi[oó]n|Fecha\s+[Ee]mis\.?)[:\s]+(\d{1,2}/\d{1,2}/\d{4})",
        r"(?:^|\n|\s)(?:FECHA|Fecha)[:\s]+(\d{1,2}/\d{1,2}/\d{4})",
    ]
    for pat in emision_pats:
        m = re.search(pat, text, re.IGNORECASE | re.MULTILINE)
        if m:
            fields["fecha"] = m.group(1).strip()
            fields["fecha_iso"] = parse_ar_date(fields["fecha"])
            break
    if not fields["fecha"]:
        m = re.search(r"(\d{1,2}/\d{1,2}/\d{4})", text)
        if m:
            fields["fecha"] = m.group(1)
            fields["fecha_iso"] = parse_ar_date(fields["fecha"])


    # ── PERCEPCIONES IIBB ───────────────────────────────────────────────────
    def _pnum(s):
        try:
            return float(normalize_amount(str(s)))
        except Exception:
            return 0.0

    _percep_total = 0.0
    # Formato 1: línea resumen (Andreani y similares)
    _pm1 = re.search(
        r"Subtotal[^\n]*IIBB[^\n]*\n\s*([\d.,]+)\s+([\d.,]+)\s+[\d.,]+\s+([\d.,]+)",
        text, re.IGNORECASE)
    if _pm1:
        _percep_total = _pnum(_pm1.group(2))

    # Formato 2: filas Per. IIBB PROVINCIA tasa importe
    if not _percep_total:
        for _pm2 in re.finditer(r"Per[.] ?IIBB\s+\S+\s+[\d.,]+\s+([\d.,]+)", text, re.IGNORECASE):
            _percep_total += _pnum(_pm2.group(1))

    if _percep_total:
        fields["percepcion_iibb"] = _percep_total
        # detalle ya disponible si llegamos por regex
        _pdet = [{"provincia": "IIBB", "importe": _percep_total}]
        fields["percepcion_iibb_detalle"] = _pdet

    # ── CONDICIONES DE VENTA ──────────────────────────────────────────────
    cond_pats = [
        r"(?:Condici[oó]n(?:es)?\s+de\s+Venta|Cond\.?\s*Vta\.?)[:\s]+([^\n]{3,80})",
    ]
    for pat in cond_pats:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            fields["condiciones_venta"] = m.group(1).strip()
            break

    # ── FECHA DE VENCIMIENTO DE PAGO ─────────────────────────────────────
    # Se calcula desde condiciones de venta (ej: "CUENTA CORRIENTE A 10 DIAS").
    # "Fecha de Vto." en facturas AFIP/CAE es el vencimiento del CAE, NO el de pago.
    cond_text = fields["condiciones_venta"] or text
    fields["dias_pago"] = parse_payment_terms(cond_text)
    if fields["dias_pago"] and fields["fecha_iso"]:
        fields["fecha_vencimiento"] = compute_vencimiento(fields["fecha_iso"], fields["dias_pago"])
        fields["fecha_vto_iso"]     = fields["fecha_vencimiento"]

    # ── IMPORTE TOTAL ─────────────────────────────────────────────────────
    total_pats = [
        r"(?:Importe\s+Total|Total\s+Factura|TOTAL\s+FACTURA)[:\s$]*\$?\s*([\d.,]+)",
        r"(?:^|\n|\s)TOTAL\s*:\s*\$?\s*([\d.,]+)",        # TOTAL: $amount
        r"(?:^|\n|\s)TOTAL\s+\$\s*([\d.,]+)",             # TOTAL $ amount
        r"(?:^|\n|\s)TOTAL\s+([\d.,]+)(?:\s|$)",          # TOTAL amount
        r"(?:^|\n|\s)PESOS\s+TOTAL[:\s$]*\$?\s*([\d.,]+)",
        r"Total\s+a\s+pagar[:\s$]*\$?\s*([\d.,]+)",
    ]
    for pat in total_pats:
        m = re.search(pat, text, re.IGNORECASE | re.MULTILINE)
        if m:
            fields["total"] = normalize_amount(m.group(1).strip())
            break

    # ── NETO GRAVADO ──────────────────────────────────────────────────────
    neto_pats = [
        # 1. Etiquetas explícitas (más seguras), incluyendo "Subt.Gravado" (abreviado)
        r"(?:Subt\.?\s*Gravado|Subtotal\s+Gravado|Neto\s+Gravado|Base\s+Imponible)[:\s$]*\$?\s*([\d.,]+)",
        r"(?:Gravado)\s*:\s*\$?\s*([\d.,]+)",
        # 2. "Subtotal:" con dos puntos
        r"(?:SUBTOTAL|Subtotal)\s*:\s*\$?\s*([\d.,]+)",
        # 3. "Subtotal" + newline opcional + "$" + monto (Odoo/columnas separadas en PDF)
        r"(?:SUBTOTAL|Subtotal)[^\n]*\n?\s*\$\s*([\d.,]+)",
        # 4. "Subtotal" + espacios + monto con separadores de miles (sin $)
        #    Exige X.XXX,XX o X,XXX.XX para no capturar "1,00" de tablas
        r"(?:SUBTOTAL|Subtotal)\s+([\d]{1,3}(?:[.,][\d]{3})+(?:[.,][\d]{2}))",
    ]
    for pat in neto_pats:
        m = re.search(pat, text, re.IGNORECASE | re.MULTILINE)
        if m:
            fields["neto"] = normalize_amount(m.group(1).strip())
            break

    # Factura C / Monotributo: sin IVA, neto = importe total
    if not fields["neto"] and fields["total"]:
        _is_monotrib = bool(re.search(
            r"(?:Responsable\s+Monotributo|MONOTRIBUTO|Factura\s+C\b|COD\.?\s*011)",
            text, re.IGNORECASE))
        if _is_monotrib:
            fields["neto"] = fields["total"]

    # ── IVA ───────────────────────────────────────────────────────────────
    iva_pats = [
        r"IVA\s*(?:21|10[.,]5|27)\s*%\s+([\d.,]+)",  # IVA 21% o IVA27% seguido de espacio + monto
        r"I\.?V\.?A[^:\n]*(?:21|10\.5|27)[^:\n]*:[:\s$]*\$?\s*([\d.,]+)",
        r"I\.?V\.?A[:\s$%\d.]*:\s*([\d.,]+)",
        r"(?:Impuesto\s+)?IVA[:\s$]*\$?\s*([\d.,]+)",
    ]
    for pat in iva_pats:
        # Usar findall y tomar el ÚLTIMO match: las filas de tabla dan el IVA por línea
        # (primer match), el resumen al final da el IVA total (último match)
        matches = re.findall(pat, text, re.IGNORECASE)
        if matches:
            fields["iva"] = normalize_amount(matches[-1].strip())
            break

    # ── RESUMEN TABULAR (fila "T o t a l", común en facturas de servicios) ────
    # Header: "Subtotal ... I.V.A. 21% ... T o t a l"
    # Fila:   "$ 262,912.50  $ 262,912.50  $ 55,211.63  $ 318,124.13"
    # Mapeo:   [0]=neto       [1]=neto       [-2]=iva      [-1]=total
    if not fields["total"] or not fields["neto"] or not fields["iva"]:
        _m_sumrow = re.search(
            r"Subtotal[^\n]*T\s+o\s+t\s+a\s+l[^\n]*\n([^\n]+)",
            text, re.IGNORECASE)
        if _m_sumrow:
            _amounts = re.findall(r"\$\s*([\d.,]+)", _m_sumrow.group(1))
            if _amounts:
                if not fields["total"]:
                    fields["total"] = normalize_amount(_amounts[-1])
                if not fields["neto"] and len(_amounts) >= 1:
                    fields["neto"] = normalize_amount(_amounts[0])
                if not fields["iva"] and len(_amounts) >= 3:
                    # IVA es el penúltimo monto (antes del total)
                    fields["iva"] = normalize_amount(_amounts[-2])

    # ── RAZÓN SOCIAL / PROVEEDOR EMISOR ──────────────────────────────────
    razon_pats = [
        r"(?:Raz[oó]n\s+[Ss]ocial|Denominaci[oó]n)[:\s]+([^\n\d][^\n]{2,79})",
        r"(?:Apellido\s+y\s+Nombre\s+o\s+Raz[oó]n\s+[Ss]ocial|Nombre\s+y\s+Apellido)[:\s]+([^\n]{3,79})",
    ]
    for pat in razon_pats:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            name = m.group(1).strip()
            name = re.sub(r'\s*\d{2}-\d{8}-\d\s*', '', name).strip()
            if len(name) >= 3:
                fields["proveedor"] = name[:80]
                break

    # Fallback: primera línea significativa que no sea keyword AFIP
    if not fields["proveedor"]:
        skip = {"FACTURA", "NOTA", "RECIBO", "REMITO", "CUIT", "AFIP", "CAE",
                "PUNTO", "FECHA", "IMPORTE", "TOTAL", "VENCIMIENTO", "INGRESOS",
                "IVA", "MONOTRIBUTO", "RESPONSABLE", "INSCRIPTO", "ORIGINAL",
                "DUPLICADO", "TRIPLICADO", "CÓDIGO", "DOMICILIO", "PROVINCIA",
                "CODIGO", "SUBTOTAL", "DESCRIPCION", "CANTIDAD", "PRECIO",
                "CONDICION", "COMPROBANTE", "PESOS", "SON"}
        for line in (l.strip() for l in text.split("\n") if l.strip()):
            if len(line) < 4 or re.match(r'^[\d$.,/\s\-()]+$', line):
                continue
            upper = line.upper()
            if any(w in upper for w in skip):
                continue
            # Descartar líneas que son claramente el número de comprobante
            if re.match(r'^[A-Z]\d{4,5}-\d{6,8}$', line):
                continue
            fields["proveedor"] = line[:80]
            break

    # ── CUIT EMISOR ───────────────────────────────────────────────────────
    # Toma el primer CUIT encontrado en el documento (suele ser el emisor)
    cuit_m = re.search(r'(?:CUIT|C\.U\.I\.T)[:\s.]*(\d{2}[-\s]?\d{8}[-\s]?\d)', text, re.IGNORECASE)
    if cuit_m:
        fields["cuit"] = re.sub(r'[\s\-]', '', cuit_m.group(1))

    # ── 4. Enriquecer con percepciones (extractor especializado) ──────────
    _p2, _pd2 = _extract_percepciones_iibb(text)
    if _p2:
        fields["percepcion_iibb"] = _p2
        fields["percepcion_iibb_detalle"] = _pd2
    _piva2, _pivad2 = _extract_percepcion_iva(text)
    if _piva2:
        fields["percepcion_iva"] = _piva2
        fields["percepcion_iva_detalle"] = _pivad2
    _m27r = re.search(r"IVA\s*27\s*%\s*([\d.,]+)", text, re.IGNORECASE)
    if _m27r:
        try:
            fields["iva_27"] = float(normalize_amount(_m27r.group(1)))
        except Exception:
            pass
    # IVA 21% explícito (facturas Telecom separan IVA27% e IVA21% en líneas distintas)
    _m21r = re.search(r"IVA\s*21\s*%\s*([\d.,]+)", text, re.IGNORECASE)
    if _m21r:
        try:
            _iva21_val = float(normalize_amount(_m21r.group(1)))
            if _iva21_val > 100:  # monto real, no un porcentaje
                fields["iva_21"] = _iva21_val
                # Total IVA = IVA21 + IVA27 (si hay ambos)
                _iva27_existing = float(fields.get("iva_27") or 0)
                fields["iva"] = f"{(_iva21_val + _iva27_existing):.2f}"
        except Exception:
            pass

    # ── 5. Si regex obtuvo los campos críticos, retornar sin llamar a la IA
    if fields.get("numero") and fields.get("total"):
        return fields, text

    # ── 6. Fallback IA: Haiku con texto truncado ──────────────────────────
    try:
        _ai_fields = _ai_extract_invoice_fields(text)
        if _ai_fields.get("proveedor") or _ai_fields.get("numero"):
            # Preservar percepciones del regex (son más confiables que la IA)
            for _k in ("percepcion_iibb", "percepcion_iibb_detalle",
                       "percepcion_iva", "percepcion_iva_detalle", "iva_27"):
                if fields.get(_k):
                    _ai_fields[_k] = fields[_k]
            # Normalizar número si la IA no lo trajo en formato estándar
            _ai_num = str(_ai_fields.get("numero") or "").strip()
            if not re.match(r"^\d{4,5}-\d{6,8}$", _ai_num):
                for _np in [
                    r"Nro\.Comprobante[:\s]*(\d{4,5}-\d{6,8})",
                    r"\b(\d{4,5}-\d{6,8})\b",
                ]:
                    _nm = re.search(_np, text, re.IGNORECASE)
                    if _nm:
                        _raw_n = _nm.group(1)
                        _parts = _raw_n.split("-")
                        if len(_parts) == 2:
                            _ai_fields["numero"] = f"{_parts[0].zfill(5)}-{_parts[1].zfill(8)}"
                        else:
                            _ai_fields["numero"] = _raw_n
                        break
            return _ai_fields, text
    except Exception:
        pass

    # ── 7. Retornar lo que haya sacado el regex ───────────────────────────
    return fields, text


def extract_arca_fields(text):
    """
    Extrae campos de una Constancia de Inscripción ARCA (AFIP).
    Devuelve dict con: nombre, cuit, forma_juridica, street, city, zip_code,
    province_name, tipo_resp (RI/MONO/EX/otro), actividad_principal.
    """
    f = {
        "nombre": "", "cuit": "", "forma_juridica": "",
        "street": "", "city": "", "zip_code": "", "province_name": "",
        "tipo_resp": "RI",   # default: Responsable Inscripto
        "actividad_principal": "",
    }
    if not text:
        return f

    lines = text.splitlines()

    # Nombre y CUIT — primera línea con "CUIT:"
    for ln in lines[:10]:
        m = re.match(r"^(.+?)\s+CUIT:\s*([\d\-]+)", ln.strip())
        if m:
            f["nombre"] = m.group(1).strip()
            f["cuit"]   = re.sub(r"[\s\-]", "", m.group(2))
            break

    # Forma jurídica
    for ln in lines:
        m = re.match(r"Forma\s+Jur[íi]dica:\s*(.+)", ln, re.I)
        if m:
            f["forma_juridica"] = m.group(1).strip()
            break

    # Tipo responsabilidad AFIP
    text_up = text.upper()
    if "MONOTRIBUTO" in text_up or "RSOC " in text_up:
        f["tipo_resp"] = "MONO"
    elif "EXENTO" in text_up and "IVA" not in text_up:
        f["tipo_resp"] = "EX"
    else:
        f["tipo_resp"] = "RI"   # IVA registrado → Responsable Inscripto

    # Actividad principal
    for ln in lines:
        m = re.search(r"Actividad\s+principal:\s*\d+\s*(?:\(F-\d+\))?\s*(.+?)(?:\s+Mes de inicio|$)", ln, re.I)
        if m:
            f["actividad_principal"] = m.group(1).strip()[:120]
            break

    # Domicilio fiscal
    # Buscar el bloque después de "DOMICILIO FISCAL - ARCA"
    try:
        idx = next(i for i, l in enumerate(lines) if "DOMICILIO FISCAL" in l.upper() and "ARCA" in l.upper())
        addr_lines = [l.strip() for l in lines[idx+1:idx+5] if l.strip()]
        if addr_lines:
            f["street"] = addr_lines[0]
        if len(addr_lines) >= 2:
            f["city"] = addr_lines[1]
        if len(addr_lines) >= 3:
            # "5963-CORDOBA" → zip=5963, province=CORDOBA
            m_cp = re.match(r"(\d+)[\s\-]+(.+)", addr_lines[2])
            if m_cp:
                f["zip_code"]      = m_cp.group(1)
                f["province_name"] = m_cp.group(2).strip().title()
            else:
                f["province_name"] = addr_lines[2].strip().title()
    except StopIteration:
        pass

    return f


def parse_alta_cliente_docx(file_bytes: bytes) -> dict:
    """Parsea el formulario interno 'ALTA DE CLIENTE' (.docx).
    Retorna dict compatible con extract_arca_fields() + campos extra:
    email, phone, website, iibb, transport_name, transport_address,
    delivery_address, forma_pago, plazos, tipo_resp_raw."""
    import zipfile as _zf
    out = {
        "nombre": "", "cuit": "", "iibb": "",
        "street": "", "city": "", "zip_code": "", "province_name": "",
        "tipo_resp": "RI", "actividad_principal": "",
        "email": "", "phone": "", "website": "",
        "transport_name": "", "transport_address": "",
        "delivery_address": "", "forma_pago": "", "plazos": "",
        "_source": "formulario_interno",
    }
    try:
        with _zf.ZipFile(BytesIO(file_bytes)) as _z:
            _xml = _z.read("word/document.xml").decode("utf-8")
        # Extraer runs de texto en orden
        _texts = re.findall(r'<w:t[^>]*>([^<]*)</w:t>', _xml)
        _lines = [t.strip() for t in _texts if t.strip()]
        _full  = "\n".join(_lines)

        # Palabras reservadas para labels: no tomar como valor
        _LABELS = {
            "razón social", "razon social", "cuit", "iibb", "domicilio",
            "correo electrónico", "correo electronico",
            "teléfono", "telefono", "int", "celular",
            "datos logísticos:", "datos logisticos:",
            "nombre de transporte", "dirección transporte", "direccion transporte",
            "web", "horario", "solicitud de turno:",
            "domicilio entrega final", "facturación:", "facturacion:",
            "tipo iva", "forma de pago", "mipymes fce", "plazos acordados",
            "si/no", "alta de cliente", "fecha:",
        }

        def _next_val(idx, lines, max_look=5):
            """Retorna el primer texto no-label después de lines[idx]."""
            for _j in range(idx + 1, min(idx + max_look, len(lines))):
                _c = lines[_j].strip()
                if _c and _c.lower() not in _LABELS:
                    return _c
            return ""

        for _i, _l in enumerate(_lines):
            _ll = _l.lower().strip()

            # Razón Social
            if re.match(r"^raz[oó]n\s+social$", _ll):
                out["nombre"] = _next_val(_i, _lines)

            # CUIT
            elif _ll == "cuit" and not out["cuit"]:
                _v = _next_val(_i, _lines)
                _digits = re.sub(r"[^\d]", "", _v)
                if len(_digits) == 11:
                    out["cuit"] = _digits

            # IIBB
            elif _ll == "iibb":
                _v = _next_val(_i, _lines)
                if re.search(r"[\d]", _v):
                    out["iibb"] = _v

            # Domicilio fiscal (solo la primera ocurrencia simple, no "entrega" ni "transporte")
            elif re.match(r"^domicilio\s*$", _ll) and not out["street"]:
                _v = _next_val(_i, _lines)
                if _v and not re.match(r'^(correo|tel[eé]|datos|nombre|direcci|factur|tipo|forma|plazo|si/no|web)', _v, re.I):
                    out["street"] = _v

            # Domicilio entrega Final
            elif re.search(r"domicilio\s+entrega\s+final", _ll):
                _v = _next_val(_i, _lines)
                if _v and not re.match(r'^(factur|tipo|forma|plazo|\*)', _v, re.I):
                    out["delivery_address"] = _v

            # Nombre de transporte
            elif re.match(r"^nombre\s+de\s+transporte$", _ll):
                out["transport_name"] = _next_val(_i, _lines)

            # Dirección Transporte
            elif re.search(r"direcci[oó]n\s+transporte", _ll):
                _v = _next_val(_i, _lines)
                if _v and not re.match(r'^(tel[eé]|correo|web|horario)', _v, re.I):
                    out["transport_address"] = _v

            # TIPO IVA
            elif re.match(r"^tipo\s+iva$", _ll):
                _v = _next_val(_i, _lines)
                if _v:
                    _vu = _v.upper()
                    if "MONOTRIBUTO" in _vu or "MONO" in _vu:
                        out["tipo_resp"] = "MONO"
                    elif "EXENTO" in _vu:
                        out["tipo_resp"] = "EX"
                    else:
                        out["tipo_resp"] = "RI"
                    out["tipo_resp_raw"] = _v

            # Forma de Pago
            elif re.match(r"^forma\s+de\s+pago$", _ll):
                _v = _next_val(_i, _lines)
                if _v and not re.match(r'^(mi|plazo|si/no|\*)', _v, re.I):
                    out["forma_pago"] = _v

            # Plazos acordados
            elif re.match(r"^plazos\s+acordados$", _ll):
                _v = _next_val(_i, _lines)
                if _v and not re.match(r'^\*', _v):
                    out["plazos"] = _v

        # Email — primer patrón válido en todo el texto
        _em = re.search(r"[\w\.\-\+]+@[\w\.\-]+\.[a-zA-Z]{2,}", _full)
        if _em:
            out["email"] = _em.group(0).strip()

        # Teléfono — primer patrón telefónico argentino
        _ph = re.search(
            r"(?:0\d{2,4}[\-\s]?\d{6,8}|\+54[\s\d\-]{8,14}|\d{4}[\-\s]\d{4,8})", _full)
        if _ph:
            out["phone"] = _ph.group(0).strip()

        # Sitio web
        _wb = re.search(r"(?:https?://|www\.)\S+", _full)
        if _wb:
            out["website"] = _wb.group(0).strip()

    except Exception:
        pass
    return out


def fetch_arca_by_cuit(cuit_str: str) -> dict:
    """Consulta datos de un contribuyente en ARCA/AFIP por CUIT.
    Intenta argentinadatos.com primero, luego TangoFactura como fallback.
    Retorna dict compatible con extract_arca_fields(), o {"_error": msg} si falla."""
    try:
        import requests as _req
        _cuit_clean = re.sub(r"[\s\-]", "", cuit_str.strip())
        if not _cuit_clean.isdigit() or len(_cuit_clean) != 11:
            return {"_error": f"CUIT inválido: debe tener 11 dígitos (recibido: '{cuit_str.strip()}')"}

        # ── Intento 1: argentinadatos.com ──────────────────────────────────
        _url1 = f"https://api.argentinadatos.com/v1/afip/personas/{_cuit_clean}"
        _resp = None
        try:
            _resp = _req.get(_url1, timeout=10)
        except Exception as _e1:
            pass

        if _resp is None or _resp.status_code != 200:
            # ── Intento 2: TangoFactura (fallback) ─────────────────────────
            _url2 = f"https://afip.tangofactura.com/Rest/GetContribuyenteFull?cuit={_cuit_clean}"
            try:
                _resp2 = _req.get(_url2, timeout=10)
                if _resp2.status_code == 200:
                    _d2 = _resp2.json()
                    _contrib = _d2.get("contribuyente") or _d2
                    _nombre2 = (_contrib.get("razonSocial") or
                                f"{_contrib.get('apellido','')} {_contrib.get('nombre','')}").strip()
                    if _nombre2:
                        _dom2 = (_contrib.get("domicilioFiscal") or
                                 _contrib.get("domicilio") or {})
                        return {
                            "nombre": _nombre2, "cuit": _cuit_clean,
                            "forma_juridica": _contrib.get("tipoPersona",""),
                            "street":   (_dom2.get("direccion") or "").strip().title(),
                            "city":     (_dom2.get("localidad") or "").strip().title(),
                            "zip_code": str(_dom2.get("codPostal") or "").strip(),
                            "province_name": (_dom2.get("descripcionProvincia") or
                                              _dom2.get("provincia") or "").strip().title(),
                            "tipo_resp": "RI", "actividad_principal": "",
                            "_source": "tangofactura",
                        }
            except Exception:
                pass

            # ── Intento 3: cuitonline.com (scraping HTML) ──────────────────
            try:
                _url3 = f"https://www.cuitonline.com/search.php?q={_cuit_clean}"
                _r3 = _req.get(_url3, timeout=10,
                               headers={"User-Agent": "Mozilla/5.0"})
                if _r3.status_code == 200:
                    # Buscar el nombre en el HTML: típicamente en <td class="nombre">...</td>
                    # o en <strong> dentro de tabla de resultados
                    _nombre3 = ""
                    _m3 = re.search(
                        r'(?:class=["\'](?:razonSocial|nombre)["\'][^>]*>|'
                        r'<td[^>]*>\s*' + re.escape(_cuit_clean) + r'\s*</td>\s*<td[^>]*>)'
                        r'\s*([^<]{3,80})',
                        _r3.text, re.I)
                    if not _m3:
                        # Fallback: buscar el CUIT formateado y tomar el siguiente td
                        _cuit_fmt = f"{_cuit_clean[:2]}-{_cuit_clean[2:10]}-{_cuit_clean[10]}"
                        _m3 = re.search(
                            r'(?:' + re.escape(_cuit_fmt) + r'|' + re.escape(_cuit_clean) + r')'
                            r'[^<]*</td>\s*<td[^>]*>\s*([^<]{3,100})',
                            _r3.text, re.I)
                    if _m3:
                        _nombre3 = re.sub(r'\s+', ' ', _m3.group(1)).strip()
                    if _nombre3:
                        return {
                            "nombre": _nombre3, "cuit": _cuit_clean,
                            "forma_juridica": "", "street": "", "city": "",
                            "zip_code": "", "province_name": "",
                            "tipo_resp": "RI", "actividad_principal": "",
                            "_source": "cuitonline",
                            "_aviso": "Solo se obtuvo el nombre (completá el resto manualmente).",
                        }
            except Exception:
                pass

            # Los tres intentos fallaron
            _status = _resp.status_code if _resp is not None else "sin respuesta"
            _cuit_fmt = f"{_cuit_clean[:2]}-{_cuit_clean[2:10]}-{_cuit_clean[10]}"
            return {
                "_error": (
                    f"CUIT {_cuit_fmt} no encontrado en las fuentes disponibles "
                    f"(argentinadatos HTTP {_status}, TangoFactura y CuitOnline fallaron). "
                    f"La constancia de ARCA requiere CAPTCHA y no puede consultarse automáticamente. "
                    f"Completá los datos manualmente o consultá en ARCA."
                ),
                "_afip_link": (
                    "https://seti.afip.gob.ar/padron-puc-constancia-internet/"
                    "ConsultaConstanciaAction.do"
                ),
                "_cuit_fmt": _cuit_fmt,
            }
        _data = _resp.json()

        _out = {
            "nombre": "", "cuit": _cuit_clean, "forma_juridica": "",
            "street": "", "city": "", "zip_code": "", "province_name": "",
            "tipo_resp": "RI", "actividad_principal": "",
        }

        # Razón social / nombre
        _tipo = _data.get("tipoPersona", "")
        if _tipo == "JURIDICA":
            _out["nombre"] = (_data.get("razonSocial") or "").strip()
            _out["forma_juridica"] = "Sociedad"
        else:
            _parts = [_data.get("apellido",""), _data.get("nombre","")]
            _out["nombre"] = " ".join(p for p in _parts if p).strip()

        # Domicilio fiscal
        _dom = _data.get("domicilioFiscal") or {}
        _out["street"]        = (_dom.get("direccion") or "").strip().title()
        _out["city"]          = (_dom.get("localidad") or "").strip().title()
        _out["zip_code"]      = str(_dom.get("codPostal") or "").strip()
        _out["province_name"] = (_dom.get("descripcionProvincia") or "").strip().title()

        # Tipo de responsabilidad
        for _c in (_data.get("caracterizaciones") or []):
            _desc = (_c.get("descripcionCaracterizacion") or "").upper()
            if "MONOTRIBUT" in _desc:
                _out["tipo_resp"] = "MONO"; break
            if "EXENTO" in _desc or "NO ALCANZADO" in _desc:
                _out["tipo_resp"] = "EX"; break
            if "RESPONSABLE INSCRIPTO" in _desc or ("IVA" in _desc and "NO" not in _desc):
                _out["tipo_resp"] = "RI"; break

        # Actividad principal (menor orden = principal)
        _acts = sorted(_data.get("actividades") or [], key=lambda a: a.get("orden", 99))
        if _acts:
            _out["actividad_principal"] = (_acts[0].get("descripcionActividad") or "")[:120]

        return _out
    except Exception:
        return {}



def _image_to_ocr_text(file_bytes):
    """
    OCR de imagen con pytesseract.
    Devuelve (text, error_msg). Si tesseract no está instalado, error_msg != "".
    """
    try:
        import pytesseract
        from PIL import Image as _PILImage
        img = _PILImage.open(BytesIO(file_bytes))
        if img.mode in ("RGBA", "LA", "P"):
            img = img.convert("RGB")
        # Mejorar resolución para OCR (mínimo 300 DPI recomendado)
        w, h = img.size
        if w < 1000:
            factor = max(2, 1200 // w)
            img = img.resize((w * factor, h * factor), _PILImage.LANCZOS)
        try:
            text = pytesseract.image_to_string(img, lang="spa+eng",
                                               config="--psm 6 --oem 3")
        except Exception:
            text = pytesseract.image_to_string(img, lang="eng",
                                               config="--psm 6 --oem 3")
        return text, ""
    except Exception as e:
        return "", str(e)

def extract_image_fields(file_bytes):
    """OCR imagen → pipeline de facturas. Intenta bot primero, luego OCR."""
    _bot = _bot_extract(file_bytes, "factura.jpg", "image/jpeg", "factura")
    if _bot.get("proveedor") or _bot.get("numero"):
        _bot["_source"] = "bot"
        return _bot, ""
    try:
        import pytesseract
        from PIL import Image as _PILImage
        img = _PILImage.open(BytesIO(file_bytes))
        if img.mode in ("RGBA", "LA", "P"):
            img = img.convert("RGB")
        try:
            pdf = pytesseract.image_to_pdf_or_hocr(img, lang="spa+eng", extension="pdf")
        except Exception:
            pdf = pytesseract.image_to_pdf_or_hocr(img, lang="eng", extension="pdf")
        return extract_pdf_fields(pdf)
    except Exception:
        return {}, ""

def extract_image_oc_fields(file_bytes):
    """
    OCR imagen → extrae campos de Orden de Compra. Intenta bot primero, luego OCR.

    Formato observado en OCs Canon:
      Línea N-1: descripción completa del producto (viene antes de la línea del precio)
      Línea N  : [ítem#] [código] [desc parcial] $ [precio]
      Línea N+1: especificación técnica (a ignorar)
    """
    _bot = _bot_extract(file_bytes, "oc.jpg", "image/jpeg", "oc")
    if _bot.get("numero_oc") or _bot.get("cuit") or _bot.get("lineas"):
        _bot["_source"] = "bot"
        return _bot, {}, ""
    text, err = _image_to_ocr_text(file_bytes)
    if err or not text.strip():
        return {}, {}, err or ""

    result = {
        "cuit": "", "numero_oc": "", "fecha": "", "fecha_iso": "",
        "condiciones_pago": "", "dias_pago": None,
        "neto": "", "iva21": "", "iva105": "", "total": "",
        "lineas": [],
    }

    lines = [ln for ln in text.split("\n")]  # conservar índices

    # ── CUIT ──────────────────────────────────────────────────────────────
    _cuit_m = re.search(r"(?:CUIT|C\.U\.I\.T\.?)[:\s#]*(\d{2}[-\.\s]?\d{8}[-\.\s]?\d)", text, re.I)
    if _cuit_m:
        result["cuit"] = re.sub(r"[\s\.]", "-", _cuit_m.group(1).strip())

    # ── Número OC ─────────────────────────────────────────────────────────
    _oc_m = re.search(
        r"(?:orden\s+de\s+compra|N[°º#.]*\s*OC|OC\s*N[°º#.]?|OC)[:\s#]*([A-Z0-9][\-A-Z0-9]{1,20})",
        text, re.I)
    if _oc_m:
        result["numero_oc"] = _oc_m.group(1).strip()

    # ── Fecha ─────────────────────────────────────────────────────────────
    _fecha_m = re.search(r"(\d{1,2})[/\-\.](\d{1,2})[/\-\.](\d{2,4})", text)
    if _fecha_m:
        d, m2, y = _fecha_m.group(1), _fecha_m.group(2), _fecha_m.group(3)
        y = "20" + y if len(y) == 2 else y
        result["fecha"] = f"{d}/{m2}/{y}"
        try:
            result["fecha_iso"] = f"{y}-{int(m2):02d}-{int(d):02d}"
        except Exception:
            pass

    # ── Total ─────────────────────────────────────────────────────────────
    for _tp in [
        r"(?:total\s+(?:general|orden|a\s+pagar)?)[:\s$]*(\d[\d.,]+)",
        r"(?:^|\n)\s*TOTAL[:\s$]*(\d[\d.,]+)",
    ]:
        _tot_m = re.search(_tp, text, re.I | re.MULTILINE)
        if _tot_m:
            result["total"] = _tot_m.group(1).strip()
            break

    # ── Líneas de productos ───────────────────────────────────────────────
    # Patrón principal: línea que tiene $ seguido de número
    # Maneja OCR artifacts: ; en vez de , en precios (77.896; → 77.896,00)
    # Formato: [ítem] [código] [desc parcial] $ [precio]
    # IMPORTANTE: precio puede terminar en ; (OCR artifact) → no requerir dígito final
    _price_line_pat = re.compile(
        r"^\s*(?:(\d{1,2})\s+)?(.+?)\s*\$\s*([\d][\d.,;:]*)\s*$"
    )

    seen_prices = set()

    for i, line in enumerate(lines):
        m = _price_line_pat.match(line)
        if not m:
            continue

        item_s   = m.group(1)   # número de ítem (puede ser None)
        code_frag = m.group(2).strip()
        price_raw = m.group(3).strip()

        # Limpiar artefactos OCR en el precio
        # 1. Quitar trailing ; : , . (OCR los agrega al final cuando corta la línea)
        price_clean = price_raw.rstrip(";:., ")
        # 2. Reemplazar ; y : internos por , (OCR confunde coma con punto y coma)
        price_clean = re.sub(r"[;:]", ",", price_clean)

        # Validar que es un precio razonable (> 100 ARS)
        price_val = safe_float(price_clean)
        if price_val < 100:
            continue

        # Evitar duplicados
        if price_clean in seen_prices:
            continue
        seen_prices.add(price_clean)

        # ── Extraer código: primer token del fragmento en la línea del precio
        parts = code_frag.split(None, 1)
        codigo    = parts[0] if parts else ""
        desc_frag = parts[1] if len(parts) > 1 else ""

        # ── Buscar descripción completa en la línea ANTERIOR (no vacía)
        desc_prev = ""
        for j in range(i - 1, max(i - 4, -1), -1):
            prev = lines[j].strip()
            # Ignorar líneas vacías o que son solo números/encabezados
            if prev and not re.match(r"^\d+$", prev) and len(prev) > 3:
                desc_prev = prev
                break

        # Usar descripción previa como descripción principal; si no hay, usar fragmento
        descripcion = desc_prev if desc_prev else desc_frag or code_frag

        # Limpiar artefactos OCR comunes en la descripción
        descripcion = re.sub(r"^[A-Z]\s+", "", descripcion)  # letra suelta al inicio (ej: "E Cabezal...")
        descripcion = descripcion.strip(" ,-—")

        result["lineas"].append({
            "codigo":      codigo,
            "descripcion": descripcion,
            "cantidad":    1,
            "precio_unit": price_clean,
            "subtotal":    price_clean,
            "iva_pct":     21,
        })

    return result, {}, text



def _claude_api_extract_oc(file_bytes: bytes, mime_type: str = "application/pdf") -> dict:
    """
    Fallback: usa Claude API directamente para extraer campos de un pedido/OC.
    Se usa cuando BOT_URL no esta configurado o el bot falla.
    """
    import base64 as _b64c2, os as _os2, json as _jc2
    _ant_key2 = _os2.getenv("ANTHROPIC_API_KEY", "")
    if not _ant_key2:
        return {}
    try:
        import anthropic as _ac2
        _client2 = _ac2.Anthropic(api_key=_ant_key2)
        _prompt2 = (
            "Analiza este documento (puede ser un pedido, orden de compra, relacion de pedidos, "
            "presupuesto o cualquier formato de solicitud de productos). "
            "Extrae los datos y devuelve SOLO un JSON con exactamente estos campos:\n"
            "{\n"
            '  "cuit": "CUIT del cliente/empresa que hace el pedido (solo digitos, sin guiones), vacio si no aparece",\n'
            '  "cliente_nombre": "nombre o razon social del cliente que hace el pedido",\n'
            '  "numero_oc": "numero de pedido, orden de compra u orden",\n'
            '  "fecha": "fecha del documento en formato DD/MM/YYYY",\n'
            '  "condiciones_pago": "texto de condicion o plazo de pago",\n'
            '  "dias_pago": numero entero de dias de pago o null,\n'
            '  "total": "monto total como texto (ej: 17.690.000)",\n'
            '  "lineas": [\n'
            '    {\n'
            '      "codigo": "codigo interno del producto",\n'
            '      "descripcion": "articulo o descripcion",\n'
            '      "marca": "marca",\n'
            '      "modelo": "modelo",\n'
            '      "cantidad": cantidad como numero entero,\n'
            '      "precio_unit": precio unitario como numero flotante,\n'
            '      "subtotal": total de la linea como numero flotante\n'
            '    }\n'
            '  ]\n'
            "}"
        )
        _resp2 = _client2.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=1024,
            messages=[{"role": "user", "content": [
                {"type": "document", "source": {
                    "type": "base64", "media_type": mime_type,
                    "data": _b64c2.b64encode(file_bytes).decode()}},
                {"type": "text", "text": _prompt2}
            ]}]
        )
        _text2 = _resp2.content[0].text.strip()
        _js2 = re.search(r'\{[\s\S]*\}', _text2)
        if not _js2:
            return {}
        _data2 = _jc2.loads(_js2.group())
        result2 = {
            "cuit": str(_data2.get("cuit", "") or "").replace("-","").replace(" ",""),
            "cliente_nombre": str(_data2.get("cliente_nombre", "") or ""),
            "numero_oc": str(_data2.get("numero_oc", "") or ""),
            "fecha": str(_data2.get("fecha", "") or ""),
            "fecha_iso": "",
            "condiciones_pago": str(_data2.get("condiciones_pago", "") or ""),
            "dias_pago": _data2.get("dias_pago"),
            "total": str(_data2.get("total", "") or ""),
            "lineas": [],
            "_source": "claude_api",
        }
        # parse fecha_iso
        _fm2 = re.match(r"(\d{1,2})/(\d{1,2})/(\d{4})", result2["fecha"])
        if _fm2:
            result2["fecha_iso"] = f"{_fm2.group(3)}-{_fm2.group(2).zfill(2)}-{_fm2.group(1).zfill(2)}"
        for _ln2 in (_data2.get("lineas") or []):
            result2["lineas"].append({
                "codigo":      str(_ln2.get("codigo", "") or ""),
                "descripcion": str(_ln2.get("descripcion", "") or ""),
                "marca":       str(_ln2.get("marca", "") or ""),
                "modelo":      str(_ln2.get("modelo", "") or ""),
                "cantidad":    float(_ln2.get("cantidad") or 0),
                "precio_unit": float(_ln2.get("precio_unit") or 0),
                "subtotal":    float(_ln2.get("subtotal") or 0),
                "iva_pct":     21.0,
            })
        return result2
    except Exception:
        return {}

def extract_oc_fields(file_bytes):
    """
    Parser para Órdenes de Compra de clientes (formato heterogéneo).
    Intenta bot primero (Claude Sonnet nativo PDF), luego pdfplumber+regex.
    Retorna (fields_dict, all_tables, raw_text).
    """
    _bot = _bot_extract(file_bytes, "oc.pdf", "application/pdf", "oc")
    if _bot.get("numero_oc") or _bot.get("cuit") or _bot.get("lineas"):
        _bot["_source"] = "bot"
        try:
            import pdfplumber
            with pdfplumber.open(BytesIO(file_bytes)) as _pdf:
                _raw = "\n".join(p.extract_text() or "" for p in _pdf.pages)
        except Exception:
            _raw = ""
        return _bot, [], _raw

    # Fallback: Claude API directo (cuando BOT_URL no esta configurado)
    _claude_r = _claude_api_extract_oc(file_bytes)
    if _claude_r.get("numero_oc") or _claude_r.get("lineas"):
        try:
            import pdfplumber
            with pdfplumber.open(BytesIO(file_bytes)) as _pdf:
                _raw = "\n".join(p.extract_text() or "" for p in _pdf.pages)
        except Exception:
            _raw = ""
        return _claude_r, [], _raw

    try:
        import pdfplumber
        with pdfplumber.open(BytesIO(file_bytes)) as pdf:
            pages_text = [page.extract_text() or "" for page in pdf.pages]
            text = "\n".join(pages_text)
            all_tables = []
            for page in pdf.pages:
                tbls = page.extract_tables()
                if tbls:
                    all_tables.extend(tbls)
    except Exception:
        return {}, [], ""
    if not text.strip():
        return {}, [], ""

    fields = {
        "cuit": "", "numero_oc": "", "fecha": "", "fecha_iso": "",
        "condiciones_pago": "", "dias_pago": None,
        "lineas": [],
        "subtotal_neto": "", "iva_21": "", "iva_105": "", "total": "",
    }

    # ── CUIT emisor ──────────────────────────────────────────────────────
    cuits_found = re.findall(r'\b(\d{2}-\d{8}-\d)\b', text)
    if not cuits_found:
        cuits_found = re.findall(r'\bCUIT[:\s.]*(\d{11})\b', text, re.IGNORECASE)
    if cuits_found:
        fields["cuit"] = re.sub(r'[-\s]', '', cuits_found[0])

    # ── Número de OC ─────────────────────────────────────────────────────
    oc_pats = [
        # CASTILLO: "Orden de Compra N 0001-0118,667"
        r"(?:Orden\s+de\s+[Cc]ompra|O\.?C\.?\s*N[°o]?|ORDEN\s+DE\s+COMPRA\s*N?\s*)[:\s]*([0-9]{4}[-/][0-9,]{4,})",
        # Carsa/MUSIMUNDO: "Orden Definitiva de Provisión Número: 4501653808"
        r"(?:Orden\s+Definitiva|Orden\s+de\s+Provisi[oó]n)\b.{0,40}N[úu]mero[:\s]+(\d{6,})",
        r"(?:N[°º]\s*(?:de\s+)?[Oo]rden|Pedido\s+N[°º]|N[°º]\s*[Pp]edido)[:\s]*([0-9]{4}[-/][0-9,]{4,}|\d{6,})",
        r"\b(0{4}[-/][0-9,]{4,})\b",
        # La Anónima / genérico: "Número: 22620313"
        r"\bN[úu]mero[:\s]+(\d{5,})\b",
    ]
    for pat in oc_pats:
        mo = re.search(pat, text, re.IGNORECASE)
        if mo:
            fields["numero_oc"] = mo.group(1).strip().replace(",", "")
            break

    # ── Fecha ─────────────────────────────────────────────────────────────
    # Soporta DD/MM/YYYY, DD.MM.YYYY, DD-MM-YYYY
    date_pats = [
        r"(?:Fecha\s+[Ee]misi[oó]n|Fecha\s+OC|Fecha)[:\s]+(\d{1,2}[/.\-]\d{1,2}[/.\-]\d{4})",
        r"(?:^|\s)(\d{1,2}[/.\-]\d{1,2}[/.\-]\d{4})(?:\s|$)",
    ]
    for pat in date_pats:
        mo = re.search(pat, text, re.IGNORECASE | re.MULTILINE)
        if mo:
            fields["fecha"] = mo.group(1).strip()
            fields["fecha_iso"] = parse_ar_date(fields["fecha"])
            break

    # ── Condiciones de pago ───────────────────────────────────────────────
    cond_pats = [
        r"(?:Condici[oó]n(?:es)?\s+de\s+[Pp]ago|Forma\s+de\s+[Pp]ago)[:\s]+([^\n]{3,80})",
        # Carsa: "Condición: 0016 - 60 Dias"
        r"(?:^|\n)\s*Condici[oó]n[:\s]+([^\n]{3,60})",
        r"(CUENTA\s+CORRIENTE[^\n]{0,60})",
    ]
    for pat in cond_pats:
        mo = re.search(pat, text, re.IGNORECASE | re.MULTILINE)
        if mo:
            fields["condiciones_pago"] = mo.group(1).strip()
            break

    # Días: buscar "Intervalo: 30", "60 Dias", "LOS XX DIAS"
    intervalo_mo = re.search(r"[Ii]ntervalo[:\s]+(\d+)", text)
    if intervalo_mo:
        fields["dias_pago"] = int(intervalo_mo.group(1))
    else:
        fields["dias_pago"] = parse_payment_terms(fields["condiciones_pago"] or text)

    # ── Totales ───────────────────────────────────────────────────────────
    mo = re.search(r"(?:Sub[-\s]?[Tt]otal\s+[Nn]eto|SUBTOTAL\s+NETO|Neto\s+Gravado)[:\s$]*\$?\s*([\d.,]+)", text, re.IGNORECASE)
    if mo:
        fields["subtotal_neto"] = normalize_amount(mo.group(1))
    # Solo capturar IVA si hay un monto en la MISMA línea (no cruzar newline)
    # Esto evita capturar EAN13 del producto siguiente como monto de IVA
    mo = re.search(r"IVA\s+21\s*%[ \t:$]*\$?[ \t]*([\d.,]+)", text, re.IGNORECASE)
    if mo:
        _iva21_raw = mo.group(1)
        # Descartar si parece un EAN13 u otro código de barras (≥12 dígitos sin coma/punto)
        if not re.match(r'^\d{12,}$', _iva21_raw.replace(',','').replace('.','')):
            fields["iva_21"] = normalize_amount(_iva21_raw)
    mo = re.search(r"IVA\s+10[.,]5\s*%[ \t:$]*\$?[ \t]*([\d.,]+)", text, re.IGNORECASE)
    if mo:
        _iva105_raw = mo.group(1)
        if not re.match(r'^\d{12,}$', _iva105_raw.replace(',','').replace('.','')):
            fields["iva_105"] = normalize_amount(_iva105_raw)
    # Carsa: "TOTAL PEDIDO DE COMPRAS ARS 35.923.359,50"
    total_pats = [
        r"(?:Total\s+OC|TOTAL\s+OC|Total\s+[Oo]rden)[:\s$]*\$?\s*([\d.,]+)",
        r"TOTAL\s+PEDIDO[^\d]+([\d.,]+)",
        r"(?:^|\n)\s*TOTAL[:\s$]*\$?\s*([\d.,]+)\s*(?:$|\n)",
    ]
    for _tp in total_pats:
        mo = re.search(_tp, text, re.IGNORECASE | re.MULTILINE)
        if mo:
            fields["total"] = normalize_amount(mo.group(1))
            break

    # ── Líneas de productos (desde tablas pdfplumber) ─────────────────────
    def _try_num(s):
        s = str(s or "").strip()
        if not s or not re.search(r'\d', s):
            return None
        try:
            return float(normalize_amount(s))
        except Exception:
            return None

    for table in all_tables:
        if not table or len(table) < 2:
            continue
        header = [str(c or "").strip().lower() for c in table[0]]
        col_map = {}
        for i, h in enumerate(header):
            if re.search(r"c[oó]d|sku|art[ií]culo|codigo", h) and "codigo" not in col_map:
                col_map["codigo"] = i
            elif re.search(r"cant|qty|cantidad", h) and "cantidad" not in col_map:
                col_map["cantidad"] = i
            elif re.search(r"detal|descri|product|nombre|item", h) and "descripcion" not in col_map:
                col_map["descripcion"] = i
            elif re.search(r"\bneto\b|p\.?\s*unit|precio\s*unit|unitario|unit\s*price", h) and "precio_unit" not in col_map:
                col_map["precio_unit"] = i
            elif re.search(r"\biva\b|tax|%\s*iva|alicuota", h) and "iva_pct" not in col_map:
                col_map["iva_pct"] = i
            elif re.search(r"sub.?total|importe|c\.?final|precio\s*final", h) and "subtotal" not in col_map:
                col_map["subtotal"] = i
        if not col_map or ("descripcion" not in col_map and "codigo" not in col_map):
            continue
        for row in table[1:]:
            if not row or all(not c for c in row):
                continue
            def _gcol(key, default=""):
                idx = col_map.get(key)
                if idx is None or idx >= len(row):
                    return default
                return str(row[idx] or "").strip()
            desc = _gcol("descripcion")
            cod  = _gcol("codigo")
            if not desc and not cod:
                continue
            if re.match(r'^(detalle|descripci[oó]n|producto|item|total|subtotal)$', desc, re.IGNORECASE):
                continue
            qty   = _try_num(_gcol("cantidad"))
            price = _try_num(_gcol("precio_unit"))
            sub   = _try_num(_gcol("subtotal"))
            iva   = _try_num(_gcol("iva_pct"))
            if qty is None and price is None and sub is None:
                continue
            fields["lineas"].append({
                "codigo":      cod,
                "descripcion": desc,
                "cantidad":    qty   if qty   is not None else 0,
                "precio_unit": price if price is not None else 0,
                "iva_pct":     iva   if iva   is not None else 21.0,
                "subtotal":    sub   if sub   is not None else (
                                   (qty * price) if (qty and price) else 0),
            })

    # ── Fallback EAN13: formato Carsa/MUSIMUNDO ──────────────────────────
    # Líneas: {EAN13} {INTCODE-DESCRIPCION} {M3} {QTY} UN {PRECIO} {SUBTOTAL}
    # pdfplumber puede wrappear la línea; combinamos hasta 3 líneas siguientes
    # para armar el registro completo antes de aplicar el regex.
    # IVA en línea posterior ("IVA 21%" / "IVA 10,5%")
    if not fields["lineas"] and text:
        _ean_lines = text.split("\n")
        _ean_pat   = re.compile(
            r'^(\d{13})\s+(.+?)\s+UN\s+([\d.]+,\d{2})\s+([\d.]+,\d{2})\s*$')
        _ean_stop  = re.compile(
            r'^(?:\d{13}\s|Entregar:|TOTAL|Vencimiento|P[aá]gina)',
            re.IGNORECASE)
        for _ei, _eln in enumerate(_ean_lines):
            _strip = _eln.strip()
            if not re.match(r'^\d{13}\b', _strip):
                continue
            # Combinar con las siguientes líneas SOLO si la línea actual no matchea completa
            # (evita agregar el modelo del producto siguiente que rompe el regex)
            _combined = _strip
            if not _ean_pat.match(_combined.strip()):
                for _fwd_idx in range(_ei + 1, min(_ei + 4, len(_ean_lines))):
                    _nxt = _ean_lines[_fwd_idx].strip()
                    if _ean_stop.match(_nxt):
                        break
                    _combined += ' ' + _nxt
                    if _ean_pat.match(_combined.strip()):
                        break
            _em = _ean_pat.match(_combined.strip())
            if not _em:
                continue
            _ean       = _em.group(1)
            _rest_ean  = _em.group(2).strip()
            _price_raw = _em.group(3)
            _sub_raw   = _em.group(4)

            # Extraer qty (último entero antes de UN — ya consumido por regex previo)
            # qty está al final de _rest_ean: "... 3,540 60"
            _qty_m = re.search(r'\s+(\d+)\s*$', _rest_ean)
            _qty   = int(_qty_m.group(1)) if _qty_m else 0
            _rest_ean = (_rest_ean[:_qty_m.start()].strip() if _qty_m else _rest_ean)

            # Remover M3 (decimal con coma como separador decimal: "3,540")
            _m3_m = re.search(r'\s+([\d]+,\d{3})\s*$', _rest_ean)
            if _m3_m:
                _rest_ean = _rest_ean[:_m3_m.start()].strip()

            # Extraer código interno: "176270-DESCRIPCION"
            _ic_m = re.match(r'^(\d{4,8})-(.+)$', _rest_ean)
            _int_code = _ic_m.group(1) if _ic_m else _ean
            _desc_ean = (_ic_m.group(2).strip() if _ic_m else _rest_ean.strip())

            # Líneas de continuación: absorber hasta encontrar "Entregar:" o otro EAN13
            _ei2 = _ei + 1
            while _ei2 < len(_ean_lines):
                _nl2 = _ean_lines[_ei2].strip()
                if (re.match(r'^\d{13}\s', _nl2)
                        or re.match(r'^Entregar:', _nl2, re.IGNORECASE)
                        or re.match(r'^TOTAL|^Vencimiento|^P[aá]gina', _nl2, re.IGNORECASE)):
                    break
                if _nl2 and re.search(r'[A-Za-z0-9]', _nl2) and not re.match(r'^IVA', _nl2, re.IGNORECASE):
                    _desc_ean += " " + _nl2
                _ei2 += 1

            # Buscar IVA en las líneas siguientes (después de Entregar:)
            _iva_ean = 21.0
            for _fwd in range(_ei + 1, min(_ei + 6, len(_ean_lines))):
                _fwd_ln = _ean_lines[_fwd].strip()
                _iva_m2 = re.match(r'IVA\s+([\d,.]+)\s*%', _fwd_ln, re.IGNORECASE)
                if _iva_m2:
                    try:
                        _iva_ean = float(normalize_amount(_iva_m2.group(1)))
                    except Exception:
                        pass
                    break

            _price_f = float(normalize_amount(_price_raw))
            _sub_f   = float(normalize_amount(_sub_raw))
            _desc_ean = re.sub(r'\s+', ' ', _desc_ean).strip()

            fields["lineas"].append({
                "codigo":      _int_code,
                "ean13":       _ean,        # código de barras EAN13 original
                "descripcion": _desc_ean,
                "cantidad":    float(_qty),
                "precio_unit": _price_f,
                "iva_pct":     _iva_ean,
                "subtotal":    _sub_f,
            })

    # ── Fallback: parser de texto cuando no hay tablas ───────────────────
    # Detecta líneas de producto del tipo:
    #   CODIGO  QTY  DESCRIPCION  NETO  IVA%  IT  CFINAL  SUBTOTAL
    # Maneja números fusionados (artefacto de pdfplumber): "297,004.1329,700,413.00"
    if not fields["lineas"] and text:
        _lines = text.split("\n")
        _hdr_idx = None
        for _i, _ln in enumerate(_lines):
            if re.search(r'\bC[oó]digo\b.{0,30}\bCant\b', _ln, re.IGNORECASE):
                _hdr_idx = _i
                break

        if _hdr_idx is not None:
            _stop = re.compile(
                r'^(?:Sub[-\s]?[Tt]otal|Totales|TOTALES|Sub-Totales|Observaciones|IMPORTANTE)',
                re.IGNORECASE)
            _num_re = r'\d{1,3}(?:[.,]\d{3})*[.,]\d{2}|\d+[.,]\d{2}'

            def _parse_num(s):
                s = str(s).strip()
                lc, ld = s.rfind(","), s.rfind(".")
                if lc > ld:
                    return float(s.replace(".", "").replace(",", "."))
                return float(s.replace(",", ""))

            _j = _hdr_idx + 1
            while _j < len(_lines):
                _raw = _lines[_j].strip()
                if _stop.match(_raw):
                    break
                _cm = re.match(r'^(\d{6,12})\s+(\d+)\s+(.+)$', _raw)
                if _cm:
                    _cod   = _cm.group(1)
                    _qty_s = _cm.group(2)
                    _rest  = _cm.group(3).strip()

                    # Absorber línea siguiente si es continuación de la descripción
                    if _j + 1 < len(_lines):
                        _nl = _lines[_j + 1].strip()
                        if (_nl and not re.match(r'^\d{6,12}\s', _nl)
                                and not _stop.match(_nl)
                                and re.search(r'[A-Za-z]', _nl)):
                            _rest += " " + _nl
                            _j += 1

                    # Separar números fusionados: ".13" seguido de dígito → ".13 "
                    _rest_fixed = re.sub(r'(\.\d{2})(\d)', r'\1 \2', _rest)

                    _raw_nums = re.findall(_num_re, _rest_fixed)
                    _nums = []
                    for _rn in _raw_nums:
                        try:
                            _nums.append(_parse_num(_rn))
                        except Exception:
                            pass

                    if len(_nums) >= 3:
                        _qty    = float(_qty_s)
                        # Estructura esperada (de izquierda a derecha):
                        #   Neto  [%Desc]  IVA%  IT  C.Final  Sub-Total
                        # El último siempre es Sub-Total, el primero es Neto
                        _sub    = _nums[-1]
                        _neto   = _nums[0]
                        # IVA%: buscar el valor típico 21.0 o 10.5 en las posiciones centrales
                        _iva    = 21.0
                        for _n in _nums[1:-1]:
                            if abs(_n - 21.0) < 0.6:
                                _iva = 21.0; break
                            if abs(_n - 10.5) < 0.6:
                                _iva = 10.5; break

                        # Descripción: texto antes del primer número + texto después del último
                        # Esto captura el modelo que queda al final de la línea (ej: G2110, G3110)
                        _fst_m  = re.search(_num_re, _rest_fixed)
                        _all_ms = list(re.finditer(_num_re, _rest_fixed))
                        _pre  = (_rest_fixed[:_fst_m.start()].strip()
                                 if _fst_m else _rest_fixed.strip())
                        _post = (_rest_fixed[_all_ms[-1].end():].strip()
                                 if _all_ms else "")
                        _desc = ((_pre + " " + _post).strip() if _post else _pre)
                        _desc = re.sub(r'\s+', ' ', _desc).strip()

                        fields["lineas"].append({
                            "codigo":      _cod,
                            "descripcion": _desc,
                            "cantidad":    _qty,
                            "precio_unit": round(_neto, 2),
                            "iva_pct":     _iva,
                            "subtotal":    round(_sub, 2),
                        })
                _j += 1

    # ── Fallback: La Anónima / tabular texto plano (Cod.Art.Prov.) ──────────
    # Header: "Cod.Art. Cod.Art.Prov. Descripción Marca Bto. Cont. U/M Cant. Costo % Bonif. % Iva Total"
    # Línea:  "2383809 LCANO00015 BOTELLA GL-190 CYA CANON 1 1 CU 25 17098.3 0.00 21.00 427458.00"
    # Números en formato US (punto como decimal): 17098.3  0.00  21.00  427458.00
    if not fields["lineas"] and text:
        _la_lines = text.split("\n")
        _la_hdr = None
        for _li, _ll in enumerate(_la_lines):
            if re.search(r'Cod\.Art\.Prov', _ll, re.IGNORECASE):
                _la_hdr = _li
                break
        if _la_hdr is not None:
            _la_stop = re.compile(
                r'^(?:Sub[-\s]?[Tt]otal|Total\s|Totales|Bonificaci[oó]n|Observaciones|'
                r'Sr\.?\s+Proveedor|Toda\s+Orden|RESERVAR)',
                re.IGNORECASE)
            for _ll in _la_lines[_la_hdr + 1:]:
                _lraw = _ll.strip()
                if not _lraw or _la_stop.match(_lraw):
                    break
                # INT_CODE PROV_CODE ... resto
                _lcm = re.match(r'^(\d{5,10})\s+([A-Z][A-Z0-9]{2,})\s+(.+)', _lraw)
                if not _lcm:
                    continue
                _lint_code  = _lcm.group(1)
                _lprov_code = _lcm.group(2)
                _lrest      = _lcm.group(3).strip()
                # Extraer los últimos 5 tokens numéricos (qty costo bonif iva% total)
                # La Anónima usa formato US: solo dígitos y punto decimal
                _ltoks     = _lrest.split()
                _lnum_toks = []
                for _lt in reversed(_ltoks):
                    if re.match(r'^\d+(?:\.\d+)?$', _lt) and len(_lnum_toks) < 5:
                        _lnum_toks.insert(0, _lt)
                    elif _lnum_toks:
                        break
                if len(_lnum_toks) < 4:
                    continue
                _lqty   = float(_lnum_toks[0])
                _lcosto = float(_lnum_toks[1])
                _liva   = float(_lnum_toks[3]) if len(_lnum_toks) > 3 else 21.0
                _ltotal = float(_lnum_toks[4]) if len(_lnum_toks) > 4 else _lqty * _lcosto
                # Descripción: tokens antes de los numéricos; limpiar U/M y Bto/Cont al final
                _lnum_start = len(_ltoks) - len(_lnum_toks)
                _ldesc = " ".join(_ltoks[:_lnum_start]).strip()
                _ldesc = re.sub(r'\s+\d+\s+\d+\s+[A-Z]{1,3}\s*$', '', _ldesc).strip()
                fields["lineas"].append({
                    "codigo":      _lprov_code,
                    "descripcion": _ldesc,
                    "cantidad":    _lqty,
                    "precio_unit": _lcosto,
                    "iva_pct":     _liva,
                    "subtotal":    _ltotal,
                })

    return fields, all_tables, text


def extract_excel_oc_fields(file_bytes, filename=""):
    """
    Parser flexible para pedidos en Excel (.xls / .xlsx).
    Soporta múltiples formatos de cliente: CANT./CANTIDAD/QTY, IMP.UNIT./PRECIO, etc.
    También extrae metadata (CUIT, razón social, fecha) de las filas de encabezado.
    """
    # ── intentar bot primero (solo xlsx) ─────────────────────────────────
    _fname = (filename or "oc.xlsx").lower()
    if _fname.endswith(".xlsx"):
        _bot = _bot_extract(
            file_bytes, "oc.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "oc")
        if _bot.get("lineas") or _bot.get("numero_oc"):
            _bot["_source"] = "bot"
            return _bot

    fields = {
        "cuit": "", "cliente": "", "numero_oc": "", "fecha": "", "fecha_iso": "",
        "condiciones_pago": "", "dias_pago": None,
        "lineas": [],
        "subtotal_neto": "", "iva_21": "", "iva_105": "", "total": "",
        "fuente": "excel",
    }

    # ── cargar el workbook (xls o xlsx) ──────────────────────────────────
    all_rows = []
    try:
        if _fname.endswith(".xls"):
            import xlrd
            wb = xlrd.open_workbook(file_contents=file_bytes)
            ws = wb.sheet_by_index(0)
            # Convertir a lista de tuplas igual que openpyxl
            for r in range(ws.nrows):
                row = []
                for c in range(ws.ncols):
                    cell = ws.cell(r, c)
                    # xlrd type 0=empty 1=text 2=number 3=date 4=bool 5=error
                    if cell.ctype == 0:
                        row.append(None)
                    elif cell.ctype == 3:
                        # Fecha serial Excel → string
                        try:
                            import datetime
                            dt = xlrd.xldate_as_datetime(cell.value, wb.datemode)
                            row.append(dt.strftime("%d/%m/%Y"))
                        except Exception:
                            row.append(cell.value)
                    else:
                        row.append(cell.value)
                all_rows.append(tuple(row))
        else:
            from openpyxl import load_workbook
            wb2 = load_workbook(BytesIO(file_bytes), data_only=True)
            ws2 = wb2.active
            all_rows = list(ws2.iter_rows(values_only=True))
    except Exception:
        return fields

    # ── Paso 0: extraer metadata de filas pre-header ──────────────────────
    # Buscar CUIT, nombre de cliente, fecha, NRO PEDIDO en filas libres
    _cuit_re = re.compile(r"(\d{2}-\d{8}-\d|\d{11})")
    _fecha_re = re.compile(r"(\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4})")
    for row in all_rows[:20]:
        vals = [str(v or "").strip() for v in row]
        joined = " ".join(vals)
        # CUIT
        if not fields["cuit"]:
            _cm = _cuit_re.search(joined)
            if _cm:
                fields["cuit"] = re.sub(r"[\s\-]", "", _cm.group(1))
        # Fecha
        if not fields["fecha"]:
            _fm = _fecha_re.search(joined)
            if _fm:
                fields["fecha"] = _fm.group(1)
        # Cliente: buscar patrón "CLIENTE <valor>" en celdas adyacentes
        for ci, v in enumerate(vals):
            vl = v.lower()
            if vl in ("cliente", "razon social", "razón social", "cliente:") and ci + 1 < len(vals):
                _cname = vals[ci + 1].strip()
                if _cname and not fields["cliente"]:
                    fields["cliente"] = _cname
            if vl in ("fecha", "fecha:", "date") and ci + 1 < len(vals):
                _fv = vals[ci + 1].strip()
                if _fv and not fields["fecha"]:
                    fields["fecha"] = _fv
            if vl.startswith("nro") and "pedido" in vl and ci + 1 < len(vals):
                _nv = vals[ci + 1].strip()
                if _nv:
                    fields["numero_oc"] = _nv
        # Si la fecha quedó como float (Excel serial) convertirla
        if fields["fecha"] and not fields["fecha_iso"]:
            try:
                _ffl = float(str(fields["fecha"]))
                import xlrd as _xlrd2
                import datetime as _dt2
                _fdt = _xlrd2.xldate_as_datetime(_ffl, 0)
                fields["fecha"] = _fdt.strftime("%d/%m/%Y")
                fields["fecha_iso"] = _fdt.strftime("%Y-%m-%d")
            except Exception:
                pass

    # ── Paso 1: detectar fila de encabezado de productos ─────────────────
    # Aliases de columnas — normalizamos quitando puntos, espacios y acentos
    def _norm(s):
        s = str(s or "").lower().strip()
        s = s.replace(".", "").replace("á","a").replace("é","e").replace("í","i")
        s  = s.replace("ó","o").replace("ú","u").replace("  "," ")
        return s

    # SKU / código
    _SKU_KW    = {"cod", "code", "codigo", "sku", "ref", "referencia", "art", "articulo"}
    # Descripción / modelo / producto
    _PROD_KW   = {"producto", "productos", "descripcion", "descripcion", "detalle",
                  "modelo", "model", "nombre", "item", "articulo", "art", "denominacion"}
    # Cantidad pedida
    _QTY_KW    = {"cant", "cantidad", "qty", "pedido", "unidades", "u", "ctd",
                  "cant pedida", "cantidad pedida", "order qty", "pedir"}
    # Precio unitario — NETO (alta prioridad: sin IVA / minorista / mayorista)
    _PRICE_KW_NETO = {"imp unit", "precio unit", "precio unitario", "p unit", "unitario",
                      "precio s/iva", "precio sin iva", "precio neto", "neto unit",
                      "precio s/ iva", "p s/iva", "imp unit s/iva", "prec unit",
                      "valor unit", "valor unitario"}
    # Prefijos de precio neto (para capturar "Precio s/IVA Minorista", "Precio s/IVA Mayorista", etc.)
    _PRICE_NETO_PREFIXES = ("precio s/iva", "precio sin iva", "precio neto", "neto ",
                             "precio unit", "p s/iva", "imp unit")
    # Precio unitario — FALLBACK (PVP, precio genérico)
    _PRICE_KW_FALLBACK = {"precio", "pvp", "price"}
    # Unión para hit-count
    _PRICE_KW  = _PRICE_KW_NETO | _PRICE_KW_FALLBACK
    # Subtotal / importe total
    _TOTAL_KW  = {"imp total", "total", "subtotal", "importe", "monto",
                  "imp tot", "total linea", "subtot",
                  "precio total c/iva", "precio total c/ iva", "total c/iva",
                  "total con iva", "precio total", "importe total",
                  "subtotal s/iva", "subtotal sin iva"}
    # Observaciones
    _OBS_KW    = {"obs", "observaciones", "observacion", "nota", "notas", "comment"}

    def _is_price_neto(n):
        """True si la celda normalizada es precio neto/sin IVA (alta prioridad)."""
        return n in _PRICE_KW_NETO or any(n.startswith(p) for p in _PRICE_NETO_PREFIXES)

    col_map = {}
    hdr_row_idx = None
    for ri, row in enumerate(all_rows[:30]):
        norms = [_norm(c) for c in row]
        hits = sum(1 for n in norms if (
            n in _SKU_KW or n in _PROD_KW or n in _QTY_KW or
            n in _PRICE_KW or n in _TOTAL_KW
        ))
        if hits >= 2:
            hdr_row_idx = ri
            # Primer paso: mapear todo excepto precio (para precio usamos lógica con prioridad)
            for ci, n in enumerate(norms):
                if n in _SKU_KW and "sku" not in col_map:
                    col_map["sku"] = ci
                if n in _PROD_KW and "modelo" not in col_map:
                    col_map["modelo"] = ci
                if n in _QTY_KW and "pedido" not in col_map:
                    col_map["pedido"] = ci
                if n in _TOTAL_KW and "subtotal" not in col_map:
                    col_map["subtotal"] = ci
                if n in _OBS_KW and "obs" not in col_map:
                    col_map["obs"] = ci
            # Segundo paso: precio — primero buscar columna neto (sin IVA), luego fallback (pvp)
            for ci, n in enumerate(norms):
                if _is_price_neto(n):
                    col_map["precio_unit"] = ci
                    break
            if "precio_unit" not in col_map:
                for ci, n in enumerate(norms):
                    if n in _PRICE_KW_FALLBACK:
                        col_map["precio_unit"] = ci
                        break
            break

    if "pedido" not in col_map and "modelo" not in col_map:
        return fields

    # Si no hay columna de cantidad pero hay modelo + precio, tomar cantidad = 1
    _has_qty = "pedido" in col_map

    # ── Paso 2: leer filas de datos ───────────────────────────────────────
    _skip_norms = _SKU_KW | _PROD_KW | _QTY_KW | _PRICE_KW | _TOTAL_KW | _OBS_KW
    for row in all_rows[hdr_row_idx + 1:]:
        if not any(c not in (None, "", 0, 0.0) for c in row):
            continue

        def _gc(ci):
            return row[ci] if ci is not None and ci < len(row) else None

        sku_val    = str(_gc(col_map.get("sku"))    or "").strip()
        modelo_val = str(_gc(col_map.get("modelo")) or "").strip()
        obs_val    = str(_gc(col_map.get("obs"))    or "").strip()

        _ident = sku_val or modelo_val
        if not _ident:
            continue
        # Saltar si es otra fila de totales o sub-encabezado
        if _norm(_ident) in _skip_norms:
            continue
        # Saltar filas de totales/descuentos (sin ident numérico ni alfanumérico de producto)
        _ident_up = _ident.upper()
        if any(kw in _ident_up for kw in ("TOTAL", "SUBTOTAL", "DESCUENTO", "DESC.", "IVA", "PLAZO", "EXPRESO", "TEL:", "CALLE")):
            continue

        # Cantidad
        if _has_qty:
            try:
                qty = float(str(_gc(col_map["pedido"]) or "0").replace(",", ".").strip())
            except Exception:
                qty = 0.0
            if qty <= 0:
                continue
        else:
            qty = 1.0

        # Precio unitario
        try:
            precio_unit = float(str(_gc(col_map.get("precio_unit")) or "0").replace(",", ".").strip())
        except Exception:
            precio_unit = 0.0

        # Subtotal
        try:
            subtotal = float(str(_gc(col_map.get("subtotal")) or "0").replace(",", ".").strip())
            if subtotal <= 0:
                subtotal = precio_unit * qty
        except Exception:
            subtotal = precio_unit * qty

        # Fallback: si no hay precio unitario pero sí subtotal y cantidad → calcular
        if precio_unit == 0.0 and subtotal > 0 and qty > 0:
            precio_unit = round(subtotal / qty, 6)
        # Fallback inverso: si no hay subtotal pero sí precio unitario
        if subtotal == 0.0 and precio_unit > 0 and qty > 0:
            subtotal = round(precio_unit * qty, 2)

        fields["lineas"].append({
            "codigo":      sku_val,
            "modelo":      modelo_val,
            "descripcion": (modelo_val or sku_val)[:200],
            "ean":         "",
            "cantidad":    qty,
            "precio_unit": precio_unit,
            "iva_pct":     21.0,
            "subtotal":    subtotal,
            "obs":         obs_val,
        })

    return fields


def extract_afip_xml_fields(xml_bytes: bytes) -> dict:
    """Extrae campos de una factura electrónica AFIP en formato XML.

    Soporta los formatos más comunes:
    - Respuesta WSFE de AFIP (FECAEDetResponse)
    - XML de comprobante genérico (Comprobante/Cabecera)
    - Datos del QR AFIP embebidos como URL en el texto del PDF

    Retorna el mismo dict que extract_pdf_fields:
    {cuit, proveedor, numero, fecha, fecha_iso, importe_total, importe_neto, iva_21, iva_105, tipo_doc}
    """
    import xml.etree.ElementTree as ET
    fields = {
        "cuit": "", "proveedor": "", "numero": "", "fecha": "", "fecha_iso": "",
        "importe_total": 0.0, "importe_neto": 0.0, "iva_21": 0.0, "iva_105": 0.0,
        "tipo_doc": "", "condiciones_venta": "", "dias_pago": None,
        "lineas": [], "_from_xml": True,
    }
    try:
        text = xml_bytes.decode("utf-8", errors="replace")
        root = ET.fromstring(text)
    except Exception:
        return fields

    def _find(tag):
        """Busca un tag en cualquier nivel del árbol, sin namespace."""
        for el in root.iter():
            local = el.tag.split("}")[-1] if "}" in el.tag else el.tag
            if local.lower() == tag.lower():
                return (el.text or "").strip()
        return ""

    # ── CUIT y razón social del emisor ────────────────────────────────────
    for tag in ("CuitEmisor", "DocNro", "cuit"):
        v = _find(tag)
        if v and len(v) >= 10:
            fields["cuit"] = re.sub(r"[^\d]", "", v)[:11]
            break
    for tag in ("RazonSocialEmisor", "RazonSocial", "Emisor", "nombre"):
        v = _find(tag)
        if v:
            fields["proveedor"] = v
            break

    # ── Número de comprobante ─────────────────────────────────────────────
    pto  = re.sub(r"[^\d]", "", _find("PtoVta") or _find("PuntoVenta") or "")
    nro  = re.sub(r"[^\d]", "", _find("CbteDesde") or _find("Numero") or _find("CbteNro") or "")
    if pto and nro:
        fields["numero"] = f"{pto.zfill(5)}-{nro.zfill(8)}"
    elif nro:
        fields["numero"] = nro

    # ── Tipo de comprobante ───────────────────────────────────────────────
    tipo = _find("CbteTipo") or _find("Tipo") or ""
    tipo_map = {"1": "FA-A", "6": "FA-B", "11": "FA-C", "51": "FA-M",
                "2": "ND-A", "7": "ND-B", "3": "NC-A", "8": "NC-B"}
    if tipo in tipo_map:
        fields["tipo_doc"] = tipo_map[tipo]

    # ── Fecha ─────────────────────────────────────────────────────────────
    fecha_raw = _find("CbteFch") or _find("FchEmision") or _find("Fecha") or ""
    if fecha_raw:
        fecha_raw = fecha_raw.strip()
        if re.match(r"\d{8}$", fecha_raw):          # YYYYMMDD
            fields["fecha_iso"] = f"{fecha_raw[:4]}-{fecha_raw[4:6]}-{fecha_raw[6:8]}"
            fields["fecha"]     = f"{fecha_raw[6:8]}/{fecha_raw[4:6]}/{fecha_raw[:4]}"
        elif re.match(r"\d{4}-\d{2}-\d{2}", fecha_raw):  # YYYY-MM-DD
            fields["fecha_iso"] = fecha_raw[:10]
            fields["fecha"]     = f"{fecha_raw[8:10]}/{fecha_raw[5:7]}/{fecha_raw[:4]}"
        else:
            fields["fecha"]     = fecha_raw
            fields["fecha_iso"] = parse_ar_date(fecha_raw)

    # ── Importes ──────────────────────────────────────────────────────────
    def _amt(tag):
        v = _find(tag)
        try: return float(v) if v else 0.0
        except: return 0.0

    fields["importe_total"] = _amt("ImpTotal") or _amt("ImporteTotal") or _amt("Total")
    fields["importe_neto"]  = _amt("ImpNeto")  or _amt("ImporteNeto")  or _amt("Neto")
    iva_total = _amt("ImpIVA") or _amt("ImporteIVA") or _amt("IVA")
    # Distribuir IVA: si hay IVA 10.5 explícito usarlo, resto a 21%
    iva105 = _amt("ImpIVA105") or _amt("Alicuota105") or 0.0
    iva21  = _amt("ImpIVA21")  or _amt("Alicuota21")  or (iva_total - iva105)
    fields["iva_21"]  = iva21
    fields["iva_105"] = iva105

    return fields


def extract_afip_qr_from_pdf_text(text: str) -> dict:
    """Extrae y decodifica el QR de AFIP embebido como URL en el texto del PDF.

    El QR de AFIP contiene una URL:
    https://www.afip.gob.ar/fe/qr/?p=<base64_json>
    El JSON tiene: ver, fecha, cuit, ptoVta, tipoCmp, nroCmp, importe, etc.
    """
    import base64, json as _json
    m = re.search(r'afip\.gob\.ar/fe/qr/\?p=([A-Za-z0-9+/=_-]+)', text)
    if not m:
        return {}
    try:
        payload = m.group(1)
        # padding
        payload += "=" * (4 - len(payload) % 4)
        data = _json.loads(base64.urlsafe_b64decode(payload).decode())
        if not isinstance(data, dict):
            return {}
        fecha = str(data.get("fecha", ""))[:10]
        cuit  = str(data.get("cuit", ""))
        pto   = str(data.get("ptoVta", "")).zfill(5)
        nro   = str(data.get("nroCmp", "")).zfill(8)
        tipo  = data.get("tipoCmp", 0)
        tipo_map = {1: "FA-A", 6: "FA-B", 11: "FA-C", 51: "FA-M"}
        return {
            "cuit":          cuit,
            "numero":        f"{pto}-{nro}" if pto and nro else "",
            "fecha":         f"{fecha[8:10]}/{fecha[5:7]}/{fecha[:4]}" if len(fecha) == 10 else "",
            "fecha_iso":     fecha,
            "importe_total": float(data.get("importe", 0)),
            "tipo_doc":      tipo_map.get(tipo, ""),
            "_from_qr":      True,
        }
    except Exception:
        return {}

def classify_document(text, carpeta_id=""):
    """
    Clasifica un documento de importación.
    Prioridad: sin-texto → no-aplica → CUIT → keyword.
    Retorna dict con: tipo, label, partner_id, journal_id, doc_type,
                      no_aplica (bool), mismatch (bool), extracted (dict).
    """
    _other = {"tipo":"other","label":"Otro comprobante","partner_id":None,
              "journal_id":10,"doc_type":None,
              "no_aplica":False,"mismatch":False,"extracted":{}}

    # ── Sin texto ─────────────────────────────────────────────────────────
    if not text.strip():
        return {**_other, "label":"Sin texto — no aplica", "no_aplica":True}

    tu = text.upper()
    extracted = {}

    # ── Extracción de campos comunes ──────────────────────────────────────
    # Fecha (DD/MM/YYYY, DD-MM-YYYY, DD.MM.YYYY)
    _fm = re.search(r'(\d{2}[/\-\.]\d{2}[/\-\.]\d{4})', text)
    if _fm:
        extracted["fecha"] = parse_ar_date(_fm.group(1))

    # CUIT (con o sin guiones): captura el primero que aparece
    _cm = re.search(r'(\d{2})[-\s]?(\d{8})[-\s]?(\d)\b', text)
    if _cm:
        extracted["cuit"]      = f"{_cm.group(1)}-{_cm.group(2)}-{_cm.group(3)}"
        extracted["cuit_norm"] = _cm.group(1) + _cm.group(2) + _cm.group(3)

    # Monto total
    _am = re.search(r'TOTAL[^\d$]*([\d\.]+,[\d]{2})', tu)
    if _am:
        extracted["monto"] = normalize_amount(_am.group(1))

    # TC desde texto: "Tipo de cambio USD 1 = ARG 1.505,18000" o similar
    _tc_m = re.search(
        r'(?:TIPO\s+DE\s+CAMBIO|T\.?\s*C\.?)\s*.*?(?:USD\s*1\s*=\s*(?:ARG\s*)?)?([\d\.]+,[\d]+)',
        tu, re.DOTALL)
    if _tc_m:
        extracted["tc_pdf"] = normalize_amount(_tc_m.group(1))

    # N° comprobante — patrón argentino: letra + 4 dígitos + guion + 8 dígitos
    _nr = re.search(r'\b([A-Z]\d{4}-\d{8})\b', text)
    if _nr:
        extracted["nro_comp"] = _nr.group(1)
    else:
        _nr2 = re.search(r'N[°º]?\s*(?:COMP\.?|FACTURA|COMPROBANTE)?[:\s]+([A-Z0-9\-]{5,20})', tu)
        if _nr2:
            _nr2_val = _nr2.group(1).strip()
            # Descartar si no tiene ningún dígito (serían palabras como "MONTO", "TOTAL"…)
            if re.search(r'\d', _nr2_val):
                extracted["nro_comp"] = _nr2_val

    # ── No aplica ─────────────────────────────────────────────────────────
    if "VOLANTE ELECTRONICO DE PAGO" in tu or (
            re.search(r'\bVEP\b', tu) and ("PAGO" in tu or "AFIP" in tu)):
        return {**_other, "label":"VEP — no aplica", "no_aplica":True, "extracted":extracted}

    if re.search(r'\bPRESUPUESTO\b', tu) and not re.search(r'\bFACTURA\b', tu):
        return {**_other, "label":"Presupuesto — no aplica", "no_aplica":True, "extracted":extracted}

    if "BILL OF LADING" in tu or ("CONOCIMIENTO" in tu and "EMBARQUE" in tu):
        return {**_other, "label":"Bill of Lading — no aplica", "no_aplica":True, "extracted":extracted}

    # ── Mismatch de carpeta ────────────────────────────────────────────────
    mismatch = False
    if carpeta_id:
        _carp_norm = re.sub(r'[_\s]', '_', carpeta_id.strip().upper())
        _refs = re.findall(r'LUMI[_\s]?\d+[A-Z]?', tu)
        for _r in _refs:
            _r_norm = re.sub(r'[_\s]', '_', _r.strip())
            if _r_norm != _carp_norm:
                mismatch = True
                extracted["mismatch_ref"] = _r_norm
                break

    # ── Líneas de producto PETDUR + N° factura uruguaya ──────────────────────
    _cuit_no_sep = tu.replace("-", "").replace(" ", "")
    if "PETDUR" in tu or "217016440010" in _cuit_no_sep:
        _p_lns = parse_petdur_invoice_lines(text)
        if _p_lns:
            extracted["lineas_petdur"] = _p_lns
        # Número de factura uruguaya: "Factura A 873", "e-Ticket A 873"
        if not extracted.get("nro_comp"):
            # Patrón directo contextual: 'FACTURA A 873' o 'E-TICKET A 873'
            _uy_direct = re.search(r'(?:FACTURA|E-TICKET)\s+([A-Z])\s+(\d+)', tu)
            if _uy_direct:
                extracted["nro_comp"] = f"{_uy_direct.group(1)}{int(_uy_direct.group(2)):04d}"
            else:
                # Fallback: letra sola al inicio de palabra + espacio + dígitos
                _uy2 = re.search(r'(?<![A-Z])([A-Z]) (\d{3,8})(?!\d)', tu)
                if _uy2:
                    extracted["nro_comp"] = f"{_uy2.group(1)}{int(_uy2.group(2)):04d}"

    # ── Clasificación por CUIT ─────────────────────────────────────────────
    cuit_norm = extracted.get("cuit_norm", "")
    # Buscar en todo el texto (por si el PDF tiene CUITs sin guiones)
    _tu_no_sep = tu.replace("-","").replace(" ","")
    for _ck, _cfg in CUIT_TO_PARTNER.items():
        if cuit_norm == _ck or _ck in _tu_no_sep:
            return {**_cfg, "no_aplica":False, "mismatch":mismatch, "extracted":extracted}

    # ── Fallback por keyword ───────────────────────────────────────────────
    _kw = [
        ("PETDUR",                {"tipo":"petdur", "label":"Bill PETDUR (Etapa 1)",  "partner_id":49328,"journal_id":71,"doc_type":None}),
        ("217016440010",          {"tipo":"petdur", "label":"Bill PETDUR (Etapa 1)",  "partner_id":49328,"journal_id":71,"doc_type":None}),
        ("26001IC",               {"tipo":"di_afip","label":"DI AFIP (Etapa 2)",      "partner_id":9,    "journal_id":10,"doc_type":66}),
        ("DECLARACION DE IMPORT", {"tipo":"di_afip","label":"DI AFIP (Etapa 2)",      "partner_id":9,    "journal_id":10,"doc_type":66}),
        ("SUBREGIMEN",            {"tipo":"di_afip","label":"DI AFIP (Etapa 2)",      "partner_id":9,    "journal_id":10,"doc_type":66}),
        ("SENASA",                {"tipo":"nac",    "label":"Bill SENASA (Etapa 2a)", "partner_id":48827,"journal_id":10,"doc_type":None}),
        ("TRICE",                 {"tipo":"nac",    "label":"Bill TRICE Transport (Etapa 2a)", "partner_id":48825,"journal_id":10,"doc_type":None}),
        ("TERMINAL 4",            {"tipo":"nac",    "label":"Bill Terminal 4 SA (Etapa 2a)",   "partner_id":48828,"journal_id":10,"doc_type":None}),
        ("MUNDO COMEX",           {"tipo":"nac",    "label":"Bill Mundo Comex (Etapa 2a)",     "partner_id":48826,"journal_id":10,"doc_type":None}),
    ]
    for _kword, _cfg in _kw:
        if _kword in tu:
            return {**_cfg, "no_aplica":False, "mismatch":mismatch, "extracted":extracted}

    return {**_other, "mismatch":mismatch, "extracted":extracted}

